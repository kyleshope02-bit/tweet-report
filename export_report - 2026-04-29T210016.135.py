import sqlite3, re, html, time, json, os, subprocess
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import requests

DB_PATH   = "tweets.db"
OUT_FILE  = "index.html"

EXCEL_FILE  = "stocks.xlsx"
# Hardcoded tickers — no Excel needed
# Major index ETFs — added directly to signal scanner (bypass market cap filter)
INDEX_TICKERS = ["SPY", "QQQ", "IWM", "DIA"]
INDEX_NAMES   = {"SPY":"S&P 500 ETF","QQQ":"Nasdaq 100 ETF","IWM":"Russell 2000 ETF","DIA":"Dow Jones ETF"}

MAG7_TICKERS = [
    ("AAPL",  "Apple Inc."),
    ("MSFT",  "Microsoft Corp."),
    ("NVDA",  "NVIDIA Corp."),
    ("AMZN",  "Amazon.com Inc."),
    ("GOOGL", "Alphabet Inc."),
    ("META",  "Meta Platforms Inc."),
    ("TSLA",  "Tesla Inc."),
]

BASE_X = "https://x.com"

# Only tweets from these handles will appear on the report
TWEET_HANDLES = {"@mr_derivatives", "@eliteoptions2"}

# Watchlist market cap minimum (USD)
WATCHLIST_MKTCAP_MIN = 10_000_000_000   # 10B

# ── Master universe — yfinance filters to ≥$10B at runtime ──────────────────
LARGE_CAP_UNIVERSE = [
    # Mag 7 (also in Mag7 tab)
    "AAPL","MSFT","NVDA","AMZN","GOOGL","META","TSLA",
    # Semiconductors
    "AVGO","TSM","ASML","QCOM","AMD","MU","AMAT","LRCX","KLAC","TXN","ADI","MRVL","INTC","ARM","NXPI",
    # Megacap Tech
    "ORCL","SAP","CRM","NOW","ADBE","IBM","INTU","CSCO","ACN","SNOW","UBER","SHOP","PLTR",
    # Financials
    "BRK-B","JPM","V","MA","BAC","WFC","GS","MS","AXP","BLK","SCHW","C","USB","PNC","COF","SPGI","MCO",
    # Healthcare / Pharma
    "LLY","JNJ","UNH","ABBV","MRK","TMO","ABT","DHR","AMGN","PFE","BMY","GILD","REGN","VRTX","SYK","ISRG","ELV","CI",
    # Consumer
    "WMT","COST","HD","MCD","SBUX","NKE","TGT","LOW","TJX","BABA","PDD","JD",
    # Energy
    "XOM","CVX","COP","SLB","EOG","PSX","VLO","MPC",
    # Industrials
    "CAT","GE","HON","UNP","RTX","LMT","DE","BA","UPS","FDX","ETN","EMR","PH","MMM",
    # Consumer Staples
    "PG","KO","PEP","PM","MO","MDLZ","CL","KMB",
    # Communication
    "NFLX","DIS","CMCSA","T","VZ","TMUS","CHTR","SPOT",
    # Real Estate / Utilities
    "NEE","DUK","SO","AEP","SRE",
    # International ADRs
    "TCEHY","TSM","ASML","NVO","AZN","SHEL","TTE","SNY","BABA","TCOM",
    # ETFs (market proxies)
    "SPY","QQQ","IWM","TLT",
    # High-growth / momentum
    "CRWD","NET","DDOG","ZS","PANW","FTNT","MELI","SE","GRAB","APP",
    # Banks / Fintech
    "PYPL","XYZ","COIN","HOOD","AFRM","NU",
    # Misc large-cap
    "BX","KKR","APO","ARES","BAM","MET","PRU","AFL","TRV","CB","MMC",
]
# Deduplicate while preserving order
_seen = set()
LARGE_CAP_UNIVERSE = [t for t in LARGE_CAP_UNIVERSE if not (t in _seen or _seen.add(t))]

# yfinance cache
YF_CACHE_FILE      = ".cache_yfinance.json"
YF_CACHE_TTL       = 6 * 3600   # 6 hours
YF_SIGNALS_CACHE   = ".cache_signals.json"
YF_SIGNALS_TTL     = 6 * 3600   # 6 hours
# Anthropic API key — get yours at https://console.anthropic.com/
import os as _os
ANTHROPIC_API_KEY  = _os.environ.get("ANTHROPIC_API_KEY", "")

# ── Open cash-secured puts (CSPs) on Mag 7 stocks ────────────────────────────
# Add new trades here when you sell puts:
OPEN_PUTS = [
    # Example — uncomment and edit:
    # {"date":"2026-03-20","ticker":"AMZN","strike":190.0,
    #  "expiry":"Apr 17, 2026","exp_yf":"2026-04-17",
    #  "contracts":1,"premium":2.75,"total_premium":275.0,
    #  "notes":"75% support CSP — pre-earnings"},
]

# ── Kyle's holdings ────────────────────────────────────────────────────────────
KYLE_HOLDINGS = [
    {
        "ticker":    "MSFT",
        "shares":    400,
        "avg_cost":  424.46,
        "col":       "#00a4ef",
        "calls": [{
            "contracts":    4,
            "strike":       440.0,
            "expiry":       "2026-05-15",
            "expiry_label": "May 15 '26",
            "premium_collected": 14.00,   # per share when sold
            "current_price":    11.08,    # current mid per share
        }],
    },
    {
        "ticker":    "PM",
        "shares":    400,
        "avg_cost":  162.71,
        "col":       "#FF9900",
        "calls": [{
            "contracts":    4,
            "strike":       175.0,
            "expiry":       "2026-06-18",
            "expiry_label": "Jun 18 '26",
            "premium_collected": 3.60,
            "current_price":    3.60,
        }],
    },
]

YF_OPTIONS_CACHE   = ".cache_options.json"
YF_OPTIONS_TTL     = 2 * 3600   # 2 hours
YF_HIST_CACHE      = ".cache_hist_signals.json"
YF_HIST_CACHE_TTL  = 6 * 3600   # 6 hours — refresh ~4x per day
WHEEL_TRACKER_FILE = "wheel_positions.json"

# ── CSS + JS constants ────────────────────────────────────────────────────────
STYLE = """<style>
body{font-family:Arial;background:#0a0f1e;margin:12px;font-size:14px;color:#e8edf5}
.card{background:#13263d;border-radius:10px;padding:14px;margin-bottom:12px;box-shadow:0 2px 8px rgba(0,0,0,.4)}
.card h3{margin:0 0 10px;font-size:16px;color:#fff}
.muted{color:#5a7fa0;font-size:12px;margin-bottom:10px}
.fresh{color:#2e7d32;font-size:12px;font-weight:400}
.tabs{display:flex;flex-wrap:wrap;gap:6px;margin-bottom:14px}
.tabbtn{padding:7px 18px;border:none;border-radius:8px;cursor:pointer;font-size:13px;font-weight:700;transition:background .2s}
.tabbtn.on{background:#1565c0;color:#fff}
.tabbtn.off{background:#1a2a3a;color:#8ab4d4}
.tabbtn:hover{background:#1976d2;color:#fff}
.hidden{display:none}
.pos{color:#00e676;font-weight:700}
.neg{color:#ef5350;font-weight:700}
.sig-bull{background:#1b5e20;color:#00e676;font-weight:700;padding:3px 8px;border-radius:8px;font-size:12px;display:inline-block}
.sig-bear{background:#b71c1c;color:#ffcdd2;font-weight:700;padding:3px 8px;border-radius:8px;font-size:12px;display:inline-block}
table{border-collapse:collapse;width:100%}
th{background:#0d2040;color:#7eb8f7;font-size:12px;padding:8px 10px;text-align:right;white-space:nowrap;border-bottom:2px solid #1e3a5f}
td{padding:7px 10px;border-bottom:1px solid #0d2040;font-size:12px;text-align:right;color:#e8edf5}
.filter{background:#0d1b2a;color:#e8edf5;border:1px solid #1e3a5f;border-radius:5px;padding:4px 8px;width:100%;font-size:11px;box-sizing:border-box}
</style>"""

JS = """<script>
function showTab(which){
  document.querySelectorAll("[data-tab='1']").forEach(function(el){
    el.classList.add("hidden");
  });
  document.querySelectorAll("[data-tabbtn='1']").forEach(function(el){
    el.classList.remove("on");
    el.classList.add("off");
  });
  var tab = document.getElementById("tab-"+which);
  if(tab){ tab.classList.remove("hidden"); }
  var btn = document.getElementById("btn-"+which);
  if(btn){ btn.classList.remove("off"); btn.classList.add("on"); }
}
function applyFilters(tableId){
  var table = document.getElementById(tableId);
  if(!table) return;
  var inputs = table.querySelectorAll(".filter");
  var rows = table.querySelectorAll("tbody tr");
  rows.forEach(function(row){
    var show = true;
    inputs.forEach(function(inp, i){
      var val = inp.value.toLowerCase();
      if(!val) return;
      var cell = row.cells[i];
      if(cell && cell.textContent.toLowerCase().indexOf(val) === -1){ show = false; }
    });
    row.style.display = show ? "" : "none";
  });
}
</script>"""


# ------------------ tweet parsing ------------------
TICKER_RE  = re.compile(r"(?<![A-Z0-9])\$(?:[A-Z]{1,6})(?![A-Z0-9])", re.I)
BUY_WORDS  = re.compile(r"\b(buy|long|bullish|accumulate|adding|breakout|highs?)\b", re.I)
SELL_WORDS = re.compile(r"\b(sell|short|bearish|trim|exit|dump|breakdown|lows?)\b", re.I)
NEGATE_RE  = re.compile(r"\b(not|no|never|don't|dont|isn't|isnt|wasn't|wasnt|no longer)\b", re.I)

def extract_tickers(text: str):
    return sorted(set([t.upper() for t in TICKER_RE.findall(text or "")]))

def classify_signal(text: str):
    if not text: return "NEUTRAL"
    words = re.split(r"\s+", text)
    buy_count = sell_count = 0
    for i, word in enumerate(words):
        window  = " ".join(words[max(0, i-4):i])
        negated = bool(NEGATE_RE.search(window))
        if BUY_WORDS.match(word):
            if not negated: buy_count += 1
        elif SELL_WORDS.match(word):
            if not negated: sell_count += 1
    if buy_count > 0 and sell_count == 0: return "BUY"
    if sell_count > 0 and buy_count == 0: return "SELL"
    if buy_count > 0 and sell_count > 0:  return "MIXED"
    return "NEUTRAL"

def esc(x):
    return html.escape("" if x is None else str(x))

def db():
    return sqlite3.connect(DB_PATH)

def table_cols(cur):
    return [r[1] for r in cur.execute("PRAGMA table_info(tweets)").fetchall()]

def parse_iso(ts: str):
    if not ts: return None
    s = ts.strip().replace("Z", "")
    try:
        return datetime.fromisoformat(s)
    except Exception:
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try: return datetime.strptime(s, fmt)
            except Exception: pass
    return None

def fmt_local_readable(dt):
    if dt is None: return ""
    if dt.tzinfo is None: return dt.strftime("%b %d, %Y %I:%M %p")
    return dt.astimezone().strftime("%b %d, %Y %I:%M %p")

def normalize_ticker(x) -> str:
    t = ("" if x is None else str(x)).strip().upper()
    if not t: return ""
    if not t.startswith("$"): t = "$" + t
    return t

def normalize_handle(x) -> str:
    s = ("" if x is None else str(x)).strip()
    if not s: return ""
    if not s.startswith("@"):
        m = re.search(r"@\w{1,32}", s)
        s = m.group(0) if m else s
    if not s.startswith("@"): s = "@" + s
    return s

# ------------------ Excel reader ------------------
def read_workbook_sheets_xlsx(path: str) -> Dict[str, List[dict]]:
    try:
        import openpyxl
    except Exception as e:
        raise RuntimeError(f"Missing dependency: openpyxl.\nInstall with: pip install openpyxl\n{e}")
    p = Path(path)
    if not p.exists():
        raise RuntimeError(f"Missing Excel file: {path}")
    wb = openpyxl.load_workbook(p, data_only=True)
    out: Dict[str, List[dict]] = {}
    for name in wb.sheetnames:
        sh   = wb[name]
        rows = list(sh.iter_rows(values_only=True))
        if not rows:
            out[name] = []
            continue
        headers = [str(c).strip() if c is not None else "" for c in rows[0]]
        data = []
        for r in rows[1:]:
            if r is None: continue
            if all((c is None or str(c).strip() == "") for c in r): continue
            d = {h: (r[i] if i < len(r) else None) for i, h in enumerate(headers) if h}
            data.append(d)
        out[name] = data
    return out

# ------------------ PNL ------------------
def as_float(x) -> Optional[float]:
    if x is None: return None
    s = str(x).strip().replace(",", "").replace("$", "")
    if s.endswith("%"): s = s[:-1].strip()
    try: return float(s)
    except Exception: return None

def fmt_money(v: Optional[float]) -> str:
    if v is None: return "$0"
    sign = "-" if v < 0 else ""
    v = abs(v)
    if v >= 1e9: return f"{sign}${v/1e9:.2f}B"
    if v >= 1e6: return f"{sign}${v/1e6:.2f}M"
    if v >= 1e3: return f"{sign}${v/1e3:.2f}K"
    return f"{sign}${v:.0f}"

def sum_pnl(sheet_rows: List[dict]) -> float:
    return sum(v for r in sheet_rows if (v := as_float(r.get("PNL"))) is not None)

# ------------------ yfinance price fetch ------------------
def fetch_yfinance_batch(tickers: List[str]) -> dict:
    """Batched yfinance price + market cap fetch with disk cache.
    On any Yahoo Finance error, falls back to the most recent cache regardless of age.
    """
    try:
        import yfinance as yf
    except ImportError:
        print("⚠️  yfinance not installed. Run: pip install yfinance")
        return {}

    clean = [t.lstrip("$") for t in tickers if t]

    def _load_cache():
        """Load cache file regardless of age — used as fallback."""
        try:
            if os.path.exists(YF_CACHE_FILE):
                cached = json.loads(Path(YF_CACHE_FILE).read_text(encoding="utf-8"))
                if cached: return cached
        except Exception:
            pass
        return None

    # Use fresh cache if within TTL
    if os.path.exists(YF_CACHE_FILE):
        age = time.time() - os.path.getmtime(YF_CACHE_FILE)
        if age < YF_CACHE_TTL:
            try:
                cached = json.loads(Path(YF_CACHE_FILE).read_text(encoding="utf-8"))
                if all(t in cached for t in clean):
                    print(f"   yfinance: using cache ({int(age/60)}m old, TTL {int(YF_CACHE_TTL/3600)}h)")
                    return cached
            except Exception:
                pass

    print(f"   yfinance: fetching {len(clean)} tickers...")
    try:
        import yfinance as yf
        df = yf.download(clean, period="6mo", group_by="ticker",
                         auto_adjust=True, progress=False, threads=True)
    except Exception as e:
        print(f"⚠️  yfinance download failed: {e}")
        stale = _load_cache()
        if stale:
            age_s = time.time() - os.path.getmtime(YF_CACHE_FILE)
            print(f"   yfinance: using stale cache as fallback ({int(age_s/3600)}h old)")
            return stale
        return {}

    results = {}
    for t in clean:
        try:
            closes = df["Close"].dropna()   if len(clean) == 1 else df[t]["Close"].dropna()
            vols   = df["Volume"].dropna()  if len(clean) == 1 else df[t]["Volume"].dropna()
            if closes.empty: continue

            def chg(n):
                return round((float(closes.iloc[-1]) / float(closes.iloc[-1-n]) - 1) * 100, 2) if len(closes) > n else None

            def vchg(n):
                idxs   = [len(vols)-1-n+j for j in range(-1, 2) if 0 <= len(vols)-1-n+j < len(vols)]
                window = [float(vols.iloc[i]) for i in idxs if float(vols.iloc[i]) > 0]
                avg    = sum(window) / len(window) if window else None
                cur    = float(vols.iloc[-1])
                return round((cur / avg - 1) * 100, 2) if avg and avg > 0 else None

            try:
                info       = yf.Ticker(t).info
                market_cap = info.get("marketCap") or info.get("market_cap")
            except Exception:
                market_cap = None

            results[t] = {
                "Price":          round(float(closes.iloc[-1]), 2),
                "Chg 1D %":      chg(1),  "Chg 5D %":      chg(5),
                "Chg 20D %":     chg(20), "Chg 30D %":     chg(30), "Chg 60D %": chg(60),
                "Volume":        int(vols.iloc[-1]),
                "Vol Chg 1D %":  vchg(1), "Vol Chg 5D %":  vchg(5),
                "Vol Chg 20D %": vchg(20),"Vol Chg 30D %": vchg(30),"Vol Chg 60D %": vchg(60),
                "market_cap":    market_cap,
            }
        except Exception as e:
            print(f"   ⚠️  yfinance: skipping {t} — {e}")

    if results:
        try: Path(YF_CACHE_FILE).write_text(json.dumps(results), encoding="utf-8")
        except Exception: pass
    elif not results:
        # Fetch returned nothing — fall back to any existing cache
        stale = _load_cache()
        if stale:
            age_s = time.time() - os.path.getmtime(YF_CACHE_FILE)
            print(f"   yfinance: empty results, using stale cache ({int(age_s/3600)}h old)")
            return stale

    print(f"   yfinance: got data for {len(results)}/{len(clean)} tickers")
    return results


# ------------------ Signal detection engine ------------------
def fetch_ohlcv_for_signals(tickers: List[str]) -> dict:
    """
    Fetch full OHLCV history needed for signal detection.
    Returns dict keyed by bare ticker with pandas Series for O/H/L/C/V.
    On any Yahoo Finance error, falls back to the most recent cache regardless of age.
    """
    try:
        import yfinance as yf
    except ImportError:
        return {}

    clean = [t.lstrip("$") for t in tickers if t]

    def _load_signals_cache():
        """Load signals cache regardless of age — fallback on Yahoo error."""
        try:
            if os.path.exists(YF_SIGNALS_CACHE):
                cached = json.loads(Path(YF_SIGNALS_CACHE).read_text(encoding="utf-8"))
                if cached:
                    result = {}
                    for t, d in cached.items():
                        result[t] = {k: d[k] for k in ("open","high","low","close","volume","dates") if k in d}
                    return result
        except Exception:
            pass
        return None

    # Use fresh cache if within TTL
    if os.path.exists(YF_SIGNALS_CACHE):
        age = time.time() - os.path.getmtime(YF_SIGNALS_CACHE)
        if age < YF_SIGNALS_TTL:
            try:
                cached = json.loads(Path(YF_SIGNALS_CACHE).read_text(encoding="utf-8"))
                if (all(t in cached for t in clean) and
                        all("dates" in cached.get(t, {}) for t in clean)):
                    print(f"   signals OHLCV: using cache ({int(age/60)}m old)")
                    result = {}
                    for t, d in cached.items():
                        result[t] = {k: d[k] for k in ("open","high","low","close","volume","dates") if k in d}
                    return result
                else:
                    print("   signals OHLCV: cache missing dates — refetching...")
            except Exception:
                pass

    print(f"   signals OHLCV: fetching {len(clean)} tickers (1y daily)...")
    try:
        df = yf.download(clean, period="1y", group_by="ticker",
                         auto_adjust=True, progress=False, threads=True)
    except Exception as e:
        print(f"⚠️  signals OHLCV fetch failed: {e}")
        stale = _load_signals_cache()
        if stale:
            age_s = time.time() - os.path.getmtime(YF_SIGNALS_CACHE)
            print(f"   signals OHLCV: using stale cache as fallback ({int(age_s/3600)}h old)")
            return stale
        return {}

    result  = {}
    to_save = {}
    for t in clean:
        try:
            sub = df if len(clean) == 1 else df[t]
            o = sub["Open"].dropna().tolist()
            h = sub["High"].dropna().tolist()
            l = sub["Low"].dropna().tolist()
            c = sub["Close"].dropna().tolist()
            v = sub["Volume"].dropna().tolist()
            if len(c) < 20: continue
            dates = [str(d)[:10] for d in sub["Close"].dropna().index.tolist()]
            result[t]  = {"open": o, "high": h, "low": l, "close": c, "volume": v, "dates": dates}
            to_save[t] = {"open": o, "high": h, "low": l, "close": c, "volume": v, "dates": dates}
        except Exception:
            continue

    if result:
        try: Path(YF_SIGNALS_CACHE).write_text(json.dumps(to_save), encoding="utf-8")
        except Exception: pass
    else:
        # No results at all — fall back to stale cache
        stale = _load_signals_cache()
        if stale:
            age_s = time.time() - os.path.getmtime(YF_SIGNALS_CACHE)
            print(f"   signals OHLCV: empty results, using stale cache ({int(age_s/3600)}h old)")
            return stale

    return result


def ema(prices: list, period: int) -> list:
    """Exponential moving average."""
    if len(prices) < period: return []
    k   = 2 / (period + 1)
    val = sum(prices[:period]) / period
    out = [None] * (period - 1) + [val]
    for p in prices[period:]:
        val = p * k + val * (1 - k)
        out.append(val)
    return out

def sma(prices: list, period: int) -> list:
    """Simple moving average."""
    out = [None] * (period - 1)
    for i in range(period - 1, len(prices)):
        out.append(sum(prices[i-period+1:i+1]) / period)
    return out

def slope_pct(ma_series: list, lookback: int = 10) -> Optional[float]:
    """Percent change of an MA over lookback bars — proxy for slope steepness."""
    valid = [x for x in ma_series if x is not None]
    if len(valid) < lookback + 1: return None
    old = valid[-lookback-1]
    new = valid[-1]
    if old == 0: return None
    return (new - old) / old * 100


def calc_atr(highs: list, lows: list, closes: list, period: int = 14) -> float:
    """Average True Range over last `period` bars."""
    trs = []
    for i in range(1, len(closes)):
        tr = max(highs[i] - lows[i],
                 abs(highs[i] - closes[i-1]),
                 abs(lows[i]  - closes[i-1]))
        trs.append(tr)
    if not trs: return 0.0
    recent = trs[-period:]
    return sum(recent) / len(recent)


def calc_levels(closes: list, highs: list, lows: list,
                bullish: bool, signal: str) -> dict:
    """
    Calculate entry, stop, and two profit targets based on ATR and pattern rules.

    Bullish signals:
      Entry  = current close (or slight pullback to nearest MA)
      Stop   = entry - 1.5x ATR  (below recent swing low if tighter)
      PT1    = entry + 1.5x ATR  (1:1 risk/reward)
      PT2    = entry + 3.0x ATR  (2:1 risk/reward)

    Bearish signals (cross/breakdown):
      Entry  = current close
      Stop   = entry + 1.5x ATR
      PT1    = entry - 1.5x ATR
      PT2    = entry - 3.0x ATR
    """
    if not closes or not highs or not lows:
        return {}

    price = closes[-1]
    atr   = calc_atr(highs, lows, closes, 14)
    if atr == 0: return {}

    # Recent swing low/high over last 10 bars for tighter stops
    swing_low  = min(lows[-10:])
    swing_high = max(highs[-10:])

    if bullish:
        entry = round(price, 2)
        stop  = round(max(entry - 1.5 * atr, swing_low - 0.01), 2)
        pt1   = round(entry + 1.5 * atr, 2)
        pt2   = round(entry + 3.0 * atr, 2)
        risk  = round(entry - stop, 2)
    else:
        entry = round(price, 2)
        stop  = round(min(entry + 1.5 * atr, swing_high + 0.01), 2)
        pt1   = round(entry - 1.5 * atr, 2)
        pt2   = round(entry - 3.0 * atr, 2)
        risk  = round(stop - entry, 2)

    rr1 = round(abs(pt1 - entry) / risk, 1) if risk else 0
    rr2 = round(abs(pt2 - entry) / risk, 1) if risk else 0

    return {
        "entry": entry,
        "stop":  stop,
        "pt1":   pt1,
        "pt2":   pt2,
        "risk":  risk,
        "atr":   round(atr, 2),
        "rr1":   rr1,
        "rr2":   rr2,
    }


def detect_signals(ticker: str, ohlcv: dict) -> List[dict]:
    """
    Undercut & Reclaim patterns for 50-day, 100-day, and 200-day SMAs.
    Each signal includes grade (A/B/C/F), momentum score, vol ratio, body%, MA dist%, undercut depth.
    """
    closes  = ohlcv["close"]
    volumes = ohlcv["volume"]
    opens   = ohlcv["open"]
    highs   = ohlcv["high"]
    lows    = ohlcv["low"]
    dates   = ohlcv.get("dates", [])

    def d(idx):
        if not dates: return ""
        try: return dates[idx]
        except IndexError: return ""

    def safe(arr, idx):
        try: return arr[idx]
        except (IndexError, TypeError): return None

    if len(closes) < 15: return []   # 5-day only needs 10+ bars; 200-day skipped if <205

    signals  = []
    _lev_bull = calc_levels(closes, highs, lows, bullish=True,  signal="")
    _lev_bear = calc_levels(closes, highs, lows, bullish=False, signal="")

    c   = closes[-1];   c1  = closes[-2]
    h   = highs[-1];    l   = lows[-1]
    o   = opens[-1]
    vol = volumes[-1]

    # Volume ratio vs 20-day avg
    avg_vol   = sum(volumes[-21:-1]) / 20 if len(volumes) >= 21 else vol
    vol_ratio = round(vol / avg_vol, 2) if avg_vol else 1.0

    # Bar quality: body as % of full range
    bar_range   = h - l
    body        = abs(c - o)
    body_pct    = round(body / bar_range * 100, 1) if bar_range > 0 else 0.0

    def _score(dist_pct_abs, vol_r, body_p):
        """0-100 score: volume 40pts, body 30pts, distance 30pts."""
        v_score = 40 if vol_r >= 2.5 else (30 if vol_r >= 2.0 else (20 if vol_r >= 1.5 else (10 if vol_r >= 1.0 else 0)))
        b_score = 30 if body_p >= 80 else (22 if body_p >= 60 else (12 if body_p >= 40 else 0))
        d_score = 30 if dist_pct_abs >= 2.0 else (20 if dist_pct_abs >= 1.0 else (10 if dist_pct_abs >= 0.3 else 0))
        return v_score + b_score + d_score

    def _grade(score):
        if score >= 70: return "A", "#00e676"
        if score >= 50: return "B", "#FFD700"
        if score >= 30: return "C", "#ffb74d"
        return "F", "#ef5350"

    for period, label in [(5,"5-Day"),(10,"10-Day"),(50,"50-Day"),(100,"100-Day"),(200,"200-Day")]:
        if len(closes) < period + 5: continue
        ma_arr  = sma(closes, period)
        ma_now  = safe(ma_arr, -1)
        ma_prev = safe(ma_arr, -2)
        if not all([ma_now, ma_prev, c, c1]): continue

        dist_pct = round((c - ma_now) / ma_now * 100, 2)
        dist_abs = abs(dist_pct)
        score    = _score(dist_abs, vol_ratio, body_pct)
        grade, grade_col = _grade(score)

        # ── Reclaim ─────────────────────────────────────────────────────────────
        if c1 < ma_prev and c > ma_now:
            # How deep below the MA was the undercut?
            undercut_depth = round((ma_prev - c1) / ma_prev * 100, 2) if ma_prev else 0

            # Scan for recent examples (look back up to 60 days for similar reclaims)
            examples = []
            for i in range(3, min(61, len(closes)-1)):
                _c  = closes[-i];    _c1 = closes[-i-1]
                _ma = safe(ma_arr, -i); _ma1 = safe(ma_arr, -i-1)
                if _ma and _ma1 and _c1 < _ma1 and _c > _ma:
                    examples.append(f"{d(-i)} (${_c:.2f})")
                if len(examples) >= 2: break

            ex_str = " | Recent examples: " + ", ".join(examples) if examples else ""
            conviction = ("High-conviction institutional reclaim — strong follow-through likely." if score >= 70
                         else "Watch for volume confirmation on follow-through session." if score >= 50
                         else "Weak reclaim — needs follow-through volume to confirm.")

            signals.append({
                "category":       f"Undercut & Reclaim — {label} SMA",
                "signal":         f"Reclaim {label} SMA",
                "description": (
                    f"Price reclaimed the {label} SMA (${ma_now:.2f}) — "
                    f"prev close ${c1:.2f} below, now ${c:.2f} above (+{dist_pct:.2f}% clear). "
                    f"Vol {vol_ratio:.1f}× avg | Body {body_pct:.0f}% | Undercut depth: {undercut_depth:.2f}% below MA. "
                    f"{conviction}"
                    f"{ex_str}"
                ),
                "bullish":        True,
                "signal_date":    d(-1),
                "levels":         _lev_bull,
                "momentum_score": score,
                "momentum_grade": grade,
                "momentum_color": grade_col,
                "vol_ratio":      vol_ratio,
                "body_pct":       body_pct,
                "ma_dist_pct":    dist_pct,
                "undercut_depth": undercut_depth,
            })

        # ── Undercut ─────────────────────────────────────────────────────────────
        elif c1 > ma_prev and c < ma_now:
            shakeout = vol_ratio < 1.0
            conviction = ("High-volume breakdown — distribution, avoid." if vol_ratio >= 1.5
                         else "Low-volume undercut — possible shakeout, watch for bounce back above MA.")

            signals.append({
                "category":       f"Undercut & Reclaim — {label} SMA",
                "signal":         f"Undercut {label} SMA",
                "description": (
                    f"Price undercut the {label} SMA (${ma_now:.2f}) — "
                    f"prev close ${c1:.2f} above, now ${c:.2f} below ({dist_pct:.2f}%). "
                    f"Vol {vol_ratio:.1f}× avg | Body {body_pct:.0f}%. "
                    f"{conviction} "
                    f"Watch for reclaim attempt next 1-3 sessions — if MA becomes resistance, downtrend accelerates."
                ),
                "bullish":        False,
                "signal_date":    d(-1),
                "levels":         _lev_bear,
                "momentum_score": score,
                "momentum_grade": grade,
                "momentum_color": grade_col,
                "vol_ratio":      vol_ratio,
                "body_pct":       body_pct,
                "ma_dist_pct":    dist_pct,
                "undercut_depth": None,
            })

    return signals


def detect_smc_signals(ticker: str, ohlcv: dict) -> List[dict]:
    """
    Smart Money Concepts (SMC / ICT) signal detection.
    Detects on the most recent completed daily bar.

    1. Liquidity Sweep + Reversal
       Price breaks a prior swing high/low (grabs stop-loss liquidity),
       then closes back on the other side — classic institutional trap.

    2. Demand Zone Bounce
       Price drops into a prior demand zone (defined as a consolidation
       before a strong impulsive move up) and reverses with volume.

    3. Supply Zone Rejection
       Price rallies into a prior supply zone (consolidation before
       strong impulsive move down) and reverses with volume.

    4. Break of Structure (BOS) — bullish
       Price closes above the most recent swing high — trend shift.

    5. Break of Structure (BOS) — bearish
       Price closes below the most recent swing low — trend shift.
    """
    closes  = ohlcv["close"]
    volumes = ohlcv["volume"]
    opens   = ohlcv["open"]
    highs   = ohlcv["high"]
    lows    = ohlcv["low"]
    dates   = ohlcv.get("dates", [])

    if len(closes) < 30: return []

    def d(idx):
        if not dates: return ""
        try: return dates[idx]
        except IndexError: return ""

    def safe(arr, idx):
        try: return arr[idx]
        except (IndexError, TypeError): return None

    signals = []
    avg_vol   = sum(volumes[-21:-1]) / 20 if len(volumes) >= 21 else volumes[-1]
    vol_ratio = round(volumes[-1] / avg_vol, 2) if avg_vol else 1.0

    c  = closes[-1];  c1 = closes[-2];  c2 = closes[-3]
    h  = highs[-1];   h1 = highs[-2]
    l  = lows[-1];    l1 = lows[-2]
    o  = opens[-1]

    _lev_bull = calc_levels(closes, highs, lows, bullish=True,  signal="")
    _lev_bear = calc_levels(closes, highs, lows, bullish=False, signal="")

    bar_range = h - l
    body_pct  = round(abs(c - o) / bar_range * 100, 1) if bar_range > 0 else 0

    def _score(vr, bp):
        v = 40 if vr>=2.5 else (30 if vr>=2.0 else (20 if vr>=1.5 else (10 if vr>=1.0 else 0)))
        b = 30 if bp>=80 else (20 if bp>=60 else (10 if bp>=40 else 0))
        return v + b

    def _grade(s):
        if s >= 70: return "A","#00e676"
        if s >= 50: return "B","#FFD700"
        if s >= 30: return "C","#ffb74d"
        return "F","#ef5350"

    # ── Find swing highs and lows (last 20 bars, n=3 each side) ──────────────
    n = 3
    swing_highs = []
    swing_lows  = []
    for i in range(n, len(highs) - n):
        if highs[i] >= max(highs[i-n:i] + highs[i+1:i+n+1]):
            swing_highs.append((i, highs[i], d(i - len(highs))))
        if lows[i] <= min(lows[i-n:i] + lows[i+1:i+n+1]):
            swing_lows.append((i, lows[i], d(i - len(lows))))

    # Most recent swing high/low before today
    prev_sh = swing_highs[-2] if len(swing_highs) >= 2 else None
    prev_sl = swing_lows[-2]  if len(swing_lows)  >= 2 else None

    # ── 1. Liquidity Sweep + Reversal (bullish) ───────────────────────────────
    # Today's low swept BELOW a prior swing low, but closed ABOVE it
    if prev_sl:
        _sl_price = prev_sl[1]
        _swept_low  = l < _sl_price           # wick went below swing low (grabbed stops)
        _recovered  = c > _sl_price           # but closed back above (reversal)
        _big_reversal = (c - l) / bar_range >= 0.6 if bar_range > 0 else False
        if _swept_low and _recovered and vol_ratio >= 1.2:
            sc = _score(vol_ratio, body_pct)
            gr, gc = _grade(sc)
            sweep_depth = round((_sl_price - l) / _sl_price * 100, 2)
            signals.append({
                "category":       "SMC — Liquidity Sweep",
                "signal":         "Sweep + Demand Bounce",
                "description": (
                    f"Today's low (${l:.2f}) swept below the prior swing low "
                    f"(${_sl_price:.2f}), grabbing stop-loss liquidity, then "
                    f"reversed to close at ${c:.2f} — {sweep_depth:.2f}% below swept level. "
                    f"Vol {vol_ratio:.1f}× avg | Body {body_pct:.0f}%. "
                    f"{'Strong reversal — institutions absorbed the sell orders.' if _big_reversal else 'Watch for follow-through above the sweep level.'}"
                ),
                "bullish":        True,
                "signal_date":    d(-1),
                "levels":         _lev_bull,
                "momentum_score": sc,
                "momentum_grade": gr,
                "momentum_color": gc,
                "vol_ratio":      vol_ratio,
                "body_pct":       body_pct,
                "ma_dist_pct":    round((c - _sl_price) / _sl_price * 100, 2),
                "undercut_depth": sweep_depth,
            })

    # ── 2. Liquidity Sweep + Reversal (bearish) ───────────────────────────────
    # Today's high swept ABOVE a prior swing high, but closed BELOW it
    if prev_sh:
        _sh_price = prev_sh[1]
        _swept_hi   = h > _sh_price
        _failed     = c < _sh_price
        _big_reject = (h - c) / bar_range >= 0.6 if bar_range > 0 else False
        if _swept_hi and _failed and vol_ratio >= 1.2:
            sc = _score(vol_ratio, body_pct)
            gr, gc = _grade(sc)
            sweep_depth = round((h - _sh_price) / _sh_price * 100, 2)
            signals.append({
                "category":       "SMC — Liquidity Sweep",
                "signal":         "Sweep + Supply Rejection",
                "description": (
                    f"Today's high (${h:.2f}) swept above the prior swing high "
                    f"(${_sh_price:.2f}), grabbing buy-stop liquidity, then "
                    f"rejected to close at ${c:.2f} — {sweep_depth:.2f}% above sweep level. "
                    f"Vol {vol_ratio:.1f}× avg | Body {body_pct:.0f}%. "
                    f"{'Strong rejection — institutions distributed into the rip.' if _big_reject else 'Watch for continued selling below the sweep level.'}"
                ),
                "bullish":        False,
                "signal_date":    d(-1),
                "levels":         _lev_bear,
                "momentum_score": sc,
                "momentum_grade": gr,
                "momentum_color": gc,
                "vol_ratio":      vol_ratio,
                "body_pct":       body_pct,
                "ma_dist_pct":    round((c - _sh_price) / _sh_price * 100, 2),
                "undercut_depth": sweep_depth,
            })

    # ── 3. Break of Structure — Bullish (BOS) ────────────────────────────────
    # Today closes ABOVE the most recent prior swing high (structure shift)
    if prev_sh and c > prev_sh[1] and c1 <= prev_sh[1] and vol_ratio >= 1.1:
        sc = _score(vol_ratio, body_pct)
        gr, gc = _grade(sc)
        signals.append({
            "category":       "SMC — Break of Structure",
            "signal":         "Bullish BOS",
            "description": (
                f"Price broke above the prior swing high (${prev_sh[1]:.2f}) "
                f"closing at ${c:.2f} — structural shift to bullish bias. "
                f"Prior resistance becomes support. "
                f"Vol {vol_ratio:.1f}× avg confirms institutional participation. "
                f"Watch for a retest of ${prev_sh[1]:.2f} as new support before next leg up."
            ),
            "bullish":        True,
            "signal_date":    d(-1),
            "levels":         _lev_bull,
            "momentum_score": sc,
            "momentum_grade": gr,
            "momentum_color": gc,
            "vol_ratio":      vol_ratio,
            "body_pct":       body_pct,
            "ma_dist_pct":    round((c - prev_sh[1]) / prev_sh[1] * 100, 2),
            "undercut_depth": None,
        })

    # ── 4. Break of Structure — Bearish (BOS) ────────────────────────────────
    if prev_sl and c < prev_sl[1] and c1 >= prev_sl[1] and vol_ratio >= 1.1:
        sc = _score(vol_ratio, body_pct)
        gr, gc = _grade(sc)
        signals.append({
            "category":       "SMC — Break of Structure",
            "signal":         "Bearish BOS",
            "description": (
                f"Price broke below the prior swing low (${prev_sl[1]:.2f}) "
                f"closing at ${c:.2f} — structural shift to bearish bias. "
                f"Prior support becomes resistance. "
                f"Vol {vol_ratio:.1f}× avg confirms institutional distribution. "
                f"Watch for a failed retest of ${prev_sl[1]:.2f} as new resistance."
            ),
            "bullish":        False,
            "signal_date":    d(-1),
            "levels":         _lev_bear,
            "momentum_score": sc,
            "momentum_grade": gr,
            "momentum_color": gc,
            "vol_ratio":      vol_ratio,
            "body_pct":       body_pct,
            "ma_dist_pct":    round((c - prev_sl[1]) / prev_sl[1] * 100, 2),
            "undercut_depth": None,
        })

    return signals


def detect_signals_historical(ticker: str, ohlcv: dict, lookback_days: int = 60) -> list:
    """Scan back to find recent historical U&R examples for context in signal cards."""
    closes  = ohlcv["close"]
    volumes = ohlcv["volume"]
    highs   = ohlcv["high"]
    lows    = ohlcv["low"]
    dates   = ohlcv.get("dates", [])

    if len(closes) < 15: return []

    def safe(arr, idx):
        try: return arr[idx]
        except (IndexError, TypeError): return None

    found = []
    scan_limit = min(lookback_days, len(closes) - 3)

    for period, label in [(200,"200-Day")]:
        if len(closes) < period + 2: continue
        ma_arr = sma(closes, period)

        for i in range(2, scan_limit + 2):
            idx      = -(i)
            c_today  = safe(closes,  idx)
            c_prev   = safe(closes,  idx - 1)
            ma_today = safe(ma_arr,  idx)
            ma_prev  = safe(ma_arr,  idx - 1)
            h_today  = safe(highs,   idx)
            lo_today = safe(lows,    idx)
            v_today  = safe(volumes, idx)
            dt       = dates[idx] if dates and abs(idx) <= len(dates) else ""

            if not all([c_today, c_prev, ma_today, ma_prev]): continue

            avg_v = sum(volumes[idx-10:idx]) / 10 if len(volumes) >= abs(idx)+10 else v_today
            vr    = round(v_today / avg_v, 2) if avg_v else 1.0
            br    = h_today - lo_today
            bq    = round((c_today - lo_today) / br, 2) if br > 0 else 0.5
            dist  = round((c_today - ma_today) / ma_today * 100, 2)

            if c_prev < ma_prev and c_today > ma_today:
                # Historical reclaim
                def _sc(d2, vr2, bq2):
                    s=0
                    if d2>=2.0: s+=3
                    elif d2>=1.0: s+=2
                    elif d2>=0.3: s+=1
                    if vr2>=2.5: s+=4
                    elif vr2>=2.0: s+=3
                    elif vr2>=1.5: s+=2
                    elif vr2>=1.0: s+=1
                    if bq2>=0.8: s+=3
                    elif bq2>=0.6: s+=2
                    elif bq2>=0.4: s+=1
                    return min(s,10)
                ms = _sc(abs(dist), vr, bq)
                ml = "Strong" if ms>=8 else ("Moderate" if ms>=5 else "Weak")
                mc = "#00e676" if ms>=8 else ("#FFD700" if ms>=5 else "#ef9a9a")
                found.append({
                    "category":       f"Undercut & Reclaim — {label} SMA",
                    "signal":         f"Reclaim {label} SMA",
                    "signal_date":    dt,
                    "bullish":        True,
                    "historical":     True,
                    "momentum_score": ms,
                    "momentum_label": ml,
                    "momentum_color": mc,
                    "vol_ratio":      vr,
                    "bar_quality":    bq,
                    "dist_pct":       dist,
                    "ma_level":       round(ma_today, 2),
                    "close_price":    round(c_today, 2),
                    "description":    (
                        f"{dt}: Reclaimed {label} SMA at ${ma_today:.2f}. "
                        f"Closed ${c_today:.2f} (+{dist:.2f}% above MA). "
                        f"Vol {vr:.1f}× avg | Bar {bq:.0%} | Score {ms}/10 ({ml})."
                    ),
                })
            elif c_prev > ma_prev and c_today < ma_today:
                found.append({
                    "category":       f"Undercut & Reclaim — {label} SMA",
                    "signal":         f"Undercut {label} SMA",
                    "signal_date":    dt,
                    "bullish":        False,
                    "historical":     True,
                    "momentum_score": None,
                    "vol_ratio":      vr,
                    "bar_quality":    bq,
                    "dist_pct":       dist,
                    "ma_level":       round(ma_today, 2),
                    "close_price":    round(c_today, 2),
                    "description":    (
                        f"{dt}: Undercut {label} SMA at ${ma_today:.2f}. "
                        f"Closed ${c_today:.2f} ({dist:.2f}% below MA). "
                        f"Vol {vr:.1f}× avg | Bar {bq:.0%}."
                    ),
                })

    # Return 2 most recent (sorted by date desc), one of each type if possible
    found.sort(key=lambda x: x.get("signal_date",""), reverse=True)
    seen_bull = seen_bear = False
    out = []
    for f in found:
        if f["bullish"] and not seen_bull:
            out.append(f); seen_bull = True
        elif not f["bullish"] and not seen_bear:
            out.append(f); seen_bear = True
        if seen_bull and seen_bear: break
    return out


def _tv_url(ticker: str, category: str, signal_date: str) -> str:
    """
    Build a TradingView chart URL pre-loaded with the right indicator for each signal category.
    Uses the 'studies' param to auto-add the relevant overlay.
    Zooms to a 3-month daily view centered on the signal date.
    """
    base   = f"https://www.tradingview.com/chart/?symbol={ticker}"
    # Map category to the most relevant TradingView built-in study id
    study_map = {
        "Undercut & Reclaim — 50-Day SMA":  "MASimple@tv-basicstudies",
        "Undercut & Reclaim — 100-Day SMA": "MASimple@tv-basicstudies",
        "Undercut & Reclaim — 200-Day SMA": "MASimple@tv-basicstudies",
    }
    study = study_map.get(category, "MASimple@tv-basicstudies")
    params = f"&interval=D&studies={study}"
    # Append signal date as a range hint (TV uses from/to as unix timestamps)
    if signal_date:
        try:
            from datetime import datetime, timedelta
            dt      = datetime.strptime(signal_date, "%Y-%m-%d")
            from_dt = dt - timedelta(days=90)
            to_dt   = dt + timedelta(days=14)
            import time as _time
            params += f"&from={int(from_dt.timestamp())}&to={int(to_dt.timestamp())}"
        except Exception:
            pass
    return base + params


def compute_csp_conviction(ohlcv: dict, confidence: str = "") -> dict:
    """
    Compute the CSP conviction score (0-100) from raw OHLCV data.
    Same criteria as the Mag 7 scoring:
      40pts — proximity to 95% annual floor (percentile of daily lows)
      30pts — confidence signal (derived from MA position if not supplied)
      30pts — Fib level confluence (nearest fib within X% of 95% floor)

    Returns dict with: floor_95, otm_pct, fib_price, fib_label,
                       fib_gap_pct, score, grade, grade_col, conviction_label
    """
    closes  = ohlcv.get("close", [])
    lows    = ohlcv.get("low",   [])
    highs   = ohlcv.get("high",  [])
    if len(lows) < 50 or len(closes) < 50:
        return {}

    px = closes[-1]

    # 95% floor = 5th percentile of daily lows over available history
    _sorted_lo = sorted(lows)
    _n = len(_sorted_lo)
    floor_95  = round(_sorted_lo[max(0, int(_n * 0.05))], 2)
    otm_pct   = round((px - floor_95) / px * 100, 1) if px else 0

    # Fibonacci levels from 52-week swing
    _hi = max(highs[-252:]) if len(highs) >= 252 else max(highs)
    _lo = min(lows[-252:])  if len(lows)  >= 252 else min(lows)
    _fib_range = _hi - _lo
    _fib_levels = []
    for _fp_pct, _fp_lbl in [
        (0.764,"23.6%"),(0.618,"38.2%"),(0.500,"50%"),
        (0.382,"61.8% — Golden Ratio"),(0.236,"76.4%"),
    ]:
        _fp = round(_lo + _fib_range * _fp_pct, 2)
        if _fp < px:   # only care about levels below current price
            _fib_levels.append((_fp, _fp_lbl))

    # Nearest Fib support below current price
    _fib_nearest = max(_fib_levels, key=lambda x: x[0]) if _fib_levels else None
    fib_price = _fib_nearest[0] if _fib_nearest else None
    fib_label = _fib_nearest[1] if _fib_nearest else ""
    fib_gap   = round(abs(fib_price - floor_95) / floor_95 * 100, 1) if fib_price else None

    # ── Score ─────────────────────────────────────────────────────────────────
    # Proximity to 95% floor
    prox_pts = (40 if otm_pct <= 15 else 25 if otm_pct <= 25
                else 10 if otm_pct <= 40 else 0)

    # Confidence (from MA position if available)
    if not confidence:
        # Derive from SMAs
        def _sma(p, n):
            return sum(p[-n:]) / n if len(p) >= n else None
        sma50  = _sma(closes, 50)
        sma200 = _sma(closes, 200) if len(closes) >= 200 else None
        if sma50 and sma200:
            if px > sma50 and sma50 > sma200:
                confidence = "HIGH"
            elif px < sma50 and px < sma200:
                confidence = "WEAK"
            elif px < sma50 and px > sma200:
                confidence = "PULLBACK"
            else:
                confidence = "RECOVERY"
        # Override if near 95% floor
        if otm_pct <= 15:
            confidence = "HIGH — CSP ZONE"

    conf_pts = (30 if "CSP ZONE" in confidence else
                20 if confidence == "HIGH" else
                10 if confidence in ("PULLBACK", "RECOVERY") else 0)

    # Fib confluence
    fib_pts = (30 if (fib_gap is not None and fib_gap <= 3) else
               20 if (fib_gap is not None and fib_gap <= 5) else 0)

    score = prox_pts + conf_pts + fib_pts

    # Grade
    if score >= 70:   grade, grade_col = "A", "#00e676"
    elif score >= 50: grade, grade_col = "B", "#FFD700"
    elif score >= 30: grade, grade_col = "C", "#ffb74d"
    else:             grade, grade_col = "F", "#ef5350"

    if score >= 70:   conv_label = "HIGH ★"
    elif score >= 40: conv_label = "MODERATE"
    else:             conv_label = "WATCH"

    return {
        "floor_95":      floor_95,
        "otm_pct":       otm_pct,
        "fib_price":     fib_price,
        "fib_label":     fib_label,
        "fib_gap_pct":   fib_gap,
        "score":         score,
        "grade":         grade,
        "grade_col":     grade_col,
        "conv_label":    conv_label,
        "confidence":    confidence,
        "prox_pts":      prox_pts,
        "conf_pts":      conf_pts,
        "fib_pts":       fib_pts,
    }


def run_all_signals(tickers: List[str]) -> List[dict]:
    """
    Run U&R signal detection across all tickers.
    Each signal includes momentum score + up to 2 historical examples from last 60 days.
    """
    clean   = [t.lstrip("$") for t in tickers if t]
    ohlcv   = fetch_ohlcv_for_signals(clean)
    results = []

    from datetime import datetime, timedelta
    cutoff_recent = (datetime.now() - timedelta(days=10)).strftime("%Y-%m-%d")

    for t in clean:
        if t not in ohlcv:
            continue
        data = ohlcv[t]

        # Recent signals (last ~3 calendar days = last 2 trading days)
        # U&R signals
        hits = detect_signals(t, data)
        # SMC signals (Sweep, BOS, Supply/Demand)
        hits += detect_smc_signals(t, data)
        # Compute CSP conviction score once per ticker (same for all hits)
        _csp_cv = compute_csp_conviction(data)
        for h in hits:
            _is_index = t in INDEX_TICKERS
            results.append({
                "ticker":              f"${t}",
                "is_index":            _is_index,
                "ticker_name":         INDEX_NAMES.get(t, ""),
                "csp_conviction":      _csp_cv,
                "category":            h["category"],
                "signal":              h["signal"],
                "description":         h["description"],
                "bullish":             h["bullish"],
                "chart_url":           _tv_url(t, h["category"], h.get("signal_date","")),
                "signal_date":         h.get("signal_date", ""),
                "momentum_score":      h.get("momentum_score"),
                "momentum_grade":      h.get("momentum_grade", "—"),
                "momentum_color":      h.get("momentum_color", "#aaa"),
                "vol_ratio":           h.get("vol_ratio"),
                "body_pct":            h.get("body_pct"),
                "ma_dist_pct":         h.get("ma_dist_pct"),
                "undercut_depth":      h.get("undercut_depth"),
                "levels":              h.get("levels", {}),
                "historical_examples": detect_signals_historical(t, data, lookback_days=60),
            })

    results = [r for r in results if r.get("bullish", False)]  # reclaims only
    results = [r for r in results if r.get("signal_date","") >= cutoff_recent]
    results = [r for r in results if (r.get("momentum_score") or 0) >= 70]  # 70+ (Grade A/B)
    results.sort(key=lambda x: (-(x.get("momentum_score") or 0), x["ticker"]))
    print(f"   signals: found {len(results)} signals scored ≥70 across {len(clean)} tickers (last 10 days)")
    return results


# ------------------ Stock table builders ------------------
def build_mag7_rows(sheet_rows: List[dict]) -> Tuple[List[dict], set]:
    ticker_meta = {}
    for r in sheet_rows:
        t = normalize_ticker(r.get("Ticker"))
        if not t: continue
        ticker_meta[t] = {"Name": r.get("Name")}
    live = fetch_yfinance_batch(list(ticker_meta.keys()))
    out_rows, tickers = [], set()
    for t, meta in ticker_meta.items():
        prices = live.get(t.lstrip("$"), {})
        out_rows.append({"Ticker": t, **meta, **prices})
        tickers.add(t)
    return out_rows, tickers


def build_watchlist_rows(sheet_rows: List[dict]) -> Tuple[List[dict], set]:
    ticker_meta = {}
    for r in sheet_rows:
        t = normalize_ticker(r.get("Ticker"))
        if not t: continue
        ticker_meta[t] = {"Name": r.get("Name")}
    live = fetch_yfinance_batch(list(ticker_meta.keys()))
    out_rows, tickers, skipped = [], set(), 0
    for t, meta in ticker_meta.items():
        prices = live.get(t.lstrip("$"), {})
        mktcap = prices.get("market_cap")
        if mktcap is not None and mktcap < WATCHLIST_MKTCAP_MIN:
            skipped += 1
            continue
        out_rows.append({"Ticker": t, **meta, **prices})
        tickers.add(t)
    if skipped:
        print(f"   Watchlist: filtered out {skipped} tickers below ${WATCHLIST_MKTCAP_MIN/1e9:.0f}B market cap")
    return out_rows, tickers

# ------------------ formatting helpers ------------------
def fmt_pct(v: Optional[float]) -> str:
    if v is None: return "—"
    return f"{v:.2f}%"

def fmt_big(n: Optional[float]) -> str:
    if n is None: return "—"
    try: x = float(n)
    except Exception: return str(n)
    absx = abs(x)
    if absx >= 1e12: return f"{x/1e12:.2f}T"
    if absx >= 1e9:  return f"{x/1e9:.2f}B"
    if absx >= 1e6:  return f"{x/1e6:.2f}M"
    if absx >= 1e3:  return f"{x/1e3:.2f}K"
    return f"{x:.0f}"

def pct_class(v: Optional[float]) -> str:
    if v is None: return ""
    if v > 0.0001:  return "pos"
    if v < -0.0001: return "neg"
    return ""

def fmt_live_percent(v) -> Tuple[str, Optional[float]]:
    if v is None: return "—", None
    try: num = float(v)
    except Exception: return str(v), None
    return f"{num:.2f}%", num

# ------------------ tweets ------------------
def render_stock_table(table_id: str, title: str, rows: List[dict]) -> str:
    out = [f'<div class="card"><h3>{esc(title)} <span class="fresh">● live</span></h3>'
           f'<table id="{esc(table_id)}" data-filter-table="1"><thead><tr>']
    for col, st in STOCK_COLUMNS:
        out.append(f'<th data-sort="{st}">{esc(col)}</th>')
    out.append('</tr><tr class="filterrow">')
    for col, _ in STOCK_COLUMNS:
        ph = {"Ticker":"$AAPL","Name":"Apple","Price":">100","Volume":">1M"}.get(col, ">0" if col.endswith("%") else "")
        out.append(f'<th><input class="filter" placeholder="{esc(ph)}" oninput="applyFilters(\'{esc(table_id)}\')"></th>')
    out.append("</tr></thead><tbody>")
    for r in rows:
        out.append("<tr>")
        for col, _ in STOCK_COLUMNS:
            v = r.get(col)
            if col == "Ticker":   out.append(f"<td>{esc(normalize_ticker(v))}</td>")
            elif col == "Name":   out.append(f"<td class='namecell'>{esc(v)}</td>")
            elif col == "Price":
                fv = as_float(v)
                out.append(f"<td>{esc('—' if fv is None else f'{fv:.2f}')}</td>")
            elif col == "Volume":
                fv = as_float(v)
                out.append(f"<td>{esc(fmt_big(fv))}</td>")
            elif col.endswith("%"):
                txt, num = fmt_live_percent(v)
                out.append(f"<td class='{pct_class(num)}'>{esc(txt)}</td>")
            else: out.append(f"<td>{esc(v)}</td>")
        out.append("</tr>")
    out.append("</tbody></table></div>")
    return "".join(out)


def render_signals_table(signals: List[dict]) -> str:
    """Render U&R signals with momentum score and historical examples."""
    from collections import defaultdict

    bull_total = sum(1 for s in signals if s["bullish"])
    bear_total = len(signals) - bull_total

    out = ['<div class="card">']
    out.append(f'<h3>Reclaim &amp; SMC Signals <span class="fresh">● live</span></h3>')
    _ur_count  = sum(1 for s in signals if not s.get("category","").startswith("SMC"))
    _smc_count = sum(1 for s in signals if s.get("category","").startswith("SMC"))
    out.append(
        f'<div style="font-size:12px;color:#8ab4d4;margin-bottom:12px">'
        f'{len(signals)} signal{"s" if len(signals)!=1 else ""} across {len(set(s["ticker"] for s in signals))} tickers — '
        f'<span style="color:#00e676;font-weight:700">{_ur_count} U&amp;R reclaim{"s" if _ur_count!=1 else ""}</span>'
        f' &nbsp;+&nbsp; '
        f'<span style="color:#7e57c2;font-weight:700">{_smc_count} SMC signal{"s" if _smc_count!=1 else ""}</span>'
        f'</div>'
    )
    # Legend
    out.append('<div style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:16px;font-size:11px">')
    out.append('<span style="background:#00e676;color:#000;padding:2px 9px;border-radius:5px;font-weight:800">U&R Reclaim</span>'
               '<span style="color:#8ab4d4;margin-right:8px"> MA undercut then reclaimed on volume</span>')
    out.append('<span style="background:#7e57c2;color:#fff;padding:2px 9px;border-radius:5px;font-weight:800">SMC Sweep</span>'
               '<span style="color:#8ab4d4;margin-right:8px"> Liquidity grab then reversal</span>')
    out.append('<span style="background:#42a5f5;color:#000;padding:2px 9px;border-radius:5px;font-weight:800">SMC BOS</span>'
               '<span style="color:#8ab4d4"> Break of structure — trend shift</span>')
    out.append('</div>')


    # ── Criteria panel ───────────────────────────────────────────────────────
    out.append('''<div style="background:#0a1628;border-radius:10px;padding:14px 18px;margin-bottom:16px;font-size:11px;line-height:1.8">
<div style="font-size:13px;font-weight:900;color:#FFD700;margin-bottom:10px">📋 How Signals Are Detected</div>
<div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
<div>
<div style="color:#00e676;font-weight:800;margin-bottom:4px">U&R Reclaim (Undercut & Reclaim)</div>
<div style="color:#8ab4d4">Previous close was <em>below</em> the MA. Today's close is <em>above</em> the MA. Price reclaimed it in a single session — institutions stepped back in.</div>
<div style="color:#5a7fa0;margin-top:4px">MAs scanned: 50-day · 100-day · 200-day SMA</div>
</div>
<div>
<div style="color:#7e57c2;font-weight:800;margin-bottom:4px">SMC Liquidity Sweep + Demand Bounce</div>
<div style="color:#8ab4d4">Today's low broke <em>below</em> a prior swing low (grabbed stop-losses), then reversed to close <em>above</em> that level. Classic institutional trap — they sell the stops then buy.</div>
<div style="color:#5a7fa0;margin-top:4px">Requires: low < prior swing low · close > swing low · vol ≥ 1.2×</div>
</div>
<div>
<div style="color:#42a5f5;font-weight:800;margin-bottom:4px">SMC Bullish BOS (Break of Structure)</div>
<div style="color:#8ab4d4">Close breaks <em>above</em> the most recent swing high for the first time — structural shift to bullish bias. Prior resistance becomes support.</div>
<div style="color:#5a7fa0;margin-top:4px">Requires: close > prior swing high · vol ≥ 1.1×</div>
</div>
<div>
<div style="color:#FFD700;font-weight:800;margin-bottom:4px">Score 0-100 (only ≥70 shown)</div>
<div style="color:#8ab4d4"><strong style="color:#e8edf5">Volume (40pts)</strong> — 2.5×+ avg = 40, 2×=30, 1.5×=20, 1×=10<br>
<strong style="color:#e8edf5">Body % (30pts)</strong> — candle body ÷ full range. ≥80%=30, ≥60%=22, ≥40%=12<br>
<strong style="color:#e8edf5">Distance (30pts)</strong> — how far above/below the MA. ≥2%=30, ≥1%=20, ≥0.3%=10</div>
</div>
</div>
<div style="color:#5a7fa0;margin-top:10px;border-top:1px solid #0d2040;padding-top:8px">
Universe: ≥$10B market cap + SPY/QQQ/IWM/DIA indexes · Lookback: last 10 trading days · 
Grade A ≥70 · B ≥50 · C ≥30 · F &lt;30
</div>
</div>''')
    if not signals:
        out.append('<div style="color:#8ab4d4;padding:16px">No signals scored ≥70 in the last 10 trading days. Markets may be range-bound — no clean reclaims or sweeps detected.</div>')
        out.append('</div>')
        return "\n".join(out)

    # Sort: bullish first, then by momentum score desc
    signals_sorted = sorted(signals, key=lambda x: (not x["bullish"], -(x.get("momentum_score") or 0)))

    for s in signals_sorted:
        is_bull  = s["bullish"]
        card_col = "#00e676" if is_bull else "#ef5350"
        sig_icon = "↑" if is_bull else "↓"
        ms       = s.get("momentum_score")
        ml       = s.get("momentum_label","") or s.get("momentum_grade","")
        mc       = s.get("momentum_color","#aaa")
        vr       = s.get("vol_ratio")
        bq       = s.get("body_pct") or s.get("bar_quality")     # new name: body_pct
        dp       = s.get("ma_dist_pct") or s.get("dist_pct")     # new name: ma_dist_pct
        hist     = s.get("historical_examples",[])

        out.append(f'<div style="background:#0d1b2a;border-left:4px solid {card_col};border-radius:10px;padding:16px 18px;margin-bottom:14px">')

        # Header row
        out.append(f'<div style="display:flex;align-items:center;gap:12px;flex-wrap:wrap;margin-bottom:10px">')
        _idx_badge = ('<span style="background:#1565c0;color:#90caf9;font-size:10px;font-weight:800;'
                       'padding:1px 7px;border-radius:4px;margin-left:4px">INDEX</span>'
                      if s.get("is_index") else "")
        _idx_name  = f'<span style="font-size:11px;color:#5a7fa0;margin-left:4px">{s.get("ticker_name","")}</span>' if s.get("ticker_name") else ""
        out.append(f'<span style="font-size:20px;font-weight:900;color:{card_col}">{esc(s["ticker"])}</span>{_idx_badge}{_idx_name}')
        # Signal type badge (SMC vs U&R)
        _is_smc = s.get("category","").startswith("SMC")
        _type_badge = '<span style="background:#7e57c2;color:#fff;font-size:10px;font-weight:800;padding:1px 7px;border-radius:4px;margin-right:4px">SMC</span>' if _is_smc else ''
        out.append(f'<span style="font-size:13px;font-weight:800;color:{card_col}">{_type_badge}{sig_icon} {esc(s["signal"])}</span>')
        out.append(f'<span style="font-size:11px;color:#8ab4d4">{esc(s.get("signal_date",""))}</span>')

        # Momentum badge (reclaims only)
        if is_bull and ms is not None:
            out.append(f'<span style="background:{mc};color:#000;font-size:11px;font-weight:900;padding:2px 10px;border-radius:6px">Momentum {ms}/10 — {ml}</span>')

        # Chart link
        if s.get("chart_url"):
            out.append(f'<a href="{esc(s["chart_url"])}" target="_blank" style="margin-left:auto;font-size:11px;color:#42a5f5;text-decoration:none">📈 Chart</a>')
        out.append('</div>')

        # Description
        out.append(f'<div style="font-size:12px;color:#d0e4f7;line-height:1.7;margin-bottom:10px">{esc(s["description"])}</div>')

        # CSP Conviction score
        _cv = s.get("csp_conviction") or {}
        if _cv:
            _cv_sc  = _cv.get("score", 0)
            _cv_gc  = _cv.get("grade_col","#aaa")
            _cv_lbl = _cv.get("conv_label","")
            _cv_f95 = _cv.get("floor_95")
            _cv_otm = _cv.get("otm_pct")
            _cv_fib = _cv.get("fib_label","")
            _cv_gap = _cv.get("fib_gap_pct")
            out.append(f'<div style="background:#0d1b2a;border-left:3px solid {_cv_gc};border-radius:7px;padding:9px 14px;margin-bottom:10px;display:flex;align-items:center;gap:14px;flex-wrap:wrap">')
            out.append(f'<span style="font-size:10px;color:#5a7fa0;font-weight:700">CSP CONVICTION</span>')
            out.append(f'<span style="background:{_cv_gc};color:#000;font-size:12px;font-weight:900;padding:2px 10px;border-radius:6px">{_cv_lbl}</span>')
            out.append(f'<span style="font-size:16px;font-weight:900;color:{_cv_gc}">{_cv_sc}/100</span>')
            if _cv_f95:
                _otm_col = "#00e676" if _cv_otm <= 15 else "#8ab4d4"
                out.append(f'<span style="font-size:11px;color:#8ab4d4">95% floor: <strong style="color:#FFD700">${_cv_f95:.2f}</strong> <span style="color:{_otm_col}">({_cv_otm:.1f}% OTM)</span></span>')
            if _cv_fib and _cv_gap is not None:
                _fib_col3 = "#FFD700" if _cv_gap <= 3 else "#8ab4d4"
                out.append(f'<span style="font-size:11px;color:{_fib_col3}">Fib: {esc(_cv_fib.split("—")[0].strip())} ({_cv_gap:.1f}% from floor)</span>')
            out.append('</div>')

        # SMC-specific metrics block
        if _is_smc and ms is not None and vr is not None:
            _smc_cat = s.get("category","")
            _sweep_d = s.get("undercut_depth")
            _ma_d    = s.get("ma_dist_pct")
            out.append('<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:8px;margin-bottom:10px">')
            _smc_metrics = [
                ("Score",       f"{ms}/100", mc,         "Combined volume + body quality score"),
                ("Vol Ratio",   f"{vr:.1f}× avg",
                 "#00e676" if vr>=1.5 else ("#FFD700" if vr>=1.0 else "#ef9a9a"),
                 ">1.5× = institutional participation confirmed"),
            ]
            if bq is not None:
                _smc_metrics.append(("Bar Body", f"{bq:.0f}%",
                    "#00e676" if bq>=60 else ("#FFD700" if bq>=40 else "#ef9a9a"),
                    ">60% = conviction reversal bar"))
            if _sweep_d is not None:
                _smc_metrics.append(("Sweep Depth", f"{_sweep_d:.2f}%",
                    "#FFD700", "How far price overshot the level before reversing"))
            if _ma_d is not None:
                _dist_lbl = "Above Level" if _ma_d >= 0 else "Below Level"
                _smc_metrics.append((_dist_lbl, f"{_ma_d:+.2f}%",
                    "#00e676" if _ma_d >= 0 else "#ef5350",
                    "Where close finished relative to the swept level"))
            # Signal-type specific label
            if "Sweep" in _smc_cat:
                _smc_metrics.append(("Signal Type", "Liquidity Sweep", "#7e57c2",
                    "Stop-hunt — institutions grab stops then reverse price"))
            elif "Break of Structure" in _smc_cat:
                _smc_metrics.append(("Signal Type", "Break of Structure", "#42a5f5",
                    "Structural shift — trend direction change confirmed"))
            for lbl, val, col, explain in _smc_metrics:
                out.append(f'<div style="background:#13263d;border-radius:7px;padding:8px 10px" title="{explain}">')
                out.append(f'<div style="font-size:10px;color:#5a7fa0;margin-bottom:2px">{lbl}</div>')
                out.append(f'<div style="font-size:13px;font-weight:800;color:{col}">{val}</div>')
                out.append('</div>')
            out.append('</div>')

        # U&R momentum metrics bar (non-SMC reclaims only)
        elif not _is_smc and is_bull and ms is not None and vr is not None and bq is not None and dp is not None:
            out.append('<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin-bottom:10px">')
            for lbl, val, col, explain in [
                ("Momentum Score", f"{ms}/10", mc,
                 "Combined score from volume, bar quality, and distance cleared"),
                ("Volume Ratio",   f"{vr:.1f}× avg",
                 "#00e676" if vr>=1.5 else ("#FFD700" if vr>=1.0 else "#ef9a9a"),
                 "Today vs 20-day avg. >1.5× = institutions buying"),
                ("Bar Quality",    f"{bq:.0f}%",
                 "#00e676" if bq>=60 else ("#FFD700" if bq>=40 else "#ef9a9a"),
                 "Candle body as % of range. >60% = conviction bar"),
                ("Distance Above", f"+{dp:.2f}%",
                 "#00e676" if dp>=1.0 else ("#FFD700" if dp>=0.3 else "#ef9a9a"),
                 "How far above the MA price closed. >1% = convincing reclaim"),
            ]:
                out.append(f'<div style="background:#13263d;border-radius:7px;padding:8px 10px" title="{explain}">')
                out.append(f'<div style="font-size:10px;color:#5a7fa0;margin-bottom:2px">{lbl}</div>')
                out.append(f'<div style="font-size:14px;font-weight:800;color:{col}">{val}</div>')
                out.append('</div>')
            out.append('</div>')

        # Historical examples (last 60 days)
        if hist:
            out.append('<div style="background:#0a1628;border-radius:7px;padding:10px 12px">')
            out.append('<div style="font-size:10px;font-weight:800;color:#5a7fa0;margin-bottom:8px">RECENT HISTORY — LAST 60 DAYS (same ticker)</div>')
            out.append('<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:8px">')
            for hx in hist:
                hc  = "#00e676" if hx["bullish"] else "#ef5350"
                hms = hx.get("momentum_score")
                hml = hx.get("momentum_label","")
                hmc = hx.get("momentum_color","#aaa")
                out.append(f'<div style="background:#0d1b2a;border-left:3px solid {hc};border-radius:6px;padding:8px 10px">')
                out.append(f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:4px">')
                out.append(f'<span style="font-size:11px;font-weight:800;color:{hc}">{"↑ Reclaim" if hx["bullish"] else "↓ Undercut"} {esc(hx.get("signal",""))}</span>')
                out.append(f'<span style="font-size:10px;color:#8ab4d4">{esc(hx.get("signal_date",""))}</span>')
                if hms is not None:
                    out.append(f'<span style="background:{hmc};color:#000;font-size:9px;font-weight:800;padding:1px 6px;border-radius:4px">{hms}/10 {hml}</span>')
                out.append('</div>')
                out.append(f'<div style="font-size:11px;color:#8ab4d4;line-height:1.5">{esc(hx["description"])}</div>')
                out.append('</div>')
            out.append('</div></div>')

        # Entry/stop levels (bullish only)
        if is_bull:
            lev = s.get("levels", {})
            if lev and lev.get("entry"):
                out.append('<div style="display:flex;gap:12px;flex-wrap:wrap;margin-top:8px">')
                for lbl, key, col in [
                    ("Entry", "entry", "#FFD700"),
                    ("Stop",  "stop",  "#ef5350"),
                    ("PT1",   "pt1",   "#42a5f5"),
                    ("PT2",   "pt2",   "#00e676"),
                ]:
                    v = lev.get(key)
                    if v:
                        out.append(f'<span style="font-size:11px;color:{col};background:#13263d;padding:2px 8px;border-radius:5px">{lbl}: ${v:.2f}</span>')
                out.append('</div>')

        out.append('</div>')

    out.append('</div>')
    return "\n".join(out)


def render_stock_table(table_id: str, title: str, rows: List[dict]) -> str:
    out = [f'<div class="card"><h3>{esc(title)} <span class="fresh">● live</span></h3>'
           f'<table id="{esc(table_id)}" data-filter-table="1"><thead><tr>']
    for col, st in STOCK_COLUMNS:
        out.append(f'<th data-sort="{st}">{esc(col)}</th>')
    out.append('</tr><tr class="filterrow">')
    for col, _ in STOCK_COLUMNS:
        ph = {"Ticker":"$AAPL","Name":"Apple","Price":">100","Volume":">1M"}.get(col, ">0" if col.endswith("%") else "")
        out.append(f'<th><input class="filter" placeholder="{esc(ph)}" oninput="applyFilters(\'{esc(table_id)}\')"></th>')
    out.append("</tr></thead><tbody>")
    for r in rows:
        out.append("<tr>")
        for col, _ in STOCK_COLUMNS:
            v = r.get(col)
            if col == "Ticker":   out.append(f"<td>{esc(normalize_ticker(v))}</td>")
            elif col == "Name":   out.append(f"<td class='namecell'>{esc(v)}</td>")
            elif col == "Price":
                fv = as_float(v)
                out.append(f"<td>{esc('—' if fv is None else f'{fv:.2f}')}</td>")
            elif col == "Volume":
                fv = as_float(v)
                out.append(f"<td>{esc(fmt_big(fv))}</td>")
            elif col.endswith("%"):
                txt, num = fmt_live_percent(v)
                out.append(f"<td class='{pct_class(num)}'>{esc(txt)}</td>")
            else: out.append(f"<td>{esc(v)}</td>")
        out.append("</tr>")
    out.append("</tbody></table></div>")
    return "".join(out)


def fetch_options_data(tickers: List[str]) -> dict:
    """
    For each ticker fetch via yfinance:
      - IV rank / IV percentile (approximated from 52wk IV history)
      - Implied move (weekly + monthly from ATM straddle)
      - Put/call volume ratio
      - Top strikes by OI for nearest two expirations
    Returns dict keyed by bare ticker.
    """
    try:
        import yfinance as yf
    except ImportError:
        print("⚠️  yfinance not installed")
        return {}

    clean = [t.lstrip("$") for t in tickers if t]

    # Check cache
    if os.path.exists(YF_OPTIONS_CACHE):
        age = time.time() - os.path.getmtime(YF_OPTIONS_CACHE)
        if age < YF_OPTIONS_TTL:
            try:
                cached = json.loads(Path(YF_OPTIONS_CACHE).read_text(encoding="utf-8"))
                if all(t in cached for t in clean):
                    print(f"   options: using cache ({int(age/60)}m old)")
                    return cached
            except Exception:
                pass

    print(f"   options: fetching data for {len(clean)} tickers...")
    results = {}

    for t in clean:
        try:
            tk   = yf.Ticker(t)
            info = tk.info or {}

            price = info.get("regularMarketPrice") or info.get("currentPrice") or 0

            # ── IV metrics ──────────────────────────────────────────────────
            # yfinance doesn't expose IV history directly; use impliedVolatility
            # from info and approximate IV rank vs 52wk range
            iv_current = info.get("impliedVolatility")  # decimal e.g. 0.35

            # Approximate historical IV from annualised price volatility (52wk)
            hist = tk.history(period="1y", interval="1d")
            if not hist.empty and len(hist) > 20:
                import math
                log_rets = [math.log(hist["Close"].iloc[i] / hist["Close"].iloc[i-1])
                            for i in range(1, len(hist))]
                hv_daily = (sum(r**2 for r in log_rets) / len(log_rets)) ** 0.5
                hv_annual = hv_daily * math.sqrt(252)
            else:
                hv_annual = None

            # IV rank: where current IV sits in 52wk IV range (approximated by HV)
            # We use IV vs historical vol as a proxy
            iv_rank = None
            iv_pct  = None
            if iv_current and hv_annual:
                # Simple proxy: iv_rank = (IV - HV_low) / (HV_high - HV_low)
                # We don't have full IV history so use ratio IV/HV as proxy
                ratio = round(iv_current / hv_annual * 100, 1) if hv_annual else None
                iv_rank = ratio  # >100 = IV > HV (elevated), <100 = IV < HV (depressed)
                iv_pct  = round(min(iv_current * 100, 999), 1)  # as %

            # ── Implied move ─────────────────────────────────────────────────
            # Expected move = IV * price * sqrt(DTE/365)
            implied_move_weekly  = None
            implied_move_monthly = None
            if iv_current and price:
                import math
                implied_move_weekly  = round(price * iv_current * math.sqrt(5/252),  2)
                implied_move_monthly = round(price * iv_current * math.sqrt(21/252), 2)
                impl_pct_weekly  = round(implied_move_weekly  / price * 100, 1)
                impl_pct_monthly = round(implied_move_monthly / price * 100, 1)
            else:
                impl_pct_weekly = impl_pct_monthly = None

            # ── Put/Call ratio + OI data from options chain ──────────────────
            pc_ratio      = None
            top_call_oi   = []
            top_put_oi    = []
            best_csp_strike   = None
            best_csp_premium  = None
            best_csp_exp      = None
            best_cc_strike    = None
            best_cc_premium   = None
            best_cc_exp       = None

            # ── Scan ALL expirations within next 12 months ──────────────────
            cutoff_12mo = datetime.now() + timedelta(days=366)
            all_exps = []
            for exp_str in (tk.options or []):
                try:
                    exp_dt = datetime.strptime(exp_str, "%Y-%m-%d")
                    if exp_dt <= cutoff_12mo:
                        all_exps.append((exp_str, exp_dt))
                except Exception:
                    continue
            all_exps.sort(key=lambda x: x[1])

            total_call_vol = total_put_vol = 0
            total_call_oi  = total_put_oi  = 0

            # Per-expiry best strike storage: list of dicts
            csp_by_exp = []
            cc_by_exp  = []

            def dte_label(dte):
                if dte <= 7:    return "Weekly"
                if dte <= 21:   return "Monthly (short)"
                if dte <= 45:   return "Monthly (sweet)"
                if dte <= 90:   return "Quarterly"
                if dte <= 180:  return "6-Month"
                return "LEAPS (9-12mo)"

            for exp_str, exp_dt in all_exps:
                try:
                    dte_val = max((exp_dt - datetime.now()).days, 1)
                    chain   = tk.option_chain(exp_str)
                    calls   = chain.calls
                    puts    = chain.puts

                    # Accumulate P/C from first 4 expirations only
                    if len(csp_by_exp) < 4:
                        total_call_vol += int(calls["volume"].fillna(0).sum())
                        total_put_vol  += int(puts["volume"].fillna(0).sum())
                        total_call_oi  += int(calls["openInterest"].fillna(0).sum())
                        total_put_oi   += int(puts["openInterest"].fillna(0).sum())

                    # For LEAPS (>180d), widen OTM range to 10-40% OTM
                    if dte_val > 180:
                        put_low, put_high   = price * 0.60, price * 0.92
                        call_low, call_high = price * 1.05, price * 1.50
                    else:
                        put_low, put_high   = price * 0.80, price * 0.97
                        call_low, call_high = price * 1.03, price * 1.20

                    # Best CSP for this expiry
                    if price and not puts.empty:
                        otm_puts = puts[(puts["strike"] > put_low) &
                                        (puts["strike"] < put_high) &
                                        (puts["lastPrice"].fillna(0) > 0)].copy()
                        if not otm_puts.empty:
                            otm_puts["pct_return"] = otm_puts["lastPrice"] / otm_puts["strike"] * 100
                            # Annualise for fair comparison across DTEs
                            otm_puts["ann_ret"] = otm_puts["pct_return"] * (365 / dte_val)
                            best = otm_puts.loc[otm_puts["ann_ret"].idxmax()]
                            csp_by_exp.append({
                                "exp":       exp_str,
                                "dte":       dte_val,
                                "label":     dte_label(dte_val),
                                "strike":    round(float(best["strike"]), 2),
                                "premium":   round(float(best["lastPrice"]), 2),
                                "pct":       round(float(best["pct_return"]), 2),
                                "ann_ret":   round(float(best["ann_ret"]), 1),
                                "otm_pct":   round((price - float(best["strike"])) / price * 100, 1),
                            })

                    # Best CC for this expiry
                    if price and not calls.empty:
                        otm_calls = calls[(calls["strike"] > call_low) &
                                          (calls["strike"] < call_high) &
                                          (calls["lastPrice"].fillna(0) > 0)].copy()
                        if not otm_calls.empty:
                            otm_calls["pct_return"] = otm_calls["lastPrice"] / price * 100
                            otm_calls["ann_ret"]    = otm_calls["pct_return"] * (365 / dte_val)
                            best = otm_calls.loc[otm_calls["ann_ret"].idxmax()]
                            cc_by_exp.append({
                                "exp":       exp_str,
                                "dte":       dte_val,
                                "label":     dte_label(dte_val),
                                "strike":    round(float(best["strike"]), 2),
                                "premium":   round(float(best["lastPrice"]), 2),
                                "pct":       round(float(best["pct_return"]), 2),
                                "ann_ret":   round(float(best["ann_ret"]), 1),
                                "otm_pct":   round((float(best["strike"]) - price) / price * 100, 1),
                            })
                except Exception:
                    continue

            if total_call_vol + total_put_vol > 0:
                pc_ratio = round(total_put_vol / total_call_vol, 2) if total_call_vol else None

            # Best single expiry for backward compat (highest ann_ret in sweet-spot DTE)
            def _best(rows):
                sweet = [r for r in rows if 21 <= r["dte"] <= 45] or rows
                return max(sweet, key=lambda r: r["ann_ret"]) if sweet else {}

            best_csp = _best(csp_by_exp)
            best_cc  = _best(cc_by_exp)

            results[t] = {
                "price":               round(float(price), 2) if price else None,
                "iv_pct":              iv_pct,
                "iv_rank":             iv_rank,
                "hv_annual":           round(hv_annual * 100, 1) if hv_annual else None,
                "impl_move_weekly":    implied_move_weekly,
                "impl_pct_weekly":     impl_pct_weekly,
                "impl_move_monthly":   implied_move_monthly,
                "impl_pct_monthly":    impl_pct_monthly,
                "pc_ratio":            pc_ratio,
                "call_oi":             total_call_oi,
                "put_oi":              total_put_oi,
                # All expirations (used by wheel tab for per-expiry rows)
                "csp_by_exp":          csp_by_exp,
                "cc_by_exp":           cc_by_exp,
                # Best single values (backward compat with confidence scorer)
                "csp_strike":          best_csp.get("strike"),
                "csp_premium":         best_csp.get("premium"),
                "csp_exp":             best_csp.get("exp"),
                "csp_pct":             best_csp.get("pct"),
                "cc_strike":           best_cc.get("strike"),
                "cc_premium":          best_cc.get("premium"),
                "cc_exp":              best_cc.get("exp"),
                "cc_pct":              best_cc.get("pct"),
            }
            print(f"   options: ✓ {t}")
        except Exception as e:
            print(f"   options: ⚠️  {t} failed — {e}")
            continue

    if results:
        try:
            Path(YF_OPTIONS_CACHE).write_text(json.dumps(results), encoding="utf-8")
        except Exception:
            pass

    print(f"   options: got data for {len(results)}/{len(clean)} tickers")
    return results


def load_wheel_positions() -> List[dict]:
    """Load wheel tracker from JSON file. Creates empty file if missing."""
    if not os.path.exists(WHEEL_TRACKER_FILE):
        empty = {"positions": [], "closed": []}
        Path(WHEEL_TRACKER_FILE).write_text(json.dumps(empty, indent=2))
        return []
    try:
        data = json.loads(Path(WHEEL_TRACKER_FILE).read_text())
        return data.get("positions", [])
    except Exception:
        return []


def _wheel_confidence(d: dict, trade_type: str) -> tuple:
    """
    Score a CSP or CC trade 0-100 and return (score, label, color, reasons).
    Factors:
      - IV/HV ratio (elevated IV = better premium, max 30pts)
      - Annualised return (higher = better, max 25pts)
      - DTE in sweet spot 21-45 days (max 20pts)
      - OTM buffer safety (max 15pts)
      - P/C ratio direction alignment (max 10pts)
    """
    score   = 0
    reasons = []

    iv_rank  = d.get("iv_rank") or 0
    pc_ratio = d.get("pc_ratio") or 1.0

    if trade_type == "CSP":
        ann_ret  = d.get("csp_ann_ret") or 0
        dte      = d.get("csp_dte") or 0
        otm_pct  = d.get("csp_otm_pct") or 0
        # IV/HV: >130 = max, 100-130 = good, <100 = poor
        if iv_rank >= 130:
            score += 30; reasons.append("IV highly elevated ✓")
        elif iv_rank >= 100:
            score += 20; reasons.append("IV above HV ✓")
        elif iv_rank >= 80:
            score += 10; reasons.append("IV near HV")
        else:
            reasons.append("IV low — thin premium")
        # Ann return
        if ann_ret >= 40:   score += 25; reasons.append(f"{ann_ret:.0f}% ann return ✓")
        elif ann_ret >= 25: score += 18; reasons.append(f"{ann_ret:.0f}% ann return")
        elif ann_ret >= 12: score += 10; reasons.append(f"{ann_ret:.0f}% ann return")
        else:               reasons.append(f"{ann_ret:.0f}% ann return — low")
        # DTE sweet spot 21-45
        if 21 <= dte <= 45:   score += 20; reasons.append(f"{dte}d DTE — optimal ✓")
        elif 14 <= dte <= 55: score += 12; reasons.append(f"{dte}d DTE — acceptable")
        elif dte > 0:         score += 5;  reasons.append(f"{dte}d DTE — outside sweet spot")
        # OTM buffer: 5-15% OTM is ideal
        if 5 <= otm_pct <= 15:   score += 15; reasons.append(f"{otm_pct:.1f}% OTM — ideal ✓")
        elif 3 <= otm_pct <= 20: score += 8;  reasons.append(f"{otm_pct:.1f}% OTM")
        else:                    reasons.append(f"{otm_pct:.1f}% OTM — too close/far")
        # P/C: low put/call = less bearish pressure (good for CSP)
        if pc_ratio <= 0.8:   score += 10; reasons.append("Bullish P/C ✓")
        elif pc_ratio <= 1.1: score += 5;  reasons.append("Neutral P/C")
        else:                 reasons.append("Bearish P/C — elevated put buying")

    else:  # CC
        ann_ret = d.get("cc_ann_ret") or 0
        dte     = d.get("cc_dte") or 0
        otm_pct = d.get("cc_otm_pct") or 0
        if iv_rank >= 130:
            score += 30; reasons.append("IV highly elevated ✓")
        elif iv_rank >= 100:
            score += 20; reasons.append("IV above HV ✓")
        elif iv_rank >= 80:
            score += 10; reasons.append("IV near HV")
        else:
            reasons.append("IV low — thin premium")
        if ann_ret >= 40:   score += 25; reasons.append(f"{ann_ret:.0f}% ann return ✓")
        elif ann_ret >= 25: score += 18; reasons.append(f"{ann_ret:.0f}% ann return")
        elif ann_ret >= 12: score += 10; reasons.append(f"{ann_ret:.0f}% ann return")
        else:               reasons.append(f"{ann_ret:.0f}% ann return — low")
        if 21 <= dte <= 45:   score += 20; reasons.append(f"{dte}d DTE — optimal ✓")
        elif 14 <= dte <= 55: score += 12; reasons.append(f"{dte}d DTE — acceptable")
        elif dte > 0:         score += 5;  reasons.append(f"{dte}d DTE — outside sweet spot")
        if 5 <= otm_pct <= 15:   score += 15; reasons.append(f"{otm_pct:.1f}% OTM — ideal ✓")
        elif 3 <= otm_pct <= 20: score += 8;  reasons.append(f"{otm_pct:.1f}% OTM")
        else:                    reasons.append(f"{otm_pct:.1f}% OTM — too close/far")
        # P/C: high put/call = bearish = bad for CC (stock may drop)
        if pc_ratio >= 1.2:  reasons.append("Bearish P/C — caution on CC")
        elif pc_ratio <= 1.0: score += 10; reasons.append("Neutral/bullish P/C ✓")
        else:                 score += 5

    score = min(score, 100)
    if score >= 75:   label, color = "⭐ Strong",  "#00e676"
    elif score >= 55: label, color = "✅ Good",    "#26a69a"
    elif score >= 35: label, color = "⚠️ Fair",    "#ffb74d"
    else:             label, color = "❌ Weak",    "#ef5350"

    return score, label, color, reasons


NOTIONAL = 100_000  # $100K per trade sizing


def _build_wheel_rows(tickers: List[str], options_data: dict):
    """Extract scored CSP and CC rows without rendering HTML. Used by AI rec tab."""
    csp_rows, cc_rows = [], []
    for t_raw in tickers:
        t = t_raw.lstrip("$")
        d = options_data.get(t, {})
        if not d: continue
        price = d.get("price") or 0
        for exp_row in d.get("csp_by_exp", []):
            dte = exp_row["dte"]; ann_ret = exp_row["ann_ret"]; otm_pct = exp_row["otm_pct"]
            d_copy = {**d, "csp_strike": exp_row["strike"], "csp_premium": exp_row["premium"],
                      "csp_pct": exp_row["pct"], "csp_dte": dte,
                      "csp_ann_ret": ann_ret, "csp_otm_pct": otm_pct, "csp_exp": exp_row["exp"]}
            score, label, s_color, reasons = _wheel_confidence(d_copy, "CSP")
            collateral = exp_row["strike"] * 100
            contracts  = max(1, int(NOTIONAL / collateral))
            csp_rows.append({"ticker": f"${t}", "price": price,
                "strike": exp_row["strike"], "premium": exp_row["premium"],
                "pct": exp_row["pct"], "ann_ret": ann_ret, "dte": dte,
                "exp": exp_row["exp"], "exp_label": exp_row["label"], "otm_pct": otm_pct,
                "contracts": contracts, "total_premium": round(exp_row["premium"]*100*contracts,2),
                "total_coll": round(collateral*contracts,2),
                "breakeven": round(exp_row["strike"]-exp_row["premium"],2),
                "score": score, "label": label, "s_color": s_color, "reasons": reasons,
                "levels": {"entry": price, "stop": round(exp_row["strike"]-exp_row["premium"],2),
                           "pt1": None, "pt2": None}})
        for exp_row in d.get("cc_by_exp", []):
            dte = exp_row["dte"]; ann_ret = exp_row["ann_ret"]; otm_pct = exp_row["otm_pct"]
            d_copy = {**d, "cc_strike": exp_row["strike"], "cc_premium": exp_row["premium"],
                      "cc_pct": exp_row["pct"], "cc_dte": dte,
                      "cc_ann_ret": ann_ret, "cc_otm_pct": otm_pct, "cc_exp": exp_row["exp"]}
            score, label, s_color, reasons = _wheel_confidence(d_copy, "CC")
            shares = max(100, int(NOTIONAL / price / 100) * 100) if price else 100
            contracts = shares // 100
            cc_rows.append({"ticker": f"${t}", "price": price,
                "strike": exp_row["strike"], "premium": exp_row["premium"],
                "pct": exp_row["pct"], "ann_ret": ann_ret, "dte": dte,
                "exp": exp_row["exp"], "exp_label": exp_row["label"], "otm_pct": otm_pct,
                "contracts": contracts, "shares": shares,
                "total_premium": round(exp_row["premium"]*100*contracts,2),
                "breakeven": round(exp_row["strike"]+exp_row["premium"],2),
                "score": score, "label": label, "s_color": s_color, "reasons": reasons})
    csp_rows.sort(key=lambda x: (-x["score"], -(x["ann_ret"] or 0)))
    cc_rows.sort( key=lambda x: (-x["score"], -(x["ann_ret"] or 0)))
    return csp_rows, cc_rows


def build_ai_rec_tab(signals: list, options_data: dict, csp_rows: list,
                     cc_rows: list, mag_rows: list, watch_rows: list) -> str:
    """
    Call Anthropic API with this week's signal + wheel data and return
    3 concrete trade recommendations rendered as an HTML card.
    """
    import json as _json

    # ── Build a compact data payload for the prompt ───────────────────────────
    top_signals = signals[:10]  # best 2-day signals
    top_csp     = csp_rows[:5]
    top_cc      = cc_rows[:5]

    # Price context for mentioned tickers
    all_price_rows = {r["Ticker"].lstrip("$"): r for r in (mag_rows + watch_rows) if r.get("Ticker")}

    sig_lines = []
    for s in top_signals:
        t = s["ticker"]
        lev = s.get("levels", {})
        sig_lines.append(
            f"{t}: {s['signal']} ({s['category']}) — "
            f"bullish={s['bullish']} date={s.get('signal_date','')} "
            f"entry=${lev.get('entry','?')} stop=${lev.get('stop','?')} "
            f"pt1=${lev.get('pt1','?')} pt2=${lev.get('pt2','?')}"
        )

    csp_lines = []
    for r in top_csp:
        csp_lines.append(
            f"{r['ticker']}: Strike ${r['strike']} exp {r['exp']} ({r['dte']}d) "
            f"premium ${r['premium']} ann_ret {r['ann_ret']}% "
            f"score {r['score']}/100 ({r['label']}) breakeven ${r['breakeven']}"
        )

    cc_lines = []
    for r in top_cc:
        cc_lines.append(
            f"{r['ticker']}: Strike ${r['strike']} exp {r['exp']} ({r['dte']}d) "
            f"premium ${r['premium']} ann_ret {r['ann_ret']}% "
            f"score {r['score']}/100 ({r['label']})"
        )

    prompt = f"""You are a professional options and equity trader. Based on the data below, 
recommend exactly 3 trades for this week. Mix trade types: 1 directional (long/short stock or option), 
1 wheel/options income trade, and 1 wildcard (your highest conviction pick of any type).

For EACH trade provide:
- ticker and trade type
- entry range (specific price or range)
- stop loss price  
- profit target 1 (conservative)
- profit target 2 (aggressive)
- position size rationale (assume $100K total capital, suggest % allocation)
- 2-sentence thesis
- key risk to the trade

SIGNALS FIRED THIS WEEK:
{chr(10).join(sig_lines) if sig_lines else 'None'}

TOP CSP OPPORTUNITIES (scored):
{chr(10).join(csp_lines) if csp_lines else 'None'}

TOP CC OPPORTUNITIES (scored):
{chr(10).join(cc_lines) if cc_lines else 'None'}

Respond ONLY with a JSON array of exactly 3 objects with these keys:
ticker, trade_type, direction, entry_low, entry_high, stop, pt1, pt2,
allocation_pct, thesis, key_risk, label (one of: "Directional", "Income", "Wildcard")

No preamble, no markdown fences, pure JSON array only."""

    # ── Call Anthropic API ────────────────────────────────────────────────────
    trades = []
    error_msg = ""
    try:
        import urllib.request
        req_body = _json.dumps({
            "model":      "claude-sonnet-4-20250514",
            "max_tokens": 1000,
            "messages":   [{"role": "user", "content": prompt}]
        }).encode("utf-8")

        import os as _os
        api_key = ANTHROPIC_API_KEY or _os.environ.get("ANTHROPIC_API_KEY", "")
        if not api_key:
            raise ValueError("No Anthropic API key — set ANTHROPIC_API_KEY in export_report.py")
        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=req_body,
            headers={
                "Content-Type":      "application/json",
                "x-api-key":         api_key,
                "anthropic-version": "2023-06-01",
            },
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = _json.loads(resp.read().decode("utf-8"))

        raw = ""
        for block in data.get("content", []):
            if block.get("type") == "text":
                raw += block["text"]

        # Strip any accidental markdown fences
        raw = raw.strip().lstrip("```json").lstrip("```").rstrip("```").strip()
        trades = _json.loads(raw)
        print(f"   AI REC: got {len(trades)} recommendations")
    except Exception as e:
        error_msg = str(e)
        print(f"   AI REC: failed — {e}")

    # ── Render ────────────────────────────────────────────────────────────────
    LABEL_COLORS = {
        "Directional": "#42a5f5",
        "Income":      "#26a69a",
        "Wildcard":    "#FF69B4",
    }
    DIR_COLORS = {"bullish": "#26a69a", "bearish": "#ef5350", "neutral": "#ffb74d"}

    out = ['<div class="card">',
           '<h3>🤖 AI Trade Recommendations <span class="fresh">● generated this run</span></h3>',
           f'<div class="muted" style="margin-bottom:16px">3 trade picks for the week based on signals, '
           f'wheel scores, and price action. Sized to $100K total capital. '
           f'<strong>Not financial advice — always do your own research.</strong></div>']

    if error_msg:
        out.append(f'<div style="color:#ef5350;padding:12px;background:#1a0a0a;border-radius:8px">'
                   f'⚠️ AI call failed: {esc(error_msg)}<br>'
                   f'Make sure the Anthropic API is accessible from your network.</div>')
    elif not trades:
        out.append('<div style="color:#999;padding:12px">No recommendations generated.</div>')
    else:
        for i, tr in enumerate(trades[:3]):
            lbl       = tr.get("label", "Trade")
            lbl_color = LABEL_COLORS.get(lbl, "#aaa")
            direction = str(tr.get("direction","")).lower()
            dir_color = DIR_COLORS.get(direction, "#fff")
            ticker    = tr.get("ticker","")
            ttype     = tr.get("trade_type","")
            alloc     = tr.get("allocation_pct", "")
            alloc_str = f"{alloc}%" if alloc else "—"
            alloc_usd = f"${int(float(alloc or 0) / 100 * 100000):,}" if alloc else ""

            out.append(
                '<div style="background:#0d1b2a;border-left:4px solid ' + lbl_color + ';'
                'border-radius:10px;padding:18px 20px;margin-bottom:18px">'
                '<div style="display:flex;align-items:center;gap:12px;flex-wrap:wrap;margin-bottom:12px">'
                '<span style="background:' + lbl_color + ';color:#000;font-size:11px;font-weight:800;'
                'padding:3px 10px;border-radius:10px">' + esc(lbl) + '</span>'
                f'<span style="font-size:22px;font-weight:900;color:#fff">{esc(ticker)}</span>'
                f'<span style="font-size:15px;color:{dir_color};font-weight:700">{esc(ttype)} — {esc(direction)}</span>'
                f'<span style="font-size:13px;color:#aaa">Allocation: {alloc_str} ({alloc_usd})</span>'
                '</div>'
                '<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:8px;margin-bottom:14px">'
                + "".join([
                    '<div style="background:#13263d;border-radius:7px;padding:9px">'
                    '<div style="font-size:10px;color:#888;margin-bottom:2px">' + lbl2 + '</div>'
                    '<div style="font-size:16px;font-weight:800;color:' + col + '">' + val + '</div>'
                    '</div>'
                    for lbl2, col, val in [
                        ("ENTRY RANGE",  "#fff",     f"${tr.get('entry_low','?')} – ${tr.get('entry_high','?')}"),
                        ("STOP LOSS",    "#ef5350",  f"${tr.get('stop','?')}"),
                        ("TARGET 1",     "#26a69a",  f"${tr.get('pt1','?')}"),
                        ("TARGET 2",     "#00e676",  f"${tr.get('pt2','?')}"),
                    ]
                ])
                + '</div>'
                f'<div style="font-size:13px;line-height:1.8;color:#ccc;margin-bottom:8px">'
                f'<strong style="color:#fff">Thesis:</strong> {esc(str(tr.get("thesis","")))}</div>'
                f'<div style="font-size:12px;color:#ef9a9a">'
                f'⚠️ <strong>Key risk:</strong> {esc(str(tr.get("key_risk","")))}</div>'
                '</div>'
            )

    out.append('</div>')
    return "".join(out)

# ------------------ GitHub publish ------------------
def render_lessons_tab() -> str:
    trades = [
        {
            "ticker": "META", "icon": "⚠️", "outcome_color": "#ffb74d",
            "category": "Early Exit", "cat_color": "#ffb74d",
            "action": "Sold 110 shares @ $735 (Jan 29)",
            "outcome": "Good timing near local top — but permanently exited a high-momentum compounder",
            "what": "META was near a short-term high. Price pulled back to ~$680–700 after the sale so timing was good. However selling shares outright means you need to re-buy higher to get back in.",
            "missed": "META's EMA 10/20 rhythm was fully intact at time of sale — no bearish crossover had fired. No distribution volume. No technical reason to exit the stock entirely.",
            "indicators": [
                ("EMA 10/20 Crossover (Oliver Kell)", "No bearish cross had occurred. A sell signal only fires when EMA 10 closes below EMA 20. Absent that, the trend is intact — hold or use a CC instead of selling."),
                ("Volume / Distribution Day", "A distribution day (unusually high volume on a down close) signals institutions are selling. None was present at time of exit."),
                ("Relative Strength vs S&P", "META was outperforming the broader market — strong RS names should stay in the portfolio until a signal says otherwise."),
            ],
            "lesson": "Never permanently exit a momentum compounder without a technical signal. EMA bearish cross or distribution volume = exit. 'It moved up nicely' is not a sell signal.",
            "better": "Sell a Covered Call at the $735 strike (30–45 DTE) instead. Collect $800–1,200 in premium per contract. If stock is called away at $735 you get your target price AND the premium. If not, you keep both the shares and the premium.",
        },
        {
            "ticker": "MSFT", "icon": "🔴", "outcome_color": "#ef5350",
            "category": "No Signal Entry", "cat_color": "#ef5350",
            "action": "Bought 60 @ $410 (Feb 3) + 60 @ $400 (Feb 5) — averaging into weakness",
            "outcome": "Flat to slight loss — two buys in 3 days into a stock in a confirmed downtrend",
            "what": "MSFT had declined ~15% YTD on AI capex concerns and slowing Azure growth. Both buys happened while the EMA 10 was below EMA 20 — a confirmed downtrend — with no reversal signal in place.",
            "missed": "The signals tracker would have shown a Bearish EMA Cross weeks before these entries. SMA 50 slope was negative. Entering without a reversal confirmation doubles down on a losing position.",
            "indicators": [
                ("EMA 10/20 Bearish Cross (Oliver Kell)", "EMA 10 below EMA 20 = downtrend confirmed. No entries until an EMA Crossback (EMA 10 reclaims above EMA 20) fires on above-average volume."),
                ("SMA 50 Slope (Slope & Crossover)", "Declining 50 SMA slope = price momentum is negative. The Signals tab flags this directly as 'Declining SMA Slope'. Wait for slope to flatten and turn up."),
                ("Pocket Pivot (Gil Morales)", "A pocket pivot above the 10 SMA on strong volume = institutional accumulation signal. This is the entry trigger to wait for — not 'the stock feels cheap'."),
            ],
            "lesson": "Every entry needs a signal trigger. Averaging into a downtrend without confirmation adds risk without edge. Check the Signals tab first — if no bullish signal is present, don't enter.",
            "better": "Sell a CSP at $370–380 strike (10–15% OTM, 30–45 DTE) to collect premium while waiting for MSFT to confirm a reversal. If assigned at $370, your effective cost basis is even lower than current price.",
        },
        {
            "ticker": "MU", "icon": "✅", "outcome_color": "#26a69a",
            "category": "Early Exit (Time-Based)", "cat_color": "#ffb74d",
            "action": "Bought 75 @ $387 (Jan 22) → Sold 75 @ $407 (Jan 27) = $30,534 gain",
            "outcome": "Profitable — but a time-based exit on a momentum name left significant gains on the table",
            "what": "Clean 5-day momentum trade for ~5%. The execution was good — but MU continued higher after the sell. Exiting after 5 days because 'enough time has passed' is not a technical reason to exit.",
            "missed": "No ATR-based profit target was used. PT1 (1.5×ATR) and PT2 (3×ATR) levels from the Signals tab would have defined the hold period by price movement, not by calendar days.",
            "indicators": [
                ("ATR Profit Targets (Signals Tab)", "PT1 = 1.5×ATR above entry (partial exit, lock in gains). PT2 = 3×ATR above entry (full exit). Both now shown on every signal in the tracker. These — not time — should trigger exits."),
                ("EMA 10 as Trail Stop", "Close below EMA 10 = exit signal on a momentum trade. MU's EMA 10 held well above price at time of sale — no exit signal had fired."),
                ("Trailing ATR Stop", "A trailing stop set at 1.5×ATR below the highest close since entry protects gains while riding the move. Moves the stop up as price rises — never down."),
            ],
            "lesson": "Use ATR-based PT1/PT2 targets from the Signals tab, not calendar time. Exit when price hits PT2 or breaks below EMA 10 — whichever comes first. Partial sell at PT1, trail stop on the rest.",
            "better": "On next MU entry: define PT1 and PT2 at entry using the Signals tab levels. Sell 50% at PT1 to lock in gains, move stop to breakeven on remaining 50%, let PT2 or EMA 10 break close it out.",
        },
        {
            "ticker": "NFLX", "icon": "🔴", "outcome_color": "#ef5350",
            "category": "No Sell Signal", "cat_color": "#ef5350",
            "action": "Bought 335 @ $89 (Jan 20) → Sold 335 @ $91.50 (Feb 27) = $30,643 over 38 days",
            "outcome": "2.8% gain over 5+ weeks on a strong momentum stock — significant underperformance vs what the trend offered",
            "what": "NFLX was in a strong uptrend throughout the entire hold period. EMA 10 above EMA 20, rising SMA 50 slope, no distribution days. The stock continued higher after the sale. There was no technical reason to exit.",
            "missed": "38 days in a momentum name with no exit signal = hold the position. The market was rewarding NFLX shareholders — selling removed you from a trend that was fully intact.",
            "indicators": [
                ("EMA 10/20 Rhythm (Oliver Kell)", "As long as EMA 10 > EMA 20 and price is above both EMAs, the trend is intact. NFLX maintained this structure throughout the entire hold. Zero sell signal."),
                ("Voodoo Pullback (Gil Morales)", "Any pullback to the 10 SMA during the hold was a re-entry/add signal, not an exit signal. Price holding the 10 SMA = institutions defending the position."),
                ("Distribution Volume", "Watch for a Distribution Day — unusually high volume on a down close signals institutions are selling. Absent this confirmed signal, hold the position regardless of how long you've been in."),
            ],
            "lesson": "Never sell a momentum stock without a technical reason. '5 weeks is enough' is not a sell signal. EMA 10/20 bearish cross or a distribution day is a sell signal. Let the market tell you when to exit.",
            "better": "Hold until EMA 10/20 bearish cross or 1.5×ATR trailing stop hits. Partial sell at PT1 (1.5×ATR above entry) to de-risk, let the rest run to PT2. Could have been 10–20%+ vs the 2.8% achieved.",
        },
        {
            "ticker": "COST", "icon": "⚠️", "outcome_color": "#ffb74d",
            "category": "Sold Compounder", "cat_color": "#ef5350",
            "action": "Bought 35 @ $889 (Jan 7) → Sold 35 @ $932 (Jan 30) = $32,612",
            "outcome": "Solid trade — but COST is a multi-year compounder that should be a permanent core holding, not a 23-day flip",
            "what": "COST was sold after a $43 move. COST has been in a relentless multi-year uptrend. Selling a compounder on a small move means you'll have to re-buy higher later — a pattern that erodes long-term returns.",
            "missed": "COST's Golden Cross (50 SMA above 200 SMA with both rising) and strong SMA slope are persistent signals the tracker would flag. These aren't 'sell' indicators — they're 'hold forever' indicators.",
            "indicators": [
                ("Golden Cross + SMA Slope (Slope & Crossover)", "COST's 50 SMA above 200 SMA with both MAs rising = textbook multi-year uptrend. The Signals tab flags Golden Cross names. Rule: never sell a golden cross name on a small move."),
                ("Launch Pad Tight Base", "COST regularly forms tight, low-volatility consolidations before breaking to new highs. These are re-entry and add signals — not exits."),
                ("ATR Stop Below 50 SMA", "A 1.5×ATR stop below the rising 50 SMA keeps you in through normal volatility without exiting on routine pullbacks. COST's ATR is low relative to trend — stop would have been very rarely hit."),
            ],
            "lesson": "Separate 'compounder' holdings from 'tactical trades'. COST, META, AMZN are compounders — keep them as core positions. Use the Wheel tab to sell CCs above resistance each month for income without exiting the position.",
            "better": "Keep 35 shares as permanent core. Sell 1x COST CC monthly at 5–8% OTM (30–45 DTE) = ~$500–900 premium per month = $6,000–10,800/year collected while still owning the stock.",
        },
        {
            "ticker": "AMZN", "icon": "🔴", "outcome_color": "#ef5350",
            "category": "No Entry Plan", "cat_color": "#42a5f5",
            "action": "Bought 120 @ $205 (Feb 6) + 120 @ $209 (Feb 24) — $50K+ with no defined stop or target",
            "outcome": "Underwater — two large buys without a signal trigger, stop loss, or profit target defined",
            "what": "AMZN purchased twice in 3 weeks with no entry signal confirmed. No stop was defined before either entry. No profit targets set. A $50K+ position with no exit framework is the highest-risk position in the portfolio.",
            "missed": "Both entries happened while AMZN was trading below its 50 SMA with a declining EMA slope. Neither entry had a technical trigger. Without a pre-defined stop, drawdowns become emotional decisions rather than mechanical ones.",
            "indicators": [
                ("EMA Entry Trigger", "Wait for EMA 10 to cross above EMA 20 on above-average volume before entering. This confirms momentum is shifting — not a dead-cat bounce. No such signal was present at either entry."),
                ("SMA 50 Reclaim", "AMZN closing above its 50 SMA on strong volume = trend resumption signal. That is the entry trigger, not 'it feels cheap at $205'."),
                ("Pre-Defined ATR Stop", "Before entering, calculate 1.5×ATR and set your stop. For AMZN at $205 with ATR ~$7, stop = ~$194.50. If it hits, exit without emotion. This must be defined BEFORE the trade, not after."),
            ],
            "lesson": "Every position needs three things defined BEFORE entry: (1) signal trigger, (2) stop loss at 1.5×ATR, (3) profit targets at PT1 and PT2. Without all three, you're guessing. The Signals tab gives you all three automatically.",
            "better": "Sell a CSP at $185–190 strike (10–15% OTM, 30–45 DTE) to collect premium while waiting for AMZN to confirm a reversal signal. If assigned, your effective cost basis is well below current price with a clear upside target.",
        },
        {
            "ticker": "SPY", "icon": "⚠️", "outcome_color": "#ffb74d",
            "category": "Idle Capital", "cat_color": "#ffb74d",
            "action": "Bought 105 @ $697 (Jan 28) = $73,241 deployed — no CC income being generated",
            "outcome": "Largest position in the portfolio sitting idle — no monthly income offsetting the drawdown",
            "what": "105 shares of SPY at $697 is the biggest single allocation in the statement. With markets pulling back, this position is under pressure with no premium income and no hedge. Dead capital.",
            "missed": "At 105 shares, selling 1x SPY CC per month (100 shares) at 30–45 DTE generates $400–600/month at current IV levels. That's $4,800–7,200/year for doing nothing extra beyond what you already own.",
            "indicators": [
                ("IV Rank on Wheel Tab", "SPY's IV rank drives CC premium. When IV spikes during sell-offs (like now), premiums are elevated — the best time to sell. The Wheel tab shows IV rank in real time."),
                ("DTE Sweet Spot (21–45 days)", "Theta decay accelerates in the last 30 days. Sell a new CC every 21–45 days for maximum income efficiency. The Wheel tab scores and ranks all expirations automatically."),
                ("OTM Buffer (3–8%)", "Sell SPY CC at 3–8% OTM to collect meaningful premium while giving the stock room to move. At $697, that's a $718–752 strike range — the Wheel tab identifies the optimal strike."),
            ],
            "lesson": "Any holding over $10K that isn't generating CC premium monthly is idle capital. The Wheel tab was built exactly for this. Check it every month and sell a CC on SPY, AMZN, and MSFT before the next report runs.",
            "better": "Sell 1x SPY CC monthly (21–45 DTE, 3–8% OTM). At current premiums: ~$400–600/month = $4,800–7,200/year. That income continues regardless of whether the market goes up, sideways, or slightly down.",
        },
        {
            "ticker": "SPYI", "icon": "⭐", "outcome_color": "#00e676",
            "category": "Best Practice", "cat_color": "#00e676",
            "action": "Bought 580 @ $52.005 (Feb 26) + 610 @ $51.765 (Mar 4) — systematic accumulation",
            "outcome": "The best-executed trade in the statement — systematic, income-generating, no emotional selling",
            "what": "SPYI is a high-income ETF that writes covered calls on the S&P internally, targeting ~12% annual distribution yield paid monthly. Two systematic buys at consistent prices = disciplined execution.",
            "missed": "Nothing to improve here. The only addition is to make this a recurring monthly purchase rather than ad-hoc — dollar-cost averaging removes timing risk entirely.",
            "indicators": [
                ("Monthly Distribution Yield (~12%)", "SPYI generates monthly income regardless of market direction via its internal options strategy. This is the income layer that all other positions are missing."),
                ("Dollar Cost Averaging", "Two buys within 1 week at similar prices = good habit. Extend to a monthly recurring buy regardless of price to fully eliminate timing risk."),
                ("Portfolio Income Offset", "SPYI income partially offsets drawdowns on SPY and MSFT positions during down months — a natural hedge without adding complexity."),
            ],
            "lesson": "This is the template for the rest of the portfolio. Systematic, income-generating, no emotional exit. Apply the same discipline to SPY and individual stock positions via the Wheel tab CC strategy every month.",
            "better": "Set a recurring monthly SPYI buy. Consider increasing the allocation relative to single-stock positions to reduce timing risk and increase consistent monthly income.",
        },
    ]

    CAT_COLORS = {
        "Early Exit": "#ffb74d", "Early Exit (Time-Based)": "#ffb74d",
        "No Signal Entry": "#ef5350", "No Sell Signal": "#ef5350",
        "Sold Compounder": "#ef5350", "No Entry Plan": "#42a5f5",
        "Idle Capital": "#ffb74d", "Best Practice": "#00e676",
    }

    good_count    = sum(1 for t in trades if t["outcome_color"] == "#00e676")
    caution_count = sum(1 for t in trades if t["outcome_color"] == "#ffb74d")
    bad_count     = sum(1 for t in trades if t["outcome_color"] == "#ef5350")

    lines = []
    def w(*parts): lines.append("".join(str(p) for p in parts))

    w('<div class="card">')
    w('<h3>📚 Lessons Learned — Portfolio Review</h3>')
    w('<div class="muted" style="margin-bottom:16px">Trade-by-trade analysis of the WE 91563 SHOPE statement. '
      'Each position reviewed against technical indicators, entry/exit discipline, and income generation opportunities.</div>')

    # Summary scorecards
    w('<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:10px;margin-bottom:22px">')
    for lbl, val, col in [
        ("TRADES REVIEWED",  len(trades),      "#fff"),
        ("WELL EXECUTED",    good_count,        "#00e676"),
        ("NEEDS ATTENTION",  caution_count,     "#ffb74d"),
        ("KEY MISTAKES",     bad_count,         "#ef5350"),
        ("INCOME MISSED/MO", "~$2,000+",        "#42a5f5"),
    ]:
        w(f'<div style="background:#13263d;border-radius:8px;padding:12px;border-left:3px solid {col}">')
        w(f'<div style="font-size:11px;color:#aaa;margin-bottom:4px">{lbl}</div>')
        w(f'<div style="font-size:26px;font-weight:900;color:{col}">{val}</div></div>')
    w('</div>')

    # 3 rules callout
    w('<div style="background:#0d1b2a;border-radius:10px;padding:16px 20px;margin-bottom:22px">')
    w('<div style="font-size:14px;font-weight:800;color:#fff;margin-bottom:12px">🏆 The 3 Rules This Portfolio Needs Going Forward</div>')
    w('<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(240px,1fr));gap:10px">')
    for rule_col, title, body in [
        ("#00e676", "Rule 1 — Never sell without a signal",
         "Only exit when: EMA 10/20 bearish cross fires, ATR trailing stop hits, or PT2 target is reached. "
         "Time in trade and gut feel are not sell signals."),
        ("#FFD700", "Rule 2 — Never buy without a signal",
         "Every entry needs: a signal trigger (EMA crossback / pocket pivot), a pre-defined stop (1.5×ATR), "
         "and profit targets (PT1 + PT2). If you can't define all three before clicking buy, don't enter."),
        ("#42a5f5", "Rule 3 — Every holding must earn its keep",
         "Any stock position over $10K that isn't generating CC premium monthly is idle capital. "
         "Use the Wheel tab on SPY, AMZN, and MSFT every single month."),
    ]:
        w(f'<div style="background:#13263d;border-radius:8px;padding:12px">')
        w(f'<div style="font-size:12px;font-weight:800;color:{rule_col};margin-bottom:5px">{rule_col and esc(title)}</div>')
        w(f'<div style="font-size:12px;color:#ccc;line-height:1.7">{esc(body)}</div></div>')
    w('</div></div>')

    # Individual trade cards
    for tr in trades:
        oc = tr["outcome_color"]
        w(f'<div style="background:#0d1b2a;border-left:4px solid {oc};border-radius:10px;padding:18px 20px;margin-bottom:16px">')
        # Header
        w('<div style="display:flex;align-items:center;gap:12px;flex-wrap:wrap;margin-bottom:10px">')
        w(f'<span style="font-size:24px;font-weight:900;color:#fff">{tr["icon"]} {esc(tr["ticker"])}</span>')
        cat_c = CAT_COLORS.get(tr["category"], "#aaa")
        w(f'<span style="background:{cat_c};color:#000;font-size:11px;font-weight:800;padding:3px 10px;border-radius:10px">{esc(tr["category"])}</span>')
        w(f'<span style="font-size:12px;color:#aaa">{esc(tr["action"])}</span>')
        w('</div>')
        # Outcome banner
        w(f'<div style="background:#13263d;border-radius:6px;padding:8px 12px;margin-bottom:12px;'
          f'font-size:13px;color:{oc};font-weight:700">{esc(tr["outcome"])}</div>')
        # Two column body
        w('<div style="display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:12px">')
        # Left col
        w('<div>')
        w('<div style="font-size:11px;color:#aaa;font-weight:800;margin-bottom:5px">WHAT HAPPENED</div>')
        w(f'<div style="font-size:12px;color:#ccc;line-height:1.7;margin-bottom:10px">{esc(tr["what"])}</div>')
        w('<div style="font-size:11px;color:#aaa;font-weight:800;margin-bottom:5px">MISSED OPPORTUNITY</div>')
        w(f'<div style="font-size:12px;color:#ccc;line-height:1.7">{esc(tr["missed"])}</div>')
        w('</div>')
        # Right col — indicators
        w('<div>')
        w('<div style="font-size:11px;color:#aaa;font-weight:800;margin-bottom:5px">INDICATORS THAT WOULD HAVE HELPED</div>')
        for ind_name, ind_body in tr["indicators"]:
            w(f'<div style="margin-bottom:9px">')
            w(f'<div style="font-size:11px;font-weight:800;color:#42a5f5;margin-bottom:2px">📊 {esc(ind_name)}</div>')
            w(f'<div style="font-size:12px;color:#bbb;line-height:1.6">{esc(ind_body)}</div></div>')
        w('</div>')
        w('</div>')
        # Lesson
        w('<div style="background:#0a1628;border-radius:8px;padding:10px 14px;margin-bottom:8px">')
        w('<div style="font-size:11px;color:#aaa;font-weight:800;margin-bottom:4px">📖 LESSON LEARNED</div>')
        w(f'<div style="font-size:13px;color:#fff;line-height:1.7">{esc(tr["lesson"])}</div></div>')
        # Better approach
        w('<div style="background:#0d2a1a;border-radius:8px;padding:10px 14px;border-left:3px solid #00e676">')
        w('<div style="font-size:11px;color:#00e676;font-weight:800;margin-bottom:4px">💡 BETTER APPROACH</div>')
        w(f'<div style="font-size:12px;color:#ccc;line-height:1.7">{esc(tr["better"])}</div></div>')
        w('</div>')

    w('</div>')
    return "\n".join(lines)


def render_ah_lessons_tab() -> str:
    """AH (After Hours / After the Fact) Lessons — live trade post-mortems."""

    trades = [
        {
            "ticker": "META",
            "type": "Cash Secured Put (Sold) — 1 Contract",
            "setup": "Sold 1x META 680P @ $19.65 premium on 3/4/2026 — expiry 3/13/2026 (9 DTE)",
            "roll_date": "03/04/2026",
            "roll_price": "$38.24 (option value at roll — 2× premium received)",
            "pnl": "-$1,760.26 (-89.57%)",
            "pnl_color": "#ef5350",
            "status": "Rolled to new expiry — still exposed at $680 strike",
            "status_color": "#ffb74d",
            "icon": "🔴",
            "context": (
                "Sold 1 contract META 680 Put on 3/4/2026, collecting $19.65/share = $1,965 total premium. "
                "Expiry was 3/13/2026 — only 9 days to expiration. Breakeven at entry: $680 - $19.65 = $660.35. "
                "META was trading near $680 at entry (effectively ATM), then dropped to ~$644 — putting the put "
                "$36 in the money and $16.35 below breakeven. The option's value rose from $19.65 to $38.24 "
                "(2× premium received), triggering the roll on 3/4 with a realised loss of $1,760.26."
            ),
            "what_went_wrong": [
                ("9 DTE is not a CSP — it is a directional bet", (
                    "The sweet spot for CSPs is 21–45 DTE where theta decay is fastest and there is time to be right. "
                    "9 DTE gives the stock almost no room to move before expiry. At 9 DTE, delta is very high on "
                    "near-the-money options — a $5 move in META moves the option $4+. This is speculation, not income selling."
                )),
                ("Strike was effectively at-the-money (0% OTM buffer)", (
                    "A $680 strike on a stock trading at ~$680 = 0% OTM. The Wheel tab scores anything under 3% OTM "
                    "as zero points. Ideal CSP buffer is 5–15% OTM — that would have been a $578–646 strike. "
                    "At $680 you were collecting premium but taking full directional exposure."
                )),
                ("No exit rule defined — option hit 2× premium with no plan", (
                    "The 2× rule: if the option reaches 2× the premium you collected, close it — do not let it run. "
                    "You collected $19.65. At $38.24 (exactly 2× premium) the roll happened, but this should have been "
                    "a pre-defined mechanical exit, not a reactive decision. Define the exit before entry: "
                    "'I will close if this option reaches $39.30 (2× $19.65).'"
                )),
                ("Sold into a confirmed downtrend", (
                    "META's 10-day MA crossed below the 50-day MA on February 19 — two weeks before this entry. "
                    "META had fallen from $796 ATH to $680, making lower highs. Selling a CSP into a downtrend "
                    "means the trend is working against you from day one."
                )),
            ],
            "indicators_missed": [
                ("DTE Warning — 9 days is too short", (
                    "The Wheel tab sweet spot is 21–45 DTE (20 pts in confidence score). 9 DTE scores near zero. "
                    "At 9 DTE, gamma risk is extreme — small price moves cause large option value swings. "
                    "Never sell a CSP with less than 21 DTE on a volatile name like META."
                )),
                ("OTM Buffer — 0% (Wheel tab would score 0 pts)", (
                    "The Wheel tab requires 5–15% OTM for full score. $680 strike / $680 stock = 0% OTM = 0 points. "
                    "The confidence screener would have given this trade a failing score and excluded it from results entirely."
                )),
                ("EMA 10/20 Bearish Cross (Feb 19, 2026)", (
                    "EMA 10 crossed below EMA 20 on February 19 — 13 days before this trade. "
                    "Primary signal on the tracker: bearish cross present = do not sell CSPs. "
                    "Wait for the EMA crossback (EMA 10 reclaims above EMA 20) before selling puts again."
                )),
                ("SMA 50 Slope — Declining", (
                    "META's 50 SMA slope was negative at entry — the Signals tab 'Declining SMA Slope' pattern "
                    "flags this directly. A declining 50 SMA means the intermediate trend is down. "
                    "Selling puts into a declining 50 SMA is fighting the trend."
                )),
                ("2× Premium Stop Rule", (
                    "Pre-define: if premium collected = $19.65, the stop is $39.30 (2× premium). "
                    "When the option hits $39.30, close it — no discussion, no hope, no roll unless the "
                    "thesis has fundamentally changed. This caps the loss at roughly 1× premium net."
                )),
            ],
            "roll_analysis": (
                "The roll was triggered when the option hit ~$38.24 — almost exactly 2× the $19.65 premium. "
                "Rolling buys time but the rolled position is still at the $680 strike, which is now $36 in the money. "
                "META needs to recover from $644 to above $680 — a +5.6% move — just to reach breakeven on the rolled position. "
                "The roll only makes sense if META shows a confirmed reversal signal (EMA crossback, SMA 50 reclaim) "
                "before the new expiry. Without that signal, the roll is deferring the loss, not avoiding it."
            ),
            "lessons": [
                "Never sell a CSP with less than 21 DTE on a volatile stock. 9 DTE is not income selling — it is directional speculation with option leverage.",
                "OTM buffer is non-negotiable: minimum 5%, target 8–12%. A $680 strike on a $680 stock = 0% buffer = full directional exposure. Use the Wheel tab — it only shows you trades with meaningful buffer.",
                "Define the 2× stop BEFORE entering: premium collected × 2 = your hard close price. At $19.65 collected, the stop was $39.30. Set a GTC buy-to-close order at that price the moment you sell.",
                "Check the trend first. EMA 10 below EMA 20 = no CSP selling on that name. Period. Wait for the crossback.",
                "The Wheel tab confidence score would have been near zero on this trade (0% OTM + 9 DTE + bearish slope). If the screener rejects it, so should you.",
            ],
            "better_approach": (
                "The correct META CSP right now: wait for EMA 10 to reclaim above EMA 20 with above-average volume. "
                "Once that fires, sell a 30–45 DTE CSP at the $580–600 strike (8–10% OTM from wherever META is at that point). "
                "Collect premium, set a GTC order to close at 2× premium, and let theta decay work. "
                "The Wheel tab will show this trade in the Strong or Good tier when conditions are right. "
                "Until then — no new META CSPs."
            ),
            "salvage_plan": {
                "heading": "🚑 Salvage Plan — What To Do Right Now",
                "current_price": 644.86,
                "strike": 680.0,
                "itm_amount": 35.14,
                "earnings_date": "April 29, 2026",
                "options": [
                    {
                        "rank": 1,
                        "label": "Roll Down + Out (Recommended)",
                        "label_color": "#00e676",
                        "action": "Buy back $680P, sell new $620–630P at 45–60 DTE (mid-to-late April, before Apr 29 earnings)",
                        "mechanics": (
                            "Rolling down $50–60 AND out 2–3 weeks drops your strike well below current price. "
                            "At $620 strike your breakeven becomes ~$600–605 — META is $45 above that. "
                            "You collect a small net credit or pay a small debit depending on IV. "
                            "Keeps the trade alive with real OTM buffer for the first time."
                        ),
                        "if_works": "META holds above $620 through April expiry → keep the premium, close the position at a small net loss vs full loss.",
                        "if_fails": "META breaks below $620 support → close it, take the defined loss, don't roll again.",
                        "key_risk": "Earnings Apr 29 — make sure your expiry is BEFORE Apr 29. A bad earnings gap could make this unrecoverable.",
                    },
                    {
                        "rank": 2,
                        "label": "Close the Position Now",
                        "label_color": "#ffb74d",
                        "action": "Buy back the rolled $680P at market, realise the full loss, move on",
                        "mechanics": (
                            "META is sitting on its weekly MA100 — the LAST major support before a potential deeper 2022-style correction. "
                            "If that level breaks (~$636), META could fall to $580–600 rapidly. "
                            "Closing now caps the loss at ~$1,760 realised. "
                            "Waiting risks the loss growing to $3,000–5,000+ if META deteriorates further."
                        ),
                        "if_works": "You free up $68,000 in collateral, take a defined $1,760 loss, and redeploy into a better setup when the signal is right.",
                        "if_fails": "N/A — closing is the definitive loss-capping move.",
                        "key_risk": "If META reverses sharply higher from here you will have closed at the bottom. But you can re-enter at a better strike when the signal confirms.",
                    },
                    {
                        "rank": 3,
                        "label": "Do NOT Roll at $680 Again",
                        "label_color": "#ef5350",
                        "action": "Do not roll the same $680 strike to a later date",
                        "mechanics": (
                            "Rolling at $680 in a downtrend with earnings Apr 29 approaching is the worst option. "
                            "You collect a small credit but remain fully exposed to a continued move lower AND earnings risk. "
                            "A bad Apr 29 earnings reaction could gap META $30–50 lower — with a $680 strike you would be catastrophically ITM."
                        ),
                        "if_works": "N/A",
                        "if_fails": "META gaps down on earnings with $680 strike in place = loss of $3,500–5,000+",
                        "key_risk": "Earnings volatility at a deep ITM strike is uncontrollable risk.",
                    },
                ],
                "watch_levels": [
                    ("$660–675", "Resistance zone — $680 CSP only safe if META reclaims this level with volume"),
                    ("$636",     "Weekly MA100 support — last line before deeper correction. Break here = close the trade"),
                    ("$620",     "Target roll-down strike — meaningful buffer from current price"),
                    ("$580–600", "Next major support if weekly MA100 breaks — where a new CSP makes sense AFTER signal confirms"),
                ],
            },
            "current_situation": (
                "META at $644.86 (Mar 8), down from $660.57 the prior day. Day range $636–$649. "
                "The $680 CSP is ~$35 ITM. Breakeven $660.35 — META is $15.50 below that. "
                "Earnings April 29 — a wildcard that could gap the stock $30–50 either way. "
                "META is sitting on its weekly MA100 — the last major technical support before a potential deeper correction. "
                "61 analysts rate Strong Buy with $862 avg price target — fundamentals are solid, this is macro/sentiment driven."
            ),
        },
    ]

    lines = []
    def w(*parts): lines.append("".join(str(p) for p in parts))

    w('<div class="card">')
    w('<h3>🔬 AH Lessons — Live Trade Post-Mortems</h3>')
    w('<div class="muted" style="margin-bottom:16px">After-the-fact analysis of real trades. '
      'What the indicators were showing, what went wrong, and what to do differently next time. '
      'Added as trades are reviewed — this is a living record.</div>')

    for tr in trades:
        oc = tr["pnl_color"]
        sc = tr["status_color"]

        # Card wrapper
        w(f'<div style="background:#0d1b2a;border-left:4px solid {oc};border-radius:12px;padding:20px 22px;margin-bottom:20px">')

        # Header
        w('<div style="display:flex;align-items:center;gap:12px;flex-wrap:wrap;margin-bottom:14px">')
        w(f'<span style="font-size:26px;font-weight:900;color:#fff">{tr["icon"]} {esc(tr["ticker"])}</span>')
        w(f'<span style="background:#1a2a3a;color:#42a5f5;font-size:12px;font-weight:700;padding:4px 12px;border-radius:10px">{esc(tr["type"])}</span>')
        w(f'<span style="background:{oc};color:#000;font-size:12px;font-weight:800;padding:4px 12px;border-radius:10px">{esc(tr["pnl"])}</span>')
        w(f'<span style="background:{sc};color:#000;font-size:11px;font-weight:700;padding:3px 10px;border-radius:10px">{esc(tr["status"])}</span>')
        w('</div>')

        # Setup summary strip
        w('<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(170px,1fr));gap:8px;margin-bottom:16px">')
        for lbl, val, col in [
            ("TRADE SETUP",   tr["setup"],       "#aaa"),
            ("ROLLED ON",     tr["roll_date"],   "#aaa"),
            ("OPTION VALUE",  tr["roll_price"],  "#ffb74d"),
            ("P&L",           tr["pnl"],         oc),
            ("OUTCOME",       tr["status"],      sc),
        ]:
            w(f'<div style="background:#13263d;border-radius:7px;padding:9px 11px">')
            w(f'<div style="font-size:10px;color:#777;margin-bottom:3px">{lbl}</div>')
            w(f'<div style="font-size:12px;font-weight:700;color:{col}">{esc(val)}</div></div>')
        w('</div>')

        # Context
        w('<div style="background:#0a1220;border-radius:8px;padding:12px 14px;margin-bottom:14px">')
        w('<div style="font-size:11px;color:#aaa;font-weight:800;margin-bottom:5px">📋 WHAT HAPPENED</div>')
        w(f'<div style="font-size:13px;color:#ccc;line-height:1.8">{esc(tr["context"])}</div></div>')

        # Two column: what went wrong + indicators missed
        w('<div style="display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:14px">')

        # Left: What went wrong
        w('<div>')
        w('<div style="font-size:11px;color:#ef5350;font-weight:800;margin-bottom:8px">❌ WHAT WENT WRONG</div>')
        for title, body in tr["what_went_wrong"]:
            w(f'<div style="background:#1a0d0d;border-left:3px solid #ef5350;border-radius:6px;padding:9px 11px;margin-bottom:8px">')
            w(f'<div style="font-size:11px;font-weight:800;color:#ef9a9a;margin-bottom:3px">{esc(title)}</div>')
            w(f'<div style="font-size:12px;color:#ccc;line-height:1.6">{esc(body)}</div></div>')
        w('</div>')

        # Right: Indicators missed
        w('<div>')
        w('<div style="font-size:11px;color:#42a5f5;font-weight:800;margin-bottom:8px">📊 INDICATORS THAT WOULD HAVE STOPPED THIS TRADE</div>')
        for title, body in tr["indicators_missed"]:
            w(f'<div style="background:#0d1a2a;border-left:3px solid #42a5f5;border-radius:6px;padding:9px 11px;margin-bottom:8px">')
            w(f'<div style="font-size:11px;font-weight:800;color:#90caf9;margin-bottom:3px">{esc(title)}</div>')
            w(f'<div style="font-size:12px;color:#ccc;line-height:1.6">{esc(body)}</div></div>')
        w('</div>')
        w('</div>')

        # Roll analysis
        w('<div style="background:#1a1200;border-left:3px solid #ffb74d;border-radius:8px;padding:12px 14px;margin-bottom:12px">')
        w('<div style="font-size:11px;color:#ffb74d;font-weight:800;margin-bottom:5px">🔄 ROLL ANALYSIS</div>')
        w(f'<div style="font-size:12px;color:#ccc;line-height:1.8">{esc(tr["roll_analysis"])}</div></div>')

        # Current situation
        w('<div style="background:#0d1b2a;border:1px solid #263238;border-radius:8px;padding:12px 14px;margin-bottom:12px">')
        w('<div style="font-size:11px;color:#aaa;font-weight:800;margin-bottom:5px">📍 CURRENT SITUATION</div>')
        w(f'<div style="font-size:12px;color:#ccc;line-height:1.8">{esc(tr["current_situation"])}</div></div>')

        # Lessons list
        w('<div style="background:#0a1628;border-radius:8px;padding:12px 14px;margin-bottom:12px">')
        w('<div style="font-size:11px;color:#fff;font-weight:800;margin-bottom:8px">📖 LESSONS LEARNED</div>')
        for i, lesson in enumerate(tr["lessons"], 1):
            w(f'<div style="display:flex;gap:8px;margin-bottom:7px">')
            w(f'<span style="background:#263238;color:#42a5f5;font-size:11px;font-weight:800;'
              f'padding:2px 7px;border-radius:10px;white-space:nowrap;height:fit-content">{i}</span>')
            w(f'<div style="font-size:12px;color:#ccc;line-height:1.7">{esc(lesson)}</div></div>')
        w('</div>')

        # Better approach
        w('<div style="background:#0d2a1a;border-left:3px solid #00e676;border-radius:8px;padding:12px 14px;margin-bottom:12px">')
        w('<div style="font-size:11px;color:#00e676;font-weight:800;margin-bottom:5px">💡 BETTER APPROACH NEXT TIME</div>')
        w(f'<div style="font-size:12px;color:#ccc;line-height:1.8">{esc(tr["better_approach"])}</div></div>')

        # Salvage plan (if present)
        sp = tr.get("salvage_plan")
        if sp:
            w(f'<div style="background:#0a1628;border:1px solid #263238;border-radius:10px;padding:16px 18px;margin-top:4px">')
            w(f'<div style="font-size:14px;font-weight:900;color:#fff;margin-bottom:4px">{esc(sp["heading"])}</div>')
            w(f'<div style="font-size:12px;color:#aaa;margin-bottom:14px">'
              f'META @ <strong style="color:#ef5350">${sp["current_price"]}</strong> &nbsp;|&nbsp; '
              f'Strike <strong style="color:#ef5350">${sp["strike"]:.0f}</strong> &nbsp;|&nbsp; '
              f'<strong style="color:#ef5350">${sp["itm_amount"]:.2f} ITM</strong> &nbsp;|&nbsp; '
              f'Earnings <strong style="color:#ffb74d">{esc(sp["earnings_date"])}</strong></div>')

            # Option cards
            for opt in sp["options"]:
                lc = opt["label_color"]
                w(f'<div style="background:#0d1b2a;border-left:3px solid {lc};border-radius:8px;padding:12px 14px;margin-bottom:10px">')
                w(f'<div style="display:flex;align-items:center;gap:10px;margin-bottom:8px">')
                w(f'<span style="background:{lc};color:#000;font-size:11px;font-weight:800;padding:2px 9px;border-radius:8px">Option {opt["rank"]}</span>')
                w(f'<span style="font-size:13px;font-weight:800;color:{lc}">{esc(opt["label"])}</span>')
                w(f'</div>')
                w(f'<div style="font-size:12px;color:#FFD700;font-weight:700;margin-bottom:5px">{esc(opt["action"])}</div>')
                w(f'<div style="font-size:12px;color:#ccc;line-height:1.7;margin-bottom:8px">{esc(opt["mechanics"])}</div>')
                w(f'<div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;font-size:11px">')
                if opt["if_works"] != "N/A":
                    w(f'<div style="background:#0d2a1a;border-radius:5px;padding:7px 9px">'
                      f'<div style="color:#00e676;font-weight:800;margin-bottom:2px">✅ If it works</div>'
                      f'<div style="color:#bbb">{esc(opt["if_works"])}</div></div>')
                w(f'<div style="background:#2a0d0d;border-radius:5px;padding:7px 9px">'
                  f'<div style="color:#ef5350;font-weight:800;margin-bottom:2px">⚠️ Key risk</div>'
                  f'<div style="color:#bbb">{esc(opt["key_risk"])}</div></div>')
                w(f'</div></div>')

            # Watch levels
            w(f'<div style="margin-top:12px">')
            w(f'<div style="font-size:11px;color:#aaa;font-weight:800;margin-bottom:7px">📍 KEY PRICE LEVELS TO WATCH</div>')
            w(f'<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:8px">')
            for level, desc in sp["watch_levels"]:
                w(f'<div style="background:#13263d;border-radius:6px;padding:8px 10px">')
                w(f'<div style="font-size:13px;font-weight:800;color:#FFD700">{esc(level)}</div>')
                w(f'<div style="font-size:11px;color:#bbb;line-height:1.6;margin-top:2px">{esc(desc)}</div></div>')
            w(f'</div></div>')
            w(f'</div>')  # end salvage plan

        w('</div>')  # end card

    # Footer note
    w('<div style="background:#13263d;border-radius:8px;padding:12px 16px;margin-top:8px;font-size:12px;color:#777;line-height:1.8">')
    w('💬 <strong style="color:#aaa">Add new trades to this tab</strong> by updating the <code>render_ah_lessons_tab()</code> function in export_report.py. '
      'Each post-mortem should include: setup details, what indicators were showing, what the mistake was, and the corrected approach.')
    w('</div>')

    w('</div>')
    return "\n".join(lines)




# ── Kyle's portfolio context (used by weekly picks) ───────────────────────────
KYLE_PORTFOLIO = {
    "current_positions": [
        {"ticker": "AMZN", "shares": 495, "cost_basis": 213.21, "market_val": 105_539,
         "note": "Buying 5 more shares NOW to reach 500 (1 more contract). Do not wait for SPY sale."},
        {"ticker": "MSFT", "shares": 287, "cost_basis": 408.96, "market_val": 117_372,
         "note": "Buying 13 more shares NOW to reach 300. Do not wait for SPY sale."},
        {"ticker": "SPY",  "shares": None, "cost_basis": None,  "market_val": 100_000,
         "note": "SELLING — submit sell order in parallel, do not wait for this to clear before buying AMZN/MSFT"},
    ],
    "open_options": [
        {"ticker": "META", "type": "CSP", "strike": 680, "status": "ROLLED — ITM",
         "note": "Salvage plan active: roll down to $620 Apr 17 or close"},
    ],
    "transition_plan": (
        "Buy AMZN and MSFT shares TODAY — do not wait for SPY to sell first. "
        "Submit the SPY sell order at the same time. "
        "AMZN: buy 5 shares now → 500 total (unlocks 5th contract). "
        "MSFT: buy 13 shares now → 300 total (unlocks 3rd contract). "
        "SPY proceeds (~$100K) replenish cash after. "
        "Sequence: (1) buy AMZN top-up, (2) buy MSFT top-up, (3) sell SPY, (4) sell covered calls on new totals."
    ),
    "contracts_after_topup": {"AMZN": 5, "MSFT": 3},
    "total_portfolio_value": 322_911,
    "notional": 100_000,
    "weekly_income_target": 3_000,
}


def fetch_kyles_positions_live() -> dict:
    """Fetch live CC option chains for AMZN and MSFT from yfinance. 30-min cache."""
    import json, os, time, datetime
    CACHE_FILE = ".cache_kyles_positions.json"
    CACHE_TTL  = 1800  # 30 minutes

    # Try cache first
    try:
        if os.path.exists(CACHE_FILE):
            with open(CACHE_FILE) as cf:
                cached = json.load(cf)
            if time.time() - cached.get("_ts", 0) < CACHE_TTL:
                print("   (using cached Kyle CC data)")
                return cached
    except Exception:
        pass

    print("   Fetching live options chains from yfinance...")
    result = {}
    try:
        import yfinance as yf
        today = datetime.date.today()
        POSITIONS = {"AMZN": {"shares": 500, "cost_basis": 213.21},
                     "MSFT": {"shares": 300, "cost_basis": 408.96}}

        for tk, pos in POSITIONS.items():
            try:
                t         = yf.Ticker(tk)
                price     = 0
                try:
                    price = float(t.fast_info.get("last_price") or t.fast_info.get("regularMarketPrice") or 0)
                except Exception:
                    pass
                if not price:
                    hist  = t.history(period="2d")
                    price = float(hist["Close"].iloc[-1]) if not hist.empty else 0

                contracts_avail = pos["shares"] // 100
                exps_raw = t.options or []
                exp_blocks = []

                for exp_str in exps_raw[:6]:
                    try:
                        exp_dt = datetime.datetime.strptime(exp_str, "%Y-%m-%d").date()
                        dte    = (exp_dt - today).days
                        if dte < 14 or dte > 60: continue
                        chain  = t.option_chain(exp_str)
                        calls  = chain.calls
                        rows_out = []
                        for _, row in calls.iterrows():
                            strike  = float(row["strike"])
                            if price <= 0: continue
                            otm_pct = (strike - price) / price * 100
                            if otm_pct < 1 or otm_pct > 15: continue
                            bid  = float(row.get("bid") or 0)
                            ask  = float(row.get("ask") or 0)
                            mid  = round((bid + ask) / 2, 2) if bid and ask else bid or ask
                            if mid <= 0 and bid <= 0: continue
                            last = float(row.get("lastPrice") or 0)
                            iv   = float(row.get("impliedVolatility") or 0) * 100
                            vol  = int(row.get("volume") or 0)
                            oi   = int(row.get("openInterest") or 0)
                            ann  = round(mid / price * (365 / max(dte, 1)) * 100, 1) if price else 0
                            be   = round(pos["cost_basis"] - mid, 2)
                            gic  = round((strike - pos["cost_basis"] + mid) * contracts_avail * 100, 2)
                            rows_out.append({
                                "strike": strike, "bid": bid, "ask": ask, "mid": mid,
                                "last": last, "otm_pct": round(otm_pct, 1), "iv_pct": round(iv, 1),
                                "volume": vol, "oi": oi, "contracts": contracts_avail,
                                "total_prem": round(mid * 100 * contracts_avail, 2),
                                "breakeven": be, "ann_ret": ann, "gain_if_called": gic,
                            })
                        if rows_out:
                            exp_blocks.append({"exp": exp_str, "dte": dte, "rows": rows_out})
                    except Exception:
                        continue

                result[tk] = {"price": price, "shares": pos["shares"],
                              "cost_basis": pos["cost_basis"], "expirations": exp_blocks}
            except Exception as e:
                result[tk] = {"price": 0, "shares": pos["shares"],
                              "cost_basis": pos["cost_basis"], "expirations": [], "error": str(e)}

        result["_ts"] = time.time()
        try:
            with open(CACHE_FILE, "w") as cf:
                json.dump(result, cf)
        except Exception:
            pass
    except Exception as e:
        print(f"   WARNING: fetch_kyles_positions_live failed: {e}")

    return result


def render_kyles_cc_tab(live_data: dict = None) -> str:
    """Kyle's Covered Calls — two tables: active tracker + trade/PNL log."""
    import datetime, json as _jk, urllib.request as _rur
    today = datetime.date.today()

    _COLORS = {"AMZN": "#FF9900", "MSFT": "#00a4ef"}
    _LIVE   = '<span style="background:#1a5c2a;color:#00e676;font-size:9px;font-weight:800;padding:1px 5px;border-radius:3px;margin-left:3px">&#10003; LIVE</span>'
    _ASMD   = '<span style="background:#b45309;color:#fff;font-size:9px;font-weight:800;padding:1px 5px;border-radius:3px;margin-left:3px">&#9888; EST</span>'

    # Earnings dates — expiries on or after these are dangerous (IV crush + gap risk)
    EARNINGS = {
        "AMZN": datetime.date(2026, 4, 30),  # unconfirmed ~Apr 30 after market
        "MSFT": datetime.date(2026, 4, 28),  # confirmed Apr 28 after close
    }

    def _earnings_safe(tk, exp_yf):
        """True if the expiry is safely before earnings."""
        e = EARNINGS.get(tk)
        if not e or not exp_yf: return True
        try:
            return datetime.datetime.strptime(exp_yf, "%Y-%m-%d").date() < e
        except Exception: return True

    def _earnings_warn(tk, exp_yf):
        """Warning string if expiry crosses or is very close to earnings."""
        e = EARNINGS.get(tk)
        if not e or not exp_yf: return ""
        try:
            exp_d = datetime.datetime.strptime(exp_yf, "%Y-%m-%d").date()
            days_b = (e - exp_d).days
            if exp_d >= e:
                return tk + " earnings " + e.strftime("%b %d") + " — expiry CROSSES earnings. Avoid."
            if days_b <= 3:
                return "Only " + str(days_b) + "d before " + tk + " earnings " + e.strftime("%b %d") + " — risky."
            return ""
        except Exception: return ""


    # ── Live stock prices ─────────────────────────────────────────────────────
    _live_px = {}
    try:
        import yfinance as _yf_k
        for _tk in ["AMZN","MSFT"]:
            try:
                _fi = _yf_k.Ticker(_tk).fast_info
                _p  = float(_fi.get("last_price") or _fi.get("regularMarketPrice") or 0)
                if not _p:
                    _h = _yf_k.Ticker(_tk).history(period="2d")
                    _p = float(_h["Close"].iloc[-1]) if not _h.empty else 0
                _live_px[_tk] = round(_p, 2) if _p else None
            except Exception:
                _live_px[_tk] = None
    except Exception:
        pass
    for _tk in ["AMZN","MSFT"]:
        if not _live_px.get(_tk):
            _ld_p = (live_data or {}).get(_tk,{}).get("price")
            if _ld_p: _live_px[_tk] = float(_ld_p)

    # ── Full trade history ────────────────────────────────────────────────────
    CC_TRADES = [
        # Round 1 — Mar 10 initial sells (CLOSED Mar 16)
        {"date":"2026-03-10","ticker":"MSFT","strike":417.0,"expiry":"2026-04-17",
         "contracts":2,"premium":9.415,"total_premium":1883.0,"exp_yf":"2026-04-17",
         "closed":True,"closed_date":"2026-03-16","closed_at":4.43,"notes":"WE 91563"},
        {"date":"2026-03-10","ticker":"AMZN","strike":220.0,"expiry":"2026-04-02",
         "contracts":3,"premium":4.60,"total_premium":1380.0,"exp_yf":"2026-04-02",
         "closed":True,"closed_date":"2026-03-16","closed_at":1.81,"notes":"WE 91563"},
        {"date":"2026-03-10","ticker":"AMZN","strike":220.0,"expiry":"2026-04-02",
         "contracts":1,"premium":4.50,"total_premium":450.0,"exp_yf":"2026-04-02",
         "closed":True,"closed_date":"2026-03-16","closed_at":1.85,"notes":"WE 91583"},
        # Round 2 — Mar 16 rolls (OPEN — Apr 10 expiry)
        # AMZN: -5 contracts total @ avg ~$4.46, current $7.10 per brokerage screenshot
        # NOTE: $1.33 in screenshot was the day CHANGE, not the price
        {"date":"2026-03-16","ticker":"AMZN","strike":None,"expiry":"2026-04-10",
         "contracts":2,"premium":4.45,"total_premium":880.00,"exp_yf":"2026-04-10",
         "closed":True,"closed_date":"2026-04-10","closed_at":0.05,
         "brokerage_current":None,"notes":"WE 91563 — bought back Apr 10"},
        {"date":"2026-03-16","ticker":"AMZN","strike":None,"expiry":"2026-04-10",
         "contracts":1,"premium":4.42,"total_premium":442.0,"exp_yf":"2026-04-10",
         "closed":True,"closed_date":"2026-04-10","closed_at":0.05,
         "brokerage_current":None,"notes":"WE 91563 — bought back Apr 10"},
        {"date":"2026-03-16","ticker":"AMZN","strike":None,"expiry":"2026-04-10",
         "contracts":1,"premium":4.52,"total_premium":437.80,"exp_yf":"2026-04-10",
         "closed":True,"closed_date":"2026-04-10","closed_at":0.05,
         "brokerage_current":None,"notes":"WE 91583 — bought back Apr 10"},
        # Mar 20 — 1x AMZN closed at 50%+ profit ($2.645 vs $4.45 sold = 40.7% gain)
        {"date":"2026-03-16","ticker":"AMZN","strike":None,"expiry":"2026-04-10",
         "contracts":1,"premium":4.45,"total_premium":440.00,"exp_yf":"2026-04-10",
         "closed":True,"closed_date":"2026-03-20","closed_at":2.645,
         "brokerage_current":None,"notes":"WE 91563 — closed 1x Mar 20 (~41% profit)"},
        # MSFT: -2 contracts, current $7.96 per brokerage screenshot
        {"date":"2026-03-16","ticker":"MSFT","strike":405.0,"expiry":"2026-04-10",
         "contracts":2,"premium":8.010,"total_premium":1587.80,"exp_yf":"2026-04-10",
         "closed":True,"closed_date":"2026-04-10","closed_at":0.05,
         "brokerage_current":None,"notes":"WE 91563 — bought back Apr 10"},
    ]

    OPEN_TRADES   = [t for t in CC_TRADES if not t.get("closed",False)]
    CLOSED_TRADES = [t for t in CC_TRADES if t.get("closed",False)]

    # Financials
    _gross_sells = sum(t["total_premium"] for t in CC_TRADES)
    _gross_buybk = sum(t.get("buyback_total") or 0 for t in CC_TRADES)
    TOTAL_NET_PREMIUM = round(_gross_sells - _gross_buybk, 2)

    # Share purchases
    SHARE_PURCHASES = [
        {"date":"2026-03-10","ticker":"AMZN","shares":14,"price_paid":None,
         "funded_by":"CC premium (Mar 10)","notes":"14 sh bought with Mar 10 CC premium"},
        {"date":"2026-03-18","ticker":"AMZN","shares":15,"price_paid":213.829,
         "funded_by":"Cash / CC income (WE 91563)","notes":"15 sh @ $213.829"},
        {"date":"2026-03-18","ticker":"AMZN","shares":10,"price_paid":213.880,
         "funded_by":"Cash / CC income (WE 91583)","notes":"10 sh @ $213.880"},
    ]
    try:
        import yfinance as _yf_sp
        _d = _yf_sp.Ticker("AMZN").history(start="2026-03-10",end="2026-03-11")
        if not _d.empty:
            SHARE_PURCHASES[0]["price_paid"] = round(float(_d["Close"].iloc[-1]),2)
    except Exception:
        pass

    # ── Fetch live option chain for open positions ────────────────────────────
    _opt_mids = {}
    _roll_chains_live = {}
    try:
        import yfinance as _yf_opt
        for _tk in set(t["ticker"] for t in OPEN_TRADES):
            try:
                _tc = _yf_opt.Ticker(_tk)
                _tc.history(period="5d")
                _px  = _live_px.get(_tk, 0) or 0
                _opt_mids[_tk]        = {}
                _roll_chains_live[_tk] = []
                for _es in (list(_tc.options or []))[:10]:
                    try:
                        _ed  = datetime.datetime.strptime(_es,"%Y-%m-%d").date()
                        _dte = (_ed - today).days
                        _ch  = _tc.option_chain(_es).calls
                        # Store mid for every strike
                        for _, _r in _ch.iterrows():
                            _b = float(_r.get("bid") or 0); _a = float(_r.get("ask") or 0)
                            _m = round((_b+_a)/2,2) if _b and _a else 0
                            if _m > 0:
                                _opt_mids[_tk][(_es, float(_r["strike"]))] = {"mid":_m,"bid":_b,"ask":_a,"dte":_dte}
                        # Roll candidates: 25-50 DTE, 4-10% OTM, real bid+ask
                        if 25 <= _dte <= 50 and _px > 0:
                            _rrows = []
                            for _, _r in _ch.iterrows():
                                _st  = float(_r["strike"])
                                _otm = (_st - _px) / _px * 100
                                if _otm < 1 or _otm > 15: continue  # wide range to find credit rolls
                                _b = float(_r.get("bid") or 0); _a = float(_r.get("ask") or 0)
                                if _b <= 0 or _a <= 0: continue
                                _m   = round((_b+_a)/2,2)
                                _ann = round(_m/_px*(365/max(_dte,1))*100,1)
                                _rrows.append({"strike":_st,"bid":_b,"ask":_a,"mid":_m,
                                               "otm_pct":round(_otm,1),"ann_ret":_ann,
                                               "dte":_dte,"exp":_es,
                                               "total_prem":round(_m*100,2)})
                            if _rrows:
                                _rrows.sort(key=lambda r: r["ann_ret"], reverse=True)
                                _roll_chains_live[_tk].append({"exp":_es,"dte":_dte,"rows":_rrows})
                    except Exception:
                        continue
            except Exception:
                pass
    except Exception:
        pass

    def _get_mid(tk, exp_yf, strike):
        if not strike or not exp_yf: return None
        d = _opt_mids.get(tk,{})
        if not d: return None
        if (exp_yf,strike) in d: return d[(exp_yf,strike)]["mid"]
        best = min(d.keys(), key=lambda k: abs(k[1]-strike) if k[0]==exp_yf else 999, default=None)
        if best and best[0]==exp_yf and abs(best[1]-strike)<=1: return d[best]["mid"]
        return None

    # Get IV for theta calc
    def _get_iv(tk, exp_yf, strike):
        if not strike or not exp_yf: return None
        d = _opt_mids.get(tk,{})
        key = (exp_yf, strike) if (exp_yf,strike) in d else None
        if key: return None  # we don't store IV separately, fall back
        return None

    # ── Per-position: compute action, theta decay, roll suggestion ────────────
    # Group open trades by ticker+expiry for combined analysis
    _pos_groups = {}
    for _t in OPEN_TRADES:
        _k = _t.get("ticker","")
        if not _k: continue
        if _k not in _pos_groups:
            _pos_groups[_k] = {"contracts":0,"total_sold":0,"weighted_prem":0,
                                "expiry":  _t.get("expiry",  _t.get("exp_yf", "")),
                                "exp_yf":  _t.get("exp_yf",  ""),
                                "strike":  _t.get("strike"),
                                "ticker":  _k,
                                "brokerage_current": _t.get("brokerage_current")}
        _pos_groups[_k]["contracts"]     += _t.get("contracts", 0)
        _pos_groups[_k]["total_sold"]    += _t.get("total_premium", 0)
        _pos_groups[_k]["weighted_prem"] += _t.get("premium", 0) * _t.get("contracts", 0)

    for _k, _pg in _pos_groups.items():
        _pg["avg_prem"] = round(_pg["weighted_prem"] / _pg["contracts"], 3) if _pg["contracts"] else 0

    # Max theta decay: typically 2-3 weeks before expiry (7-21 DTE) for short-dated options
    # For a CC seller the BEST theta decay is actually 30→21 DTE window
    # After 21 DTE gamma accelerates — that's when to close/roll
    def _theta_analysis(exp_yf, avg_prem):
        try:
            _exp_d = datetime.datetime.strptime(exp_yf,"%Y-%m-%d").date()
            _dte   = (_exp_d - today).days
            _day21 = _exp_d - datetime.timedelta(days=21)
            _days_to_21 = (_day21 - today).days
            return {"dte":_dte, "day21":_day21.strftime("%b %d"), "days_to_21":_days_to_21,
                    "decay_pct": round((1 - (_dte/45)**0.5) * 100, 0) if _dte <= 45 else 0}
        except Exception:
            return {"dte":0,"day21":"—","days_to_21":0,"decay_pct":0}

    # Best roll per ticker (top ann_ret in 30-45 DTE window, else best available)
    def _best_roll(tk, current_mid, contracts):
        chains  = _roll_chains_live.get(tk,[])
        buyback = current_mid or 0
        # Flatten all rows, sort by net credit desc (new_prem - buyback)
        all30_45 = sorted([r for c in chains if 30<=c["dte"]<=45 for r in c["rows"]],
                           key=lambda r: r["mid"]-buyback, reverse=True)
        all_any  = sorted([r for c in chains for r in c["rows"]],
                           key=lambda r: r["mid"]-buyback, reverse=True)
        best = all30_45[0] if all30_45 else (all_any[0] if all_any else None)
        if best:
            # Net = new premium collected MINUS cost to buy back current position
            # Positive = net credit (you receive cash to roll)
            # Negative = net debit (rolling costs you money — current calls too expensive)
            buyback_cost = current_mid or 0
            new_prem     = best["mid"]
            _net = round((new_prem - buyback_cost) * 100 * contracts, 2)
            best["net_credit"]      = _net
            best["buyback_cost"]    = round(buyback_cost * 100 * contracts, 2)
            best["new_prem_total"]  = round(new_prem * 100 * contracts, 2)
            # Label: credit vs debit
            if _net >= 0:
                best["net_credit_str"]   = f"+${_net:,.0f} credit"
                best["net_credit_label"] = "NET CREDIT"
                best["net_credit_color"] = "#00e676"
                best["net_explain"]      = (
                    f"Collect ${new_prem:.2f} new × {contracts*100} shares = "
                    f"${best['new_prem_total']:,.0f}, "
                    f"pay ${buyback_cost:.2f} × {contracts*100} = "
                    f"${best['buyback_cost']:,.0f} to close. "
                    f"Net cash IN: +${_net:,.0f}."
                )
            else:
                best["net_credit_str"]   = f"${_net:,.0f} debit"
                best["net_credit_label"] = "NET DEBIT"
                best["net_credit_color"] = "#ef5350"
                best["net_explain"]      = (
                    f"Closing costs ${buyback_cost:.2f} × {contracts*100} = "
                    f"${best['buyback_cost']:,.0f}, "
                    f"but new premium is only ${new_prem:.2f} × {contracts*100} = "
                    f"${best['new_prem_total']:,.0f}. "
                    f"Rolling NOW costs ${abs(_net):,.0f} out of pocket. "
                    f"Wait for the current calls to decay closer to "
                    f"${new_prem:.2f} before rolling."
                )
            _dte_lbl = "✓ Optimal" if 30<=best["dte"]<=45 else ("⚠ Suboptimal" if 25<=best["dte"]<30 else "⚠ Sub-30")
            best["dte_label"] = _dte_lbl
            best["dte_col"] = "#00e676" if "Optimal" in _dte_lbl else "#ffb74d"
        return best

    # Action recommendation
    def _recommend(avg_prem, current_mid, dte, ok_assigned=True):
        """
        For CC sellers: positive gain_pct = option decayed = good (bought back cheaper).
        Negative gain_pct = option gained value = bad (costs more to close than you collected).
        """
        if current_mid is None: return "HOLD","#aaa","Insufficient data to advise."
        gain_pct = round((avg_prem - current_mid)/avg_prem*100,1) if avg_prem else 0
        loss_mult = current_mid / avg_prem if avg_prem else 1  # how many times over sold price

        # ── Losing scenarios (option worth MORE than you sold it for) ──────────
        if loss_mult >= 2.0:
            return ("CLOSE","#ef5350",
                    f"Option has doubled to ${current_mid:.2f} (sold at ${avg_prem:.2f}, "
                    f"now {gain_pct:.0f}%). This is the 2× stop-loss rule — close the position "
                    f"to prevent further losses. Stock has moved sharply against your strike.")
        if loss_mult >= 1.5:
            return ("REVIEW — LOSING","#ef5350",
                    f"Option is {abs(gain_pct):.0f}% above your sale price "
                    f"(sold ${avg_prem:.2f}, now ${current_mid:.2f}). "
                    f"Stock is moving against you. Consider closing now or rolling up+out "
                    f"if you can do so for a net credit. Watch for the 2× level at ${avg_prem*2:.2f}.")
        if loss_mult >= 1.1:
            return ("HOLD — MONITOR","#ffb74d",
                    f"Option slightly above your sale price "
                    f"(sold ${avg_prem:.2f}, now ${current_mid:.2f}, {gain_pct:.1f}%). "
                    f"Small unrealised loss — theta is still working. "
                    f"Watch the 2× stop level at ${avg_prem*2:.2f}. No action yet.")

        # ── Winning scenarios (option decayed below your sale price) ──────────
        if gain_pct >= 75:
            return ("CLOSE","#00e676",
                    f"Up {gain_pct:.0f}% — option nearly worthless. "
                    f"Close it now, collect the remaining value, and sell a new CC. "
                    f"No point waiting for the last few cents while carrying gamma risk.")
        if gain_pct >= 50:
            return ("CLOSE & ROLL","#00e676",
                    f"Up {gain_pct:.0f}% — at the 50% profit threshold. "
                    f"Buy back now (${current_mid:.2f}), sell a new CC further out for fresh income. "
                    f"This is the optimal time to roll.")
        if dte <= 21:
            action = "ROLL"
            return (action,"#ffb74d",
                    f"{dte} DTE — inside 21 days, gamma is accelerating even though "
                    f"you are up {gain_pct:.0f}%. Roll to the next 30-45 DTE window now "
                    f"to reset theta and avoid gamma exposure.")
        if gain_pct >= 30:
            return ("HOLD / WATCH","#42a5f5",
                    f"Up {gain_pct:.0f}% with {dte} DTE remaining. "
                    f"Theta still working — wait for 50% profit threshold before rolling.")
        return ("HOLD","#42a5f5",
                f"Up {gain_pct:.0f}% with {dte} DTE. Early in the trade, let theta decay. "
                f"Check again at 50% profit (${avg_prem*0.5:.2f}) or 21 DTE.")

    # ── Call AI for roll reasoning ────────────────────────────────────────────
    _ai_rolls = {}
    try:
        import datetime as _rdt
        _sig_ctx = ""
        try:
            import yfinance as _yf_sg
            for _sk in ["AMZN","MSFT"]:
                _sh2 = _yf_sg.Ticker(_sk).history(period="30d")
                if not _sh2.empty:
                    _cl2 = list(_sh2["Close"])
                    def _eq2(c,n,_k=None,_e=None):
                        k=2/(n+1); e=c[0]
                        for x in c[1:]: e=x*k+e*(1-k)
                        return round(e,2)
                    _e8_=_eq2(_cl2,8); _e21_=_eq2(_cl2,21); _cur_=round(_cl2[-1],2)
                    if _cur_>_e8_>_e21_:
                        _sig_ctx+=f"{_sk}: Bullish (EMA Cloud, price ${_cur_} above 8/${_e8_} above 21/{_e21_}). "
                    elif _cur_<_e8_<_e21_:
                        _sig_ctx+=f"{_sk}: Bearish (price ${_cur_} below both EMAs). "
                    else:
                        _sig_ctx+=f"{_sk}: Mixed (price ${_cur_}, 8EMA ${_e8_}, 21EMA ${_e21_}). "
        except Exception:
            _sig_ctx = "unavailable"

        _ctx_lines = []
        for _tk, _pg in _pos_groups.items():
            _chain_mid = _get_mid(_tk, _pg["exp_yf"], _pg["strike"])
            _brok_mid  = _pg.get("brokerage_current")
            _cm        = _chain_mid or _brok_mid  # prefer live chain, fall back to brokerage
            _cm_src    = "live chain" if _chain_mid else ("brokerage screenshot Mar 20" if _brok_mid else "unavailable")
            _th = _theta_analysis(_pg["exp_yf"], _pg["avg_prem"])
            _br = _best_roll(_tk, _cm, _pg["contracts"])
            _avg = _pg["avg_prem"]
            _gp  = round((_avg - _cm) / _avg * 100, 1) if _cm else None
            _gp_str = f"{_gp:+.1f}%" if _gp is not None else "unknown"
            # Plain string formatting — no format spec in conditional
            _cm_str = ("$" + f"{_cm:.2f}") if _cm else "unavailable"
            _l = (
                f"{_tk} Apr 10, 2026 ({_th['dte']} DTE) — {_pg['contracts']}x contracts\n"
                f"  Sold at (avg): ${_avg:.2f}\n"
                f"  Current option price: {_cm_str} (source: {_cm_src})\n"
                f"  Gain on position: {_gp_str} "
                f"({'option decayed = WINNING' if (_gp or 0)>0 else 'option gained = LOSING'} for CC seller)\n"
                f"  Stock price now: ${_live_px.get(_tk, 'unknown')}\n"
                f"  Stop-loss level (2x sold): ${_avg*2:.2f}"
            )
            if _br:
                _l += (
                    f"\n  Best roll: {_br['exp']} {_br['dte']}DTE"
                    f" ${_br['strike']:.0f} strike @ ${_br['mid']:.2f}"
                    f" — net {_br['net_credit_str']}"
                )
            _ctx_lines.append(_l)

        _tds = today.strftime("%B %d, %Y")
        _prompt = (
            f"You are Kyle's covered call advisor. Today is {_tds}.\n\n"
            f"SIGNALS (live EMA data):\n{_sig_ctx}\n\n"
            "CONFIRMED POSITION DATA (use these numbers — do not say data is unavailable):\n"
            + "\n\n".join(_ctx_lines) + "\n\n"
            "CONTEXT:\n"
            "- Kyle sold COVERED CALLS on stock he owns (500 AMZN + 300 MSFT)\n"
            "- He is FINE being assigned/called away — focus on income, not share preservation\n"
            "- For a CC seller: gain% POSITIVE = option decayed = good. NEGATIVE = option gained value = losing\n"
            "- Key rules: close at 50% profit, stop at 2x sold price, roll inside 21 DTE\n"
            "- Earnings: MSFT Apr 28, AMZN Apr 30 — do NOT suggest expiries on or after those dates\n\n"
            "For EACH ticker give ONE clear recommendation using the position data above.\n"
            "Reasoning must be 3-4 sentences using the ACTUAL numbers provided. "
            "State the gain%, what it means for a CC seller, whether to act now or hold, "
            "and reference the signal.\n"
            'JSON only: {"rolls":[{"ticker":"AMZN","action":"HOLD","roll_to_strike":null,'
            '"roll_to_expiry":null,"roll_to_dte":null,"net_credit":null,"urgency":"Hold",'
            '"reasoning":"...","dte_quality":"Optimal"}]}'
        )
        _pl = _jk.dumps({"model":"claude-sonnet-4-20250514","max_tokens":700,
                          "messages":[{"role":"user","content":_prompt}]}).encode()
        _rq = _rur.Request("https://api.anthropic.com/v1/messages",data=_pl,
                            headers={"x-api-key":ANTHROPIC_API_KEY,"anthropic-version":"2023-06-01",
                                     "Content-Type":"application/json"},method="POST")
        with _rur.urlopen(_rq,timeout=30) as _rs:
            _rx = _jk.loads(_rs.read())
            _tx = "".join(b.get("text","") for b in _rx.get("content",[]))
            _tx = _tx.strip().lstrip("```json").lstrip("```").rstrip("```").strip()
            for _r in _jk.loads(_tx).get("rolls",[]):
                _ai_rolls[_r.get("ticker","")] = _r
    except Exception:
        pass

    # ── RENDER ────────────────────────────────────────────────────────────────
    lines = []
    def w(*parts): lines.append("".join(str(p) for p in parts))

    w('<div class="card">')
    w(f'<h3>Kyle\'s Covered Calls <span class="fresh">{today.strftime("%b %d, %Y")}</span></h3>')

    # ════════════════════════════════════════════════════════════════════════════
    # TABLE 1 — Active Position Tracker with action, theta, roll suggestion
    # ════════════════════════════════════════════════════════════════════════════
    w('<div style="background:#0a1628;border-radius:12px;padding:18px 22px;margin-bottom:18px">')
    w('<div style="font-size:14px;font-weight:900;color:#fff;margin-bottom:14px">Active Positions</div>')

    for _tk, _pg in _pos_groups.items():
        _col    = _COLORS[_tk]
        _px_now = _live_px.get(_tk)
        _exp_yf = _pg["exp_yf"]
        _strike = _pg.get("strike")
        _cts    = _pg["contracts"]
        _avg_pm = _pg["avg_prem"]
        _total_s= _pg["total_sold"]

        # Use brokerage screenshot price as primary, fallback to live chain
        _brok_mid = _pg.get("brokerage_current")
        _chain_mid = _get_mid(_tk, _exp_yf, _strike)
        _cur_mid = _brok_mid or _chain_mid

        _th     = _theta_analysis(_exp_yf, _avg_pm)
        _dte    = _th["dte"]
        _action, _act_col, _act_reason = _recommend(_avg_pm, _cur_mid, _dte, ok_assigned=True)
        _best_r = _best_roll(_tk, _cur_mid, _cts)
        _ai_r   = _ai_rolls.get(_tk, {})

        # Gain calc
        _gain_pct = round((_avg_pm - _cur_mid) / _avg_pm * 100, 1) if _cur_mid else None
        _gain_col = "#00e676" if (_gain_pct or 0) >= 0 else "#ef5350"
        _pnl_est  = round((_avg_pm - _cur_mid) * 100 * _cts, 2) if _cur_mid else None

        # OTM status
        _otm_str, _otm_col = "—", "#aaa"
        if _strike and _px_now:
            _otm_v = (_strike - _px_now) / _px_now * 100
            _otm_str = f"{'+'if _otm_v>=0 else ''}{_otm_v:.1f}% to strike"
            _otm_col = "#00e676" if _otm_v >= 0 else "#ef5350"

        # DTE color
        _dte_col  = "#00e676" if 30<=_dte<=45 else ("#ffb74d" if 21<=_dte<30 else "#ef5350")
        _dte_warn = " ⚠ manage now" if _dte <= 21 else (" — inside sweet spot" if _dte <= 45 else "")

        w(f'<div style="background:#0d1b2a;border-left:4px solid {_col};border-radius:10px;padding:16px 18px;margin-bottom:14px">')

        # Position header row
        w(f'<div style="display:flex;align-items:center;gap:12px;flex-wrap:wrap;margin-bottom:12px">')
        w(f'<span style="font-size:18px;font-weight:900;color:{_col}">{_tk}</span>')
        w(f'<span style="font-size:13px;color:#8ab4d4">Apr 10, 2026 · {_cts}x contracts</span>')
        if _px_now:
            w(f'<span style="font-size:12px;color:{_col}">Stock ${_px_now:.2f}{_LIVE}</span>')
        w(f'<span style="font-size:12px;color:{_otm_col}">{_otm_str}</span>')
        # Earnings safety badge
        _earn_dt  = EARNINGS.get(_tk)
        _exp_safe = _earnings_safe(_tk, _pg["exp_yf"])
        if _earn_dt:
            _ecol = "#00e676" if _exp_safe else "#ef5350"
            _esafe = "safe ✓" if _exp_safe else "CROSSES ✗"
            _elbl = "Earnings " + _earn_dt.strftime("%b %d") + " — expiry " + _esafe
            w(f'<span style="font-size:11px;color:{_ecol};background:#0d1b2a;padding:2px 8px;border-radius:5px">{_elbl}</span>')
        # Action badge
        w(f'<span style="margin-left:auto;background:{_act_col};color:#000;font-size:12px;font-weight:900;padding:3px 14px;border-radius:8px">{_action}</span>')
        w('</div>')

        # Stats row
        w('<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(120px,1fr));gap:8px;margin-bottom:12px">')
        for _lbl, _val, _vc in [
            ("Avg Sold At",    f"${_avg_pm:.2f}",                       "#e8edf5"),
            ("Current Mid",    f"${_cur_mid:.2f}" if _cur_mid else "—", "#e8edf5"),
            ("Gain %",         f"{_gain_pct:+.1f}%" if _gain_pct is not None else "—", _gain_col),
            ("Est P&L",        f"${_pnl_est:+,.0f}" if _pnl_est else "—", _gain_col),
            ("DTE",            f"{_dte}d{_dte_warn}",                   _dte_col),
            ("Roll by",        _th["day21"],                            "#ffb74d"),
        ]:
            w(f'<div style="background:#13263d;border-radius:7px;padding:8px 10px">')
            w(f'<div style="font-size:10px;color:#5a7fa0;margin-bottom:2px">{_lbl}</div>')
            w(f'<div style="font-size:13px;font-weight:800;color:{_vc}">{esc(_val)}</div></div>')
        w('</div>')

        # AI reasoning (primary) or rule-based fallback
        _reasoning = _ai_r.get("reasoning","") or _act_reason
        w(f'<div style="background:#0a1628;border-left:3px solid {_act_col};border-radius:6px;padding:9px 13px;margin-bottom:10px">')
        w(f'<div style="font-size:11px;font-weight:800;color:{_act_col};margin-bottom:3px">RECOMMENDATION</div>')
        w(f'<div style="font-size:12px;color:#dde8f8;line-height:1.6">{esc(_reasoning)}</div>')
        w('</div>')

        # Roll suggestion (one row only)
        _roll_target = _best_r
        if _ai_r.get("roll_to_expiry"):
            # Prefer AI suggestion if action involves rolling
            _roll_target = {
                "exp":       _ai_r.get("roll_to_expiry",""),
                "dte":       _ai_r.get("roll_to_dte") or 0,
                "strike":    _ai_r.get("roll_to_strike") or 0,
                "mid":       None,
                "otm_pct":   _ai_r.get("roll_to_otm_pct") or 0,
                "ann_ret":   0,
                "net_credit": _ai_r.get("net_credit"),
                "net_credit_str": (f"+${_ai_r['net_credit']:,.0f} credit" if (_ai_r.get("net_credit") or 0) >= 0
                                   else f"${_ai_r['net_credit']:,.0f} debit"),
                "net_credit_label": ("NET CREDIT" if (_ai_r.get("net_credit") or 0) >= 0 else "NET DEBIT"),
                "net_credit_color": ("#00e676" if (_ai_r.get("net_credit") or 0) >= 0 else "#ef5350"),
                "net_explain": (f"AI estimate — collect new premium, pay ~${abs(_ai_r.get('net_credit',0)/100):.2f}/share to close. "
                                f"{'Cash flows IN.' if (_ai_r.get('net_credit') or 0) >= 0 else 'Rolling costs cash OUT — only do this if stock is trending against you and you want to extend/adjust.'}"),
                "dte_col":   "#00e676" if _ai_r.get("dte_quality","")=="Optimal" else "#ffb74d",
                "total_prem": None,
            }

        if _roll_target and _roll_target.get("exp"):
            _rt = _roll_target
            _rt_dte = int(_rt.get("dte") or 0)
            _rt_col = "#00e676" if 30<=_rt_dte<=45 else "#ffb74d"
            _rt_st  = _rt.get("strike")
            _rt_nc  = _rt.get("net_credit_str","—")
            _nc_val = _rt.get("net_credit") or 0
            _nc_col = "#00e676" if (_nc_val or 0)>=0 else "#ef5350"
            _rt_mid = _rt.get("mid")

            # Look up live mid if we have strike
            if _rt_st and not _rt_mid:
                # find exp_yf from the exp string
                _rt_exp_yf = _rt["exp"] if len(_rt["exp"])==10 else None
                if _rt_exp_yf:
                    _rt_mid = _get_mid(_tk, _rt_exp_yf, float(_rt_st))

            w(f'<div style="background:#0d2a18;border-radius:8px;padding:10px 14px">')
            w(f'<div style="font-size:11px;font-weight:800;color:#00e676;margin-bottom:6px">SUGGESTED ROLL TARGET</div>')
            w(f'<div style="display:flex;align-items:center;gap:12px;flex-wrap:wrap">')
            w(f'<span style="font-size:14px;font-weight:900;color:#FFD700">')
            w(f'${_rt_st:.0f} strike' if _rt_st else 'Strike TBD')
            w(f' &nbsp;·&nbsp; {_rt["exp"]} &nbsp;·&nbsp; <span style="color:{_rt_col}">{_rt_dte}d DTE</span>')
            w('</span>')
            if _rt_mid:
                w(f'<span style="font-size:12px;color:#00e676">Mid ${_rt_mid:.2f}</span>')
            _nl  = _rt.get("net_credit_label","NET")
            _nc  = _rt.get("net_credit_color", _nc_col)
            w(f'<span style="background:{_nc};color:#000;font-size:10px;font-weight:900;padding:2px 9px;border-radius:5px">{_nl}: {_rt_nc}</span>')
            if _rt.get("dte_label"):
                w(f'<span style="background:{_rt_col};color:#000;font-size:10px;font-weight:800;padding:1px 8px;border-radius:5px">{_rt["dte_label"]}</span>')
            w('</div>')
            # Earnings warning on the roll target expiry
            _rt_expyf2 = _rt.get("exp","") if len(_rt.get("exp","")) == 10 else ""
            _ewarn = _earnings_warn(_tk, _rt_expyf2) if _rt_expyf2 else ""
            if _ewarn:
                w(f'<div style="background:#2a1000;border-left:3px solid #ef5350;border-radius:5px;padding:6px 10px;margin-top:6px;font-size:11px;color:#ef9a9a;font-weight:700">\u26a0 {esc(_ewarn)}</div>')
            # Explanation line
            if _rt.get("net_explain"):
                w(f'<div style="font-size:11px;color:#8ab4d4;margin-top:6px;line-height:1.5">{esc(_rt["net_explain"])}</div>')
            w('</div>')

        w('</div>')  # end ticker card

    w('</div>')  # end Table 1

    # ════════════════════════════════════════════════════════════════════════════
    # TABLE 2 — Trade log, P&L, and shares purchased
    # ════════════════════════════════════════════════════════════════════════════
    w('<div style="background:#0a1628;border-radius:12px;padding:18px 22px">')
    w('<div style="font-size:14px;font-weight:900;color:#fff;margin-bottom:4px">Trade Log &amp; P&amp;L</div>')

    # Summary strip
    _locked = sum(round((t["premium"]-(t.get("closed_at") or 0))*100*t["contracts"],2) for t in CLOSED_TRADES)
    _open_v = sum(t["total_premium"] for t in OPEN_TRADES)
    _open_cost = sum((t.get("brokerage_current") or 0)*100*t["contracts"] for t in OPEN_TRADES)
    _open_pnl  = round(sum((t["premium"]-(t.get("brokerage_current") or t["premium"]))*100*t["contracts"] for t in OPEN_TRADES),2)
    w('<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:8px;margin-bottom:16px">')
    for _sl,_sv,_sc in [
        ("Locked Profit",  f"${_locked:,.0f}",          "#00e676"),
        ("Open Sold",      f"${_open_v:,.0f}",           "#42a5f5"),
        ("Open P&L",       f"${_open_pnl:+,.0f}",       "#00e676" if _open_pnl>=0 else "#ef5350"),
        ("Total Net",      f"${TOTAL_NET_PREMIUM:,.0f}", "#FFD700"),
    ]:
        w(f'<div style="background:#0d1b2a;border-radius:7px;padding:9px 12px">')
        w(f'<div style="font-size:10px;color:#5a7fa0;margin-bottom:2px">{_sl}</div>')
        w(f'<div style="font-size:16px;font-weight:900;color:{_sc}">{_sv}</div></div>')
    w('</div>')

    # Closed trades
    w('<div style="font-size:11px;font-weight:800;color:#5a7fa0;margin-bottom:5px">CLOSED</div>')
    w('<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;font-size:12px;margin-bottom:14px">')
    w('<thead><tr style="background:#0d2040;border-bottom:2px solid #1e3a5f">')
    for _h in ["Date","Ticker","Strike","Expiry","Qty","Sold","Closed At","Net/sh","Net Total","Notes"]:
        w(f'<th style="padding:7px 10px;color:#7eb8f7;font-weight:700;text-align:right;white-space:nowrap">{_h}</th>')
    w('</tr></thead><tbody>')
    for _t in CLOSED_TRADES:
        _col = _COLORS.get(_t["ticker"],"#fff")
        _ca  = _t.get("closed_at"); _net = round(_t["premium"]-(_ca or 0),2) if _ca else _t["premium"]
        _nt  = round(_net*100*_t["contracts"],2); _nc2 = "#00e676" if _nt>=0 else "#ef5350"
        w('<tr style="border-bottom:1px solid #0d2040">')
        w(f'<td style="padding:7px 10px;color:#8ab4d4;text-align:right">{_t["date"]}</td>')
        w(f'<td style="padding:7px 10px;color:{_col};font-weight:800;text-align:right">{_t["ticker"]}</td>')
        w(f'<td style="padding:7px 10px;color:#FFD700;text-align:right">${_t["strike"]:.0f}</td>' if _t.get("strike") else '<td style="padding:7px 10px;color:#555;text-align:right">—</td>')
        w(f'<td style="padding:7px 10px;color:#8ab4d4;text-align:right;white-space:nowrap">{_t["expiry"][:10]}</td>')
        w(f'<td style="padding:7px 10px;color:#e8edf5;text-align:right">{_t["contracts"]}x</td>')
        w(f'<td style="padding:7px 10px;color:#00e676;text-align:right">${_t["premium"]:.2f}</td>')
        w(f'<td style="padding:7px 10px;color:#ef9a9a;text-align:right">${_ca:.2f}</td>' if _ca else '<td style="text-align:right;color:#555">—</td>')
        w(f'<td style="padding:7px 10px;color:{_nc2};font-weight:700;text-align:right">${_net:.2f}</td>')
        w(f'<td style="padding:7px 10px;color:{_nc2};font-weight:800;text-align:right">${_nt:,.0f}</td>')
        w(f'<td style="padding:7px 10px;color:#8ab4d4;font-size:11px;text-align:right">{_t.get("notes","")}</td>')
        w('</tr>')
    w(f'<tr style="background:#0d2040;border-top:2px solid #1e3a5f">')
    w(f'<td colspan="8" style="padding:7px 10px;color:#7eb8f7;font-weight:900;text-align:right">LOCKED IN</td>')
    w(f'<td style="padding:7px 10px;color:#00e676;font-weight:900;text-align:right">${_locked:,.0f}</td>')
    w('<td></td></tr>')
    w('</tbody></table></div>')

    # Open trades
    w('<div style="font-size:11px;font-weight:800;color:#5a7fa0;margin-bottom:5px">OPEN — APR 10 (prices as of Mar 20)</div>')
    w('<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;font-size:12px;margin-bottom:14px">')
    w('<thead><tr style="background:#0d2040;border-bottom:2px solid #1e3a5f">')
    for _h in ["Date","Ticker","Strike","Expiry","Qty","Sold At","Broker Price","Gain %","Est P&L","Notes"]:
        w(f'<th style="padding:7px 10px;color:#7eb8f7;font-weight:700;text-align:right;white-space:nowrap">{_h}</th>')
    w('</tr></thead><tbody>')
    _open_total = 0
    for _t in OPEN_TRADES:
        _col = _COLORS.get(_t["ticker"],"#fff")
        _bm  = _t.get("brokerage_current")
        _gp  = round((_t["premium"]-_bm)/_t["premium"]*100,1) if _bm else None
        _gpc = "#00e676" if (_gp or 0)>=0 else "#ef5350"
        _ep  = round((_t["premium"]-(_bm or _t["premium"]))*100*_t["contracts"],2)
        _tc  = _t["total_premium"]; _open_total += _tc
        _st  = f"${_t['strike']:.0f}" if _t.get("strike") else "TBD"
        w('<tr style="border-bottom:1px solid #0d2040">')
        w(f'<td style="padding:7px 10px;color:#8ab4d4;text-align:right">{_t["date"]}</td>')
        w(f'<td style="padding:7px 10px;color:{_col};font-weight:800;text-align:right">{_t["ticker"]}</td>')
        w(f'<td style="padding:7px 10px;color:#FFD700;text-align:right">{_st}</td>')
        w(f'<td style="padding:7px 10px;color:#8ab4d4;text-align:right;white-space:nowrap">{_t["expiry"][:10]}</td>')
        w(f'<td style="padding:7px 10px;color:#e8edf5;text-align:right">{_t["contracts"]}x</td>')
        w(f'<td style="padding:7px 10px;color:#00e676;text-align:right">${_t["premium"]:.2f}</td>')
        w(f'<td style="padding:7px 10px;color:#e8edf5;text-align:right">${_bm:.2f}</td>' if _bm else '<td style="text-align:right;color:#555">—</td>')
        w(f'<td style="padding:7px 10px;color:{_gpc};font-weight:800;text-align:right">{_gp:+.1f}%</td>' if _gp is not None else '<td style="text-align:right;color:#555">—</td>')
        w(f'<td style="padding:7px 10px;color:{_gpc};font-weight:800;text-align:right">${_ep:+,.0f}</td>')
        w(f'<td style="padding:7px 10px;color:#8ab4d4;font-size:11px;text-align:right">{_t.get("notes","")}</td>')
        w('</tr>')
    w(f'<tr style="background:#0d2040;border-top:2px solid #1e3a5f">')
    w(f'<td colspan="5" style="padding:7px 10px;color:#7eb8f7;font-weight:900;text-align:right">OPEN TOTAL</td>')
    w(f'<td style="padding:7px 10px;color:#00e676;font-weight:900;text-align:right">${_open_total:,.0f}</td>')
    w(f'<td colspan="2"></td>')
    w(f'<td style="padding:7px 10px;color:{"#00e676" if _open_pnl>=0 else "#ef5350"};font-weight:900;text-align:right">${_open_pnl:+,.0f}</td>')
    w('<td></td></tr>')
    w('</tbody></table></div>')

    # Shares purchased
    w('<div style="font-size:11px;font-weight:800;color:#5a7fa0;margin-bottom:5px">SHARES PURCHASED WITH CC INCOME</div>')
    w('<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;font-size:12px">')
    w('<thead><tr style="background:#0d2040;border-bottom:2px solid #1e3a5f">')
    for _h in ["Date","Ticker","Shares","Price Paid","Cost","Current Price","Value Now","Gain $","Gain %","Notes"]:
        w(f'<th style="padding:7px 10px;color:#7eb8f7;font-weight:700;text-align:right;white-space:nowrap">{_h}</th>')
    w('</tr></thead><tbody>')
    for _sp in SHARE_PURCHASES:
        _col = _COLORS.get(_sp["ticker"],"#fff")
        _px_n = _live_px.get(_sp["ticker"])
        _pp   = _sp.get("price_paid")
        _cost = round(_pp*_sp["shares"],2) if _pp else None
        _cv   = round(_px_n*_sp["shares"],2) if _px_n else None
        _gain = round(_cv-_cost,2) if (_cv and _cost) else None
        _gp2  = round(_gain/_cost*100,2) if (_gain and _cost) else None
        _gpc2 = "#00e676" if (_gain or 0)>=0 else "#ef5350"
        w('<tr style="border-bottom:1px solid #0d2040">')
        w(f'<td style="padding:7px 10px;color:#8ab4d4;text-align:right">{_sp["date"]}</td>')
        w(f'<td style="padding:7px 10px;color:{_col};font-weight:800;text-align:right">{_sp["ticker"]}</td>')
        w(f'<td style="padding:7px 10px;color:#e8edf5;text-align:right">{_sp["shares"]}</td>')
        w(f'<td style="padding:7px 10px;color:#e8edf5;text-align:right">{"$"+str(_pp) if _pp else "—"}</td>')
        w(f'<td style="padding:7px 10px;color:#e8edf5;text-align:right">{"$"+f"{_cost:,.2f}" if _cost else "—"}</td>')
        w(f'<td style="padding:7px 10px;color:{_col};text-align:right">{"$"+str(_px_n) if _px_n else "—"}{_LIVE if _px_n else ""}</td>')
        w(f'<td style="padding:7px 10px;color:{_gpc2};text-align:right">{"$"+f"{_cv:,.2f}" if _cv else "—"}</td>')
        w(f'<td style="padding:7px 10px;color:{_gpc2};font-weight:800;text-align:right">{"$"+f"{_gain:+,.2f}" if _gain else "—"}</td>')
        w(f'<td style="padding:7px 10px;color:{_gpc2};font-weight:900;text-align:right">{f"{_gp2:+.1f}%" if _gp2 else "—"}</td>')
        w(f'<td style="padding:7px 10px;color:#8ab4d4;font-size:11px;text-align:right">{_sp["notes"]}</td>')
        w('</tr>')
    w('</tbody></table></div>')
    w('</div>')  # end Table 2

    w('</div>')  # end card
    return "\n".join(lines)


def render_azam_tab() -> str:
    """Azam's covered call screener for CMG, MSFT, AMZN, NTR."""
    import json as _json, os, time, math
    AZAM_TICKERS = ["CMG", "MSFT", "AMZN", "NTR"]
    AZAM_SHARES  = {"CMG": 100, "MSFT": 100, "AMZN": 100, "NTR": 200}
    NAMES = {"CMG":"Chipotle Mexican Grill","MSFT":"Microsoft Corp","AMZN":"Amazon.com Inc","NTR":"Nutrien Ltd"}
    TICKER_COLORS = {"CMG":"#c0392b","MSFT":"#00a4ef","AMZN":"#FF9900","NTR":"#27ae60"}

    lines = []
    def w(*parts): lines.append("".join(str(p) for p in parts))

    # ── Fetch live CC chains ──────────────────────────────────────────────────
    CACHE_FILE = ".cache_azam_positions.json"
    CACHE_TTL  = 1800
    azam_live = {}
    try:
        if os.path.exists(CACHE_FILE):
            with open(CACHE_FILE) as cf:
                _c = _json.load(cf)
            if time.time() - _c.get("_ts", 0) < CACHE_TTL:
                azam_live = _c
    except Exception:
        pass

    if not azam_live:
        import yfinance as yf, datetime
        today = datetime.date.today()
        for tk in AZAM_TICKERS:
            try:
                t = yf.Ticker(tk)
                price = 0
                try:
                    price = float(t.fast_info.get("last_price") or t.fast_info.get("regularMarketPrice") or 0)
                except Exception:
                    pass
                if not price:
                    hist = t.history(period="2d")
                    price = float(hist["Close"].iloc[-1]) if not hist.empty else 0
                shares = AZAM_SHARES.get(tk, 100)
                contracts_avail = max(1, shares // 100)
                exps_raw = t.options
                exp_blocks = []
                for exp_str in (exps_raw or [])[:5]:
                    try:
                        exp_dt = datetime.datetime.strptime(exp_str, "%Y-%m-%d").date()
                        dte = (exp_dt - today).days
                        if dte < 7 or dte > 90: continue
                        chain = t.option_chain(exp_str)
                        calls = chain.calls
                        rows_out = []
                        for _, row in calls.iterrows():
                            strike = float(row["strike"])
                            if price <= 0: continue
                            otm_pct = (strike - price) / price * 100
                            if otm_pct < 1 or otm_pct > 18: continue
                            bid  = float(row.get("bid") or 0)
                            ask  = float(row.get("ask") or 0)
                            mid  = round((bid + ask) / 2, 2) if bid and ask else bid or ask
                            if mid <= 0: continue
                            iv   = float(row.get("impliedVolatility") or 0) * 100
                            vol  = int(row.get("volume") or 0)
                            oi   = int(row.get("openInterest") or 0)
                            last = float(row.get("lastPrice") or 0)
                            ann  = round(mid / price * (365 / max(dte, 1)) * 100, 1)
                            be   = round(price - mid, 2)
                            gic  = round((strike - price + mid) * contracts_avail * 100, 2)
                            rows_out.append({
                                "strike": strike, "bid": bid, "ask": ask, "mid": mid,
                                "last": last, "otm_pct": round(otm_pct, 1), "iv_pct": round(iv, 1),
                                "volume": vol, "oi": oi, "contracts": contracts_avail,
                                "total_prem": round(mid * 100 * contracts_avail, 2),
                                "breakeven": be, "ann_ret": ann, "gain_if_called": gic,
                            })
                        if rows_out:
                            exp_blocks.append({"exp": exp_str, "dte": dte, "rows": rows_out})
                    except Exception:
                        continue
                azam_live[tk] = {"price": price, "shares": shares, "expirations": exp_blocks}
            except Exception:
                azam_live[tk] = {}
        azam_live["_ts"] = time.time()
        try:
            with open(CACHE_FILE, "w") as cf:
                _json.dump(azam_live, cf)
        except Exception:
            pass

    # ── AI recommendations ────────────────────────────────────────────────────
    import urllib.request as _ureq
    ai_recs = None
    try:
        pos_lines = "\n".join(
            f"  {tk}: {AZAM_SHARES.get(tk,100)} shares, price ~${azam_live.get(tk,{}).get('price',0):.2f}"
            for tk in AZAM_TICKERS
        )
        import datetime as _dt
        _today_str = _dt.date.today().strftime("%B %d, %Y")
        _target_exp = (_dt.date.today() + _dt.timedelta(days=35)).strftime("%b %d, %Y")
        # Build highest ann-ret option per ticker so AI can explain why it skipped it
        _max_ret_lines = []
        for _tk in AZAM_TICKERS:
            _exps = azam_live.get(_tk, {}).get("expirations", [])
            _all_rows = [r for eb in _exps for r in eb.get("rows", [])]
            if _all_rows:
                _best = max(_all_rows, key=lambda r: r.get("ann_ret", 0))
                _max_ret_lines.append(
                    f"  {_tk}: Highest ann-ret option is ${_best['strike']:.0f} strike, "
                    f"{_best['otm_pct']:.1f}% OTM, {_best['ann_ret']:.1f}% ann ret, "
                    f"${_best['total_prem']:.0f} total prem"
                )
        _max_ret_block = "\n".join(_max_ret_lines) if _max_ret_lines else "  (unavailable)"
        _prompt = (
            f"You are Azam's options advisor. Today's date is {_today_str}.\n"
            "Azam holds:\n" + pos_lines + "\n\n"
            "He wants to sell covered calls for income on ALL four stocks.\n"
            "For each, recommend the single BEST covered call to sell right now.\n"
            f"Target: 30-45 DTE from today ({_today_str}), 4-8% OTM, meaningful premium.\n"
            f"Suggest real expiry dates that exist AFTER {_today_str} (standard monthly/weekly expirations).\n"
            "NTR = Nutrien Ltd (Canadian fertilizer company).\n\n"
            "For context, the highest annualized-return option per stock (usually near-ATM or ITM) is:\n"
            + _max_ret_block + "\n\n"
            "IMPORTANT: For the 'vs_max_return' field, explain in 1-2 sentences WHY you did NOT pick "
            "the highest ann-ret option above — e.g. assignment risk if stock rallies, near-ITM danger, "
            "cap on upside, poor premium quality. Be specific about the tradeoff.\n\n"
            'Return ONLY valid JSON (no markdown):\n'
            '{"recommendations":[{"ticker":"CMG","strike_suggestion":58.0,"otm_pct":5.5,'
            f'"expiry_suggestion":"~{_target_exp}","dte":35,"rationale":"...","income_thesis":"...",'
            '"risk":"...","tier":"Balanced","vs_max_return":"Why highest ann-ret was skipped..."}],'
            '"portfolio_note":"..."}'
        )
        _payload = _json.dumps({
            "model": "claude-sonnet-4-20250514",
            "max_tokens": 1200,
            "messages": [{"role": "user", "content": _prompt}]
        }).encode()
        _req = _ureq.Request(
            "https://api.anthropic.com/v1/messages",
            data=_payload,
            headers={"x-api-key": ANTHROPIC_API_KEY,
                     "anthropic-version": "2023-06-01",
                     "Content-Type": "application/json"},
            method="POST"
        )
        with _ureq.urlopen(_req, timeout=30) as _resp:
            _raw  = _json.loads(_resp.read())
            _text = "".join(b.get("text", "") for b in _raw.get("content", []))
            _text = _text.strip().lstrip("```json").lstrip("```").rstrip("```").strip()
            ai_recs = _json.loads(_text)
    except Exception:
        ai_recs = None

    # ── Render ────────────────────────────────────────────────────────────────
    w('<div class="card">')
    w('<h3>Azam\'s Covered Calls <span class="fresh"> live from yfinance</span></h3>')
    w('<div class="muted" style="margin-bottom:16px">Live covered call chains for CMG, MSFT, AMZN, and NTR. '
      'AI picks the best strike per stock. Not financial advice.</div>')

    # Summary strip
    w('<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:10px;margin-bottom:20px">')
    for tk in AZAM_TICKERS:
        d      = azam_live.get(tk, {})
        price  = d.get("price", 0)
        shares = d.get("shares", AZAM_SHARES.get(tk, 100))
        col    = TICKER_COLORS.get(tk, "#aaa")
        val    = price * shares if price else 0
        w(f'<div style="background:#13263d;border-radius:8px;padding:11px 14px;border-left:3px solid {col}">')
        w(f'<div style="font-size:18px;font-weight:900;color:{col}">{tk}</div>')
        w(f'<div style="font-size:12px;color:#c8d8e8">{shares} shares @ ${price:,.2f}</div>')
        w(f'<div style="font-size:13px;font-weight:800;color:#fff">${val:,.0f} position</div>')
        w(f'<div style="font-size:11px;color:#FFD700">{max(1,shares//100)}x contracts available</div></div>')
    w('</div>')

    # AI portfolio note
    if ai_recs and ai_recs.get("portfolio_note"):
        w('<div style="background:#0a2010;border-left:3px solid #00e676;border-radius:8px;padding:13px 16px;margin-bottom:18px">')
        w('<div style="font-size:11px;color:#00e676;font-weight:800;margin-bottom:5px">AI STRATEGY NOTE</div>')
        w(f'<div style="font-size:12px;color:#dde6f5;line-height:1.7">{esc(ai_recs["portfolio_note"])}</div></div>')

    recs = {}
    if ai_recs:
        recs = {r["ticker"]: r for r in ai_recs.get("recommendations", [])}

    # Per-ticker sections
    for tk in AZAM_TICKERS:
        d        = azam_live.get(tk, {})
        price    = d.get("price", 0)
        shares   = d.get("shares", AZAM_SHARES.get(tk, 100))
        col      = TICKER_COLORS.get(tk, "#aaa")
        exp_list = d.get("expirations", [])
        rec      = recs.get(tk)

        w(f'<div style="background:#0d1b2a;border-left:4px solid {col};border-radius:12px;padding:18px 20px;margin-bottom:22px">')
        w(f'<div style="display:flex;align-items:center;gap:12px;flex-wrap:wrap;margin-bottom:12px">')
        w(f'<span style="font-size:26px;font-weight:900;color:{col}">{tk}</span>')
        w(f'<span style="font-size:14px;color:#c0cfe4">{NAMES.get(tk,"")}</span>')
        if price:
            w(f'<span style="background:#1e3a5f;color:#7eb8f7;font-size:12px;padding:4px 12px;border-radius:10px;font-weight:700">${price:,.2f} current</span>')
        w('</div>')

        # AI pick card
        if rec:
            tier_col = {"Conservative":"#26a69a","Balanced":"#FFD700","Aggressive":"#ef5350"}.get(rec.get("tier","Balanced"),"#FFD700")
            w(f'<div style="background:#0a1628;border-radius:10px;padding:13px 16px;margin-bottom:14px;border:1px solid {tier_col}">')
            w(f'<div style="display:flex;align-items:center;gap:10px;flex-wrap:wrap;margin-bottom:8px">')
            w(f'<span style="background:{tier_col};color:#000;font-size:10px;font-weight:900;padding:2px 9px;border-radius:8px">AI PICK - {esc(rec.get("tier",""))}</span>')
            w(f'<span style="font-size:15px;font-weight:900;color:{tier_col}">Strike ${rec.get("strike_suggestion","?")} &nbsp;{rec.get("expiry_suggestion","?")} &nbsp;{rec.get("dte","?")} DTE &nbsp;{rec.get("otm_pct","?")}% OTM</span>')
            w('</div>')
            w(f'<div style="font-size:12px;color:#e0e8f5;margin-bottom:5px"><strong style="color:#00e676">Income:</strong> {esc(rec.get("income_thesis",""))}</div>')
            w(f'<div style="font-size:12px;color:#e0e8f5;margin-bottom:5px"><strong style="color:#FFD700">Why:</strong> {esc(rec.get("rationale",""))}</div>')
            w(f'<div style="font-size:12px;color:#ffb3b3;margin-bottom:5px"><strong style="color:#ff6b6b">Risk:</strong> {esc(rec.get("risk",""))}</div>')
            if rec.get("vs_max_return"):
                w(f'<div style="margin-top:8px;background:#0d1a2e;border-left:3px solid #7c6f00;border-radius:6px;padding:8px 12px">')
                w(f'<div style="font-size:10px;font-weight:900;color:#b8a800;margin-bottom:3px">⚖️ WHY NOT THE HIGHEST ANN RETURN?</div>')
                w(f'<div style="font-size:11px;color:#e8eef8;line-height:1.6">{esc(rec.get("vs_max_return",""))}</div>')
                w('</div>')
            w('</div>')

        # Live chain tables
        if exp_list:
            w(f'<div style="font-size:11px;font-weight:800;color:#7eb8f7;margin-bottom:6px">LIVE CC CHAIN</div>')
            for eb in exp_list:
                dte_col = "#00e676" if 21 <= eb["dte"] <= 45 else ("#ffb74d" if 14 <= eb["dte"] <= 60 else "#ef5350")
                w(f'<div style="font-size:12px;color:#fff;margin:10px 0 5px">Expiry: <span style="color:#42a5f5">{eb["exp"]}</span> <span style="color:{dte_col}">({eb["dte"]} DTE)</span></div>')
                w('<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;font-size:11px;margin-bottom:10px">')
                w('<thead><tr style="border-bottom:1px solid #263238">')
                for hdr in ["Strike","Bid","Ask","Mid","OTM%","IV%","Volume","OI","Contracts","Total Prem","Breakeven","Ann Ret%","Gain if Called"]:
                    w(f'<th style="padding:5px 8px;color:#e0e6ef;font-weight:700;text-align:right;white-space:nowrap;background:#0d1e30">{hdr}</th>')
                w('</tr></thead><tbody>')
                for row in eb["rows"]:
                    otm    = row["otm_pct"]
                    border = "#ef5350" if otm <= 2 else ("#FFD700" if otm <= 5 else ("#26a69a" if otm <= 10 else "#555"))
                    ann_c  = "#00e676" if row["ann_ret"] >= 20 else ("#ffb74d" if row["ann_ret"] >= 10 else "#c8d4e8")
                    gic_c  = "#00e676" if row["gain_if_called"] >= 0 else "#ff6b6b"
                    is_pick = rec and abs(row["strike"] - float(rec.get("strike_suggestion", -999))) < 0.5
                    row_style = '' if is_pick else ""
                    w(f'<tr{row_style}>')
                    w(f'<td style="padding:6px 8px;color:{border};font-weight:800;text-align:right;border-left:3px solid {border}">${row["strike"]:.0f}{"  AI" if is_pick else ""}</td>')
                    w(f'<td style="padding:6px 8px;color:#e8edf5;text-align:right">${row["bid"]:.2f}</td>')
                    w(f'<td style="padding:6px 8px;color:#e8edf5;text-align:right">${row["ask"]:.2f}</td>')
                    w(f'<td style="padding:6px 8px;color:#00e676;font-weight:700;text-align:right">${row["mid"]:.2f}</td>')
                    w(f'<td style="padding:6px 8px;color:#d0daea;text-align:right">{row["otm_pct"]:.1f}%</td>')
                    w(f'<td style="padding:6px 8px;color:#d0daea;text-align:right">{row["iv_pct"]:.0f}%</td>')
                    w(f'<td style="padding:6px 8px;color:#d0daea;text-align:right">{row["volume"]:,}</td>')
                    w(f'<td style="padding:6px 8px;color:#d0daea;text-align:right">{row["oi"]:,}</td>')
                    w(f'<td style="padding:6px 8px;color:#fff;font-weight:700;text-align:right">{row["contracts"]}x</td>')
                    w(f'<td style="padding:6px 8px;color:#FFD700;font-weight:800;text-align:right">${row["total_prem"]:,.0f}</td>')
                    w(f'<td style="padding:6px 8px;color:#ffb74d;text-align:right">${row["breakeven"]:.2f}</td>')
                    w(f'<td style="padding:6px 8px;color:{ann_c};font-weight:700;text-align:right">{row["ann_ret"]:.1f}%</td>')
                    w(f'<td style="padding:6px 8px;color:{gic_c};text-align:right">${row["gain_if_called"]:,.0f}</td>')
                    w('</tr>')
                w('</tbody></table></div>')
        else:
            w(f'<div style="color:#ffb74d;font-size:12px;padding:10px;background:#1a1200;border-radius:7px">No live chain data for {tk}.</div>')

        w('</div>')  # end ticker card
    w('</div>')  # end main card
    return "\n".join(lines)


def render_mag7_tab(mag_rows: list) -> str:
    """Mag 7 rich cards — confidence, MA levels, convergence, U&R signals."""
    import datetime
    today = datetime.date.today()
    lines = []
    def w(*parts): lines.append("".join(str(p) for p in parts))

    # ── Fetch 1-year daily OHLCV for each Mag7 ticker ────────────────────────
    MA_DATA = {}
    try:
        import yfinance as _yf_m
        for _row in mag_rows:
            _tk = _row.get("Ticker","").lstrip("$")
            if not _tk: continue
            try:
                _h = _yf_m.Ticker(_tk).history(period="1y", interval="1d")
                if _h.empty or len(_h) < 55: continue
                _c  = list(_h["Close"]); _v  = list(_h["Volume"])
                _hi = list(_h["High"]);  _lo = list(_h["Low"])
                _dt = [str(d.date()) for d in _h.index]

                def _sma(p, n): return round(sum(p[-n:])/n, 2) if len(p) >= n else None
                def _ema_v(p, n):
                    if len(p) < n: return None
                    k=2/(n+1); v=sum(p[:n])/n
                    for x in p[n:]: v=x*k+v*(1-k)
                    return round(v,2)

                px = round(_c[-1],2); px1 = round(_c[-2],2)
                sma5  = _sma(_c,5);  sma10 = _sma(_c,10)
                sma20 = _sma(_c,20); sma50 = _sma(_c,50)
                sma100= _sma(_c,100); sma200= _sma(_c,200)
                ema8  = _ema_v(_c,8); ema21 = _ema_v(_c,21)

                wk52_hi = round(max(_hi),2); wk52_lo = round(min(_lo),2)
                pct_from_hi = round((px-wk52_hi)/wk52_hi*100,1)
                pct_from_lo = round((px-wk52_lo)/wk52_lo*100,1)
                avg_vol = sum(_v[-21:-1])/20 if len(_v)>=21 else _v[-1]
                vol_ratio = round(_v[-1]/avg_vol,2) if avg_vol else 1.0

                def _d(ma): return round((px-ma)/ma*100,2) if ma else None

                # ── MA convergence ────────────────────────────────────────────
                long_vals  = [m for m in [sma50,sma100,sma200] if m]
                short_vals = [m for m in [sma5,sma10,sma20,sma50] if m]
                long_spread  = round((max(long_vals)-min(long_vals))/min(long_vals)*100,2) if len(long_vals)>=2 else None
                short_spread = round((max(short_vals)-min(short_vals))/min(short_vals)*100,2) if len(short_vals)>=2 else None

                conv_signals = []
                if long_spread is not None and long_spread <= 4.0:
                    conv_signals.append(f"⚡ Long-term MAs converging (50/100/200 within {long_spread:.1f}%) — coiling for a major move")
                if short_spread is not None and short_spread <= 1.0:
                    conv_signals.append(f"⚡ Short-term MAs converging (5/10/20/50 within {short_spread:.1f}%) — breakout or breakdown imminent")
                # Price approaching key MA cluster
                if long_vals and abs(_d(sum(long_vals)/len(long_vals))) < 2.0:
                    conv_signals.append(f"📍 Price within 2% of long-term MA cluster — key decision zone")

                # ── Confidence level ──────────────────────────────────────────
                above = sum(1 for m in [sma5,sma10,sma20,sma50,sma100,sma200,ema8,ema21] if m and px>m)
                total_ma = len([m for m in [sma5,sma10,sma20,sma50,sma100,sma200,ema8,ema21] if m])
                trend_score = round(above/total_ma*100) if total_ma else 0

                # Confidence has two components:
                # 1. Trend structure (MA alignment)
                # 2. CSP zone — if price is near its 95% statistical floor, that is
                #    HIGH CONVICTION for selling puts regardless of MA position.
                #    The floor held 95% of the time = maximum statistical backing.
                #
                # We compute the 95% floor here (before _sr_levels is built below)
                # to inform the confidence label correctly.
                _sorted_lo_cf = sorted(_lo)
                _n_cf         = len(_sorted_lo_cf)
                _floor_95     = round(_sorted_lo_cf[max(0, int(_n_cf*0.05))], 2)
                _otm_from_95  = round((px - _floor_95) / px * 100, 1)
                _near_95_cf   = _otm_from_95 <= 15  # within 15% of the 95% floor

                if _near_95_cf:
                    # Near 95% statistical floor = HIGH conviction CSP zone
                    # Stock has almost never been this cheap on an annual basis
                    confidence, conf_col = "HIGH — CSP ZONE", "#00e676"
                    conf_desc = (f"Price is {_otm_from_95:.1f}% above its 95% annual support floor (${_floor_95:.2f}). "
                                 f"This is the statistical bottom — held on 95% of trading days over the past year. "
                                 f"High conviction zone to sell long-dated cash-secured puts.")
                elif sma50 and sma200 and px>sma50 and sma50>sma200 and pct_from_hi>-15:
                    confidence, conf_col = "HIGH", "#00e676"
                    conf_desc = (f"Uptrend intact — price above 50 & 200 SMA ({trend_score}% of MAs above). "
                                 f"Structure healthy. Not yet near the 95% CSP floor (${_floor_95:.2f}, "
                                 f"{_otm_from_95:.1f}% OTM).")
                elif sma50 and sma200 and px<sma50 and px<sma200:
                    confidence, conf_col = "WEAK", "#ef5350"
                    conf_desc = (f"Below 50 & 200 SMA — downtrend structure ({trend_score}% of MAs above price). "
                                 f"Wait for stabilisation before selling puts. 95% floor: ${_floor_95:.2f} "
                                 f"({_otm_from_95:.1f}% OTM).")
                elif sma50 and px<sma50 and sma200 and px>sma200:
                    confidence, conf_col = "PULLBACK", "#ffb74d"
                    conf_desc = (f"Below 50 SMA but above 200 — pullback within longer uptrend. "
                                 f"Watch for 50 SMA reclaim. 95% floor: ${_floor_95:.2f} ({_otm_from_95:.1f}% OTM).")
                elif sma50 and px>sma50 and sma200 and px<sma200:
                    confidence, conf_col = "RECOVERY", "#FFD700"
                    conf_desc = (f"Above 50 but below 200 SMA — recovery attempt, {trend_score}% of MAs supportive. "
                                 f"95% floor: ${_floor_95:.2f} ({_otm_from_95:.1f}% OTM).")
                else:
                    confidence, conf_col = "NEUTRAL", "#8ab4d4"
                    conf_desc = f"Insufficient MA data. 95% floor: ${_floor_95:.2f} ({_otm_from_95:.1f}% OTM)."

                # ── Recent U&R signals (last 5 days) ─────────────────────────
                recent_sigs = []
                for _i in range(1,6):
                    if _i+1 >= len(_c): break
                    _cc=_c[-_i]; _cc1=_c[-_i-1]
                    for _ma, _name in [(sma200,"200 SMA"),(sma100,"100 SMA"),(sma50,"50 SMA"),(sma20,"20 SMA"),(sma10,"10 SMA")]:
                        if not _ma: continue
                        if _cc1 < _ma and _cc > _ma:
                            recent_sigs.append(("RECLAIM", _name, _dt[-_i], _cc, _ma, True))
                        elif _cc1 > _ma and _cc < _ma:
                            recent_sigs.append(("UNDERCUT", _name, _dt[-_i], _cc, _ma, False))

                # ── Multi-level Support & Resistance ─────────────────────
                # Each level = percentile of 1-year daily lows/highs
                # 95% = price held above/below 95% of days (most conservative)
                # 75% = held 75% of days (moderate cushion)
                # 50% = median — price was above/below half the time
                _sorted_lo = sorted(_lo)
                _sorted_hi = sorted(_hi)
                _n_bars    = len(_sorted_lo)

                def _pct_idx(pct): return max(0, min(_n_bars-1, int(_n_bars * pct)))

                def _ma_confluence(price):
                    near = []
                    for _mn, _mv in [("50 SMA",sma50),("100 SMA",sma100),("200 SMA",sma200)]:
                        if _mv and abs(_mv - price) / price < 0.02:
                            near.append(_mn)
                    return near

                _sr_levels = []
                for _conf_pct, _lo_pct, _hi_pct, _label, _col in [
                    (95, 0.05, 0.95, "95%", "#00e676"),
                    (75, 0.25, 0.75, "75%", "#FFD700"),
                    (50, 0.50, 0.50, "50%", "#ffb74d"),
                ]:
                    _sp = round(_sorted_lo[_pct_idx(_lo_pct)], 2)
                    _rp = round(_sorted_hi[_pct_idx(_hi_pct)], 2)
                    _sp_otm  = round((px - _sp) / px * 100, 1)
                    _rp_dist = round((_rp - px) / px * 100, 1)
                    _sp_mas  = _ma_confluence(_sp)
                    _rp_mas  = _ma_confluence(_rp)
                    _days_br = sum(1 for l in _lo if l < _sp)
                    _days_ex = sum(1 for h in _hi if h > _rp)
                    _sr_levels.append({
                        "conf":       _conf_pct,
                        "label":      _label,
                        "color":      _col,
                        "sup_price":  _sp,
                        "res_price":  _rp,
                        "sup_otm":    _sp_otm,
                        "res_dist":   _rp_dist,
                        "sup_mas":    _sp_mas,
                        "res_mas":    _rp_mas,
                        "days_below": _days_br,
                        "days_above": _days_ex,
                    })

                # Best CSP = 75% support (enough cushion, realistic premium)
                _best_sup = _sr_levels[1]  # 75%
                _best_res = _sr_levels[0]  # 95%

                # ── Fibonacci Retracement Levels ──────────────────────────────
                # Use the 52-week high and 52-week low as the swing range.
                # Standard Fib levels: 0%, 23.6%, 38.2%, 50%, 61.8%, 78.6%, 100%
                # Price below 50% Fib = in the lower half of the range (support territory)
                # Price near 61.8% (golden ratio) = strongest retracement support
                _fib_swing_hi = wk52_hi
                _fib_swing_lo = wk52_lo
                _fib_range    = _fib_swing_hi - _fib_swing_lo

                # Compute avg volume for ratio baseline
                _avg_vol_fib = sum(_v) / len(_v) if _v else 1

                def _inst_touches(fib_price, tolerance=0.02):
                    """
                    Scan 1y of daily bars for days where:
                      - The low traded within tolerance% of the fib level
                      - Volume was >= 1.2x average (institutions present)
                    Then classify each touch:
                      - BOUNCE: closed above the level that day AND the level has held since
                      - BOUNCE (later broken): closed above that day BUT price eventually broke below
                      - FAILED: closed below the level that day
                    """
                    touches = []
                    _opens  = list(_h.get("Open", _h["Close"]))
                    _closes = list(_h["Close"])
                    _lows   = list(_h["Low"])
                    _vols   = list(_h["Volume"])
                    _dates  = [str(d.date()) for d in _h.index]
                    _level_now_broken = px < fib_price  # current price is below this level

                    for _ii in range(len(_closes)):
                        _lo_i = _lows[_ii]
                        _cl_i = _closes[_ii]
                        _op_i = _opens[_ii]
                        _vl_i = _vols[_ii]
                        _dt_i = _dates[_ii]
                        _vr_i = round(_vl_i / _avg_vol_fib, 2) if _avg_vol_fib else 1
                        if abs(_lo_i - fib_price) / fib_price > tolerance: continue
                        if _vr_i < 1.2: continue
                        _closed_above = _cl_i > fib_price * (1 - tolerance)
                        if _closed_above:
                            # Check if it was eventually broken AFTER this date
                            _later_break = any(
                                _closes[_jj] < fib_price * (1 - tolerance)
                                for _jj in range(_ii + 1, len(_closes))
                            )
                            if _later_break or _level_now_broken:
                                _action  = "Bounce (later broken)"
                                _bullish = False  # ultimately failed
                            else:
                                _action  = "Bounce ✓"
                                _bullish = True
                        else:
                            _action  = "Failed"
                            _bullish = False
                        touches.append({
                            "date":     _dt_i,
                            "low":      round(_lo_i, 2),
                            "close":    round(_cl_i, 2),
                            "vol_ratio":_vr_i,
                            "action":   _action,
                            "bullish":  _bullish,
                        })
                    return sorted(touches, key=lambda x: x["date"], reverse=True)

                _fib_levels = []
                for _fib_pct, _fib_lbl in [
                    (1.000, "0% — 52wk High"),
                    (0.764, "23.6%"),
                    (0.618, "38.2%"),
                    (0.500, "50%"),
                    (0.382, "61.8% — Golden Ratio"),
                    (0.236, "76.4%"),
                    (0.000, "100% — 52wk Low"),
                ]:
                    _fp      = round(_fib_swing_lo + _fib_range * _fib_pct, 2)
                    _fd      = round((px - _fp) / px * 100, 1)
                    _touches = _inst_touches(_fp)
                    _fib_levels.append({
                        "pct":     _fib_pct,
                        "label":   _fib_lbl,
                        "price":   _fp,
                        "dist":    _fd,
                        "above":   px >= _fp,
                        "touches": _touches,        # institutional volume events
                        "n_bounce": sum(1 for t in _touches if t["bullish"]),
                        "n_fail":   sum(1 for t in _touches if not t["bullish"]),
                    })

                # Find which Fib level price is currently closest to
                _fib_nearest = min(_fib_levels, key=lambda f: abs(f["dist"]))

                # Key Fib supports below price (for CSP consideration)
                _fib_supports = [f for f in _fib_levels if f["above"] and f["pct"] < 1.0]
                # Nearest Fib support below price
                _fib_next_sup = max(_fib_supports, key=lambda f: f["price"]) if _fib_supports else None

                MA_DATA[_tk] = {
                    "px":px,"px1":px1,"wk52_hi":wk52_hi,"wk52_lo":wk52_lo,
                    "pct_from_hi":pct_from_hi,"pct_from_lo":pct_from_lo,
                    "vol_ratio":vol_ratio,"trend_score":trend_score,
                    "sma5":sma5,"sma10":sma10,"sma20":sma20,"sma50":sma50,
                    "sma100":sma100,"sma200":sma200,"ema8":ema8,"ema21":ema21,
                    "confidence":confidence,"conf_col":conf_col,"conf_desc":conf_desc,
                    "long_spread":long_spread,"short_spread":short_spread,
                    "conv_signals":conv_signals,"recent_sigs":recent_sigs,
                    "name": _row.get("Name",""),
                    "sr_levels":      _sr_levels,
                    "best_support":   _best_sup,
                    "best_resistance":_best_res,
                    "n_bars":         _n_bars,
                    "fib_levels":     _fib_levels,
                    "fib_nearest":    _fib_nearest,
                    "fib_next_sup":   _fib_next_sup,
                    "fib_swing_hi":   _fib_swing_hi,
                    "fib_swing_lo":   _fib_swing_lo,
                }
            except Exception:
                continue
    except Exception:
        pass

    # ── Render ────────────────────────────────────────────────────────────────
    COLORS = {"AAPL":"#555","MSFT":"#00a4ef","NVDA":"#76b900","AMZN":"#FF9900",
              "GOOGL":"#4285F4","META":"#1877F2","TSLA":"#cc0000"}

    w('<div class="card">')
    w(f'<h3>Mag 7 <span class="fresh">{today.strftime("%b %d, %Y")} — live</span></h3>')

    if not MA_DATA:
        w('<div style="color:#ffb74d;padding:20px">No MA data available — yfinance needed.</div>')
        w('</div>')
        return "\n".join(lines)

    # ── Strategy Summary Table — all 7 stocks ───────────────────────────────
    # For each stock compute:
    #   - Best option strategy (same picker as per-card)
    #   - CSP conviction score (40+30+30)
    #   - Best put strike (95% floor)
    #   - Live premium for that strike (fetched below per-card, so we do it here too)
    import datetime as _dt_top

    def _strat_score(strat_label, d2, l95, l75, fib_next_sup):
        """
        0-100 score measuring conviction in the RECOMMENDED strategy, not just CSP.
        Three components (each 0-33pts, rounded to 100):

        1. Setup quality (33pts) — how clean is the entry signal?
           CSP/BullPutSpread: proximity to 95% floor (nearer = better)
           CoveredCall:       proximity to 52wk high (nearer = better)
           BullCallSpread:    trend strength above 200 SMA
           ProtectivePut:     how far below 200 SMA (deeper = more urgent)
           IronCondor:        how neutral/range-bound (low ADX proxy via trend_score)
           Wait:              always low (by definition not acting)

        2. Confirmation (34pts) — MA structure + Fib confluence agree with setup
           33pts if both trend and Fib support the strategy direction
           20pts if only one confirms
           0pts if neither

        3. Timing (33pts) — how well-timed is the entry?
           Earnings window: penalise if pre-earnings for premium-selling strategies
           CSP: reward if near floor AND IV typically elevated at lows
        """
        _px2      = d2["px"]
        _sma200   = d2.get("sma200")
        _above200 = _sma200 and _px2 > _sma200
        _pct_hi   = d2.get("pct_from_hi", -30)
        _trend    = d2.get("trend_score", 50)
        _conf     = d2.get("confidence","")
        _earn_dt2 = None
        for _etk, _edt in _earn_top.items():
            if _etk == list(d2.keys())[0] if d2 else "": _earn_dt2 = _edt
        _otm95 = l95["sup_otm"] if l95 else 50
        _otm75 = l75["sup_otm"] if l75 else 20

        # ── Component 1: Setup quality ────────────────────────────────────────
        if strat_label in ("CASH-SECURED PUT", "WAIT / DEEP CSP"):
            # Nearer to 95% floor = better setup
            _q = (33 if _otm95 <= 10 else 25 if _otm95 <= 15 else
                  18 if _otm95 <= 25 else 10 if _otm95 <= 35 else 3)
        elif strat_label == "COVERED CALL":
            # Nearer to 52wk high = better setup (capping near resistance)
            _q = (33 if _pct_hi >= -3 else 25 if _pct_hi >= -6 else
                  18 if _pct_hi >= -10 else 10)
        elif strat_label == "BULL CALL SPREAD":
            # Strong trend above 200 SMA = better setup
            _q = (33 if _trend >= 75 else 22 if _trend >= 55 else
                  12 if _trend >= 40 else 5)
        elif strat_label == "PROTECTIVE PUT":
            # Deep below 200 SMA = more urgent need
            _dist200 = abs((_px2 - _sma200) / _sma200 * 100) if _sma200 else 0
            _q = (30 if _dist200 >= 10 else 20 if _dist200 >= 5 else 12)
        elif strat_label == "BULL PUT SPREAD":
            _q = (28 if _otm75 <= 8 else 20 if _otm75 <= 12 else 12)
        elif strat_label == "IRON CONDOR":
            # Range-bound = low trend score is good for condor
            _q = (28 if _trend <= 40 else 18 if _trend <= 55 else 8)
        else:
            _q = 5

        # ── Component 2: Confirmation ─────────────────────────────────────────
        _fib_confirms = False
        if fib_next_sup and l95:
            _fg = abs(fib_next_sup["price"] - l95["sup_price"]) / l95["sup_price"] * 100
            _fib_confirms = _fg <= 5
        _ma_confirms = (
            (_above200 and strat_label in ("COVERED CALL","BULL CALL SPREAD","BULL PUT SPREAD","IRON CONDOR"))
            or (not _above200 and strat_label in ("PROTECTIVE PUT","WAIT / DEEP CSP"))
            or ("CSP ZONE" in _conf and strat_label == "CASH-SECURED PUT")
        )
        _c = (34 if (_fib_confirms and _ma_confirms) else
              20 if (_fib_confirms or _ma_confirms) else 5)

        # ── Component 3: Timing ───────────────────────────────────────────────
        # Premium-selling strategies are penalised if pre-earnings (<21d)
        # Directional strategies are rewarded pre-earnings
        _earn_days = min((_edt - _today_top).days for _edt in _earn_top.values()
                         if hasattr(_edt, "strftime")) if _earn_top else 999
        # Get this stock's earnings specifically
        _stk_earn  = 999
        for _etk2, _edt2 in _earn_top.items():
            pass  # we'll use per-card earn_dt; use 999 as safe fallback
        _pre_earn_t = _earn_days <= 21

        if strat_label in ("CASH-SECURED PUT","COVERED CALL","IRON CONDOR","BULL PUT SPREAD"):
            # Selling premium: penalise pre-earnings (IV spike = risk)
            _t = 5 if _pre_earn_t else 33
        elif strat_label in ("BULL CALL SPREAD","PROTECTIVE PUT"):
            # Directional: reward pre-earnings (IV = defined cost, big move expected)
            _t = 30 if _pre_earn_t else 15
        else:
            _t = 10

        return min(100, _q + _c + _t)

    _earn_top = {
        "AMZN": _dt_top.date(2026,4,30), "MSFT": _dt_top.date(2026,4,28),
        "AAPL": _dt_top.date(2026,5,1),  "NVDA": _dt_top.date(2026,5,28),
        "GOOGL":_dt_top.date(2026,4,29), "META": _dt_top.date(2026,4,23),
        "TSLA": _dt_top.date(2026,4,22),
    }
    _today_top = _dt_top.date.today()

    def _pick_strat(tk2, d2, l95, l75):
        """Return (icon, label, color, why_short) for best strategy."""
        _px2        = d2["px"]
        _earn_dt2   = _earn_top.get(tk2)
        _days_earn2 = (_earn_dt2 - _today_top).days if _earn_dt2 else 999
        _pre_earn2  = _days_earn2 <= 21
        _above_200  = d2.get("sma200") and _px2 > d2["sma200"]
        _pct_hi2    = d2.get("pct_from_hi", -30)
        _trend2     = d2.get("trend_score", 50)
        _near_95_2  = l95 and l95["sup_otm"] <= 15
        if _near_95_2 and not _pre_earn2:
            return "💰","CASH-SECURED PUT","#00e676",f"Near 95% floor ({l95['sup_otm']:.1f}% OTM) — highest conviction CSP zone"
        elif _above_200 and _pct_hi2 >= -8 and not _pre_earn2:
            return "📞","COVERED CALL","#42a5f5",f"Near 52wk high ({_pct_hi2:+.1f}%) — sell calls to capture premium at the top"
        elif _pre_earn2 and _above_200:
            return "🚀","BULL CALL SPREAD","#FFD700",f"Earnings {_earn_dt2.strftime('%b %d')} in {_days_earn2}d — defined risk play into bullish print"
        elif _pre_earn2 and not _above_200:
            return "🛡️","PROTECTIVE PUT","#ffb74d",f"Earnings {_earn_dt2.strftime('%b %d')} in {_days_earn2}d — below 200 SMA, hedge binary risk"
        elif not _above_200 and _trend2 < 40:
            return "⏳","WAIT / DEEP CSP","#ef9a9a","Below 200 SMA, weak trend — wait or use deep OTM put only"
        elif _above_200 and _pct_hi2 < -15:
            return "📊","BULL PUT SPREAD","#7e57c2",f"Pullback in uptrend ({_pct_hi2:+.1f}% off high) — defined risk income at support"
        else:
            return "🦅","IRON CONDOR","#8ab4d4","Range-bound — sell premium on both sides"

    _strat_rows = []
    for _row in mag_rows:
        _tk2 = _row.get("Ticker","").lstrip("$")
        _d2  = MA_DATA.get(_tk2)
        if not _d2 or not _d2.get("sr_levels"): continue
        _srl2  = _d2["sr_levels"]
        _l95_2 = next((l for l in _srl2 if l["conf"]==95), None)
        _l75_2 = next((l for l in _srl2 if l["conf"]==75), None)
        _fns2  = _d2.get("fib_next_sup")
        _si2_tmp, _sl2_tmp, _, _ = _pick_strat(_tk2, _d2, _l95_2, _l75_2)
        _sc2   = _strat_score(_sl2_tmp, _d2, _l95_2, _l75_2, _fns2)
        _si2, _sl2, _sc_col2, _sw2 = _pick_strat(_tk2, _d2, _l95_2, _l75_2)
        _gc2 = "#00e676" if _sc2 >= 70 else ("#FFD700" if _sc2 >= 40 else "#ef5350")
        _strat_rows.append({
            "tk":_tk2,"px":_d2["px"],"conf":_d2["confidence"],"cc":_d2["conf_col"],
            "l95":_l95_2,"l75":_l75_2,"score":_sc2,"score_col":_gc2,
            "icon":_si2,"strat":_sl2,"strat_col":_sc_col2,"why":_sw2,
        })
    # Sort: highest CSP score first
    _strat_rows.sort(key=lambda x: -x["score"])

    w('<div style="background:#0a1628;border-radius:12px;padding:16px 20px;margin-bottom:20px">')
    w('<div style="font-size:14px;font-weight:900;color:#FFD700;margin-bottom:4px">'
      '📊 Mag 7 — Options Strategy Summary</div>')
    w('<div style="font-size:11px;color:#5a7fa0;margin-bottom:12px">'
      'Best strategy for each stock based on price position, MA structure, earnings proximity, and support levels. '
      'Score 0-100: setup quality (33pts) + confirmation signals (34pts) + timing (33pts). '
      'Scroll down for full detail on each stock.</div>')
    w('<div style="overflow-x:auto">')
    w('<table style="width:100%;border-collapse:collapse;font-size:12px">')
    w('<thead><tr style="background:#0d2040;border-bottom:2px solid #1e3a5f">')
    for _h in ["#","Ticker","Price","Best Strategy","Score","95% Put Strike","OTM %","75% Strike","Why","Chart"]:
        w(f'<th style="padding:7px 10px;color:#7eb8f7;font-weight:700;text-align:right;white-space:nowrap">{_h}</th>')
    w('</tr></thead><tbody>')

    for _ri2, _sr in enumerate(_strat_rows, 1):
        _col2 = COLORS.get(_sr["tk"], "#aaa")
        _tv2  = f"https://www.tradingview.com/chart/?symbol={_sr['tk']}&interval=D"
        _l95r = _sr["l95"]; _l75r = _sr["l75"]
        _row_bg2 = "background:#0a1200;" if _sr["score"] >= 70 else ("background:#0a1628;" if _sr["score"] >= 40 else "")
        w(f'<tr style="border-bottom:1px solid #0d2040;{_row_bg2}">')
        w(f'<td style="padding:7px 10px;font-weight:900;color:#5a7fa0;text-align:right">#{_ri2}</td>')
        w(f'<td style="padding:7px 10px;font-weight:900;color:{_col2};text-align:right;font-size:14px">{_sr["tk"]}</td>')
        w(f'<td style="padding:7px 10px;color:#e8edf5;font-weight:700;text-align:right">${_sr["px"]:.2f}</td>')
        w(f'<td style="padding:7px 10px;text-align:right">'
          f'<span style="background:{_sr["strat_col"]};color:#000;font-size:11px;font-weight:900;padding:3px 10px;border-radius:5px">'
          f'{_sr["icon"]} {_sr["strat"]}</span></td>')
        w(f'<td style="padding:7px 10px;color:{_sr["score_col"]};font-weight:900;font-size:16px;text-align:right">{_sr["score"]}</td>')
        if _l95r:
            _otm_col2 = "#00e676" if _l95r["sup_otm"] <= 15 else "#8ab4d4"
            w(f'<td style="padding:7px 10px;color:#00e676;font-weight:900;text-align:right">${_l95r["sup_price"]:.2f}</td>')
            w(f'<td style="padding:7px 10px;color:{_otm_col2};font-weight:700;text-align:right">{_l95r["sup_otm"]:.1f}%</td>')
        else:
            w('<td style="color:#555;text-align:right">—</td><td style="color:#555;text-align:right">—</td>')
        if _l75r:
            w(f'<td style="padding:7px 10px;color:#FFD700;text-align:right">${_l75r["sup_price"]:.2f} ({_l75r["sup_otm"]:.1f}%)</td>')
        else:
            w('<td style="color:#555;text-align:right">—</td>')
        w(f'<td style="padding:7px 10px;color:#8ab4d4;font-size:11px;text-align:right;max-width:240px">{esc(_sr["why"])}</td>')
        w(f'<td style="padding:7px 10px;text-align:right"><a href="{_tv2}" target="_blank" style="color:#42a5f5;font-size:11px">↗</a></td>')
        w('</tr>')

    w('</tbody></table></div></div>')
    w('<div style="font-size:11px;color:#5a7fa0;margin-bottom:16px;padding:0 4px">'
      '⚡ Score ≥70 = HIGH conviction (green). 40-69 = MODERATE. &lt;40 = WATCH. '
      'Scroll down for full analysis on each stock.</div>')

    # ── Open CSP positions ────────────────────────────────────────────────────
    _open_puts = [p for p in OPEN_PUTS if not p.get("closed", False)]
    if _open_puts:
        _put_colors = {"AAPL":"#555","MSFT":"#00a4ef","NVDA":"#76b900","AMZN":"#FF9900",
                       "GOOGL":"#4285F4","META":"#1877F2","TSLA":"#cc0000"}
        w('<div style="background:#0a1628;border-radius:12px;padding:16px 20px;margin-bottom:20px">')
        w('<div style="font-size:14px;font-weight:900;color:#00e676;margin-bottom:12px">Open Cash-Secured Puts</div>')
        w('<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;font-size:12px">')
        w('<thead><tr style="background:#0d2040;border-bottom:2px solid #1e3a5f">')
        for _h in ["Date","Ticker","Strike","Expiry","Qty","Premium","Total","Option Now","Gain %","P&L","Notes"]:
            w(f'<th style="padding:7px 10px;color:#7eb8f7;font-weight:700;text-align:right;white-space:nowrap">{_h}</th>')
        w('</tr></thead><tbody>')
        for _p in _open_puts:
            _ptk  = _p["ticker"]
            _pcol = _put_colors.get(_ptk, "#aaa")
            _ppx  = MA_DATA.get(_ptk, {}).get("px")
            _pmid = None
            try:
                import yfinance as _yf_put
                _ptc = _yf_put.Ticker(_ptk)
                _ptc.history(period="2d")
                _pex = _p["exp_yf"]
                if _pex in (list(_ptc.options or [])):
                    _pch = _ptc.option_chain(_pex).puts
                    if not _pch.empty:
                        _prow = _pch[abs(_pch["strike"] - _p["strike"]) < 0.5]
                        if not _prow.empty:
                            _pb = float(_prow.iloc[0].get("bid",0) or 0)
                            _pa = float(_prow.iloc[0].get("ask",0) or 0)
                            if _pb > 0 and _pa > 0: _pmid = round((_pb+_pa)/2, 2)
            except Exception:
                pass
            _pgain = round((_p["premium"] - _pmid) / _p["premium"] * 100, 1) if _pmid else None
            _ppnl  = round((_p["premium"] - (_pmid or _p["premium"])) * 100 * _p["contracts"], 2)
            _pgc   = "#00e676" if (_pgain or 0) >= 0 else "#ef5350"
            _otm   = round((_ppx - _p["strike"]) / _ppx * 100, 1) if _ppx else None
            w(f'<tr style="border-bottom:1px solid #0d2040">')
            w(f'<td style="padding:7px 10px;color:#8ab4d4;text-align:right">{_p["date"]}</td>')
            w(f'<td style="padding:7px 10px;color:{_pcol};font-weight:900;text-align:right">{_ptk}</td>')
            _otm_badge = f' <span style="font-size:10px;color:#8ab4d4">({_otm:.1f}% OTM)</span>' if _otm else ""
            w(f'<td style="padding:7px 10px;color:#FFD700;font-weight:900;text-align:right">${_p["strike"]:.2f}{_otm_badge}</td>')
            w(f'<td style="padding:7px 10px;color:#8ab4d4;text-align:right;white-space:nowrap">{_p["expiry"]}</td>')
            w(f'<td style="padding:7px 10px;color:#e8edf5;text-align:right">{_p["contracts"]}x</td>')
            w(f'<td style="padding:7px 10px;color:#00e676;text-align:right">${_p["premium"]:.2f}</td>')
            w(f'<td style="padding:7px 10px;color:#00e676;font-weight:700;text-align:right">${_p["total_premium"]:,.0f}</td>')
            if _pmid is not None:
                w(f'<td style="padding:7px 10px;color:#e8edf5;text-align:right">${_pmid:.2f}</td>')
                w(f'<td style="padding:7px 10px;color:{_pgc};font-weight:800;text-align:right">{_pgain:+.1f}%</td>')
                w(f'<td style="padding:7px 10px;color:{_pgc};font-weight:800;text-align:right">${_ppnl:+,.0f}</td>')
            else:
                w('<td style="padding:7px 10px;color:#555;text-align:right">—</td>')
                w('<td style="padding:7px 10px;color:#555;text-align:right">—</td>')
                w('<td style="padding:7px 10px;color:#555;text-align:right">—</td>')
            w(f'<td style="padding:7px 10px;color:#8ab4d4;font-size:11px;text-align:right">{_p.get("notes","")}</td>')
            w('</tr>')
        w('</tbody></table></div></div>')

    for _row in mag_rows:
        _tk = _row.get("Ticker","").lstrip("$")
        d   = MA_DATA.get(_tk)
        if not d: continue

        col = COLORS.get(_tk,"#aaa")
        px  = d["px"]

        w(f'<div style="background:#0d1b2a;border-left:4px solid {col};border-radius:12px;padding:18px 22px;margin-bottom:16px">')

        # ── Header ────────────────────────────────────────────────────────────
        w(f'<div style="display:flex;align-items:center;gap:14px;flex-wrap:wrap;margin-bottom:14px">')
        w(f'<span style="font-size:24px;font-weight:900;color:{col}">{_tk}</span>')
        w(f'<span style="font-size:14px;color:#8ab4d4">{esc(d["name"])}</span>')
        w(f'<span style="font-size:20px;font-weight:900;color:#fff">${px:.2f}</span>')
        chg = round((px-d["px1"])/d["px1"]*100,2)
        chg_col = "#00e676" if chg>=0 else "#ef5350"
        w(f'<span style="font-size:13px;color:{chg_col};font-weight:700">{chg:+.2f}% today</span>')
        # Confidence badge
        w(f'<span style="background:{d["conf_col"]};color:#000;font-size:12px;font-weight:900;padding:3px 14px;border-radius:10px;margin-left:auto">{d["confidence"]} CONFIDENCE</span>')
        w('</div>')

        # ── Confidence description ─────────────────────────────────────────────
        w(f'<div style="background:#0a1628;border-left:3px solid {d["conf_col"]};border-radius:7px;padding:9px 14px;margin-bottom:12px;font-size:12px;color:#dde8f8">{esc(d["conf_desc"])}</div>')

        # ── Stats grid ────────────────────────────────────────────────────────
        w('<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:8px;margin-bottom:12px">')
        # 52-week range
        hi_col = "#00e676" if d["pct_from_hi"] > -5 else ("#ffb74d" if d["pct_from_hi"] > -20 else "#ef5350")
        lo_col = "#ef5350" if d["pct_from_lo"] < 20 else "#ffb74d"
        for lbl,val,vc in [
            ("52-wk High",  f"${d['wk52_hi']:.2f} ({d['pct_from_hi']:+.1f}%)", hi_col),
            ("52-wk Low",   f"${d['wk52_lo']:.2f} (+{d['pct_from_lo']:.1f}%)", lo_col),
            ("Vol Ratio",   f"{d['vol_ratio']:.2f}×", "#00e676" if d["vol_ratio"]>=1.5 else ("#FFD700" if d["vol_ratio"]>=1.0 else "#ef5350")),
            ("MAs Above",   f"{d['trend_score']}%", "#00e676" if d["trend_score"]>=70 else ("#FFD700" if d["trend_score"]>=40 else "#ef5350")),
        ]:
            w(f'<div style="background:#13263d;border-radius:7px;padding:8px 10px">')
            w(f'<div style="font-size:10px;color:#5a7fa0;margin-bottom:2px">{lbl}</div>')
            w(f'<div style="font-size:13px;font-weight:800;color:{vc}">{val}</div></div>')
        w('</div>')



        # ── MA Convergence alerts ─────────────────────────────────────────────
        if d["conv_signals"]:
            for cs in d["conv_signals"]:
                w(f'<div style="background:#1a1200;border-left:3px solid #FFD700;border-radius:7px;padding:8px 13px;margin-bottom:8px;font-size:12px;color:#FFD700;font-weight:700">{esc(cs)}</div>')

        # ── Support & Resistance levels ───────────────────────────────────────
        _sr  = d.get("sr_levels", [])
        _tv  = f"https://www.tradingview.com/chart/?symbol={_tk}&interval=D"
        _nb  = d.get("n_bars", 252)

        if _sr:
            w('<div style="margin-bottom:12px">')
            w('<div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:8px">')
            w('<span style="font-size:12px;font-weight:900;color:#e8edf5">Support &amp; Resistance — Confidence Intervals</span>')
            w(f'<a href="{_tv}" target="_blank" style="font-size:11px;color:#42a5f5;text-decoration:none;background:#0d2040;padding:3px 10px;border-radius:5px">📈 Chart ↗</a>')
            w('</div>')

            # Multi-level table
            w('<div style="overflow-x:auto">')
            w('<table style="width:100%;border-collapse:collapse;font-size:12px;margin-bottom:14px">')
            w('<thead><tr style="background:#0d2040;border-bottom:2px solid #1e3a5f">')
            for _h in ["Confidence","Support (floor)","OTM %","Days below","Resistance (ceiling)","Dist %","Days above","MA Confluence"]:
                w(f'<th style="padding:7px 10px;color:#7eb8f7;font-weight:700;text-align:right;white-space:nowrap">{_h}</th>')
            w('</tr></thead><tbody>')

            for _lvl in _sr:
                _c   = _lvl["color"]
                _sp  = _lvl["sup_price"]; _rp = _lvl["res_price"]
                _csp_flag = "✅ " if _lvl["conf"] == 75 else ""
                _mas_str  = ", ".join(_lvl["sup_mas"] + _lvl["res_mas"]) or "—"
                w(f'<tr style="border-bottom:1px solid #0d2040">')
                w(f'<td style="padding:7px 10px;text-align:right"><span style="background:{_c};color:#000;font-size:11px;font-weight:900;padding:2px 9px;border-radius:5px">{_lvl["label"]}</span></td>')
                w(f'<td style="padding:7px 10px;color:#FFD700;font-weight:900;text-align:right">{_csp_flag}${_sp:.2f}</td>')
                w(f'<td style="padding:7px 10px;color:{_c};text-align:right">{_lvl["sup_otm"]:.1f}%</td>')
                w(f'<td style="padding:7px 10px;color:#8ab4d4;text-align:right">{_lvl["days_below"]}d / {_nb}d</td>')
                w(f'<td style="padding:7px 10px;color:#ef9a9a;font-weight:700;text-align:right">${_rp:.2f}</td>')
                w(f'<td style="padding:7px 10px;color:#ef5350;text-align:right">+{_lvl["res_dist"]:.1f}%</td>')
                w(f'<td style="padding:7px 10px;color:#8ab4d4;text-align:right">{_lvl["days_above"]}d / {_nb}d</td>')
                w(f'<td style="padding:7px 10px;color:#FFD700;font-size:11px;text-align:right">{esc(_mas_str)}</td>')
                w('</tr>')
            w('</tbody></table></div>')

            # Explanation
            w('<div style="font-size:11px;color:#5a7fa0;margin-bottom:4px">')
            w('95% = price held above support / below resistance 95% of the past year. ')
            w('75% = held 75% of days — closer to current price, better premium, slightly less safe. ')
            w('✅ = recommended CSP strike level.')
            w('</div>')
            w('</div>')

        # ── Recent U&R signals ────────────────────────────────────────────────
        if d["recent_sigs"]:
            w('<div style="margin-bottom:4px">')
            for sig_type, ma_name, sig_date, sig_px, ma_v, bullish in d["recent_sigs"][:3]:
                sc = "#00e676" if bullish else "#ef5350"
                icon = "🟢" if bullish else "🔴"
                w(f'<div style="background:{"#0a2010" if bullish else "#1a0a0a"};border-left:3px solid {sc};border-radius:6px;padding:7px 12px;margin-bottom:5px;font-size:11px">')
                w(f'<span style="color:{sc};font-weight:900">{icon} {sig_type} {ma_name}</span>')
                w(f' <span style="color:#8ab4d4">on {sig_date} at ${sig_px:.2f} (MA: ${ma_v:.2f})</span>')
                w('</div>')
            w('</div>')

        # ── Fibonacci Retracement ─────────────────────────────────────────────
        _fibs     = d.get("fib_levels", [])
        _fib_near = d.get("fib_nearest")
        _fib_nsup = d.get("fib_next_sup")
        _fhi      = d.get("fib_swing_hi", 0)
        _flo      = d.get("fib_swing_lo", 0)

        if _fibs:
            w('<div style="margin-bottom:12px">')
            w(f'<div style="font-size:12px;font-weight:900;color:#e8edf5;margin-bottom:6px">'
              f'Fibonacci Retracements '
              f'<span style="font-size:11px;font-weight:400;color:#5a7fa0">'
              f'52wk swing: ${_flo:.2f} → ${_fhi:.2f}</span></div>')

            w('<div style="overflow-x:auto">')
            w('<table style="width:100%;border-collapse:collapse;font-size:12px">')
            w('<thead><tr style="background:#0d2040;border-bottom:2px solid #1e3a5f">')
            for _fh in ["Level","Price","Dist from Price","Position","Confluence","Bounces/Fails","Institutional Evidence (vol ≥1.2× avg)"]:
                w(f'<th style="padding:6px 10px;color:#7eb8f7;font-weight:700;text-align:right;white-space:nowrap">{_fh}</th>')
            w('</tr></thead><tbody>')

            for _fl in _fibs:
                _is_near   = abs(_fl["dist"]) <= 3.0
                _is_above  = _fl["above"]
                _is_golden = "61.8" in _fl["label"]
                _is_50     = _fl["label"] == "50%"
                _touches   = _fl.get("touches", [])
                _n_b       = _fl.get("n_bounce", 0)
                _n_f       = _fl.get("n_fail", 0)

                # Color: green = support (below price), red = resistance (above price), gold = golden ratio
                if _is_golden:
                    _fc = "#FFD700"
                elif _is_50:
                    _fc = "#42a5f5"
                elif _is_above:
                    _fc = "#00e676"  # support below price
                else:
                    _fc = "#ef9a9a"  # resistance above price

                _row_bg = "background:#0d2a18;" if (_is_near and _is_above) else (
                          "background:#1a0a0a;" if (_is_near and not _is_above) else "")

                # MA confluence
                _fib_mas = []
                for _mn, _mv in [("50 SMA",d.get("sma50")),("100 SMA",d.get("sma100")),
                                  ("200 SMA",d.get("sma200")),("21 EMA",d.get("ema21"))]:
                    if _mv and abs(_mv - _fl["price"]) / _fl["price"] < 0.015:
                        _fib_mas.append(_mn)

                _pos_lbl = "SUPPORT ▲" if _is_above else "RESISTANCE ▼"
                _pos_col = "#00e676" if _is_above else "#ef9a9a"

                # Build institutional evidence bullet string
                _inst_bullets = []
                for _t in _touches[:4]:
                    if _t["action"] == "Bounce ✓":
                        _ic = "#00e676"; _sym = "✓"
                    elif _t["action"] == "Bounce (later broken)":
                        _ic = "#ffb74d"; _sym = "⚠"
                    else:
                        _ic = "#ef5350"; _sym = "✗"
                    _inst_bullets.append(
                        f'<span style="color:{_ic}">{_sym} {_t["date"]}: '
                        f'low ${_t["low"]:.2f} → close ${_t["close"]:.2f} '
                        f'({_t["vol_ratio"]:.1f}× vol) — {_t["action"]}</span>'
                    )
                _inst_html = "<br>".join(_inst_bullets) if _inst_bullets else "—"
                _touch_summary = ""
                if _touches:
                    _n_clean  = sum(1 for t in _touches if t["action"] == "Bounce ✓")
                    _n_broken = sum(1 for t in _touches if t["action"] == "Bounce (later broken)")
                    _n_fail   = sum(1 for t in _touches if t["action"] == "Failed")
                    _parts = []
                    if _n_clean:  _parts.append(f'<span style="color:#00e676;font-weight:700">{_n_clean} held</span>')
                    if _n_broken: _parts.append(f'<span style="color:#ffb74d;font-weight:700">{_n_broken} broken</span>')
                    if _n_fail:   _parts.append(f'<span style="color:#ef5350;font-weight:700">{_n_fail} failed</span>')
                    _touch_summary = '<span style="color:#5a7fa0"> / </span>'.join(_parts)

                w(f'<tr style="border-bottom:1px solid #0d2040;{_row_bg}">')
                w(f'<td style="padding:6px 10px;font-weight:800;color:{_fc};text-align:right;white-space:nowrap">{_fl["label"]}</td>')
                w(f'<td style="padding:6px 10px;font-weight:900;color:#FFD700;text-align:right">${_fl["price"]:.2f}</td>')
                w(f'<td style="padding:6px 10px;color:{"#00e676" if _fl["dist"]>=0 else "#ef5350"};text-align:right">'
                  f'{"+" if _fl["dist"]>=0 else ""}{_fl["dist"]:.1f}%</td>')
                w(f'<td style="padding:6px 10px;text-align:right">'
                  f'<span style="background:{_pos_col};color:#000;font-size:10px;font-weight:800;padding:1px 7px;border-radius:4px">{_pos_lbl}</span></td>')
                w(f'<td style="padding:6px 10px;color:#FFD700;font-size:11px;text-align:right">'
                  f'{"★ " + ", ".join(_fib_mas) if _fib_mas else "—"}</td>')
                w(f'<td style="padding:6px 10px;text-align:right;font-size:11px">{_touch_summary}</td>')
                w(f'<td style="padding:6px 10px;font-size:11px;min-width:220px;line-height:1.7">{_inst_html}</td>')
                w('</tr>')

            w('</tbody></table></div>')

            # Nearest support summary
            if _fib_nsup:
                _fib_sup_mas = []
                for _mn, _mv in [("50 SMA",d.get("sma50")),("200 SMA",d.get("sma200"))]:
                    if _mv and abs(_mv - _fib_nsup["price"]) / _fib_nsup["price"] < 0.02:
                        _fib_sup_mas.append(_mn)
                _fib_note_col = "#FFD700" if "61.8" in _fib_nsup["label"] else "#00e676"
                w(f'<div style="background:#0d1b2a;border-left:3px solid {_fib_note_col};border-radius:6px;padding:8px 12px;margin-top:8px;font-size:11px;color:#d0e4f7">')
                w(f'<strong style="color:{_fib_note_col}">Nearest Fib support:</strong> '
                  f'${_fib_nsup["price"]:.2f} ({_fib_nsup["label"]}) — '
                  f'{abs(_fib_nsup["dist"]):.1f}% below current price.')
                if _fib_sup_mas:
                    w(f' Confluent with {", ".join(_fib_sup_mas)} — double confirmation.')
                w(' This is the next logical level to watch for a bounce or CSP placement.')
                w('</div>')
            w('</div>')

        # ── Options Strategy Recommendation ──────────────────────────────────
        # Picks the best strategy based on current price vs key levels, trend, earnings.
        # Strategies: Cash-Secured Put · Covered Call · Bull Call Spread ·
        #             Bull Put Spread · Protective Put · Iron Condor ·
        #             Long Call (breakout) · Short Straddle (pre-earnings IV sell)
        import datetime as _dt_put

        _lvl95 = next((l for l in d.get("sr_levels",[]) if l["conf"]==95), None)
        _lvl75 = next((l for l in d.get("sr_levels",[]) if l["conf"]==75), None)
        _today_dt = _dt_put.date.today()
        _earn_dates = {
            "AMZN": _dt_put.date(2026, 4, 30), "MSFT": _dt_put.date(2026, 4, 28),
            "AAPL": _dt_put.date(2026, 5, 1),  "NVDA": _dt_put.date(2026, 5, 28),
            "GOOGL":_dt_put.date(2026, 4, 29), "META": _dt_put.date(2026, 4, 23),
            "TSLA": _dt_put.date(2026, 4, 22),
        }
        _earn_dt     = _earn_dates.get(_tk)
        _days_to_earn = (_earn_dt - _today_dt).days if _earn_dt else 999
        _trend_sc    = d.get("trend_score", 50)
        _near_95     = _lvl95 and _lvl95["sup_otm"] <= 15
        _near_75     = _lvl75 and _lvl75["sup_otm"] <= 10
        _pct_from_hi = d.get("pct_from_hi", -20)
        _above_200   = (d.get("sma200") and px > d["sma200"])
        _pre_earn    = 0 < _days_to_earn <= 21           # within 3 weeks of earnings
        _vol_ratio   = d.get("vol_ratio", 1.0) or 1.0

        # ── Strategy picker ───────────────────────────────────────────────────
        if _near_95 and not _pre_earn:
            _strat = "CASH-SECURED PUT"
            _strat_col = "#00e676"
            _strat_icon = "💰"
            _strat_why = (f"{_tk} is within 15% of its 95% annual support floor — the statistical bottom. "
                          f"Sell a 120-180 DTE put at the 95% strike to collect maximum premium "
                          f"at the strongest support level. Assignment means buying at a deep discount.")
            _dte_target = 150; _otm_target = _lvl95["sup_otm"] if _lvl95 else 10
            _use_puts = True

        elif _above_200 and _pct_from_hi >= -5 and not _pre_earn:
            # Near 52wk high, strong uptrend — stock breaking out
            _strat = "LONG CALL (BREAKOUT)"
            _strat_col = "#00e676"
            _strat_icon = "🚀"
            _strat_why = (f"{_tk} is {_pct_from_hi:+.1f}% from its 52-week high and above the 200 SMA — "
                          f"breakout momentum. Buy a 30-60 DTE slightly OTM call to participate in the "
                          f"continuation with defined risk. Cheaper than buying shares outright and profits "
                          f"accelerate if the stock pushes to new highs.")
            _dte_target = 45; _otm_target = 3
            _use_puts = False

        elif _above_200 and _pct_from_hi >= -8 and not _pre_earn:
            _strat = "COVERED CALL"
            _strat_col = "#42a5f5"
            _strat_icon = "📞"
            _strat_why = (f"{_tk} is near its 52-week high ({_pct_from_hi:+.1f}%) and above the 200 SMA — "
                          f"strong uptrend but extended. Sell a 30-45 DTE covered call "
                          f"5-8% OTM to collect premium while giving the position room to run.")
            _dte_target = 38; _otm_target = 6
            _use_puts = False

        elif _pre_earn and _above_200 and _days_to_earn >= 7:
            # Pre-earnings with bullish bias — buy the call spread, don't sell premium naked
            _strat = "BULL CALL SPREAD"
            _strat_col = "#FFD700"
            _strat_icon = "📈"
            _strat_why = (f"Earnings in {_days_to_earn} days ({_earn_dt.strftime('%b %d') if _earn_dt else '—'}). "
                          f"{_tk} is above its 200 SMA — bullish bias into the print. "
                          f"Buy ATM call + sell OTM call to cap cost while participating in an upside beat. "
                          f"Avoid selling naked premium into the IV spike — IV crush will hurt sellers after the print.")
            _dte_target = _days_to_earn + 7; _otm_target = 8
            _use_puts = False

        elif _pre_earn and _days_to_earn <= 7:
            # Very close to earnings — sell the IV spike via straddle if IV is elevated
            _strat = "WAIT FOR POST-EARNINGS"
            _strat_col = "#ffb74d"
            _strat_icon = "⏸️"
            _strat_why = (f"Earnings in {_days_to_earn} day{'s' if _days_to_earn!=1 else ''} "
                          f"({_earn_dt.strftime('%b %d') if _earn_dt else '—'}). "
                          f"Too close to initiate new positions — IV is maximally inflated and "
                          f"any strategy has binary risk. Wait until the day after earnings, "
                          f"then sell calls (if gap up) or cash-secured puts (if gap down) into elevated IV.")
            _dte_target = 38; _otm_target = 8
            _use_puts = True

        elif _pre_earn and not _above_200:
            _strat = "PROTECTIVE PUT"
            _strat_col = "#ef9a9a"
            _strat_icon = "🛡️"
            _strat_why = (f"Earnings in {_days_to_earn} days — {_tk} is below the 200 SMA (weak structure). "
                          f"If you hold shares, buy a short-dated protective put at the 75% support level "
                          f"to cap downside through the binary event. "
                          f"Avoid selling premium — IV is elevated and actual downside risk is real.")
            _dte_target = _days_to_earn + 14; _otm_target = _lvl75["sup_otm"] if _lvl75 else 12
            _use_puts = True

        elif not _above_200 and _trend_sc < 40:
            _strat = "WAIT — CASH-SECURED PUT (DEEP OTM)"
            _strat_col = "#ef9a9a"
            _strat_icon = "⏳"
            _strat_why = (f"{_tk} is below the 200 SMA with weak trend ({_trend_sc}% of MAs bullish). "
                          f"Don't sell calls — you'd cap upside on a recovering stock. "
                          f"Only consider a deep OTM put at the 95% floor if you'd be happy owning at that price. "
                          f"Otherwise wait for the 200 SMA reclaim as confirmation before taking a directional position.")
            _dte_target = 45; _otm_target = _lvl95["sup_otm"] if _lvl95 else 20
            _use_puts = True

        elif _above_200 and _pct_from_hi < -15:
            _strat = "BULL PUT SPREAD"
            _strat_col = "#7e57c2"
            _strat_icon = "📊"
            _strat_why = (f"{_tk} is {_pct_from_hi:.1f}% off its high but above the 200 SMA — "
                          f"pullback within an uptrend. Sell a put at the 75% support and "
                          f"buy a lower put as a hedge. Collects net premium with defined max loss. "
                          f"Better risk profile than a naked CSP when the stock is mid-range.")
            _dte_target = 45; _otm_target = _lvl75["sup_otm"] if _lvl75 else 12
            _use_puts = True

        else:
            _strat = "IRON CONDOR"
            _strat_col = "#8ab4d4"
            _strat_icon = "🦅"
            _strat_why = (f"{_tk} is range-bound — not near a major floor or extended to new highs. "
                          f"An iron condor (sell OTM call + sell OTM put, buy further wings) "
                          f"profits if the stock stays in its current range through expiry. "
                          f"Target the 75% support as put wing and equivalent distance as call wing.")
            _dte_target = 30; _otm_target = 8
            _use_puts = True

        # Fetch live option data — both legs for multi-leg strategies
        _sug_exp = None; _sug_dte = None; _sug_exp_yf = None
        _leg1 = {}   # buy leg (long)
        _leg2 = {}   # sell leg (short) — populated for multi-leg strategies
        _is_multileg = _strat in ("BULL CALL SPREAD","BULL PUT SPREAD","IRON CONDOR","PROTECTIVE PUT / HOLD")

        try:
            import yfinance as _yf_sg
            _tsg = _yf_sg.Ticker(_tk)
            _tsg.history(period="2d")
            _all_exps = list(_tsg.options or [])
            _best_exp = None; _best_dte_v = None
            for _ex in _all_exps:
                try:
                    _ed = _dt_put.datetime.strptime(_ex, "%Y-%m-%d").date()
                    _dte_v = (_ed - _today_dt).days
                    if _dte_v < 7: continue
                    if _strat in ("BULL CALL SPREAD","PROTECTIVE PUT / HOLD"):
                        if _dte_v < _days_to_earn: continue
                        if _dte_v > _days_to_earn + 21: continue
                    elif _strat in ("CASH-SECURED PUT","BULL PUT SPREAD"):
                        if _dte_v < 90 or _dte_v > 210: continue
                    else:
                        if _dte_v < 21 or _dte_v > 60: continue
                    if _best_dte_v is None or abs(_dte_v - _dte_target) < abs(_best_dte_v - _dte_target):
                        _best_exp = _ex; _best_dte_v = _dte_v
                except Exception: continue

            if _best_exp:
                _sug_exp_yf = _best_exp
                _sug_dte    = _best_dte_v
                _sug_exp    = _dt_put.datetime.strptime(_best_exp,"%Y-%m-%d").date().strftime("%b %d, %Y")
                _chain      = _tsg.option_chain(_best_exp)
                _calls_df   = _chain.calls if not _chain.calls.empty else None
                _puts_df    = _chain.puts  if not _chain.puts.empty  else None

                def _get_leg(df, target_strike):
                    if df is None or df.empty: return {}
                    _diff = abs(df["strike"] - target_strike)
                    _nr = df.loc[_diff.idxmin()]
                    _pb = float(_nr.get("bid",0) or 0)
                    _pa = float(_nr.get("ask",0) or 0)
                    if _pb <= 0 or _pa <= 0: return {}
                    return {
                        "strike": float(_nr["strike"]),
                        "mid":    round((_pb+_pa)/2, 2),
                        "bid":    _pb, "ask": _pa,
                        "iv":     round(float(_nr.get("impliedVolatility",0) or 0)*100, 1),
                        "oi":     int(_nr.get("openInterest",0) or 0),
                    }

                if _strat == "CASH-SECURED PUT":
                    _sk1 = round(px * (1 - _otm_target/100), 0)
                    _leg1 = _get_leg(_puts_df, _sk1)  # sell this put

                elif _strat == "COVERED CALL":
                    _sk1 = round(px * (1 + _otm_target/100), 0)
                    _leg1 = _get_leg(_calls_df, _sk1)  # sell this call

                elif _strat == "BULL CALL SPREAD":
                    # Buy ATM call, sell OTM call ~8% higher
                    _sk1 = round(px, 0)                       # buy leg (ATM)
                    _sk2 = round(px * 1.08, 0)                # sell leg (OTM)
                    _leg1 = _get_leg(_calls_df, _sk1)         # long call
                    _leg2 = _get_leg(_calls_df, _sk2)         # short call

                elif _strat == "BULL PUT SPREAD":
                    # Sell put at 75% level, buy put 8% lower
                    _sk1 = round(px * (1 - (_otm_target)/100), 0)       # sell leg (upper)
                    _sk2 = round(px * (1 - (_otm_target+8)/100), 0)     # buy leg (lower wing)
                    _leg1 = _get_leg(_puts_df, _sk1)          # short put (higher)
                    _leg2 = _get_leg(_puts_df, _sk2)          # long put (lower)

                elif _strat == "IRON CONDOR":
                    # Sell OTM call + OTM put, buy further wings
                    _call_sell = round(px * 1.08, 0)
                    _call_buy  = round(px * 1.13, 0)
                    _put_sell  = round(px * 0.92, 0)
                    _put_buy   = round(px * 0.87, 0)
                    _lc1 = _get_leg(_calls_df, _call_sell)    # short call
                    _lc2 = _get_leg(_calls_df, _call_buy)     # long call wing
                    _lp1 = _get_leg(_puts_df,  _put_sell)     # short put
                    _lp2 = _get_leg(_puts_df,  _put_buy)      # long put wing
                    # Store as leg1=short side net, leg2=long side net
                    _ic_credit = round(
                        (_lc1.get("mid",0) - _lc2.get("mid",0)) +
                        (_lp1.get("mid",0) - _lp2.get("mid",0)), 2)
                    _leg1 = {"strike": _call_sell, "mid": _lc1.get("mid",0),
                             "strike2": _put_sell, "mid2": _lp1.get("mid",0),
                             "wing_call": _call_buy, "wing_put": _put_buy,
                             "net_credit": _ic_credit,
                             "lc1":_lc1,"lc2":_lc2,"lp1":_lp1,"lp2":_lp2}

                elif _strat == "PROTECTIVE PUT / HOLD":
                    # Buy ATM or slight OTM put as hedge
                    _sk1 = round(px * 0.97, 0)
                    _leg1 = _get_leg(_puts_df, _sk1)

                elif _strat in ("WAIT / CASH-SECURED PUT (LOWER STRIKE)", "WAIT / DEEP CSP"):
                    _sk1 = round(px * (1 - (_otm_target)/100), 0)
                    _leg1 = _get_leg(_puts_df, _sk1)

        except Exception: pass

        # ── Render strategy box ───────────────────────────────────────────────
        _s_border = f"1px solid {_strat_col}44"
        w(f'<div style="background:#0d1b2a;border:{_s_border};border-left:4px solid {_strat_col};border-radius:10px;padding:14px 16px;margin-top:4px">')
        w(f'<div style="display:flex;align-items:center;gap:10px;margin-bottom:12px;flex-wrap:wrap">')
        w(f'<span style="font-size:11px;font-weight:700;color:#5a7fa0">📐 OPTIONS STRATEGY</span>')
        w(f'<span style="font-size:14px;font-weight:900;color:{_strat_col}">{_strat_icon} {_strat}</span>')
        if _earn_dt:
            _ec = "#ef5350" if _pre_earn else "#5a7fa0"
            w(f'<span style="font-size:11px;color:{_ec}">Earnings: {_earn_dt.strftime("%b %d")} ({_days_to_earn}d away)</span>')
        if _sug_exp:
            w(f'<span style="font-size:11px;color:#8ab4d4">Expiry: <strong style="color:#e8edf5">{_sug_exp}</strong> ({_sug_dte}d)</span>')
        w('</div>')

        # ── Strategy-specific metrics ─────────────────────────────────────────
        if _strat == "BULL CALL SPREAD" and _leg1 and _leg2:
            _net_debit = round(_leg1["mid"] - _leg2["mid"], 2)
            _spread_w  = round(_leg2["strike"] - _leg1["strike"], 0)
            _max_profit= round((_spread_w - _net_debit) * 100, 0)
            _max_loss  = round(_net_debit * 100, 0)
            _be_price  = round(_leg1["strike"] + _net_debit, 2)
            _rr        = round(_max_profit / _max_loss, 1) if _max_loss else 0
            w('<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:7px;margin-bottom:12px">')
            for _ml,_mv,_mc,_ms in [
                ("Buy (long) call",  f"${_leg1['strike']:.0f} call @ ${_leg1['mid']:.2f}", "#00e676", "11px"),
                ("Sell (short) call",f"${_leg2['strike']:.0f} call @ ${_leg2['mid']:.2f}", "#ef9a9a", "11px"),
                ("Net debit",        f"${_net_debit:.2f}/sh = ${_max_loss:.0f}/contract", "#FFD700", "12px"),
                ("Spread width",     f"${_spread_w:.0f}",                                 "#8ab4d4", "12px"),
                ("Max profit",       f"${_max_profit:.0f}/contract",                      "#00e676", "13px"),
                ("Max loss",         f"${_max_loss:.0f}/contract",                        "#ef5350", "13px"),
                ("Breakeven",        f"${_be_price:.2f}",                                 "#e8edf5", "12px"),
                ("Risk/Reward",      f"1 : {_rr:.1f}",                                   "#FFD700", "12px"),
            ]:
                w(f'<div style="background:#0a1628;border-radius:6px;padding:7px 10px">')
                w(f'<div style="font-size:10px;color:#5a7fa0;margin-bottom:2px">{_ml}</div>')
                w(f'<div style="font-size:{_ms};font-weight:800;color:{_mc}">{_mv}</div>')
                w('</div>')
            w('</div>')
            # Payoff summary
            w('<div style="background:#0a1200;border-radius:6px;padding:9px 12px;margin-bottom:10px;font-size:11px;line-height:1.8">')
            w(f'<div style="color:#5a7fa0;font-weight:700;margin-bottom:4px">PAYOFF AT EXPIRY</div>')
            w(f'<span style="color:#ef5350">▼ Below ${_leg1["strike"]:.0f}:</span> Max loss <strong style="color:#ef5350">-${_max_loss:.0f}</strong> &nbsp;|&nbsp; ')
            w(f'<span style="color:#8ab4d4">◆ Breakeven:</span> <strong style="color:#e8edf5">${_be_price:.2f}</strong> &nbsp;|&nbsp; ')
            w(f'<span style="color:#00e676">▲ Above ${_leg2["strike"]:.0f}:</span> Max profit <strong style="color:#00e676">+${_max_profit:.0f}</strong>')
            w('</div>')

        elif _strat == "BULL PUT SPREAD" and _leg1 and _leg2:
            _net_credit= round(_leg1["mid"] - _leg2["mid"], 2)
            _spread_w  = round(_leg1["strike"] - _leg2["strike"], 0)
            _max_profit= round(_net_credit * 100, 0)
            _max_loss  = round((_spread_w - _net_credit) * 100, 0)
            _be_price  = round(_leg1["strike"] - _net_credit, 2)
            _rr        = round(_max_profit / _max_loss, 1) if _max_loss else 0
            w('<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:7px;margin-bottom:12px">')
            for _ml,_mv,_mc,_ms in [
                ("Sell (short) put",  f"${_leg1['strike']:.0f} put @ ${_leg1['mid']:.2f}", "#00e676", "11px"),
                ("Buy (long) put",    f"${_leg2['strike']:.0f} put @ ${_leg2['mid']:.2f}", "#ef9a9a", "11px"),
                ("Net credit",        f"${_net_credit:.2f}/sh = ${_max_profit:.0f}/contract", "#00e676", "12px"),
                ("Max profit",        f"${_max_profit:.0f} (keep credit if above ${_leg1['strike']:.0f})", "#00e676", "11px"),
                ("Max loss",          f"${_max_loss:.0f} (if below ${_leg2['strike']:.0f})",  "#ef5350", "11px"),
                ("Breakeven",         f"${_be_price:.2f}",                                    "#e8edf5", "12px"),
                ("Spread width",      f"${_spread_w:.0f}",                                    "#8ab4d4", "12px"),
                ("Risk/Reward",       f"1 : {_rr:.1f}",                                       "#FFD700", "12px"),
            ]:
                w(f'<div style="background:#0a1628;border-radius:6px;padding:7px 10px">')
                w(f'<div style="font-size:10px;color:#5a7fa0;margin-bottom:2px">{_ml}</div>')
                w(f'<div style="font-size:{_ms};font-weight:800;color:{_mc}">{_mv}</div>')
                w('</div>')
            w('</div>')
            w('<div style="background:#0a1200;border-radius:6px;padding:9px 12px;margin-bottom:10px;font-size:11px;line-height:1.8">')
            w(f'<div style="color:#5a7fa0;font-weight:700;margin-bottom:4px">PAYOFF AT EXPIRY</div>')
            w(f'<span style="color:#ef5350">▼ Below ${_leg2["strike"]:.0f}:</span> Max loss <strong style="color:#ef5350">-${_max_loss:.0f}</strong> &nbsp;|&nbsp; ')
            w(f'<span style="color:#8ab4d4">◆ Breakeven:</span> <strong style="color:#e8edf5">${_be_price:.2f}</strong> &nbsp;|&nbsp; ')
            w(f'<span style="color:#00e676">▲ Above ${_leg1["strike"]:.0f}:</span> Max profit <strong style="color:#00e676">+${_max_profit:.0f}</strong>')
            w('</div>')

        elif _strat == "IRON CONDOR" and _leg1 and _leg1.get("net_credit") is not None:
            _ic = _leg1
            _nc = _ic.get("net_credit",0)
            _lc1d = _ic.get("lc1",{}); _lc2d = _ic.get("lc2",{})
            _lp1d = _ic.get("lp1",{}); _lp2d = _ic.get("lp2",{})
            _call_w = round((_ic.get("wing_call",0) - _ic.get("strike",0)), 0)
            _put_w  = round((_ic.get("strike2",0)   - _ic.get("wing_put",0)), 0)
            _max_loss_ic = round((max(_call_w,_put_w) - _nc) * 100, 0)
            _max_prof_ic = round(_nc * 100, 0)
            w('<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:7px;margin-bottom:12px">')
            for _ml,_mv,_mc,_ms in [
                ("Sell call",  f"${_ic.get('strike',0):.0f} @ ${_lc1d.get('mid',0):.2f}",  "#ef9a9a", "11px"),
                ("Buy call ▲", f"${_ic.get('wing_call',0):.0f} @ ${_lc2d.get('mid',0):.2f}","#8ab4d4","11px"),
                ("Sell put",   f"${_ic.get('strike2',0):.0f} @ ${_lp1d.get('mid',0):.2f}",  "#ef9a9a","11px"),
                ("Buy put ▼",  f"${_ic.get('wing_put',0):.0f} @ ${_lp2d.get('mid',0):.2f}", "#8ab4d4","11px"),
                ("Net credit", f"${_nc:.2f}/sh = ${_max_prof_ic:.0f}/contract",               "#00e676","12px"),
                ("Max profit", f"${_max_prof_ic:.0f} (stock stays in range)",                  "#00e676","11px"),
                ("Max loss",   f"${_max_loss_ic:.0f} (breaks outside wings)",                  "#ef5350","11px"),
                ("Profit zone",f"${_ic.get('strike2',0):.0f} — ${_ic.get('strike',0):.0f}",   "#FFD700","11px"),
            ]:
                w(f'<div style="background:#0a1628;border-radius:6px;padding:7px 10px">')
                w(f'<div style="font-size:10px;color:#5a7fa0;margin-bottom:2px">{_ml}</div>')
                w(f'<div style="font-size:{_ms};font-weight:800;color:{_mc}">{_mv}</div>')
                w('</div>')
            w('</div>')

        elif _leg1:
            # Single-leg strategies: CSP, Covered Call, Protective Put
            _mid = _leg1["mid"]; _sk = _leg1["strike"]
            _is_credit = _strat in ("CASH-SECURED PUT","COVERED CALL",
                                    "WAIT / CASH-SECURED PUT (LOWER STRIKE)","WAIT / DEEP CSP")
            _ann = round(_mid / _sk * (365 / _sug_dte) * 100, 1) if (_sug_dte and _is_credit) else None
            _be  = round(_sk - _mid, 2) if _strat in ("CASH-SECURED PUT","WAIT / CASH-SECURED PUT (LOWER STRIKE)","WAIT / DEEP CSP") else round(_sk + _mid, 2)
            _iv  = _leg1.get("iv",0)
            _action = "Sell" if _is_credit else "Buy"
            _leg_type = "put" if _strat in ("CASH-SECURED PUT","PROTECTIVE PUT / HOLD","WAIT / CASH-SECURED PUT (LOWER STRIKE)","WAIT / DEEP CSP") else "call"
            w(f'<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(120px,1fr));gap:7px;margin-bottom:10px">')
            for _ml,_mv,_mc in [
                ("Action",      f"{_action} {_tk} ${_sk:.0f} {_leg_type}", _strat_col),
                ("Mid premium", f"${_mid:.2f}/share",         "#00e676" if _is_credit else "#ef9a9a"),
                ("Per contract",f"${_mid*100:.0f}",           "#00e676" if _is_credit else "#ef9a9a"),
                ("Breakeven",   f"${_be:.2f}",                "#e8edf5"),
                ("Ann. return", f"~{_ann:.1f}%" if _ann else "—", "#FFD700" if _ann else "#555"),
                ("IV",          f"{_iv:.0f}%" if _iv else "—","#8ab4d4"),
            ]:
                w(f'<div style="background:#0a1628;border-radius:6px;padding:7px 10px">')
                w(f'<div style="font-size:10px;color:#5a7fa0;margin-bottom:2px">{_ml}</div>')
                w(f'<div style="font-size:12px;font-weight:800;color:{_mc}">{_mv}</div>')
                w('</div>')
            w('</div>')

        else:
            w('<div style="color:#ffb74d;font-size:12px;margin-bottom:10px">Live option prices unavailable — check broker for current chain.</div>')

        w(f'<div style="font-size:11px;color:#8ab4d4;line-height:1.7;border-top:1px solid #0d2040;padding-top:8px">')
        w(f'<strong style="color:{_strat_col}">Why this strategy:</strong> {esc(_strat_why)}')
        w('</div></div>')

        w('</div>')  # end ticker card

    w('</div>')  # end card
    return "\n".join(lines), MA_DATA



def run_historical_signals(tickers: List[str], months_back: int = 3, min_score: int = 90) -> List[dict]:
    """
    Scan the last N months of daily bars for all tickers.
    Returns every bullish signal (U&R Reclaim + SMC Sweep/BOS) scored >= min_score.
    Uses the same 0-100 scoring as live signals.
    Pulls 2y of data so MAs have enough history even for signals 3 months ago.
    """
    import datetime as _hdt
    clean  = [t.lstrip("$") for t in tickers if t]

    # Use disk cache if fresh (avoids re-fetching 2y for 200 tickers every run)
    try:
        import json as _jh2
        if os.path.exists(YF_HIST_CACHE):
            _hage = time.time() - os.path.getmtime(YF_HIST_CACHE)
            if _hage < YF_HIST_CACHE_TTL:
                _hcached = _jh2.loads(Path(YF_HIST_CACHE).read_text(encoding="utf-8"))
                # Filter by score AND ensure Jan 1 2026 through today
                _today_str = _hdt.date.today().isoformat()
                _hcached = [r for r in _hcached
                            if (r.get("momentum_score") or 0) >= min_score
                            and r.get("signal_date","") >= "2026-01-01"
                            and r.get("signal_date","") <= _today_str]
                print(f"   historical: using cache ({int(_hage/60)}m old) — {len(_hcached)} signals")
                return _hcached
    except Exception:
        pass

    # Fetch 2y so we have enough data for 200-SMA even at the start of the lookback window
    print(f"   historical: fetching 2y OHLCV for {len(clean)} tickers...")
    try:
        import yfinance as _yf_h
        import pandas as _pd_h
        _df_h = _yf_h.download(
            clean, period="2y", group_by="ticker",
            auto_adjust=True, progress=False, threads=True,
        )
    except Exception as _e:
        print(f"   historical: fetch failed — {_e}")
        return []

    # Date range: Jan 1 2026 through today (inclusive)
    _cutoff_start = _hdt.date(2026, 1, 1)
    _cutoff_end   = _hdt.date.today() + _hdt.timedelta(days=1)  # inclusive of today

    def _extract(tk):
        try:
            if len(clean) == 1:
                df = _df_h
            else:
                df = _df_h[tk] if tk in _df_h.columns.get_level_values(0) else None
            if df is None or df.empty: return None
            df = df.dropna(subset=["Close"])
            return {
                "close":  list(df["Close"]),
                "open":   list(df["Open"]),
                "high":   list(df["High"]),
                "low":    list(df["Low"]),
                "volume": list(df["Volume"]),
                "dates":  [str(d.date()) for d in df.index],
            }
        except Exception:
            return None

    # 0-100 scoring matching live signals
    def _score(vr, bp, dist_abs):
        v = 40 if vr>=2.5 else (30 if vr>=2.0 else (20 if vr>=1.5 else (10 if vr>=1.0 else 0)))
        b = 30 if bp>=80  else (22 if bp>=60  else (12 if bp>=40  else 0))
        d = 30 if dist_abs>=2.0 else (20 if dist_abs>=1.0 else (10 if dist_abs>=0.3 else 0))
        return v + b + d

    def _grade(s):
        if s >= 70: return "A","#00e676"
        if s >= 50: return "B","#FFD700"
        if s >= 30: return "C","#ffb74d"
        return "F","#ef5350"

    results = []

    for tk in clean:
        data = _extract(tk)
        if not data: continue

        closes  = data["close"]
        opens   = data["open"]
        highs   = data["high"]
        lows    = data["low"]
        volumes = data["volume"]
        dates   = data["dates"]
        n       = len(closes)

        if n < 210: continue  # need enough for 200-SMA

        avg_vol_all = sum(volumes) / n if n else 1

        # ── Scan every bar in the 2026 window ─────────────────────────────────
        # Compute CSP conviction for this ticker once
        _hcsp = compute_csp_conviction({
            "close": closes, "low": lows, "high": highs, "open": opens
        })
        for i in range(1, n - 1):
            dt_str = dates[i] if i < len(dates) else ""
            try:
                bar_dt = _hdt.datetime.strptime(dt_str, "%Y-%m-%d").date()
            except Exception:
                continue
            if bar_dt < _cutoff_start or bar_dt > _cutoff_end:
                continue

            c  = closes[i];   c1 = closes[i-1]
            h  = highs[i];    l  = lows[i]
            o  = opens[i]
            v  = volumes[i]

            avg_v = sum(volumes[max(0,i-21):i]) / min(20, i) if i > 0 else v
            vr    = round(v / avg_v, 2) if avg_v else 1.0
            br    = h - l
            bp    = round(abs(c - o) / br * 100, 1) if br > 0 else 0

            # ── U&R Reclaims ──────────────────────────────────────────────────
            for period, label in [(50,"50-Day"),(100,"100-Day"),(200,"200-Day")]:
                if i < period + 2: continue
                # Use full history up to and including today's bar for accurate SMA
                _ma_arr = sma(closes[:i+1], period)
                if not _ma_arr or len(_ma_arr) < 2: continue
                ma_now  = _ma_arr[-1]
                ma_prev = _ma_arr[-2]
                if not ma_now or not ma_prev: continue
                dist = round((c - ma_now) / ma_now * 100, 2)
                # Reclaim: prev close was below prev MA, today's close above today's MA
                if c1 < ma_prev and c > ma_now:
                    sc = _score(vr, bp, abs(dist))
                    if sc < min_score: continue
                    gr, gc = _grade(sc)
                    ud = round((ma_prev - c1) / ma_prev * 100, 2)
                    _nxt_o = round(opens[i+1], 2) if i+1 < len(opens) else None
                    _nxt_d = dates[i+1] if i+1 < len(dates) else None
                    _gap_p = round((_nxt_o - c) / c * 100, 2) if _nxt_o else None
                    # 5-day % change: close on day i+5 vs close on signal day
                    _5d_close = round(closes[i+5], 2) if i+5 < len(closes) else None
                    _5d_date  = dates[i+5] if i+5 < len(dates) else None
                    _5d_chg   = round((_5d_close - c) / c * 100, 2) if _5d_close else None
                    _30d_close = round(closes[i+21], 2) if i+21 < len(closes) else None
                    _30d_date  = dates[i+21] if i+21 < len(dates) else None
                    _30d_chg   = round((_30d_close - c) / c * 100, 2) if _30d_close else None
                    results.append({
                        "ticker":         f"${tk}",
                        "is_index":       tk in INDEX_TICKERS,
                        "ticker_name":    INDEX_NAMES.get(tk, ""),
                        "csp_conviction": _hcsp,
                        "category":       f"Undercut & Reclaim — {label} SMA",
                        "signal":         f"Reclaim {label} SMA",
                        "signal_date":    dt_str,
                        "trigger_time":   "4:00 PM EST",
                        "trigger_note":   f"Daily close — {label} MA reclaim confirmed",
                        "next_open":      _nxt_o,
                        "next_open_date": _nxt_d,
                        "gap_pct":        _gap_p,
                        "fiveday_close":  _5d_close,
                        "fiveday_date":   _5d_date,
                        "fiveday_chg":    _5d_chg,
                        "thirtyDay_close": _30d_close,
                        "thirtyDay_date":  _30d_date,
                        "thirtyDay_chg":   _30d_chg,
                        "bullish":        True,
                        "momentum_score": sc,
                        "momentum_grade": gr,
                        "momentum_color": gc,
                        "vol_ratio":      vr,
                        "body_pct":       bp,
                        "ma_dist_pct":    dist,
                        "undercut_depth": ud,
                        "ma_level":       round(ma_now, 2),
                        "close_price":    round(c, 2),
                        "description": (
                            f"Reclaimed {label} SMA (${ma_now:.2f}) — "
                            f"prev close ${c1:.2f} below, closed ${c:.2f} above (+{dist:.2f}%). "
                            f"Vol {vr:.1f}× avg | Body {bp:.0f}% | Undercut depth {ud:.2f}%."
                        ),
                        "smc": False,
                    })

            # ── SMC: Liquidity Sweep + Demand Bounce ──────────────────────────
            # Find prior swing low (last 5 bars, 3-bar pivot)
            if i >= 5:
                _swing_lows = []
                for _si in range(max(0,i-40), i-1):  # 40-bar lookback
                    _n = 3
                    if _si < _n or _si >= len(lows)-_n: continue
                    if lows[_si] <= min(lows[max(0,_si-_n):_si] + lows[_si+1:_si+_n+1]):
                        _swing_lows.append(lows[_si])
                if _swing_lows:
                    _sl = _swing_lows[-1]
                    if l < _sl and c > _sl and vr >= 1.2:
                        _big_rev = (c - l) / br >= 0.6 if br > 0 else False
                        # For sweeps use wick_ratio instead of body% —
                        # a long lower wick is the signal, not a fat body
                        _wick_ratio = round((c - l) / br * 100, 1) if br > 0 else bp
                        sc = _score(vr, _wick_ratio, round((c - _sl) / _sl * 100, 2))
                        if sc >= min_score:
                            gr, gc = _grade(sc)
                            _nxt_os = round(opens[i+1], 2) if i+1 < len(opens) else None
                            _nxt_ds = dates[i+1] if i+1 < len(dates) else None
                            _gap_ps = round((_nxt_os - c) / c * 100, 2) if _nxt_os else None
                            _5d_cs  = round(closes[i+5], 2) if i+5 < len(closes) else None
                            _5d_dts = dates[i+5] if i+5 < len(dates) else None
                            _5d_chs = round((_5d_cs - c) / c * 100, 2) if _5d_cs else None
                            _30d_cs  = round(closes[i+21], 2) if i+21 < len(closes) else None
                            _30d_dts = dates[i+21] if i+21 < len(dates) else None
                            _30d_chs = round((_30d_cs - c) / c * 100, 2) if _30d_cs else None
                            results.append({
                                "ticker":         f"${tk}",
                                "is_index":       tk in INDEX_TICKERS,
                                "ticker_name":    INDEX_NAMES.get(tk, ""),
                                "csp_conviction": _hcsp,
                                "category":       "SMC — Liquidity Sweep",
                                "signal":         "Sweep + Demand Bounce",
                                "signal_date":    dt_str,
                                "trigger_time":   "4:00 PM EST",
                                "trigger_note":   "Daily close — sweep + reversal confirmed",
                                "next_open":      _nxt_os,
                                "next_open_date": _nxt_ds,
                                "gap_pct":        _gap_ps,
                                "fiveday_close":  _5d_cs,
                                "fiveday_date":   _5d_dts,
                                "fiveday_chg":    _5d_chs,
                                "thirtyDay_close": _30d_cs,
                                "thirtyDay_date":  _30d_dts,
                                "thirtyDay_chg":   _30d_chs,
                                "bullish":        True,
                                "momentum_score": sc,
                                "momentum_grade": gr,
                                "momentum_color": gc,
                                "vol_ratio":      vr,
                                "body_pct":       _wick_ratio,  # wick ratio for sweeps
                                "ma_dist_pct":    round((c - _sl) / _sl * 100, 2),
                                "undercut_depth": round((_sl - l) / _sl * 100, 2),
                                "close_price":    round(c, 2),
                                "description": (
                                    f"Low (${l:.2f}) swept below prior swing low (${_sl:.2f}), "
                                    f"reversed to close ${c:.2f}. "
                                    f"Vol {vr:.1f}× avg | Wick recovery {_wick_ratio:.0f}% of bar. "
                                    f"{'Strong reversal — institutions absorbed sellers.' if _big_rev else 'Reversal — watch for follow-through.'}"
                                ),
                                "smc": True,
                            })

    results.sort(key=lambda x: (x["signal_date"], -(x["momentum_score"] or 0)), reverse=True)
    # Cache results to disk
    try:
        import json as _jh
        Path(YF_HIST_CACHE).write_text(_jh.dumps(results), encoding="utf-8")
    except Exception:
        pass
    print(f"   historical: found {len(results)} signals scored >={min_score} in 2026")
    return results


def render_historical_signals_table(signals: List[dict], min_score: int = 90) -> str:
    """Historical signal table — all bullish signals scored >= min_score in 2026."""
    lines = []
    def w(*parts): lines.append("".join(str(p) for p in parts))

    w('<div class="card">')
    w(f'<h3>&#128200; Historical Signals — 2026 &nbsp;'
      f'<span class="fresh">Score &ge;70 | Bullish only | Jan 1 – today</span></h3>')


    w('''<div style="background:#0a1628;border-radius:10px;padding:14px 18px;margin-bottom:16px;font-size:11px;line-height:1.8">
<div style="font-size:13px;font-weight:900;color:#FFD700;margin-bottom:10px">📋 Historical Signal Criteria</div>
<div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px">
<div><div style="color:#00e676;font-weight:800;margin-bottom:4px">What is scanned</div>
<div style="color:#8ab4d4">Every trading day from Jan 1 2026 to today. Same signal types as the live Signals tab: U&R Reclaims on 50/100/200 SMA and SMC Liquidity Sweeps.</div></div>
<div><div style="color:#FFD700;font-weight:800;margin-bottom:4px">Score threshold: ≥70 (Grade A/B)</div>
<div style="color:#8ab4d4">Volume (40pts) + Body% (30pts) + Distance from MA (30pts). Only strong, high-volume reclaims and sweeps make the cut.</div></div>
<div><div style="color:#42a5f5;font-weight:800;margin-bottom:4px">5D % Change column</div>
<div style="color:#8ab4d4">The close 5 trading days after the signal vs the signal-day close. <span style="color:#00e676">▲ Green</span> = stock followed through. <span style="color:#ef5350">▼ Red</span> = failed. <em>Pending</em> = signal too recent.</div></div>
<div><div style="color:#e8edf5;font-weight:800;margin-bottom:4px">Trigger Time: 4:00 PM EST</div>
<div style="color:#8ab4d4">All signals are confirmed at the daily close. You would have seen this setup after 4pm on the signal date. The Next Day Open is your earliest actionable entry.</div></div>
<div><div style="color:#e8edf5;font-weight:800;margin-bottom:4px">Next Day Open</div>
<div style="color:#8ab4d4">The opening price the following morning — your real entry if you acted on the signal. The Gap column shows overnight drift from close to open.</div></div>
<div><div style="color:#5a7fa0;font-weight:800;margin-bottom:4px">Use this tab to</div>
<div style="color:#8ab4d4">Back-test signal quality. How many ≥70 signals in 2026 were followed by a positive 5D return? That's your win rate for the strategy.</div></div>
</div></div>''')
    if not signals:
        w(f'<div style="color:#8ab4d4;padding:20px">No signals scored &ge;{min_score} found in 2026.</div>')
        w('</div>')
        return "\n".join(lines)

    # ── Portfolio P&L summary ─────────────────────────────────────────────────
    # Assume equal-size position in each signal ($10,000 per trade)
    _pos_size  = 10000
    _resolved  = [s for s in signals if s.get("thirtyDay_chg") is not None]
    _wins      = [s for s in _resolved if (s.get("thirtyDay_chg") or 0) > 0]
    _losses    = [s for s in _resolved if (s.get("thirtyDay_chg") or 0) <= 0]
    _win_rate  = round(len(_wins) / len(_resolved) * 100, 1) if _resolved else 0
    _avg_win   = round(sum(s["thirtyDay_chg"] for s in _wins) / len(_wins), 2) if _wins else 0
    _avg_loss  = round(sum(s["thirtyDay_chg"] for s in _losses) / len(_losses), 2) if _losses else 0
    _total_pnl = round(sum((s.get("thirtyDay_chg") or 0) / 100 * _pos_size for s in _resolved), 0)
    _total_deployed = len(_resolved) * _pos_size
    _total_return = round(_total_pnl / _total_deployed * 100, 2) if _total_deployed else 0
    _pending   = len(signals) - len(_resolved)

    w('<div style="background:#0a1628;border-radius:10px;padding:14px 18px;margin-bottom:16px">')
    w('<div style="font-size:13px;font-weight:900;color:#FFD700;margin-bottom:10px">'
      '💼 Portfolio Performance — $10,000 equal-size per signal</div>')
    w('<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:8px;margin-bottom:8px">')
    _pnl_col = "#00e676" if _total_pnl >= 0 else "#ef5350"
    _wr_col  = "#00e676" if _win_rate >= 60 else ("#FFD700" if _win_rate >= 50 else "#ef5350")
    for _sl, _sv, _sc in [
        ("Total Signals",      str(len(signals)),                "#e8edf5"),
        ("Resolved (30d)",     str(len(_resolved)),              "#8ab4d4"),
        ("Pending",            str(_pending),                    "#5a7fa0"),
        ("Win Rate (30d)",     f"{_win_rate:.1f}%",              _wr_col),
        ("Wins",               str(len(_wins)),                  "#00e676"),
        ("Losses",             str(len(_losses)),                "#ef5350"),
        ("Avg Win",            f"+{_avg_win:.2f}%" if _wins else "—",  "#00e676"),
        ("Avg Loss",           f"{_avg_loss:.2f}%" if _losses else "—","#ef5350"),
        ("Total P&L",          f"${_total_pnl:,.0f}",           _pnl_col),
        ("Total Return",       f"{_total_return:+.2f}%",        _pnl_col),
    ]:
        w(f'<div style="background:#0d1b2a;border-radius:7px;padding:9px 12px">')
        w(f'<div style="font-size:10px;color:#5a7fa0;margin-bottom:2px">{_sl}</div>')
        w(f'<div style="font-size:17px;font-weight:900;color:{_sc}">{_sv}</div></div>')
    w('</div>')
    w(f'<div style="font-size:11px;color:#5a7fa0">'
      f'Equal-size model: ${_pos_size:,} per signal · {len(_resolved)} resolved trades · '
      f'${_total_deployed:,} total deployed · Win = stock up after 30 trading days vs signal close.</div>')
    w('</div>')

    # Summary strip
    _tickers  = sorted(set(s["ticker"] for s in signals))
    _ur_count = sum(1 for s in signals if not s.get("smc"))
    _smc_count= sum(1 for s in signals if s.get("smc"))
    _a_count  = sum(1 for s in signals if s.get("momentum_grade")=="A")
    w('<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:8px;margin-bottom:16px">')
    for _sl, _sv, _sc in [
        ("Total Signals",    str(len(signals)),   "#e8edf5"),
        ("U&R Reclaims",     str(_ur_count),      "#00e676"),
        ("SMC Sweeps",       str(_smc_count),     "#7e57c2"),
        ("Grade A (90-100)", str(_a_count),        "#FFD700"),
        ("Unique Tickers",   str(len(_tickers)),  "#42a5f5"),
    ]:
        w(f'<div style="background:#0d1b2a;border-radius:7px;padding:9px 12px">')
        w(f'<div style="font-size:10px;color:#5a7fa0;margin-bottom:2px">{_sl}</div>')
        w(f'<div style="font-size:18px;font-weight:900;color:{_sc}">{_sv}</div></div>')
    w('</div>')

    # Table
    w('<div style="overflow-x:auto">')
    w('<table style="width:100%;border-collapse:collapse;font-size:12px">')
    w('<thead><tr style="background:#0d2040;border-bottom:2px solid #1e3a5f">')
    for _h in ["Signal Date","Trigger EST","Next Day Open","Gap","5D %","30D %","W/L","Ticker","Signal","Type","Score","Grade","Vol","Body%","MA Dist%","Close","Description"]:
        w(f'<th style="padding:7px 10px;color:#7eb8f7;font-weight:700;text-align:right;white-space:nowrap">{_h}</th>')
    w('</tr></thead><tbody>')

    for s in signals:
        _sc2   = s.get("momentum_score", 0)
        _gr    = s.get("momentum_grade","—")
        _gc    = s.get("momentum_color","#aaa")
        _vr    = s.get("vol_ratio")
        _bp    = s.get("body_pct")
        _dist  = s.get("ma_dist_pct")
        _is_smc= s.get("smc", False)
        _next_o = s.get("next_open")
        _next_d = s.get("next_open_date","")
        _gap    = s.get("gap_pct")
        _gap_col= "#00e676" if (_gap or 0) >= 0 else "#ef5350"
        _5d_chg = s.get("fiveday_chg")
        _5d_dt  = s.get("fiveday_date","")
        _5d_col = "#00e676" if (_5d_chg or 0) >= 0 else "#ef5350"
        _tnote  = s.get("trigger_note","Daily close")
        _type_badge = ('<span style="background:#7e57c2;color:#fff;font-size:10px;font-weight:800;'
                       'padding:1px 7px;border-radius:4px">SMC</span>' if _is_smc else
                       '<span style="background:#1a5c2a;color:#00e676;font-size:10px;font-weight:800;'
                       'padding:1px 7px;border-radius:4px">U&R</span>')
        _tv = f'<a href="https://www.tradingview.com/chart/?symbol={s["ticker"].lstrip("$")}&interval=D" target="_blank" style="color:#42a5f5">↗</a>'
        _sc2_val = s.get("momentum_score", 0) or 0
        _row_bg = "background:#0a1200;border-left:3px solid #FFD700;" if _sc2_val >= 90 else ""
        w(f'<tr style="border-bottom:1px solid #0d2040;{_row_bg}">')
        # Date + trigger note
        w(f'<td style="padding:7px 10px;text-align:right">')
        w(f'  <div style="font-weight:700;color:#e8edf5">{s["signal_date"]}</div>')
        w(f'  <div style="font-size:10px;color:#5a7fa0">{esc(_tnote)}</div>')
        w('</td>')
        # 4:00 PM EST
        w('<td style="padding:7px 10px;text-align:right">')
        w('  <div style="font-weight:800;color:#FFD700">4:00 PM EST</div>')
        w('  <div style="font-size:10px;color:#5a7fa0">close confirms signal</div>')
        w('</td>')
        # Next day open
        w('<td style="padding:7px 10px;text-align:right">')
        if _next_o:
            w(f'  <div style="font-weight:800;color:#42a5f5">${_next_o:.2f}</div>')
            w(f'  <div style="font-size:10px;color:#5a7fa0">{_next_d} open</div>')
        else:
            w('  <div style="color:#555">—</div>')
        w('</td>')
        # Gap
        if _gap is not None:
            w(f'<td style="padding:7px 10px;color:{_gap_col};font-weight:700;text-align:right">{_gap:+.2f}%</td>')
        else:
            w('<td style="padding:7px 10px;color:#555;text-align:right">—</td>')
        # 5-day % change
        w('<td style="padding:7px 10px;text-align:right">')
        if _5d_chg is not None:
            _5d_icon = "▲" if _5d_chg >= 0 else "▼"
            w(f'  <div style="font-weight:800;color:{_5d_col}">{_5d_icon} {_5d_chg:+.2f}%</div>')
            w(f'  <div style="font-size:10px;color:#5a7fa0">{_5d_dt}</div>')
        else:
            w('  <div style="color:#555">pending</div>')
        w('</td>')
        # 30-day % change
        _30d_chg = s.get("thirtyDay_chg")
        _30d_dt  = s.get("thirtyDay_date","")
        _30d_col = "#00e676" if (_30d_chg or 0) >= 0 else "#ef5350"
        w('<td style="padding:7px 10px;text-align:right">')
        if _30d_chg is not None:
            _30d_icon = "▲" if _30d_chg >= 0 else "▼"
            w(f'  <div style="font-weight:800;color:{_30d_col}">{_30d_icon} {_30d_chg:+.2f}%</div>')
            w(f'  <div style="font-size:10px;color:#5a7fa0">{_30d_dt}</div>')
        else:
            w('  <div style="color:#555;font-size:11px">pending</div>')
        w('</td>')
        # Win / Loss badge
        w('<td style="padding:7px 10px;text-align:center">')
        if _30d_chg is not None:
            if _30d_chg > 0:
                w('<span style="background:#00e676;color:#000;font-size:11px;font-weight:900;padding:2px 8px;border-radius:5px">WIN</span>')
            else:
                w('<span style="background:#ef5350;color:#fff;font-size:11px;font-weight:900;padding:2px 8px;border-radius:5px">LOSS</span>')
        else:
            w('<span style="color:#5a7fa0;font-size:11px">—</span>')
        w('</td>')
        _hidx = ('<span style="background:#1565c0;color:#90caf9;font-size:10px;font-weight:800;'
                  'padding:1px 6px;border-radius:4px;margin-left:3px">INDEX</span>'
                 if s.get("is_index") else "")
        _hname = f'<div style="font-size:10px;color:#5a7fa0">{s.get("ticker_name","")}</div>' if s.get("ticker_name") else ""
        w(f'<td style="padding:7px 10px;font-weight:900;color:#e8edf5;text-align:right">'
          f'{s["ticker"]} {_tv}{_hidx}{_hname}</td>')
        w(f'<td style="padding:7px 10px;color:#00e676;font-weight:700;text-align:right;white-space:nowrap">{esc(s["signal"])}</td>')
        w(f'<td style="padding:7px 10px;text-align:right">{_type_badge}</td>')
        w(f'<td style="padding:7px 10px;color:{_gc};font-weight:900;font-size:14px;text-align:right">{_sc2}</td>')
        w(f'<td style="padding:7px 10px;text-align:right"><span style="background:{_gc};color:#000;font-size:11px;font-weight:900;padding:2px 8px;border-radius:5px">{_gr}</span></td>')
        w(f'<td style="padding:7px 10px;color:{"#00e676" if (_vr or 0)>=1.5 else "#FFD700"};text-align:right">{f"{_vr:.1f}&times;" if _vr else "—"}</td>')
        w(f'<td style="padding:7px 10px;color:#d0daea;text-align:right">{f"{_bp:.0f}%" if _bp else "—"}</td>')
        w(f'<td style="padding:7px 10px;color:#00e676;text-align:right">{f"+{_dist:.2f}%" if _dist else "—"}</td>')
        _cp = s.get("close_price")
        w(f'<td style="padding:7px 10px;color:#FFD700;font-weight:700;text-align:right">${_cp:.2f}</td>' if _cp else '<td style="color:#555;text-align:right">—</td>')
        w(f'<td style="padding:7px 10px;color:#8ab4d4;font-size:11px;max-width:320px;line-height:1.5">{esc(s["description"])}</td>')
        w('</tr>')
    w('</tbody></table></div>')
    w('</div>')
    return "\n".join(lines)



def render_dog_tab(mag_rows: list, MA_DATA: dict = None) -> str:
    """Dog of the Mag 7 — rebound ranking + earnings preview."""
    lines = []
    def w(*parts): lines.append("".join(str(p) for p in parts))
    import datetime as _ddt

    MAG7_COLORS = {"AAPL":"#a8c7fa","MSFT":"#00a4ef","NVDA":"#76b900",
                   "AMZN":"#FF9900","GOOGL":"#4285F4","META":"#1877F2","TSLA":"#cc0000"}

    FUNDAMENTALS = {
        "AAPL":  {"eps_growth":19.0,"rev_growth":16.0,"fwd_pe":24,"ytd_pct":-12.0,"sector":"Consumer Tech / Services"},
        "MSFT":  {"eps_growth":20.1,"rev_growth":15.2,"fwd_pe":27,"ytd_pct":-8.0, "sector":"Cloud / AI"},
        "NVDA":  {"eps_growth":66.8,"rev_growth":73.0,"fwd_pe":35,"ytd_pct":+5.0, "sector":"AI Infrastructure"},
        "AMZN":  {"eps_growth":29.9,"rev_growth":11.0,"fwd_pe":34,"ytd_pct":+9.4, "sector":"Cloud / E-Commerce"},
        "GOOGL": {"eps_growth":34.3,"rev_growth":12.0,"fwd_pe":18,"ytd_pct":-6.0, "sector":"Search / Cloud"},
        "META":  {"eps_growth":20.0,"rev_growth":20.7,"fwd_pe":19,"ytd_pct":-3.0, "sector":"Social / AI"},
        "TSLA":  {"eps_growth":-4.0,"rev_growth":0.0, "fwd_pe":85,"ytd_pct":-25.0,"sector":"EV / Energy"},
    }

    EARNINGS = {
        "TSLA":  {"date":"2026-04-22","time":"After Close","conf":True,
                  "eps_est":0.44,"rev_est":21.3,"eps_yoy":-4.0,"rev_yoy":0.0,
                  "beat_rate":68,"our_pred":"MISS","pred_col":"#ef5350",
                  "catalyst":"Delivery numbers already missed badly (-13% QoQ). Watch: Robotaxi update, Optimus production timeline, any guidance revision. Bar is low but sentiment is toxic.",
                  "actual_eps":None,"actual_rev":None,"beat_miss":None},
        "META":  {"date":"2026-04-23","time":"After Close","conf":True,
                  "eps_est":5.25,"rev_est":42.2,"eps_yoy":20.0,"rev_yoy":15.0,
                  "beat_rate":85,"our_pred":"BEAT","pred_col":"#00e676",
                  "catalyst":"Advantage+ AI ad suite at $60B run rate. Watch: capex guide ($115-135B) — if maintained stock may sell off despite beat. Morgan Stanley top pick into print.",
                  "actual_eps":None,"actual_rev":None,"beat_miss":None},
        "GOOGL": {"date":"2026-04-29","time":"After Close","conf":True,
                  "eps_est":2.11,"rev_est":89.3,"eps_yoy":34.3,"rev_yoy":12.0,
                  "beat_rate":78,"our_pred":"BEAT","pred_col":"#00e676",
                  "catalyst":"Search resilient — paid search high-teens growth. Cloud 60% YoY, 56% incremental margins. Watch: AI disruption impact to query volume.",
                  "actual_eps":None,"actual_rev":None,"beat_miss":None},
        "MSFT":  {"date":"2026-04-28","time":"After Close","conf":True,
                  "eps_est":3.22,"rev_est":68.4,"eps_yoy":20.1,"rev_yoy":15.2,
                  "beat_rate":82,"our_pred":"BEAT","pred_col":"#00e676",
                  "catalyst":"Azure growth key — watching for re-acceleration to 40%+ after Q4 dipped to 39%. Copilot monetisation ramp. Bar reset lower after 20% drawdown.",
                  "actual_eps":None,"actual_rev":None,"beat_miss":None},
        "AMZN":  {"date":"2026-04-30","time":"After Close","conf":False,
                  "eps_est":1.35,"rev_est":158.5,"eps_yoy":29.9,"rev_yoy":11.0,
                  "beat_rate":80,"our_pred":"BEAT","pred_col":"#00e676",
                  "catalyst":"AWS key — Morgan Stanley models 29% growth. Advertising accelerating. Watch: $125B capex commentary, retail margins. Morgan Stanley #2 pick.",
                  "actual_eps":None,"actual_rev":None,"beat_miss":None},
        "AAPL":  {"date":"2026-05-01","time":"After Close","conf":False,
                  "eps_est":1.62,"rev_est":94.2,"eps_yoy":19.0,"rev_yoy":16.0,
                  "beat_rate":76,"our_pred":"IN-LINE","pred_col":"#FFD700",
                  "catalyst":"Services ($26.7B, 75% margin) is the story. Watch: China demand, FX headwinds, AI/Siri roadmap with Google Gemini.",
                  "actual_eps":None,"actual_rev":None,"beat_miss":None},
        "NVDA":  {"date":"2026-05-28","time":"After Close","conf":False,
                  "eps_est":0.88,"rev_est":43.0,"eps_yoy":66.8,"rev_yoy":73.0,
                  "beat_rate":92,"our_pred":"BEAT","pred_col":"#00e676",
                  "catalyst":"Vera Rubin 6 months ahead of schedule. Data Center $62B last Q (+75% YoY). Watch: export restriction commentary. Historically beats estimates every quarter.",
                  "actual_eps":None,"actual_rev":None,"beat_miss":None},
    }

    if MA_DATA is None: MA_DATA = {}
    today_dt = _ddt.date.today()

    w('<div class="card">')
    w(f'<h3>🐕 Dog of the Mag 7 <span class="fresh">{today_dt.strftime("%b %d, %Y")} — rebound ranking</span></h3>')

    w('''<div style="background:#0a1628;border-radius:10px;padding:14px 18px;margin-bottom:16px;font-size:11px;line-height:1.8">
<div style="font-size:13px;font-weight:900;color:#FFD700;margin-bottom:10px">📋 Dog of the Mag 7 — How Stocks Are Ranked</div>
<div style="display:grid;grid-template-columns:repeat(5,1fr);gap:10px;margin-bottom:10px">
<div style="background:#0d1b2a;border-radius:7px;padding:8px 10px">
<div style="color:#FFD700;font-weight:800">V — Valuation (25pts)</div>
<div style="color:#8ab4d4;margin-top:3px">How far off the 52-week high. ≥30% off=25, ≥20%=20, ≥12%=14, ≥6%=8. More beaten = more upside potential.</div></div>
<div style="background:#0d1b2a;border-radius:7px;padding:8px 10px">
<div style="color:#FFD700;font-weight:800">MA — Structure (25pts)</div>
<div style="color:#8ab4d4;margin-top:3px">Below 200 SMA but above 50 SMA = 25pts (classic dog setup — big-picture downtrend but short-term recovering). Below both = 18pts. Above both = 5pts.</div></div>
<div style="background:#0d1b2a;border-radius:7px;padding:8px 10px">
<div style="color:#FFD700;font-weight:800">Fib — Support (20pts)</div>
<div style="color:#8ab4d4;margin-top:3px">Price sitting on the 61.8% golden ratio Fibonacci retracement = 20pts. 50% level = 15pts. These are the levels institutions historically defend.</div></div>
<div style="background:#0d1b2a;border-radius:7px;padding:8px 10px">
<div style="color:#FFD700;font-weight:800">EPS — Growth (15pts)</div>
<div style="color:#8ab4d4;margin-top:3px">2026 consensus EPS growth rate. ≥50%=15, ≥25%=12, ≥10%=8. High growth + beaten down price = classic value opportunity.</div></div>
<div style="background:#0d1b2a;border-radius:7px;padding:8px 10px">
<div style="color:#FFD700;font-weight:800">RW — Relative Weakness (15pts)</div>
<div style="color:#8ab4d4;margin-top:3px">Worst YTD performer in the Mag 7 = 15pts. Pure contrarian factor — the stock the market hates most often mean-reverts hardest.</div></div>
</div>
<div style="display:grid;grid-template-columns:1fr 1fr;gap:10px">
<div><div style="color:#00e676;font-weight:800;margin-bottom:3px">The Dog Strategy Logic</div>
<div style="color:#8ab4d4">Based on "Dogs of the Dow" — buy the highest-quality, most beaten-down names in a blue-chip group. Applied here to Mag 7: these are world-class businesses with secular growth. When they sell off, it's usually sentiment-driven not fundamental. The recovery tends to be sharp once the catalyst reverses.</div></div>
<div><div style="color:#FFD700;font-weight:800;margin-bottom:3px">Earnings Preview</div>
<div style="color:#8ab4d4">Beat Rate = historical % of quarters where company beat EPS consensus. Our Call = our directional prediction based on setup, guidance trends, and peer signals. After each earnings print, update the EARNINGS dict with actual results to track prediction accuracy over time.</div></div>
</div></div>''')

    w('<div style="font-size:12px;color:#8ab4d4;margin-bottom:16px;line-height:1.6">')
    w('"Dogs of the Dow" applied to Mag 7 — the most beaten-down, highest-quality names tend to mean-revert hardest. ')
    w('Ranked by long-entry conviction: valuation discount + technical setup + Fib support + earnings momentum + relative weakness.')
    w('</div>')

    # Build ranking
    all_ytd   = [FUNDAMENTALS[t]["ytd_pct"] for t in FUNDAMENTALS]
    worst_ytd = min(all_ytd); best_ytd = max(all_ytd)
    dog_rows  = []

    for _row in mag_rows:
        _tk  = _row.get("Ticker","").lstrip("$")
        d    = MA_DATA.get(_tk)
        fun  = FUNDAMENTALS.get(_tk, {})
        if not d: continue
        _px     = d["px"]; _hi52 = d["wk52_hi"]; _lo52 = d["wk52_lo"]
        _sma200 = d.get("sma200"); _sma50 = d.get("sma50")
        _pct_hi = d.get("pct_from_hi", -30)
        _fib_ns = d.get("fib_next_sup")
        _trend  = d.get("trend_score", 50)
        _ytd    = fun.get("ytd_pct", 0)
        _eg     = fun.get("eps_growth", 0)

        _disc_pct = abs(_pct_hi)
        _v_pts  = (25 if _disc_pct>=30 else 20 if _disc_pct>=20 else 14 if _disc_pct>=12 else 8 if _disc_pct>=6 else 3)
        _bel200 = _sma200 and _px < _sma200
        _ab50   = _sma50  and _px > _sma50
        _ma_pts = (25 if _bel200 and _ab50 else 18 if _bel200 else 12 if not _bel200 and _trend<50 else 5)
        _fib_pts = 0
        if _fib_ns:
            _flab = _fib_ns.get("label",""); _fdist = abs(_fib_ns.get("dist",20))
            if "61.8" in _flab and _fdist<=5: _fib_pts=20
            elif "50%" in _flab and _fdist<=5: _fib_pts=15
            elif _fdist<=3: _fib_pts=12
            elif _fdist<=8: _fib_pts=7
        _eg_pts = (15 if _eg>=50 else 12 if _eg>=25 else 8 if _eg>=10 else 4 if _eg>=0 else 0)
        _ytd_range = best_ytd-worst_ytd if best_ytd!=worst_ytd else 1
        _rw_pts = round((1-(_ytd-worst_ytd)/_ytd_range)*15)
        _total  = _v_pts + _ma_pts + _fib_pts + _eg_pts + _rw_pts

        dog_rows.append({"tk":_tk,"px":_px,"hi52":_hi52,"lo52":_lo52,"pct_hi":_pct_hi,
                         "ytd":_ytd,"trend":_trend,"sma200":_sma200,"sma50":_sma50,
                         "fib_ns":_fib_ns,"fun":fun,"score":_total,
                         "v_pts":_v_pts,"ma_pts":_ma_pts,"fib_pts":_fib_pts,
                         "eg_pts":_eg_pts,"rw_pts":_rw_pts})
    dog_rows.sort(key=lambda x: -x["score"])

    # ── Ranking table ─────────────────────────────────────────────────────────
    w('<div style="background:#0a1628;border-radius:12px;padding:16px 20px;margin-bottom:20px">')
    w('<div style="font-size:14px;font-weight:900;color:#FFD700;margin-bottom:4px">🏆 Long Entry Conviction Ranking</div>')
    w('<div style="font-size:11px;color:#5a7fa0;margin-bottom:12px">'
      'Score 0-100: Valuation Discount (25pts) + MA Setup (25pts) + Fib Support (20pts) + EPS Growth (15pts) + Relative Weakness (15pts). '
      'Higher = stronger rebound candidate. Updates on every daily run.</div>')
    w('<div style="overflow-x:auto">')
    w('<table style="width:100%;border-collapse:collapse;font-size:12px">')
    w('<thead><tr style="background:#0d2040;border-bottom:2px solid #1e3a5f">')
    for _h in ["Rank","Ticker","Price","Score","V","MA","Fib","EPS","RW",
               "vs 52wk Hi","YTD","200 SMA","50 SMA","Fib Level","EPS Grwth","Fwd P/E","Rebound Thesis"]:
        w(f'<th style="padding:7px 10px;color:#7eb8f7;font-weight:700;text-align:right;white-space:nowrap">{_h}</th>')
    w('</tr></thead><tbody>')

    for _ri, _dr in enumerate(dog_rows, 1):
        _col2   = MAG7_COLORS.get(_dr["tk"],"#aaa")
        _gc     = "#00e676" if _dr["score"]>=70 else ("#FFD700" if _dr["score"]>=45 else "#ef9a9a")
        _rb     = "background:#0a2010;" if _dr["score"]>=70 else ""
        _tv2    = f'https://www.tradingview.com/chart/?symbol={_dr["tk"]}&interval=D'
        _fun    = _dr["fun"]
        _bel200 = _dr["sma200"] and _dr["px"] < _dr["sma200"]
        _ab50   = _dr["sma50"]  and _dr["px"] > _dr["sma50"]
        _tp     = []
        if abs(_dr["pct_hi"])>=20: _tp.append(f'{abs(_dr["pct_hi"]):.0f}% off highs')
        if _bel200 and _ab50:      _tp.append("below 200 / above 50 SMA")
        if _dr["fib_ns"] and "61.8" in _dr["fib_ns"].get("label",""): _tp.append("at golden ratio Fib")
        if _fun.get("eps_growth",0)>=25: _tp.append(f'{_fun["eps_growth"]:.0f}% EPS growth')
        if _dr["ytd"]<=-15: _tp.append("worst YTD in group")
        _thesis  = " · ".join(_tp) if _tp else "Monitoring"
        _fib_str = (f'${_dr["fib_ns"]["price"]:.2f} ({_dr["fib_ns"]["label"].split("—")[0].strip()})'
                    if _dr["fib_ns"] else "—")
        _ytd_col = "#00e676" if _dr["ytd"]>=0 else "#ef5350"
        _hi_col  = "#ef5350" if _dr["pct_hi"]<-15 else ("#FFD700" if _dr["pct_hi"]<-5 else "#00e676")
        _200c    = "#ef5350" if _bel200 else "#00e676"
        _50c     = "#00e676" if _ab50   else "#ef5350"
        _eg2     = _fun.get("eps_growth",0)
        _egc     = "#00e676" if _eg2>=20 else ("#FFD700" if _eg2>=5 else "#ef5350")
        _s200    = f'${_dr["sma200"]:.2f}' if _dr["sma200"] else "—"
        _s50     = f'${_dr["sma50"]:.2f}'  if _dr["sma50"]  else "—"

        w(f'<tr style="border-bottom:1px solid #0d2040;{_rb}">')
        w(f'<td style="padding:7px 10px;font-size:15px;font-weight:900;color:{_gc};text-align:right">#{_ri}</td>')
        w(f'<td style="padding:7px 10px;font-weight:900;color:{_col2};font-size:14px;text-align:right">'
          f'<a href="{_tv2}" target="_blank" style="color:{_col2};text-decoration:none">{_dr["tk"]}</a></td>')
        w(f'<td style="padding:7px 10px;color:#e8edf5;font-weight:700;text-align:right">${_dr["px"]:.2f}</td>')
        w(f'<td style="padding:7px 10px;color:{_gc};font-weight:900;font-size:15px;text-align:right">{_dr["score"]}</td>')
        for _cp, _cv in [(_dr["v_pts"],25),(_dr["ma_pts"],25),(_dr["fib_pts"],20),(_dr["eg_pts"],15),(_dr["rw_pts"],15)]:
            _cc2 = "#00e676" if _cp>=_cv*0.8 else ("#FFD700" if _cp>=_cv*0.5 else "#ef9a9a")
            w(f'<td style="padding:7px 10px;color:{_cc2};font-weight:700;text-align:right">{_cp}</td>')
        w(f'<td style="padding:7px 10px;color:{_hi_col};font-weight:700;text-align:right">{_dr["pct_hi"]:+.1f}%</td>')
        w(f'<td style="padding:7px 10px;color:{_ytd_col};text-align:right">{_dr["ytd"]:+.1f}%</td>')
        w(f'<td style="padding:7px 10px;text-align:right">'
          f'<span style="color:{_200c};font-size:10px">{"▼" if _bel200 else "▲"}</span> '
          f'<span style="color:#5a7fa0">{_s200}</span></td>')
        w(f'<td style="padding:7px 10px;text-align:right">'
          f'<span style="color:{_50c};font-size:10px">{"▼" if not _ab50 else "▲"}</span> '
          f'<span style="color:#5a7fa0">{_s50}</span></td>')
        w(f'<td style="padding:7px 10px;color:#FFD700;font-size:11px;text-align:right">{esc(_fib_str)}</td>')
        w(f'<td style="padding:7px 10px;color:{_egc};font-weight:700;text-align:right">{_eg2:+.1f}%</td>')
        w(f'<td style="padding:7px 10px;color:#8ab4d4;text-align:right">{_fun.get("fwd_pe","—")}x</td>')
        w(f'<td style="padding:7px 10px;color:#8ab4d4;font-size:11px;max-width:280px;line-height:1.5">{esc(_thesis)}</td>')
        w('</tr>')

    w('</tbody></table></div>')
    w('<div style="display:grid;grid-template-columns:repeat(5,1fr);gap:6px;margin-top:10px;font-size:10px">')
    for _lbl, _pts, _desc in [
        ("V — Valuation","25pts","% off 52wk high"),
        ("MA — Structure","25pts","Below 200 / above 50 SMA"),
        ("Fib — Support","20pts","Near golden ratio"),
        ("EPS — Growth","15pts","2026 consensus EPS growth"),
        ("RW — Weakness","15pts","Worst YTD = most contrarian"),
    ]:
        w(f'<div style="background:#0d1b2a;border-radius:6px;padding:6px 8px">'
          f'<span style="color:#FFD700;font-weight:800">{_lbl}</span> '
          f'<span style="color:#5a7fa0">({_pts})</span>'
          f'<div style="color:#5a7fa0;margin-top:2px">{_desc}</div></div>')
    w('</div></div>')

    # ── Earnings Preview ──────────────────────────────────────────────────────
    w('<div style="background:#0a1628;border-radius:12px;padding:16px 20px;margin-bottom:20px">')
    w('<div style="font-size:14px;font-weight:900;color:#FFD700;margin-bottom:4px">'
      '📅 Q1 2026 Earnings Preview — Beat Predictions</div>')
    w('<div style="font-size:11px;color:#5a7fa0;margin-bottom:12px">'
      'Consensus estimates and our call. Actual columns populate when results come in — '
      'run the report after each print to log the result and track prediction accuracy.</div>')
    w('<div style="overflow-x:auto">')
    w('<table style="width:100%;border-collapse:collapse;font-size:12px">')
    w('<thead><tr style="background:#0d2040;border-bottom:2px solid #1e3a5f">')
    for _h in ["Ticker","Date","EPS Est","EPS YoY","Rev Est","Rev YoY",
               "Beat Rate","Our Call","Actual EPS","Beat/Miss","Key Catalyst"]:
        w(f'<th style="padding:7px 10px;color:#7eb8f7;font-weight:700;text-align:right;white-space:nowrap">{_h}</th>')
    w('</tr></thead><tbody>')

    earn_sorted = sorted(EARNINGS.items(), key=lambda x: x[1]["date"])
    for _etk, _e in earn_sorted:
        _ecol    = MAG7_COLORS.get(_etk, "#aaa")
        _edt     = _ddt.datetime.strptime(_e["date"], "%Y-%m-%d").date()
        _days    = (_edt - today_dt).days
        _is_past = _days < 0
        _soon_bg = "background:#0a1200;" if (not _is_past and _days <= 7) else ""
        _days_lbl = "reported" if _is_past else f"in {_days}d"
        _date_lbl = f'{_edt.strftime("%b %d")} ({_days_lbl})'
        _conf_ico = '<span style="color:#00e676">✓</span>' if _e["conf"] else '<span style="color:#ffb74d">~</span>'
        _eg_c    = "#00e676" if _e["eps_yoy"] >= 0 else "#ef5350"
        _rv_c    = "#00e676" if _e["rev_yoy"] >= 0 else "#ef5350"
        _tv3     = f'https://www.tradingview.com/chart/?symbol={_etk}&interval=D'

        w(f'<tr style="border-bottom:1px solid #0d2040;{_soon_bg}">')
        w(f'<td style="padding:7px 10px;font-weight:900;color:{_ecol};text-align:right">'
          f'<a href="{_tv3}" target="_blank" style="color:{_ecol};text-decoration:none">{_etk}</a></td>')
        w(f'<td style="padding:7px 10px;text-align:right">'
          f'<div style="color:#e8edf5;font-weight:700">{_date_lbl}</div>'
          f'<div style="color:#5a7fa0;font-size:10px">{_conf_ico} {"Confirmed" if _e["conf"] else "Estimated"}</div></td>')
        w(f'<td style="padding:7px 10px;color:#FFD700;font-weight:800;text-align:right">${_e["eps_est"]:.2f}</td>')
        w(f'<td style="padding:7px 10px;color:{_eg_c};font-weight:700;text-align:right">{_e["eps_yoy"]:+.1f}%</td>')
        w(f'<td style="padding:7px 10px;color:#e8edf5;text-align:right">${_e["rev_est"]:.1f}B</td>')
        w(f'<td style="padding:7px 10px;color:{_rv_c};text-align:right">{_e["rev_yoy"]:+.1f}%</td>')
        w(f'<td style="padding:7px 10px;color:#8ab4d4;text-align:right">{_e["beat_rate"]}%</td>')
        w(f'<td style="padding:7px 10px;text-align:right">'
          f'<span style="background:{_e["pred_col"]};color:#000;font-size:11px;font-weight:900;'
          f'padding:2px 9px;border-radius:5px">{_e["our_pred"]}</span></td>')
        if _e["actual_eps"] is not None:
            _sur = round((_e["actual_eps"]-_e["eps_est"])/_e["eps_est"]*100,1) if _e["eps_est"] else 0
            _ac  = "#00e676" if _e["actual_eps"] > _e["eps_est"] else "#ef5350"
            w(f'<td style="padding:7px 10px;color:{_ac};font-weight:800;text-align:right">'
              f'${_e["actual_eps"]:.2f} ({_sur:+.1f}%)</td>')
            _bmc = ("#00e676" if _e["beat_miss"]=="BEAT" else
                    "#ef5350" if _e["beat_miss"]=="MISS" else "#FFD700")
            w(f'<td style="padding:7px 10px;text-align:right">'
              f'<span style="background:{_bmc};color:#000;font-size:11px;font-weight:900;'
              f'padding:2px 9px;border-radius:5px">{_e["beat_miss"] or "—"}</span></td>')
        else:
            w('<td style="padding:7px 10px;color:#555;text-align:right;font-style:italic">pending</td>')
            w('<td style="padding:7px 10px;color:#555;text-align:right">—</td>')
        w(f'<td style="padding:7px 10px;color:#8ab4d4;font-size:11px;max-width:280px;line-height:1.5">'
          f'{esc(_e["catalyst"])}</td>')
        w('</tr>')

    w('</tbody></table></div>')
    w('<div style="font-size:11px;color:#5a7fa0;margin-top:8px">'
      'Estimates from FactSet / analyst consensus as of Apr 15 2026. '
      '✓ = confirmed date. ~ = estimated. '
      'Add actual_eps, actual_rev, beat_miss to each EARNINGS entry after reporting to track accuracy.</div>')
    w('</div>')
    w('</div>')
    return "\n".join(lines)



def _compute_ripster(closes: list, highs: list, lows: list, volumes: list) -> dict:
    """
    Compute Ripster cloud states and corroborating indicators from daily OHLCV.

    Ripster clouds (all EMA-based):
      Cloud 1 — 8/9 EMA   (fast, intraday rhythm — approximated on daily)
      Cloud 2 — 5/13 EMA  (Daily MTF1)
      Cloud 3 — 34/50 EMA (intermediate trend)
      Cloud 4 — 5/12 EMA  weekly — approximated as 25/60 EMA on daily (5d*5, 12d*5)

    Cloud state: BULLISH if faster EMA > slower EMA, BEARISH if faster < slower.

    Corroborating indicators:
      - Price vs 20/50/100/200 SMA (above/below)
      - RSI(14) — >60 bullish, <40 bearish
      - Volume ratio vs 20-day avg
      - ATR(14) normalised as % of price
      - Distance from 52-week high/low
      - Trend score: % of 8 MAs price is above

    Returns dict with all computed values.
    """
    n = len(closes)
    if n < 210:
        return {}

    px = closes[-1]

    def _ema(prices, period):
        if len(prices) < period: return []
        k = 2 / (period + 1)
        val = sum(prices[:period]) / period
        out = [None]*(period-1) + [val]
        for p in prices[period:]:
            val = p*k + val*(1-k)
            out.append(val)
        return out

    def _sma(prices, period):
        out = [None]*(period-1)
        for i in range(period-1, len(prices)):
            out.append(sum(prices[i-period+1:i+1])/period)
        return out

    def _last(arr):
        if not arr: return None
        v = arr[-1]
        return round(v, 2) if v is not None else None

    # Ripster EMAs
    e8   = _last(_ema(closes, 8))
    e9   = _last(_ema(closes, 9))
    e5   = _last(_ema(closes, 5))
    e13  = _last(_ema(closes, 13))
    e34  = _last(_ema(closes, 34))
    e50  = _last(_ema(closes, 50))
    e25  = _last(_ema(closes, 25))
    e60  = _last(_ema(closes, 60))

    def _cloud(fast, slow):
        if fast is None or slow is None: return "N/A"
        return "BULLISH" if fast > slow else "BEARISH"

    c1_89   = _cloud(e8,  e9)
    c2_513  = _cloud(e5,  e13)
    c3_3450 = _cloud(e34, e50)
    c4_week = _cloud(e25, e60)

    clouds = [c1_89, c2_513, c3_3450, c4_week]
    bull_count = sum(1 for c in clouds if c == "BULLISH")
    bear_count = sum(1 for c in clouds if c == "BEARISH")

    if bull_count == 4:    alignment = "FULL BULL"
    elif bull_count == 3:  alignment = "MOSTLY BULL"
    elif bear_count == 4:  alignment = "FULL BEAR"
    elif bear_count == 3:  alignment = "MOSTLY BEAR"
    else:                  alignment = "MIXED"

    # SMAs
    sma20  = _last(_sma(closes, 20))
    sma50  = _last(_sma(closes, 50))
    sma100 = _last(_sma(closes, 100))
    sma200 = _last(_sma(closes, 200))

    def _vs(ma):
        if ma is None: return "—"
        return "▲" if px >= ma else "▼"

    # ── RSI(14) — current and 10 bars ago for divergence ─────────────────────
    def _calc_rsi(price_slice):
        if len(price_slice) < 15: return None
        gains, losses = [], []
        for i in range(1, 15):
            diff = price_slice[-i] - price_slice[-i-1]
            if diff > 0: gains.append(diff)
            else: losses.append(abs(diff))
        avg_g = sum(gains)/14 if gains else 0
        avg_l = sum(losses)/14 if losses else 0.0001
        return round(100 - 100/(1 + avg_g/avg_l), 1)

    rsi     = _calc_rsi(closes)
    rsi_10  = _calc_rsi(closes[:-10]) if n > 25 else None  # RSI 10 bars ago

    # RSI Divergence detection (last 20 bars)
    # Bullish div: price lower low, RSI higher low
    # Bearish div: price higher high, RSI lower high
    rsi_div       = "NONE"
    rsi_div_col   = "#8ab4d4"
    rsi_div_score = 0
    if n >= 30 and rsi is not None:
        # Compare current bar vs 10 bars ago
        px_10ago  = closes[-11]
        rsi_10ago = rsi_10 or rsi
        # Also check the low vs 10d low for divergence
        lo_now    = min(lows[-5:])
        lo_10     = min(lows[-15:-10])
        hi_now    = max(highs[-5:])
        hi_10     = max(highs[-15:-10])
        if lo_now < lo_10 and rsi > rsi_10ago + 3:
            rsi_div = "BULL DIV"; rsi_div_col = "#00e676"; rsi_div_score = 20
        elif hi_now > hi_10 and rsi < rsi_10ago - 3:
            rsi_div = "BEAR DIV"; rsi_div_col = "#ef5350"; rsi_div_score = -10
        elif lo_now > lo_10 * 0.995 and rsi > rsi_10ago + 2:
            rsi_div = "HIDDEN BULL"; rsi_div_col = "#00e676"; rsi_div_score = 12
        elif hi_now < hi_10 * 1.005 and rsi < rsi_10ago - 2:
            rsi_div = "HIDDEN BEAR"; rsi_div_col = "#ef9a9a"; rsi_div_score = -5

    # ── ATR(14) + ATR Squeeze ─────────────────────────────────────────────────
    atr     = None
    atr_20  = None   # 20-bar avg of ATR(14) — baseline
    atr_squeeze     = False
    atr_squeeze_pct = None
    if n >= 35:
        def _atr14(idx_end):
            trs = []
            for i in range(1, 15):
                ix = idx_end - i
                if ix < 1: break
                tr = max(highs[ix]-lows[ix],
                         abs(highs[ix]-closes[ix-1]),
                         abs(lows[ix]-closes[ix-1]))
                trs.append(tr)
            return sum(trs)/len(trs) if trs else None

        atr_val = _atr14(n - 1)
        if atr_val:
            atr = round(atr_val / px * 100, 2)

        # Average ATR over past 20 sessions (each session's ATR14)
        atr_history = [_atr14(n - 1 - i) for i in range(20) if n - 1 - i >= 14]
        atr_history = [a for a in atr_history if a]
        if atr_history and atr_val:
            atr_20_val = sum(atr_history) / len(atr_history)
            atr_20 = round(atr_20_val / px * 100, 2)
            atr_squeeze_pct = round((atr_val - atr_20_val) / atr_20_val * 100, 1)
            # Squeeze: current ATR is 20%+ below its own 20-day avg → coiling
            atr_squeeze = atr_squeeze_pct < -20

    # ── OBV (On-Balance Volume) trend ─────────────────────────────────────────
    # Rising OBV = accumulation (buying pressure), falling = distribution
    obv_trend   = "FLAT"
    obv_trend_col = "#8ab4d4"
    obv_pts     = 0
    if n >= 21:
        obv = 0
        obv_series = []
        for i in range(len(closes)):
            if i == 0:
                obv_series.append(0)
            elif closes[i] > closes[i-1]:
                obv += volumes[i]
                obv_series.append(obv)
            elif closes[i] < closes[i-1]:
                obv -= volumes[i]
                obv_series.append(obv)
            else:
                obv_series.append(obv)
        # Compare current OBV vs 10 and 20 days ago
        obv_now  = obv_series[-1]
        obv_10d  = obv_series[-11] if len(obv_series) > 11 else obv_series[0]
        obv_20d  = obv_series[-21] if len(obv_series) > 21 else obv_series[0]
        _obv_chg_10 = (obv_now - obv_10d) / (abs(obv_10d) or 1) * 100
        _obv_chg_20 = (obv_now - obv_20d) / (abs(obv_20d) or 1) * 100
        if _obv_chg_10 > 5 and _obv_chg_20 > 5:
            obv_trend = "RISING"; obv_trend_col = "#00e676"; obv_pts = 15
        elif _obv_chg_10 < -5 and _obv_chg_20 < -5:
            obv_trend = "FALLING"; obv_trend_col = "#ef5350"; obv_pts = -10
        elif _obv_chg_10 > 2:
            obv_trend = "RISING"; obv_trend_col = "#00e676"; obv_pts = 8
        elif _obv_chg_10 < -2:
            obv_trend = "FALLING"; obv_trend_col = "#ef9a9a"; obv_pts = -5
        else:
            obv_trend = "FLAT"

    # ── Cloud compression (fast cloud squeezing into slow cloud) ──────────────
    # Compression = fast EMAs (8/9, 5/13) are converging with slow EMAs (34/50)
    cloud_compression  = False
    compression_label  = ""
    compression_col    = "#8ab4d4"
    compression_pts    = 0
    if all(v is not None for v in [e8, e9, e34, e50]):
        _fast_spread = abs(e8 - e9)
        _slow_spread = abs(e34 - e50)
        _fast_to_slow = abs(((e8 + e9) / 2) - ((e34 + e50) / 2))
        # Compare to 10 bars ago
        _e8_10  = _last(_ema(closes[:-10], 8))  if n > 20 else None
        _e34_10 = _last(_ema(closes[:-10], 34)) if n > 44 else None
        if _e8_10 and _e34_10:
            _gap_10 = abs(_e8_10 - _e34_10)
            _gap_now = _fast_to_slow
            if _gap_now < _gap_10 * 0.7:   # fast closing in on slow by 30%+
                cloud_compression = True
                if bull_count >= 2:
                    compression_label = "BULL COIL"; compression_col = "#00e676"; compression_pts = 15
                else:
                    compression_label = "BEAR COIL"; compression_col = "#ef9a9a"; compression_pts = 10

    # 52-week
    hi52 = max(highs[-252:]) if n >= 252 else max(highs)
    lo52 = min(lows[-252:])  if n >= 252 else min(lows)
    pct_from_hi = round((px-hi52)/hi52*100, 1)
    pct_from_lo = round((px-lo52)/lo52*100, 1)

    # Volume ratio
    avg_vol   = sum(volumes[-21:-1])/20 if n >= 21 else volumes[-1]
    vol_ratio = round(volumes[-1]/avg_vol, 2) if avg_vol else 1.0

    # Trend score
    all_mas = [e8, e9, e5, e13, e34, e50, sma50, sma200]
    valid   = [m for m in all_mas if m is not None]
    trend_score = round(sum(1 for m in valid if px > m)/len(valid)*100) if valid else 50

    # Overall bias
    if alignment in ("FULL BULL","MOSTLY BULL") and (rsi or 50) > 50 and trend_score >= 60:
        bias = "BULLISH"; bias_col = "#00e676"
    elif alignment in ("FULL BEAR","MOSTLY BEAR") and (rsi or 50) < 50 and trend_score <= 40:
        bias = "BEARISH"; bias_col = "#ef5350"
    elif alignment in ("FULL BULL","MOSTLY BULL"):
        bias = "LEANING BULL"; bias_col = "#00e676"
    elif alignment in ("FULL BEAR","MOSTLY BEAR"):
        bias = "LEANING BEAR"; bias_col = "#ef9a9a"
    else:
        bias = "NEUTRAL"; bias_col = "#8ab4d4"

    # ── Conviction Score (0-100) — REVISED ────────────────────────────────────
    # Now includes leading indicators: RSI divergence, ATR squeeze, OBV, cloud compression
    #
    # Component          Max pts   What it measures
    # Cloud alignment      30      Core Ripster signal
    # Trend score          15      % of MAs confirming direction
    # RSI level            10      Momentum confirmation
    # RSI divergence       20      LEADING — exhaustion signal before price moves
    # ATR squeeze          10      LEADING — coiling before explosive move
    # OBV                  15      Smart money accumulation/distribution
    # SMA stack            10      Structural confirmation
    # Volume               10      Institutional participation
    # Cloud compression    10      LEADING — energy building in clouds
    # ─────────────────────────────────────────────────────────────
    # Total possible      130 → capped at 100

    _dom_bull  = bull_count >= bear_count
    _dom_count = bull_count if _dom_bull else bear_count

    # 1. Cloud alignment (30pts)
    _align_pts = {4:30, 3:20, 2:8, 1:3, 0:0}.get(_dom_count, 0)

    # 2. Trend score (15pts) — toward dominant direction
    _trend_pts = round((trend_score if _dom_bull else 100-trend_score) / 100 * 15)

    # 3. RSI level (10pts)
    _rsi_val = rsi or 50
    if _dom_bull:
        _rsi_pts = 10 if _rsi_val > 65 else (7 if _rsi_val > 55 else (3 if _rsi_val > 45 else 0))
    else:
        _rsi_pts = 10 if _rsi_val < 35 else (7 if _rsi_val < 45 else (3 if _rsi_val < 55 else 0))

    # 4. RSI divergence (20pts — highest weight, most leading)
    _div_pts = max(0, rsi_div_score) if _dom_bull else max(0, -rsi_div_score)

    # 5. ATR squeeze (10pts) — present = energy coiling
    _squeeze_pts = 10 if atr_squeeze else 0

    # 6. OBV (15pts)
    _obv_pts_scored = max(0, obv_pts) if _dom_bull else max(0, -obv_pts)

    # 7. SMA stack (10pts)
    _sma_agree = sum(1 for m in [sma20,sma50,sma100,sma200] if m and (px>m if _dom_bull else px<m))
    _sma_pts   = {4:10, 3:7, 2:4, 1:1, 0:0}.get(_sma_agree, 0)

    # 8. Volume (10pts)
    _vol_pts = 10 if vol_ratio >= 1.5 else (6 if vol_ratio >= 1.2 else (3 if vol_ratio >= 1.0 else 0))

    # 9. Cloud compression (10pts — bonus for coiling)
    _comp_pts = compression_pts if cloud_compression else 0

    conviction_score = min(100, _align_pts + _trend_pts + _rsi_pts + _div_pts +
                           _squeeze_pts + _obv_pts_scored + _sma_pts + _vol_pts + _comp_pts)

    return {
        "px": round(px, 2),
        # Ripster clouds
        "c1_89":   c1_89,   "e8":  e8,   "e9":  e9,
        "c2_513":  c2_513,  "e5":  e5,   "e13": e13,
        "c3_3450": c3_3450, "e34": e34,  "e50": e50,
        "c4_week": c4_week, "e25": e25,  "e60": e60,
        "bull_count": bull_count, "bear_count": bear_count,
        "alignment": alignment,
        # SMAs
        "sma20": sma20, "sma50": sma50, "sma100": sma100, "sma200": sma200,
        "vs20": _vs(sma20), "vs50": _vs(sma50),
        "vs100": _vs(sma100), "vs200": _vs(sma200),
        # Leading indicators
        "rsi": rsi,
        "rsi_div": rsi_div, "rsi_div_col": rsi_div_col,
        "atr": atr, "atr_20": atr_20,
        "atr_squeeze": atr_squeeze, "atr_squeeze_pct": atr_squeeze_pct,
        "obv_trend": obv_trend, "obv_trend_col": obv_trend_col,
        "cloud_compression": cloud_compression,
        "compression_label": compression_label, "compression_col": compression_col,
        # Other
        "vol_ratio": vol_ratio,
        "hi52": round(hi52,2), "lo52": round(lo52,2),
        "pct_from_hi": pct_from_hi, "pct_from_lo": pct_from_lo,
        "trend_score": trend_score, "bias": bias, "bias_col": bias_col,
        "conviction_score": conviction_score,
        # Score breakdown for transparency
        "_pts": {"align":_align_pts,"trend":_trend_pts,"rsi":_rsi_pts,
                 "div":_div_pts,"squeeze":_squeeze_pts,"obv":_obv_pts_scored,
                 "sma":_sma_pts,"vol":_vol_pts,"comp":_comp_pts},
    }


def render_ripster_tab(tickers: list, ohlcv_data: dict) -> str:
    """
    Ripster Cloud Screener — scans all tickers and shows:
    - Ripster cloud alignment (4 clouds: 8/9, 5/13, 34/50, 5/12weekly)
    - Full Bull / Full Bear / Mixed
    - Corroborating: price vs 20/50/100/200 SMA, RSI, volume, trend score
    - Sortable by alignment then trend score
    """
    lines = []
    def w(*parts): lines.append("".join(str(p) for p in parts))
    import datetime as _ddt
    today_dt = _ddt.date.today()

    w('<div class="card">')
    w(f'<h3>🌊 Ripster Cloud Screener <span class="fresh">{today_dt.strftime("%b %d, %Y")} — daily bars</span></h3>')
    w('<div style="font-size:12px;color:#8ab4d4;margin-bottom:16px;line-height:1.6">'
      'Screens all large-cap tickers for Ripster cloud alignment. '
      'Cloud 1: 8/9 EMA (short-term) · Cloud 2: 5/13 EMA (Daily MTF1) · '
      'Cloud 3: 34/50 EMA (intermediate) · Cloud 4: 25/60 EMA (weekly 5/12 approx). '
      'Full Bull = all 4 clouds bullish. Full Bear = all 4 bearish. '
      'Corroborating columns show additional confluence.</div>')

    # Compute ripster for every ticker
    rows = []
    clean = [t.lstrip("$") for t in tickers if t]
    for tk in clean:
        data = ohlcv_data.get(tk)
        if not data: continue
        r = _compute_ripster(data["close"], data["high"], data["low"], data["volume"])
        if not r: continue
        rows.append({"tk": tk, **r})


    w('''<div style="background:#0a1628;border-radius:10px;padding:14px 18px;margin-bottom:16px;font-size:11px;line-height:1.8">
<div style="font-size:13px;font-weight:900;color:#FFD700;margin-bottom:10px">📋 Ripster Cloud Screener — How It Works</div>
<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:10px">
<div style="background:#0d1b2a;border-radius:7px;padding:8px 10px">
<div style="color:#00e676;font-weight:800">Cloud 1: 8/9 EMA</div>
<div style="color:#8ab4d4;margin-top:3px">Short-term momentum. Fast cloud — first to flip. If 8 EMA &gt; 9 EMA = bullish.</div></div>
<div style="background:#0d1b2a;border-radius:7px;padding:8px 10px">
<div style="color:#42a5f5;font-weight:800">Cloud 2: 5/13 EMA</div>
<div style="color:#8ab4d4;margin-top:3px">Daily MTF1. Medium momentum — same as the Ripster Daily MTF1 signal on your TradingView chart.</div></div>
<div style="background:#0d1b2a;border-radius:7px;padding:8px 10px">
<div style="color:#FFD700;font-weight:800">Cloud 3: 34/50 EMA</div>
<div style="color:#8ab4d4;margin-top:3px">Intermediate trend. Slower to flip — confirms the larger trend direction. Most weight for swing trades.</div></div>
<div style="background:#0d1b2a;border-radius:7px;padding:8px 10px">
<div style="color:#ef9a9a;font-weight:800">Cloud 4: 25/60 EMA</div>
<div style="color:#8ab4d4;margin-top:3px">Weekly 5/12 proxy. Long-term bias. When this flips bullish after being bearish, major trend reversals occur.</div></div>
</div>
<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:10px">
<div style="background:#051a0a;border-radius:7px;padding:8px 10px;border:1px solid #00e67633">
<div style="color:#00e676;font-weight:900">🟢 FULL BULL</div>
<div style="color:#8ab4d4;margin-top:3px">All 4 clouds bullish. Maximum conviction long. Historically the strongest setups for continuation moves.</div></div>
<div style="background:#1a050a;border-radius:7px;padding:8px 10px;border:1px solid #ef535033">
<div style="color:#ef5350;font-weight:900">🔴 FULL BEAR</div>
<div style="color:#8ab4d4;margin-top:3px">All 4 clouds bearish. Maximum conviction short/avoid. Stocks in sustained downtrends.</div></div>
<div style="background:#0d1b2a;border-radius:7px;padding:8px 10px">
<div style="color:#FFD700;font-weight:900">🔺 RSI Divergence</div>
<div style="color:#8ab4d4;margin-top:3px">Price makes lower low, RSI makes higher low = sellers exhausted. One of the most reliable reversal signals before a move.</div></div>
<div style="background:#0d1b2a;border-radius:7px;padding:8px 10px">
<div style="color:#FFD700;font-weight:900">🔥 ATR Squeeze</div>
<div style="color:#8ab4d4;margin-top:3px">Current ATR is 20%+ below its 20-day average — volatility has compressed. Price is coiling. Breakouts from squeezes tend to be sharp and sustained.</div></div>
</div>
<div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px">
<div><div style="color:#00e676;font-weight:800;margin-bottom:3px">▲ OBV Rising</div>
<div style="color:#8ab4d4">On-Balance Volume trending up means volume is flowing in on up days. Institutions are accumulating before price reflects it. Most powerful when OBV rises while price is flat.</div></div>
<div><div style="color:#FFD700;font-weight:800;margin-bottom:3px">🌀 Cloud Coil</div>
<div style="color:#8ab4d4">Fast EMAs (8/9) compressing into slow EMAs (34/50). Energy building inside the cloud structure. Direction of the breakout usually matches the dominant cloud bias.</div></div>
<div><div style="color:#e8edf5;font-weight:800;margin-bottom:3px">Score 0-100 breakdown</div>
<div style="color:#8ab4d4">Clouds 30 · Trend 15 · RSI 10 · <span style="color:#FFD700">RSI Div 20</span> · <span style="color:#FFD700">ATR Sqz 10</span> · <span style="color:#FFD700">OBV 15</span> · SMA Stack 10 · Vol 10 · Coil 10<br>
Yellow = leading indicators (fire before the move).</div></div>
</div></div>''')
    if not rows:
        w('<div style="color:#ffb74d;padding:20px">No data available — run again after market close.</div>')
        w('</div>')
        return "\n".join(lines)

    # Always include Mag 7 regardless of alignment
    MAG7_SET = {t for t, _ in MAG7_TICKERS}
    mag7_rows_rip = [r for r in rows if r["tk"] in MAG7_SET]
    # Filter others to Full Bull / Full Bear only
    filtered_rows = [r for r in rows if r["tk"] not in MAG7_SET
                     and r["alignment"] in ("FULL BULL","FULL BEAR")]
    # Sort by conviction score desc within each alignment group
    _align_order = {"FULL BULL":0,"FULL BEAR":1}
    filtered_rows.sort(key=lambda x: (_align_order.get(x["alignment"],0), -x.get("conviction_score",0)))
    mag7_rows_rip.sort(key=lambda x: -x.get("conviction_score",0))
    rows = mag7_rows_rip + filtered_rows

    # Summary counts
    full_bull  = sum(1 for r in rows if r["alignment"]=="FULL BULL")
    mostly_bull= sum(1 for r in rows if r["alignment"]=="MOSTLY BULL")
    mixed      = sum(1 for r in rows if r["alignment"]=="MIXED")
    mostly_bear= sum(1 for r in rows if r["alignment"]=="MOSTLY BEAR")
    full_bear  = sum(1 for r in rows if r["alignment"]=="FULL BEAR")

    w('<div style="display:grid;grid-template-columns:repeat(5,1fr);gap:8px;margin-bottom:16px">')
    for _lbl, _cnt, _col in [
        ("🟢 Full Bull",   full_bull,   "#00e676"),
        ("↗ Mostly Bull",  mostly_bull, "#00e676"),
        ("➡ Mixed",        mixed,       "#8ab4d4"),
        ("↘ Mostly Bear",  mostly_bear, "#ef9a9a"),
        ("🔴 Full Bear",   full_bear,   "#ef5350"),
    ]:
        w(f'<div style="background:#0d1b2a;border-radius:8px;padding:10px 14px;text-align:center">')
        w(f'<div style="font-size:20px;font-weight:900;color:{_col}">{_cnt}</div>')
        w(f'<div style="font-size:11px;color:#5a7fa0">{_lbl}</div></div>')
    w('</div>')

    # Cloud colour helper
    def _cc(state):
        if state == "BULLISH": return "#00e676", "🟢"
        if state == "BEARISH": return "#ef5350", "🔴"
        return "#8ab4d4", "➖"

    # ── Best Pick recommendation ──────────────────────────────────────────
    _bull_picks = [r for r in rows if r["alignment"]=="FULL BULL" and r["tk"] not in MAG7_SET]
    if _bull_picks:
        _bp = _bull_picks[0]  # highest conviction
        _bp2 = _bull_picks[1] if len(_bull_picks)>1 else None
        _bp_col = "#00e676"
        w('<div style="background:#051a0a;border:1px solid #00e67644;border-radius:12px;padding:16px 20px;margin-bottom:16px">')
        w('<div style="font-size:13px;font-weight:900;color:#00e676;margin-bottom:10px">')
        w('🏆 BEST BULLISH PICK — If you can only choose one:</div>')
        w(f'<div style="display:flex;align-items:center;gap:16px;flex-wrap:wrap;margin-bottom:10px">')
        w(f'<span style="font-size:26px;font-weight:900;color:#00e676">{_bp["tk"]}</span>')
        w(f'<span style="font-size:18px;color:#e8edf5;font-weight:700">${_bp["px"]:.2f}</span>')
        w(f'<span style="background:#00e67622;color:#00e676;font-size:12px;font-weight:900;padding:3px 12px;border-radius:6px">Score: {_bp["conviction_score"]}/100</span>')
        w(f'<span style="background:#00e67222;color:#00e672;font-size:11px;font-weight:800;padding:2px 10px;border-radius:6px">Trend: {_bp.get("trend_score",0)}%</span>')
        w(f'<span style="color:#8ab4d4;font-size:11px">RSI: {_bp.get("rsi","—")}</span>')
        w('</div>')
        # Why this one
        _why_parts = []
        _why_parts.append("All 4 Ripster clouds bullish — complete alignment across every timeframe")
        if (_bp.get("rsi") or 0) > 60:
            _why_parts.append(f'RSI {_bp["rsi"]:.0f} confirming momentum (>60 = bullish)')
        _sma_count = sum(1 for k in ["vs20","vs50","vs100","vs200"] if _bp.get(k)=="▲")
        if _sma_count == 4:
            _why_parts.append("Price above all 4 SMAs (20/50/100/200) — full trend alignment")
        elif _sma_count >= 3:
            _why_parts.append(f"Price above {_sma_count}/4 SMAs — strong trend structure")
        if (_bp.get("pct_from_hi") or 0) > -10:
            _why_parts.append(f'Only {abs(_bp["pct_from_hi"]):.1f}% off 52-week high — near breakout territory')
        elif (_bp.get("pct_from_hi") or 0) > -20:
            _why_parts.append(f'Pulling back {abs(_bp["pct_from_hi"]):.1f}% from highs while clouds stay bullish — healthy retracement')
        if _bp.get("vol_ratio",1) >= 1.3:
            _why_parts.append(f'Volume {_bp["vol_ratio"]:.1f}× average — institutional participation')
        if _bp2:
            _why_parts.append(f'Runner-up: {_bp2["tk"]} (score {_bp2["conviction_score"]}) — consider if {_bp["tk"]} is already in your portfolio')
        for _wp in _why_parts:
            w(f'<div style="font-size:11px;color:#8ab4d4;padding:2px 0;line-height:1.6">✓ {esc(_wp)}</div>')
        w('</div>')

    # Full table
    w('<div style="overflow-x:auto">')
    w('<table style="width:100%;border-collapse:collapse;font-size:11px">')
    w('<thead><tr style="background:#0d2040;border-bottom:2px solid #1e3a5f">')
    _headers = [
        "Rank","Ticker","Price","Alignment","Score","Bias",
        "8/9 EMA","5/13 EMA","34/50 EMA","25/60 EMA",
        "Trend%","vs 20","vs 50","vs 100","vs 200",
        "RSI","RSI Div","ATR Sqz","OBV","Coil",
        "Vol","52w Hi%","Chart"
    ]
    for _h in _headers:
        w(f'<th style="padding:6px 8px;color:#7eb8f7;font-weight:700;text-align:center;white-space:nowrap">{_h}</th>')
    w('</tr></thead><tbody>')

    _last_align = None
    _bull_rank  = 0
    _bear_rank  = 0

    # ── Render in three explicit sections: Mag7, Full Bull, Full Bear ─────────
    _sections = [
        ("✨ Mag 7 — Always Shown", "#7eb8f7", mag7_rows_rip),
        ("🟢 FULL BULL",            "#00e676", [r for r in filtered_rows if r["alignment"]=="FULL BULL"]),
        ("🔴 FULL BEAR",            "#ef5350", [r for r in filtered_rows if r["alignment"]=="FULL BEAR"]),
    ]

    for _sec_label, _sec_col, _sec_rows in _sections:
        if not _sec_rows: continue
        # Section header row
        w(f'<tr><td colspan="24" style="background:#0a1628;padding:7px 12px;'
          f'color:{_sec_col};font-weight:900;font-size:12px;border-top:2px solid {_sec_col}44">'
          f'{_sec_label} ({len(_sec_rows)} stocks)</td></tr>')

        for r in _sec_rows:
            # Rank counter
            _is_m7 = r["tk"] in MAG7_SET
            if not _is_m7:
                if r["alignment"] == "FULL BULL":  _bull_rank += 1;  _row_rank = f"#{_bull_rank}"
                elif r["alignment"] == "FULL BEAR": _bear_rank += 1; _row_rank = f"#{_bear_rank}"
                else: _row_rank = "—"
            else: _row_rank = "M7"
            _conv = r.get("conviction_score", 0)
            _conv_col = "#00e676" if _conv>=70 else ("#FFD700" if _conv>=50 else "#ef9a9a")

            bc = r["bull_count"]
            bg = ("background:#051a0a;" if bc==4 else
                  "background:#0a1628;" if bc==3 else
                  "background:#1a050a;" if bc==0 else
                  "background:#100c0a;" if bc==1 else "")
            if _is_m7:
                bg = "background:#0a1220;"  # distinct tint for Mag7
            _tv = f'https://www.tradingview.com/chart/?symbol={r["tk"]}&interval=D'
            _bias_col = r.get("bias_col","#8ab4d4")
            _ts = r.get("trend_score",50)
            _ts_col = "#00e676" if _ts>=70 else ("#FFD700" if _ts>=50 else "#ef5350")
            _rsi = r.get("rsi")
            _rsi_col = "#00e676" if (_rsi or 50)>60 else ("#ef5350" if (_rsi or 50)<40 else "#FFD700")
            _vr = r.get("vol_ratio",1)
            _vr_col = "#00e676" if _vr>=1.5 else ("#FFD700" if _vr>=1.0 else "#5a7fa0")

            def _sma_cell(vs_str):
                col = "#00e676" if vs_str=="▲" else ("#ef5350" if vs_str=="▼" else "#555")
                return f'<span style="color:{col};font-weight:700">{vs_str}</span>'

            def _cloud_cell(state):
                _c, _icon = _cc(state)
                return f'<span style="background:{_c}22;color:{_c};font-size:10px;font-weight:800;padding:1px 6px;border-radius:4px">{_icon} {state}</span>'

            _ac = {"FULL BULL":"#00e676","MOSTLY BULL":"#00e676","MIXED":"#8ab4d4",
                   "MOSTLY BEAR":"#ef9a9a","FULL BEAR":"#ef5350"}.get(r["alignment"],"#8ab4d4")
            _rk_col = "#FFD700" if _row_rank.startswith("#1") else ("#7eb8f7" if _is_m7 else "#e8edf5")

            w(f'<tr style="border-bottom:1px solid #0d2040;{bg}">')
            w(f'<td style="padding:6px 8px;font-weight:900;color:{_rk_col};text-align:center">{_row_rank}</td>')
            w(f'<td style="padding:6px 8px;font-weight:900;color:#e8edf5;white-space:nowrap">'
              f'<a href="{_tv}" target="_blank" style="color:#e8edf5;text-decoration:none">{r["tk"]}</a></td>')
            w(f'<td style="padding:6px 8px;color:#e8edf5;font-weight:700;text-align:right">${r["px"]:.2f}</td>')
            w(f'<td style="padding:6px 8px;text-align:center">'
              f'<span style="background:{_ac}22;color:{_ac};font-size:10px;font-weight:900;padding:2px 7px;border-radius:4px">'
              f'{r["alignment"]}</span></td>')
            w(f'<td style="padding:6px 8px;text-align:center">'
              f'<span style="color:{_bias_col};font-weight:800;font-size:10px">{r.get("bias","—")}</span></td>')
            w(f'<td style="padding:6px 8px;text-align:center;color:{_conv_col};font-weight:900;font-size:13px">{_conv}</td>')
            for _ck in ["c1_89","c2_513","c3_3450","c4_week"]:
                w(f'<td style="padding:6px 8px;text-align:center">{_cloud_cell(r.get(_ck,"—"))}</td>')
            w(f'<td style="padding:6px 8px;text-align:center;color:{_ts_col};font-weight:800">{_ts}%</td>')
            for _vs_key in ["vs20","vs50","vs100","vs200"]:
                w(f'<td style="padding:6px 8px;text-align:center">{_sma_cell(r.get(_vs_key,"—"))}</td>')
            w(f'<td style="padding:6px 8px;text-align:center;color:{_rsi_col};font-weight:700">'
              f'{f"{_rsi:.0f}" if _rsi else "—"}</td>')
            # RSI Divergence
            _rdiv     = r.get("rsi_div","NONE")
            _rdiv_col = r.get("rsi_div_col","#8ab4d4")
            _rdiv_lbl = {"BULL DIV":"🔺 BULL","BEAR DIV":"🔻 BEAR",
                         "HIDDEN BULL":"↗ H.BULL","HIDDEN BEAR":"↘ H.BEAR","NONE":"—"}.get(_rdiv,"—")
            w(f'<td style="padding:6px 8px;text-align:center">'
              f'<span style="color:{_rdiv_col};font-size:10px;font-weight:800">{_rdiv_lbl}</span></td>')
            # ATR Squeeze
            _sqz     = r.get("atr_squeeze", False)
            _sqz_pct = r.get("atr_squeeze_pct")
            _sqz_lbl = f'🔥 {_sqz_pct:.0f}%' if _sqz else ("—" if _sqz_pct is None else f'{_sqz_pct:.0f}%')
            _sqz_col = "#FFD700" if _sqz else "#5a7fa0"
            w(f'<td style="padding:6px 8px;text-align:center;color:{_sqz_col};font-size:10px;font-weight:700">{_sqz_lbl}</td>')
            # OBV trend
            _obv_t    = r.get("obv_trend","FLAT")
            _obv_col2 = r.get("obv_trend_col","#8ab4d4")
            _obv_icon = {"RISING":"▲","FALLING":"▼","FLAT":"➖"}.get(_obv_t,"➖")
            w(f'<td style="padding:6px 8px;text-align:center">'
              f'<span style="color:{_obv_col2};font-size:10px;font-weight:800">{_obv_icon} {_obv_t}</span></td>')
            # Cloud compression / coil
            _coil     = r.get("cloud_compression", False)
            _coil_lbl = r.get("compression_label","")
            _coil_col = r.get("compression_col","#8ab4d4")
            w(f'<td style="padding:6px 8px;text-align:center">'
              f'<span style="color:{_coil_col};font-size:10px;font-weight:800">{"🌀 " + _coil_lbl if _coil else "—"}</span></td>')
            # Vol ratio
            w(f'<td style="padding:6px 8px;text-align:center;color:{_vr_col}">{_vr:.1f}×</td>')
            # 52wk hi
            _hi_col = "#ef5350" if r.get("pct_from_hi",0)<-20 else ("#FFD700" if r.get("pct_from_hi",0)<-10 else "#00e676")
            w(f'<td style="padding:6px 8px;text-align:center;color:{_hi_col}">{r.get("pct_from_hi",0):+.1f}%</td>')
            # Chart
            w(f'<td style="padding:6px 8px;text-align:center">'
              f'<a href="{_tv}" target="_blank" style="color:#42a5f5;font-size:11px">↗</a></td>')
            w('</tr>')

    w('</tbody></table></div>')
    w('<div style="font-size:11px;color:#5a7fa0;margin-top:10px;line-height:1.8">'
      '<strong style="color:#e8edf5">Score components (total 100):</strong> '
      'Clouds 30 · Trend 15 · RSI 10 · '
      '<strong style="color:#FFD700">RSI Div 20</strong> · '
      '<strong style="color:#FFD700">ATR Squeeze 10</strong> · '
      '<strong style="color:#FFD700">OBV 15</strong> · '
      'SMA Stack 10 · Vol 10 · Coil 10. '
      'Yellow = leading indicators (fire before price moves). '
      '🔺 BULL DIV = price lower low / RSI higher low (exhaustion). '
      '🔥 ATR Squeeze = volatility contracted — explosive move imminent. '
      '▲ OBV RISING = institutions accumulating beneath the surface. '
      '🌀 Coil = fast cloud compressing into slow cloud — energy building.</div>')
    w('</div>')
    return "\n".join(lines)




def render_mag7_options_tab() -> str:
    """
    Mag 7 Options — Covered Call Income Tab.
    META + MSFT May 15 expiry. Full strike ladder from deep ITM to OTM.
    Premiums calculated via Black-Scholes using real IV data.
    Interactive P&L table for every strike.
    """
    lines = []
    def w(*parts): lines.append("".join(str(p) for p in parts))
    import math, datetime as _dt

    today    = _dt.date.today()
    exp_date = _dt.date(2026, 5, 15)
    dte      = (exp_date - today).days
    T        = dte / 365.0
    rf       = 0.045   # risk-free rate — named rf to avoid collision with loop var r

    def _N(x):
        return 0.5 * (1 + math.erf(x / math.sqrt(2)))

    def bs_call(S, K, iv):
        if T <= 0: return max(S - K, 0)
        d1 = (math.log(S/K) + (rf + 0.5*iv**2)*T) / (iv*math.sqrt(T))
        d2 = d1 - iv*math.sqrt(T)
        return round(S*_N(d1) - K*math.exp(-rf*T)*_N(d2), 2)

    def bs_delta(S, K, iv):
        if T <= 0: return 1.0 if S >= K else 0.0
        d1 = (math.log(S/K) + (rf + 0.5*iv**2)*T) / (iv*math.sqrt(T))
        return round(_N(d1), 3)

    # Live data as of Apr 19 2026 — updated each run via constants
    # IV from optionsamurai / alphaquery: MSFT 36.4% (96th pct), META ~38%
    STOCKS = [
        {
            "tk": "META", "name": "Meta Platforms", "col": "#1877F2",
            "px": 686.97, "iv": 0.38,
            "earn": _dt.date(2026,4,29), "earn_lbl": "Apr 29",
            "hi52": 796.25, "lo52": 479.80,
            "strikes": [
                (600, "Deep ITM"), (630, "Deep ITM"), (650, "ITM"),
                (660, "ITM"),      (670, "ITM"),      (680, "Slight ITM"),
                (690, "ATM"),      (700, "OTM"),      (710, "OTM"),
                (720, "OTM"),      (730, "OTM"),
            ],
        },
        {
            "tk": "MSFT", "name": "Microsoft Corp", "col": "#00a4ef",
            "px": 422.79, "iv": 0.364,
            "earn": _dt.date(2026,4,28), "earn_lbl": "Apr 28",
            "hi52": 555.45, "lo52": 355.67,
            "strikes": [
                (370, "Deep ITM"), (385, "Deep ITM"), (395, "ITM"),
                (405, "ITM"),      (410, "ITM"),      (415, "ITM"),
                (420, "Slight ITM"),(425, "ATM"),     (430, "OTM"),
                (440, "OTM"),      (450, "OTM"),
            ],
        },
    ]

    # Build option rows for each stock
    def build_rows(st):
        S, iv = st["px"], st["iv"]
        rows = []
        for K, typ in st["strikes"]:
            mid   = bs_call(S, K, iv)
            bid   = round(mid * 0.93, 2)
            yld   = round(mid / S * 100, 2)
            ann   = round(yld * 365 / dte, 1)
            be    = round(S - mid, 2)
            eff   = round(K + mid, 2)
            mp    = round((eff - S) * 100, 0)
            delta = bs_delta(S, K, iv)
            pct   = round((K - S) / S * 100, 1)
            itm   = K < S
            rows.append({
                "K": K, "typ": typ, "mid": mid, "bid": bid,
                "yld": yld, "ann": ann, "be": be, "eff": eff,
                "mp": mp, "delta": delta, "pct": pct, "itm": itm,
            })
        return rows

    # ── Page output ────────────────────────────────────────────────────────────
    w('<div class="card">')
    w(f'<h3>📞 Mag 7 Covered Calls — Income Strategy '
      f'<span class="fresh">May 15 · {dte}d DTE · {today.strftime("%b %d, %Y")}</span></h3>')

    # Criteria panel
    w('<div style="background:#0a1628;border-radius:10px;padding:14px 18px;margin-bottom:20px">')
    w('<div style="font-size:13px;font-weight:900;color:#FFD700;margin-bottom:10px">📋 How to read this tab</div>')
    w('<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:12px;font-size:11px;line-height:1.7">')
    w('<div style="background:#0d1b2a;border-radius:7px;padding:9px 11px">'
      '<div style="color:#00e676;font-weight:800;margin-bottom:3px">Deep ITM calls</div>'
      '<div style="color:#8ab4d4">Strike well below current price. Premium is mostly intrinsic value. '
      'High yield, maximum downside cushion, but you cap upside very close to today\'s price. '
      'Best if neutral-to-bearish short term. Rarely get called away at a loss if held to expiry.</div></div>')
    w('<div style="background:#0d1b2a;border-radius:7px;padding:9px 11px">'
      '<div style="color:#FFD700;font-weight:800;margin-bottom:3px">Slight ITM / ATM calls</div>'
      '<div style="color:#8ab4d4">Strike near current price. Mix of intrinsic + time value. '
      '1-4% yield for the month. Good balance: meaningful income, modest upside cap. '
      'Sweet spot for a monthly income strategy on quality stocks you want to hold.</div></div>')
    w('<div style="background:#0d1b2a;border-radius:7px;padding:9px 11px">'
      '<div style="color:#42a5f5;font-weight:800;margin-bottom:3px">OTM calls</div>'
      '<div style="color:#8ab4d4">Strike above current price. Pure time value — no intrinsic. '
      'Lower yield (1-3%) but stock keeps full upside up to the strike. '
      'Best if bullish but want some income. Stock must rally past strike to get called away.</div></div>')
    w('<div style="background:#0d1b2a;border-radius:7px;padding:9px 11px">'
      '<div style="color:#e8edf5;font-weight:800;margin-bottom:3px">Premiums (Black-Scholes)</div>'
      f'<div style="color:#8ab4d4">Calculated using real IV data: MSFT {0.364*100:.0f}% IV (96th pct — very elevated), '
      f'META {0.38*100:.0f}% IV (pre-earnings elevated). '
      'Mid = theoretical fair value. Bid is what you\'ll realistically fill at when selling. '
      'Always verify live chain in your broker before placing.</div></div>')
    w('<div style="background:#0d1b2a;border-radius:7px;padding:9px 11px">'
      '<div style="color:#ef9a9a;font-weight:800;margin-bottom:3px">⚠ Earnings risk</div>'
      '<div style="color:#8ab4d4">Both earnings are before May 15 expiry. '
      'META: Apr 29 · MSFT: Apr 28. '
      'A strong beat can gap the stock above your strike → you miss upside. '
      'A miss can gap it below → losses mount. Premiums are fat partly because '
      'the market is pricing in this binary risk.</div></div>')
    w('<div style="background:#0d1b2a;border-radius:7px;padding:9px 11px">'
      '<div style="color:#e8edf5;font-weight:800;margin-bottom:3px">Key columns</div>'
      '<div style="color:#8ab4d4">'
      '<strong style="color:#e8edf5">Yield%</strong> = premium ÷ stock price<br>'
      '<strong style="color:#e8edf5">Breakeven</strong> = stock price − premium (your floor)<br>'
      '<strong style="color:#e8edf5">Eff. Sale</strong> = strike + premium (if called away)<br>'
      '<strong style="color:#e8edf5">Max P&L</strong> = (eff. sale − purchase price) × 100<br>'
      '<strong style="color:#e8edf5">Delta</strong> = probability call is ITM at expiry</div></div>')
    w('</div></div>')

    # ── Per-stock section ──────────────────────────────────────────────────────
    for st in STOCKS:
        tk    = st["tk"]
        S     = st["px"]
        col   = st["col"]
        earn  = st["earn"]
        days_earn = (earn - today).days
        rows  = build_rows(st)

        w(f'<div style="background:#0d1b2a;border-left:4px solid {col};border-radius:12px;'
          f'padding:18px 22px;margin-bottom:24px">')

        # Header row
        w(f'<div style="display:flex;align-items:center;gap:14px;flex-wrap:wrap;margin-bottom:4px">')
        w(f'<span style="font-size:26px;font-weight:900;color:{col}">{tk}</span>')
        w(f'<span style="font-size:14px;color:#8ab4d4">{st["name"]}</span>')
        w(f'<span style="font-size:22px;font-weight:900;color:#fff">${S:,.2f}</span>')
        w(f'<span style="font-size:12px;color:#8ab4d4">52wk ${st["lo52"]:.0f}–${st["hi52"]:.0f}</span>')
        w(f'<span style="font-size:12px;color:#ef9a9a;background:#1a0505;padding:3px 10px;border-radius:6px">'
          f'⚠ Earnings {earn.strftime("%b %d")} ({days_earn}d) — before May 15</span>')
        w(f'<span style="font-size:12px;color:#FFD700;background:#1a1400;padding:3px 10px;border-radius:6px">'
          f'IV {st["iv"]*100:.0f}% — elevated pre-earnings</span>')
        w(f'<a href="https://www.tradingview.com/chart/?symbol={tk}&interval=D" target="_blank" '
          f'style="color:#42a5f5;font-size:12px;margin-left:auto">↗ Chart</a>')
        w('</div>')
        w(f'<div style="font-size:11px;color:#5a7fa0;margin-bottom:14px">'
          f'May 15 expiry · {dte} DTE · Black-Scholes mid at {st["iv"]*100:.0f}% IV · '
          f'Fill at bid (shown) when selling — mid is theoretical fair value</div>')

        # Strike comparison table
        w('<div style="overflow-x:auto">')
        w('<table style="width:100%;border-collapse:collapse;font-size:11px;min-width:900px">')
        w('<thead>')
        # Group header
        w('<tr style="background:#08111e">')
        w(f'<td colspan="4" style="padding:5px 10px;color:#5a7fa0;font-size:10px;border-bottom:0.5px solid #1e3a5f">STRIKE</td>')
        w(f'<td colspan="3" style="padding:5px 10px;color:#00e676;font-size:10px;border-bottom:0.5px solid #1e3a5f">PREMIUM COLLECTED</td>')
        w(f'<td colspan="4" style="padding:5px 10px;color:#FFD700;font-size:10px;border-bottom:0.5px solid #1e3a5f">POSITION ECONOMICS</td>')
        w(f'<td colspan="1" style="padding:5px 10px;color:#8ab4d4;font-size:10px;border-bottom:0.5px solid #1e3a5f">GREEK</td>')
        w('</tr>')
        # Column headers
        w('<tr style="background:#0a1220;border-bottom:1px solid #1e3a5f">')
        for h in ["Strike","Type","vs Px","Rec?",
                  "Mid/sh","Bid/sh","Per Lot","Yield%","Ann%","Breakeven","Eff.Sale","Max P&L","Delta"]:
            w(f'<th style="padding:7px 9px;color:#7eb8f7;font-weight:700;text-align:right;white-space:nowrap">{h}</th>')
        w('</tr></thead><tbody>')

        for row in rows:
            # Highlight best 1-2% ITM zone
            is_target = 0.8 <= row["yld"] <= 2.5 and row["itm"]
            is_sweet  = 2.5 < row["yld"] <= 5.0 and row["itm"]
            row_bg    = ("background:#0a2010;" if is_target
                         else "background:#051508;" if is_sweet else "")
            # Rec column
            if row["yld"] >= 4.0 and row["itm"]:
                rec = "★★★"; rec_col = "#00e676"
            elif row["yld"] >= 2.0 and row["itm"]:
                rec = "★★"; rec_col = "#00e676"
            elif 1.0 <= row["yld"] < 2.0:
                rec = "★"; rec_col = "#FFD700"
            elif row["yld"] >= 1.0 and not row["itm"]:
                rec = "OTM✓"; rec_col = "#42a5f5"
            else:
                rec = "—"; rec_col = "#5a7fa0"

            yc  = "#00e676" if row["yld"] >= 3 else ("#FFD700" if row["yld"] >= 1.5 else "#8ab4d4")
            tkc = "#ef9a9a" if row["itm"] else "#42a5f5"
            w(f'<tr style="border-bottom:0.5px solid #0d2040;{row_bg}">')
            w(f'<td style="padding:7px 9px;color:{col};font-weight:900;text-align:right">${row["K"]:,}</td>')
            w(f'<td style="padding:7px 9px;color:{tkc};font-size:10px;text-align:right">{row["typ"]}</td>')
            pct_col = "#ef9a9a" if row["pct"] < 0 else "#42a5f5"
            w(f'<td style="padding:7px 9px;color:{pct_col};text-align:right">{row["pct"]:+.1f}%</td>')
            w(f'<td style="padding:7px 9px;color:{rec_col};font-weight:700;text-align:right">{rec}</td>')
            w(f'<td style="padding:7px 9px;color:#00e676;font-weight:700;text-align:right">${row["mid"]:.2f}</td>')
            w(f'<td style="padding:7px 9px;color:#8ab4d4;text-align:right">${row["bid"]:.2f}</td>')
            w(f'<td style="padding:7px 9px;color:#00e676;font-weight:700;text-align:right">${row["mid"]*100:,.0f}</td>')
            w(f'<td style="padding:7px 9px;color:{yc};font-weight:900;text-align:right">{row["yld"]:.2f}%</td>')
            w(f'<td style="padding:7px 9px;color:{yc};text-align:right">{row["ann"]:.0f}%</td>')
            w(f'<td style="padding:7px 9px;color:#FFD700;text-align:right">${row["be"]:,.2f}</td>')
            w(f'<td style="padding:7px 9px;color:#00e676;text-align:right">${row["eff"]:,.2f}</td>')
            mp_col = "#00e676" if row["mp"] > 0 else "#ef5350"
            w(f'<td style="padding:7px 9px;color:{mp_col};font-weight:700;text-align:right">${row["mp"]:,.0f}</td>')
            d_col = "#ef5350" if row["delta"] > 0.7 else ("#FFD700" if row["delta"] > 0.5 else "#42a5f5")
            w(f'<td style="padding:7px 9px;color:{d_col};text-align:right">{row["delta"]:.2f}</td>')
            w('</tr>')

        w('</tbody></table></div>')

        # Legend
        w('<div style="display:flex;gap:12px;flex-wrap:wrap;margin-top:8px;font-size:10px;color:#5a7fa0">')
        w('<span style="background:#0a2010;padding:2px 7px;border-radius:4px;color:#8ab4d4">Dark green = 1–2.5% target zone</span>')
        w('<span style="background:#051508;padding:2px 7px;border-radius:4px;color:#8ab4d4">Green = 2.5–5% high-yield ITM</span>')
        w('<span style="color:#8ab4d4">★★★ = high-conviction pick · ★★ = good · ★ = conservative</span>')
        w('<span style="color:#8ab4d4">OTM✓ = OTM option, keeps upside</span>')
        w('</div>')

        # P&L scenarios for 3 key strikes: 1% ITM, ATM, OTM
        key_strikes = []
        for row in rows:
            if row["itm"] and 0.8 <= row["yld"] <= 1.5 and len(key_strikes) == 0:
                key_strikes.append(("1% ITM target", row))
            if not row["itm"] and row["pct"] <= 1.0 and len(key_strikes) == 1:
                key_strikes.append(("ATM/slight OTM", row))
            if not row["itm"] and row["pct"] >= 3.0 and len(key_strikes) == 2:
                key_strikes.append(("OTM — keep upside", row))
            if len(key_strikes) == 3:
                break

        if key_strikes:
            w('<div style="margin-top:16px">')
            w('<div style="font-size:12px;color:#5a7fa0;margin-bottom:8px">'
              'P&L at May 15 expiry — three strikes compared (100 shares, buy @ market)</div>')
            w('<div style="overflow-x:auto">')
            w('<table style="width:100%;border-collapse:collapse;font-size:11px;min-width:700px">')
            w('<thead><tr style="background:#0a1220;border-bottom:1px solid #1e3a5f">')
            w('<th style="padding:7px 9px;color:#7eb8f7;font-weight:700;text-align:left">META at Expiry</th>')
            for lbl, r in key_strikes:
                w(f'<th colspan="2" style="padding:7px 9px;color:#7eb8f7;font-weight:700;text-align:center;border-left:1px solid #1e3a5f">'
                  f'${r["K"]:,} Call ({lbl})<br>'
                  f'<span style="color:#00e676;font-size:10px">Prem ${r["mid"]:.2f} · Yield {r["yld"]:.1f}%</span></th>')
            w('</tr>')
            w('<tr style="background:#08111e;border-bottom:1px solid #0d2040">')
            w('<td style="padding:5px 9px;color:#5a7fa0;font-size:10px">Scenario</td>')
            for lbl, r in key_strikes:
                w('<td style="padding:5px 9px;color:#5a7fa0;font-size:10px;border-left:1px solid #0d2040">P&L $</td>')
                w('<td style="padding:5px 9px;color:#5a7fa0;font-size:10px">Return %</td>')
            w('</tr></thead><tbody>')

            price_scenarios = [
                (f"−20% crash",      round(S * 0.80)),
                (f"−12%",            round(S * 0.88)),
                (f"−6%",             round(S * 0.94)),
                (f"Flat",            round(S * 1.00)),
                (f"+4% modest beat", round(S * 1.04)),
                (f"+8% strong beat", round(S * 1.08)),
                (f"+12% blow-out",   round(S * 1.12)),
                (f"+20% rip",        round(S * 1.20)),
            ]

            for sc_lbl, price in price_scenarios:
                is_up   = price > S
                sc_col  = "#00e676" if is_up else "#ef5350"
                w(f'<tr style="border-bottom:0.5px solid #0d2040">')
                w(f'<td style="padding:7px 9px;color:{sc_col}">{sc_lbl} (${price:,})</td>')
                for lbl, r in key_strikes:
                    K    = r["K"]
                    prem = r["mid"]
                    # Stock P&L
                    spnl = (price - S) * 100
                    # Option P&L (short call)
                    if price >= K:
                        opnl = (prem - (price - K)) * 100
                    else:
                        opnl = prem * 100
                    total = round(spnl + opnl)
                    ror   = total / ((S - prem) * 100) * 100
                    tc    = "#00e676" if total >= 0 else "#ef5350"
                    w(f'<td style="padding:7px 9px;color:{tc};font-weight:700;text-align:right;border-left:1px solid #0d2040">'
                      f'{"+" if total>=0 else ""}${total:,}</td>')
                    w(f'<td style="padding:7px 9px;color:{tc};text-align:right">{ror:+.1f}%</td>')
                w('</tr>')

            w('</tbody></table></div>')

            # Key insight callout
            atm_r = key_strikes[1][1] if len(key_strikes) > 1 else key_strikes[0][1]
            drop_buf = round((S - atm_r["be"]) / S * 100, 1)
            w(f'<div style="background:#0a1220;border-radius:8px;padding:10px 14px;margin-top:10px;'
              f'font-size:11px;color:#8ab4d4;line-height:1.7">')
            w(f'<strong style="color:{col}">{tk} sweet spot:</strong> '
              f'The <strong style="color:#FFD700">${atm_r["K"]:,} call</strong> '
              f'({atm_r["typ"]}, {atm_r["yld"]:.1f}% yield) gives you '
              f'<strong style="color:#00e676">${atm_r["mid"]*100:,.0f}</strong> income per contract, '
              f'a breakeven of <strong>${atm_r["be"]:,.2f}</strong> '
              f'({drop_buf:.1f}% downside cushion before losing money), '
              f'and a max profit of <strong style="color:#00e676">${atm_r["mp"]:,.0f}</strong> '
              f'if {tk} closes at or above ${atm_r["K"]:,} on May 15. '
              f'Earnings <strong style="color:#ef9a9a">{st["earn_lbl"]}</strong> is the main event — '
              f'a beat could push {tk} above the strike (you\'re capped but still profitable), '
              f'a miss could push through your breakeven (premium softens but doesn\'t eliminate the loss).')
            w('</div></div>')

        w('</div>')  # end stock card

    # Footer
    w('<div style="font-size:11px;color:#5a7fa0;margin-top:4px;padding:0 4px;line-height:1.7">'
      'Premiums calculated via Black-Scholes. MSFT IV 36.4% (96th percentile as of Apr 18), '
      'META IV 38% (pre-earnings elevated). Bid ≈ mid × 0.93 — fills typically between bid and mid. '
      'Always verify the live chain in your broker before placing orders. '
      'Not financial advice — model your own position size and risk tolerance.</div>')
    w('</div>')
    return "\n".join(lines)


def render_kyles_holdings_tab() -> str:
    """
    Kyle's Holdings — live P&L, covered call status, technical analysis,
    and actionable roll/close/hold recommendations for each position.
    """
    lines = []
    def w(*parts): lines.append("".join(str(p) for p in parts))
    import datetime as _dt
    import yfinance as _yf

    today    = _dt.date.today()
    now_str  = today.strftime("%b %d, %Y")

    w('<div class="card">')
    w(f'<h3>💼 Kyle\'s Holdings <span class="fresh">{now_str} — live analysis</span></h3>')

    # Criteria / legend panel
    w('<div style="background:#0a1628;border-radius:10px;padding:14px 18px;margin-bottom:20px;font-size:11px;line-height:1.8">')
    w('<div style="font-size:13px;font-weight:900;color:#FFD700;margin-bottom:10px">📋 How to Read This Tab</div>')
    w('<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:10px">')
    for _lbl, _col, _desc in [
        ("50% Rule",        "#00e676", "When the short call has decayed to 50% of the original premium — buy it back and re-sell. Risk/reward flips at this point."),
        ("Roll Down & Out", "#42a5f5", "Buy back current call, sell a new call at a lower strike / further expiry for a net credit. Reduces cap, collects more premium."),
        ("Roll Up & Out",   "#FFD700", "Buy back current call, sell a new call at a higher strike / further expiry. Gives stock more room to run after a rally."),
        ("Let Expire",      "#8ab4d4", "If call is near zero and expiry is close — let it expire worthless, then sell fresh next cycle."),
    ]:
        w(f'<div style="background:#0d1b2a;border-radius:7px;padding:8px 10px">'
          f'<div style="color:{_col};font-weight:800;margin-bottom:3px">{_lbl}</div>'
          f'<div style="color:#8ab4d4">{_desc}</div></div>')
    w('</div></div>')

    # ── Per holding ────────────────────────────────────────────────────────────
    for h in KYLE_HOLDINGS:
        tk     = h["ticker"]
        shares = h["shares"]
        avg    = h["avg_cost"]
        col    = h["col"]
        calls  = h["calls"]

        # Fetch live data
        px         = avg   # fallback
        sma20      = None; sma50 = None; sma200 = None
        rsi        = None; vol_ratio = None
        pct_from_hi= None; wk52_hi = None; wk52_lo = None
        trend_lbl  = "—"
        try:
            _t    = _yf.Ticker(tk)
            _hist = _t.history(period="1y")
            if not _hist.empty:
                closes  = list(_hist["Close"])
                volumes = list(_hist["Volume"])
                px      = round(float(closes[-1]), 2)
                wk52_hi = round(max(closes[-252:] if len(closes)>=252 else closes), 2)
                wk52_lo = round(min(closes[-252:] if len(closes)>=252 else closes), 2)
                pct_from_hi = round((px - wk52_hi) / wk52_hi * 100, 1)
                def _sma(n):
                    if len(closes) < n: return None
                    return round(sum(closes[-n:]) / n, 2)
                sma20  = _sma(20); sma50 = _sma(50); sma200 = _sma(200)
                # RSI 14
                if len(closes) >= 15:
                    gains = [max(closes[-i]-closes[-i-1],0) for i in range(1,15)]
                    losses= [max(closes[-i-1]-closes[-i],0) for i in range(1,15)]
                    ag = sum(gains)/14; al = sum(losses)/14 or 0.0001
                    rsi = round(100 - 100/(1+ag/al), 1)
                # Vol ratio
                if len(volumes) >= 21:
                    vol_ratio = round(volumes[-1] / (sum(volumes[-21:-1])/20), 2)
                # Trend label
                _mas_above = sum(1 for m in [sma20,sma50,sma200] if m and px > m)
                trend_lbl  = ["BEARISH","WEAK","MIXED","BULLISH"][_mas_above]
        except Exception:
            pass

        stock_pnl       = round((px - avg) * shares, 0)
        stock_pnl_pct   = round((px - avg) / avg * 100, 2)
        stock_val       = round(px * shares, 0)
        cost_basis_total= round(avg * shares, 0)
        pnl_col = "#00e676" if stock_pnl >= 0 else "#ef5350"

        w(f'<div style="background:#0d1b2a;border-left:4px solid {col};border-radius:12px;padding:18px 22px;margin-bottom:20px">')

        # ── Header ─────────────────────────────────────────────────────────────
        w(f'<div style="display:flex;align-items:center;gap:14px;flex-wrap:wrap;margin-bottom:14px">')
        w(f'<span style="font-size:28px;font-weight:900;color:{col}">{tk}</span>')
        w(f'<span style="font-size:22px;font-weight:900;color:#fff">${px:,.2f}</span>')
        _px_vs = ("▲" if px >= avg else "▼")
        _pxc   = "#00e676" if px >= avg else "#ef5350"
        w(f'<span style="font-size:14px;color:{_pxc}">{_px_vs} {stock_pnl_pct:+.2f}% vs avg</span>')
        w(f'<span style="font-size:13px;color:#8ab4d4">{shares} shares · avg ${avg:,.2f}</span>')
        w(f'<a href="https://www.tradingview.com/chart/?symbol={tk}&interval=D" target="_blank" '
          f'style="color:#42a5f5;font-size:12px;margin-left:auto">↗ Chart</a>')
        w('</div>')

        # ── P&L summary cards ──────────────────────────────────────────────────
        # Total premium across all calls
        total_prem_collected = sum(c["premium_collected"] * c["contracts"] * 100 for c in calls)
        total_prem_current   = sum(c["current_price"]     * c["contracts"] * 100 for c in calls)
        total_prem_pnl       = round(total_prem_collected - total_prem_current, 0)
        eff_cost             = round(avg - sum(c["premium_collected"] for c in calls), 2)
        combined_pnl         = round(stock_pnl + total_prem_pnl, 0)
        combined_col         = "#00e676" if combined_pnl >= 0 else "#ef5350"

        w('<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:8px;margin-bottom:16px">')
        for _sl, _sv, _sc in [
            ("Stock value",       f"${stock_val:,.0f}",                    "#e8edf5"),
            ("Stock P&L",         f"${stock_pnl:+,.0f} ({stock_pnl_pct:+.1f}%)", pnl_col),
            ("Premium collected", f"${total_prem_collected:,.0f}",         "#00e676"),
            ("Premium P&L",       f"${total_prem_pnl:+,.0f}",             "#00e676" if total_prem_pnl>=0 else "#ef5350"),
            ("Combined P&L",      f"${combined_pnl:+,.0f}",               combined_col),
            ("Eff. cost basis",   f"${eff_cost:,.2f}/sh",                 "#FFD700"),
            ("52wk High",         f"${wk52_hi:,.2f}" if wk52_hi else "—", "#8ab4d4"),
            ("vs 52wk High",      f"{pct_from_hi:+.1f}%" if pct_from_hi else "—",
             "#ef5350" if (pct_from_hi or 0)<-15 else "#FFD700" if (pct_from_hi or 0)<-5 else "#00e676"),
        ]:
            w(f'<div style="background:#0a1220;border-radius:7px;padding:9px 12px">')
            w(f'<div style="font-size:10px;color:#5a7fa0;margin-bottom:2px">{_sl}</div>')
            w(f'<div style="font-size:16px;font-weight:900;color:{_sc}">{_sv}</div></div>')
        w('</div>')

        # ── Technical analysis ─────────────────────────────────────────────────
        w('<div style="background:#0a1220;border-radius:8px;padding:12px 14px;margin-bottom:14px">')
        w('<div style="font-size:12px;font-weight:900;color:#FFD700;margin-bottom:8px">📊 Technical Analysis</div>')
        w('<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(110px,1fr));gap:7px">')
        def _sma_cell(label, sma_val):
            if not sma_val: return f'<div style="background:#0d1b2a;border-radius:6px;padding:7px 9px"><div style="font-size:10px;color:#5a7fa0">{label}</div><div style="color:#555">—</div></div>'
            above = px >= sma_val
            c2 = "#00e676" if above else "#ef5350"
            return (f'<div style="background:#0d1b2a;border-radius:6px;padding:7px 9px">'
                    f'<div style="font-size:10px;color:#5a7fa0;margin-bottom:2px">{label}</div>'
                    f'<div style="color:{c2};font-weight:700">{"▲" if above else "▼"} ${sma_val:,.2f}</div></div>')
        w(_sma_cell("20 SMA", sma20))
        w(_sma_cell("50 SMA", sma50))
        w(_sma_cell("200 SMA", sma200))
        _rc = "#00e676" if (rsi or 50)>60 else ("#ef5350" if (rsi or 50)<40 else "#FFD700")
        w(f'<div style="background:#0d1b2a;border-radius:6px;padding:7px 9px">'
          f'<div style="font-size:10px;color:#5a7fa0;margin-bottom:2px">RSI(14)</div>'
          f'<div style="color:{_rc};font-weight:700">{rsi:.0f}' if rsi else '<div style="color:#555">—')
        w('</div></div>')
        _vc = "#00e676" if (vol_ratio or 1)>=1.5 else ("#FFD700" if (vol_ratio or 1)>=1.0 else "#5a7fa0")
        w(f'<div style="background:#0d1b2a;border-radius:6px;padding:7px 9px">'
          f'<div style="font-size:10px;color:#5a7fa0;margin-bottom:2px">Vol Ratio</div>'
          f'<div style="color:{_vc};font-weight:700">{vol_ratio:.1f}×' if vol_ratio else '<div style="color:#555">—')
        w('</div></div>')
        _tc2 = {"BULLISH":"#00e676","MIXED":"#FFD700","WEAK":"#ef9a9a","BEARISH":"#ef5350"}.get(trend_lbl,"#8ab4d4")
        w(f'<div style="background:#0d1b2a;border-radius:6px;padding:7px 9px">'
          f'<div style="font-size:10px;color:#5a7fa0;margin-bottom:2px">Trend</div>'
          f'<div style="color:{_tc2};font-weight:800">{trend_lbl}</div></div>')
        w('</div></div>')

        # ── Per-call analysis ──────────────────────────────────────────────────
        for c in calls:
            K          = c["strike"]
            exp_str    = c["expiry"]
            exp_lbl    = c["expiry_label"]
            contracts  = c["contracts"]
            prem_coll  = c["premium_collected"]
            prem_curr  = c["current_price"]
            exp_dt     = _dt.datetime.strptime(exp_str, "%Y-%m-%d").date()
            dte        = (exp_dt - today).days
            call_pnl   = round((prem_coll - prem_curr) * contracts * 100, 0)
            pct_captured = round((prem_coll - prem_curr) / prem_coll * 100, 1) if prem_coll else 0
            breakeven  = round(avg - prem_coll, 2)
            eff_exit   = round(K + prem_coll, 2)
            itm        = px > K
            otm_pct    = round((K - px) / px * 100, 1)
            call_pnl_col = "#00e676" if call_pnl >= 0 else "#ef5350"

            w(f'<div style="background:#08111e;border:1px solid {"#ef5350" if itm else "#1e3a5f"};'
              f'border-radius:10px;padding:14px 16px;margin-bottom:12px">')
            w(f'<div style="display:flex;align-items:center;gap:12px;flex-wrap:wrap;margin-bottom:12px">')
            w(f'<span style="font-size:13px;font-weight:900;color:#e8edf5">'
              f'{contracts}× {tk} ${K:.0f} Call — {exp_lbl}</span>')
            w(f'<span style="font-size:11px;color:#8ab4d4">{dte}d to expiry</span>')
            _itm_badge = (f'<span style="background:#ef535033;color:#ef5350;font-size:11px;font-weight:800;'
                          f'padding:2px 8px;border-radius:4px">⚠ ITM by {abs(otm_pct):.1f}%</span>' if itm else
                          f'<span style="background:#00e67622;color:#00e676;font-size:11px;font-weight:700;'
                          f'padding:2px 8px;border-radius:4px">OTM {otm_pct:.1f}%</span>')
            w(_itm_badge)
            w('</div>')

            # Call metrics
            w('<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(120px,1fr));gap:7px;margin-bottom:12px">')
            for _sl, _sv, _sc in [
                ("Sold at",         f"${prem_coll:.2f}/sh",              "#8ab4d4"),
                ("Current value",   f"${prem_curr:.2f}/sh",              "#e8edf5"),
                ("Call P&L",        f"${call_pnl:+,.0f}",               call_pnl_col),
                ("% Captured",      f"{pct_captured:.1f}%",              "#00e676" if pct_captured>=50 else "#FFD700"),
                ("Breakeven",       f"${breakeven:,.2f}",                "#FFD700"),
                ("Eff. exit",       f"${eff_exit:,.2f}",                 "#00e676"),
                ("DTE",             f"{dte} days",                       "#8ab4d4"),
                ("Total collected", f"${prem_coll*contracts*100:,.0f}",  "#00e676"),
            ]:
                w(f'<div style="background:#0a1220;border-radius:6px;padding:7px 10px">')
                w(f'<div style="font-size:10px;color:#5a7fa0;margin-bottom:2px">{_sl}</div>')
                w(f'<div style="font-size:14px;font-weight:800;color:{_sc}">{_sv}</div></div>')
            w('</div>')

            # ── Recommendation ────────────────────────────────────────────────
            # Logic:
            # 1. >50% captured + >21 DTE → close and re-sell
            # 2. ITM + stock strong → roll up and out
            # 3. <21 DTE + <$0.50 → let expire
            # 4. <21 DTE + ITM → roll out immediately
            # 5. Stock below key SMA + ITM call → roll down and out
            # 6. Default → hold

            if dte <= 7 and prem_curr < 1.0:
                _action = "LET EXPIRE"; _action_col = "#8ab4d4"; _action_icon = "⏳"
                _reason = (f"Only {dte} days left and the call is worth ${prem_curr:.2f} — nearly zero. "
                           f"Let it expire worthless and sell a fresh call next cycle. "
                           f"Not worth paying commissions to close a $0 position.")
            elif dte <= 14 and itm:
                _action = "ROLL OUT NOW"; _action_col = "#ef5350"; _action_icon = "🚨"
                _reason = (f"Call is ITM (stock ${px:.2f} vs strike ${K:.0f}) with only {dte} days left. "
                           f"If you let it expire you get called away at ${K:.0f} + ${prem_coll:.2f} premium = ${eff_exit:.2f} effective exit. "
                           f"Roll to a further expiry at the same or higher strike to avoid assignment and collect more premium.")
            elif pct_captured >= 50 and dte > 21:
                _action = "CLOSE & RE-SELL"; _action_col = "#00e676"; _action_icon = "✅"
                _reason = (f"{pct_captured:.0f}% of premium captured with {dte} days remaining. "
                           f"The 50% rule says close it now — the remaining ${prem_curr:.2f} takes {dte} days to decay "
                           f"but your upside risk stays fully capped. Buy back at ${prem_curr:.2f} and sell a new call "
                           f"at a higher strike or further expiry for fresh premium.")
            elif itm and (sma50 and px > sma50) and dte > 21:
                _action = "ROLL UP & OUT"; _action_col = "#FFD700"; _action_icon = "📈"
                _reason = (f"Stock is above your ${K:.0f} strike (ITM) but technical structure is strong "
                           f"(above 50 SMA at ${sma50:,.2f}). Roll up to a higher strike to give the stock "
                           f"more room to run while still collecting net credit. "
                           f"Target: buy back ${K:.0f} call, sell ${K+10:.0f} or ${K+15:.0f} at further expiry.")
            elif not itm and pct_captured < 30 and dte > 30:
                _action = "HOLD"; _action_col = "#8ab4d4"; _action_icon = "⏸"
                _reason = (f"Only {pct_captured:.0f}% of premium captured with {dte} days left. "
                           f"Time decay is working in your favour — hold and let theta do the work. "
                           f"Re-evaluate when you hit 50% captured or {dte-14} more days pass.")
            elif sma200 and px < sma200 and itm:
                _action = "ROLL DOWN & OUT"; _action_col = "#42a5f5"; _action_icon = "🔄"
                _reason = (f"Stock below 200 SMA (${sma200:,.2f}) — bearish structure. Call is ITM. "
                           f"Roll down to a lower strike at a further expiry to collect more premium "
                           f"and lower your effective cost basis further. "
                           f"This protects you if the stock continues to drift lower.")
            else:
                _action = "HOLD"; _action_col = "#8ab4d4"; _action_icon = "⏸"
                _reason = (f"Position is well structured. {pct_captured:.0f}% of premium captured, "
                           f"{dte} days remaining. Stock is {'above' if not itm else 'below'} your ${K:.0f} strike. "
                           f"Let time decay work. Next decision point: when 50% of premium is captured "
                           f"or {max(dte-14,0)} more days pass.")

            w(f'<div style="background:{"#051a0a" if "HOLD" in _action else "#0a0d1a"};'
              f'border:1px solid {_action_col}44;border-radius:8px;padding:12px 14px">')
            w(f'<div style="display:flex;align-items:center;gap:10px;margin-bottom:8px">')
            w(f'<span style="font-size:13px;font-weight:900;color:{_action_col}">'
              f'{_action_icon} {_action}</span>')
            w(f'<span style="font-size:11px;color:#5a7fa0">{pct_captured:.0f}% captured · '
              f'{dte}d remaining · ${prem_curr:.2f} current value</span>')
            w('</div>')
            w(f'<div style="font-size:11px;color:#8ab4d4;line-height:1.7">{esc(_reason)}</div>')
            w('</div></div>')  # end call card

        # ── P&L scenarios at expiry ───────────────────────────────────────────
        if calls:
            c0     = calls[0]
            K0     = c0["strike"]
            pc0    = c0["premium_collected"]
            nc0    = c0["contracts"]
            exp_l0 = c0["expiry_label"]

            w(f'<div style="margin-top:14px">')
            w(f'<div style="font-size:12px;color:#5a7fa0;margin-bottom:8px">'
              f'P&L at expiry — {exp_l0} · {nc0}× ${K0:.0f} call · {shares} shares</div>')
            w('<div style="overflow-x:auto">')
            w('<table style="width:100%;border-collapse:collapse;font-size:11px">')
            w('<thead><tr style="background:#0a1220;border-bottom:1px solid #1e3a5f">')
            for _h in ["Scenario","Price","Stock P&L","Call P&L","Total P&L","Return"]:
                w(f'<th style="padding:6px 9px;color:#7eb8f7;font-weight:700;text-align:right">{_h}</th>')
            w('</tr></thead><tbody>')

            _scens = [
                ("Crash −20%",      round(px*0.80)),
                ("Drop −12%",       round(px*0.88)),
                ("Drop −6%",        round(px*0.94)),
                (f"Breakeven (${avg-pc0:.0f})", round(avg - pc0)),
                ("Flat",            round(px)),
                (f"Strike ${K0:.0f}", int(K0)),
                ("Rally +6%",       round(px*1.06)),
                ("Rally +12%",      round(px*1.12)),
                ("Rip +20%",        round(px*1.20)),
            ]
            _scens.sort(key=lambda x: x[1])

            for _sl2, _pr2 in _scens:
                _sp2 = round((_pr2 - avg) * shares, 0)
                if _pr2 >= K0:
                    _op2 = round((pc0 - (_pr2 - K0)) * nc0 * 100, 0)
                else:
                    _op2 = round(pc0 * nc0 * 100, 0)
                _tot2 = _sp2 + _op2
                _ror2 = _tot2 / (avg * shares) * 100
                _tc3  = "#00e676" if _tot2 >= 0 else "#ef5350"
                _rb2  = "background:#051a0a;" if _pr2 == int(K0) else (
                        "background:#1a050a;" if _tot2 < 0 else "")
                w(f'<tr style="border-bottom:0.5px solid #0d2040;{_rb2}">')
                w(f'<td style="padding:6px 9px;color:#8ab4d4">{_sl2}</td>')
                w(f'<td style="padding:6px 9px;font-weight:700;text-align:right">${_pr2:,}</td>')
                w(f'<td style="padding:6px 9px;color:{"#00e676" if _sp2>=0 else "#ef5350"};text-align:right">${_sp2:+,.0f}</td>')
                w(f'<td style="padding:6px 9px;color:{"#00e676" if _op2>=0 else "#ef5350"};text-align:right">${_op2:+,.0f}</td>')
                w(f'<td style="padding:6px 9px;color:{_tc3};font-weight:800;text-align:right">${_tot2:+,.0f}</td>')
                w(f'<td style="padding:6px 9px;color:{_tc3};text-align:right">{_ror2:+.1f}%</td>')
                w('</tr>')
            w('</tbody></table></div></div>')

        w('</div>')  # end holding card

    w('<div style="font-size:11px;color:#5a7fa0;margin-top:8px;padding:0 4px;line-height:1.7">'
      'Update KYLE_HOLDINGS at the top of export_report.py whenever you add/close positions or roll calls. '
      'Technical analysis pulls live daily data from yfinance each run.</div>')
    w('</div>')
    return "\n".join(lines)

def publish_to_github(out_file: str = OUT_FILE):
    try:
        print("\n🚀 Publishing to GitHub...")
        subprocess.run(["git", "add", out_file], check=True)
        subprocess.run(["git", "commit", "-m", f"report: {datetime.now().strftime('%Y-%m-%d %I:%M %p')}"], check=True)
        subprocess.run(["git", "push", "--set-upstream", "origin", "main"], check=True)
        print("✅ Live at: https://kyleshope02-bit.github.io/tweet-report/")
    except subprocess.CalledProcessError as e:
        print(f"⚠️  Git publish failed: {e}")
    except FileNotFoundError:
        print("⚠️  git not found — make sure Git is in your PATH")


# ------------------ main ------------------
def main():
    print("📈 Fetching Mag7 prices from yfinance...")
    _mag7_rows = [{"Ticker": t, "Name": n} for t, n in MAG7_TICKERS]
    mag_rows, mag_tickers = build_mag7_rows(_mag7_rows)
    _mag7_html = ""; _mag7_data = {}  # populated when tab renders

    # Signals run across full large-cap universe (>$50B) for U&R patterns
    print("🔍 Building universe for signal detection (>$10B mkt cap)...")
    _universe_rows = [{"Ticker": t, "Name": t} for t in LARGE_CAP_UNIVERSE]
    _watch_rows_sig, _watch_tickers_sig = build_watchlist_rows(_universe_rows)
    # Add index ETFs directly — they bypass the $10B market cap filter
    all_tickers = list(mag_tickers | _watch_tickers_sig | set(INDEX_TICKERS))

    print(f"🔍 Running signal detection across {len(all_tickers)} tickers (incl. indexes)...")
    ohlcv_data = fetch_ohlcv_for_signals([t.lstrip("$") for t in all_tickers])
    signals = run_all_signals(all_tickers)

    print("📅 Running historical signal scan (2026, score ≥90)...")
    hist_signals = run_historical_signals(all_tickers, months_back=3, min_score=70)


    # Convert UTC to Eastern Time (ET = UTC-4 EDT / UTC-5 EST)
    import time as _time
    _utc_now = datetime.utcnow()
    # EST is UTC-5, EDT is UTC-4. Use UTC-4 (EDT) Apr-Nov, UTC-5 (EST) otherwise
    _month = _utc_now.month
    _et_offset = -4 if 3 <= _month <= 11 else -5
    _et_now = _utc_now + timedelta(hours=_et_offset)
    _et_label = "EDT" if _et_offset == -4 else "EST"
    updated = _et_now.strftime(f"%Y-%m-%d %I:%M %p") + f" {_et_label}"
    yf_cache_age = ""
    if os.path.exists(YF_CACHE_FILE):
        age_min = int((time.time() - os.path.getmtime(YF_CACHE_FILE)) / 60)
        yf_cache_age = f"yfinance cache {age_min}m old (TTL {int(YF_CACHE_TTL/3600)}h)"
    else:
        yf_cache_age = "yfinance live fetch"

    bull_count  = len(signals)  # all results are now reclaims only
    meta = (
        f"Updated {updated} • {yf_cache_age} • "
        f"Reclaims: {bull_count} across {len(all_tickers)} tickers"
    )

    page = f"""<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Stock Dashboard</title>
""" + STYLE + """
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
</head>
<body>

<h2>Stock Dashboard</h2>
<div class="muted">""" + esc(meta) + f"""</div>

<div class="tabs">
  <button id="btn-signals" class="tabbtn on"  data-tabbtn="1" onclick="showTab('signals')">📊 Signals ({len(signals)})</button>
  <button id="btn-hist"    class="tabbtn off" data-tabbtn="1" onclick="showTab('hist')">📅 Historical ({len(hist_signals)}) ≥80</button>
  <button id="btn-rip"     class="tabbtn off" data-tabbtn="1" onclick="showTab('rip')">🌊 Ripster Screener</button>
  <button id="btn-m7opt"   class="tabbtn off" data-tabbtn="1" onclick="showTab('m7opt')">📞 Mag 7 Options</button>
  <button id="btn-kyle"    class="tabbtn off" data-tabbtn="1" onclick="showTab('kyle')">💼 Kyle's Holdings</button>
</div>

<div id="tab-signals" data-tab="1">
""" + render_signals_table(signals) + """
</div>

<div id="tab-hist" class="hidden" data-tab="1">
""" + render_historical_signals_table(hist_signals) + """
</div>

<div id="tab-rip" class="hidden" data-tab="1">
""" + render_ripster_tab(all_tickers, ohlcv_data) + """
</div>

<div id="tab-m7opt" class="hidden" data-tab="1">
""" + render_mag7_options_tab() + """
</div>

<div id="tab-kyle" class="hidden" data-tab="1">
""" + render_kyles_holdings_tab() + """
</div>


""" + JS + """

<script>showTab('signals');</script>
</body>
</html>
"""

    Path(OUT_FILE).write_text(page, encoding="utf-8")
    print(f"\n✅ Wrote {OUT_FILE}")
    print("Open it: double-click index.html (or right click → Open with → Edge).")

    publish_to_github()

if __name__ == "__main__":
    main()
